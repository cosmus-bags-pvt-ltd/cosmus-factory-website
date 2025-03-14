from collections import defaultdict
import datetime
import decimal
from email import message
from io import BytesIO
from itertools import chain
from operator import attrgetter, itemgetter
import os
import select
from sys import prefix
import uuid
from django.conf import Settings, settings
from django.core.exceptions import ValidationError , ObjectDoesNotExist
import json
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q, Sum, ProtectedError, Avg, Count,F,Exists, OuterRef, Subquery, DecimalField
from django.db.models.functions import Round
from django.db import DatabaseError, IntegrityError, transaction
from django.contrib.auth.decorators import login_required, user_passes_test
import logging
import urllib.parse
from django.contrib import messages
# from networkx import difference
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.forms import inlineformset_factory, modelformset_factory
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseServerError, JsonResponse
from django.views.decorators.cache import cache_control
from django.db.models.functions import Coalesce
import pandas as pd
from openpyxl.drawing.image import Image  
from PIL import Image as PILImage
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import requests
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.db.models.functions import Cast
from django.db.models import CharField
import base64
from django.contrib.staticfiles import finders
from django.contrib.auth.decorators import permission_required
from django.core.exceptions import PermissionDenied

from .models import (AccountGroup, AccountSubGroup, Color, DeliveryChallanMaster, DeliveryChallanProducts, Fabric_Group_Model,
                    FabricFinishes, Finished_goods_Stock_TransferMaster, Finished_goods_transfer_records, Finished_goods_warehouse, Godown_finished_goods, Godown_raw_material,
                    Item_Creation, Ledger, MainCategory, PProduct_Creation, Picklist_process_in_outward, Picklist_products_list, Picklist_voucher_master, Product,
                    Product2SubCategory, Product_bin_quantity_through_table, Product_warehouse_quantity_through_table,  ProductImage, RawStockTransferMaster, RawStockTrasferRecords, SalesVoucherDeliveryChallan, Salesman_info, StockItem,
                    SubCategory, Unit_Name_Create, account_credit_debit_master_table, bin_for_raw_material, cutting_room, factory_employee,
                    finished_goods_warehouse_racks, finished_goods_warehouse_zone, 
                    finished_product_warehouse_bin, finishedgoodsbinallocation, godown_item_report_for_cutting_room,
                    gst, item_color_shade, item_godown_quantity_through_table,
                    item_purchase_voucher_master, labour_work_in_master, labour_work_in_product_to_item,
                    labour_workin_approval_report, labour_workout_childs, labour_workout_cutting_items,
                    labour_workout_master, ledgerTypes, opening_shade_godown_quantity, outward_product_master, outward_products,
                    packaging, product_2_item_through_table, product_delivery_challan_quantity_through_table, product_godown_quantity_through_table, 
                    product_purchase_voucher_items, product_purchase_voucher_master, product_to_item_labour_child_workout,
                    product_to_item_labour_workout, purchase_order, purchase_order_for_puchase_voucher_rm, 
                    purchase_order_for_raw_material, purchase_order_master_for_puchase_voucher_rm, 
                    purchase_order_raw_material_cutting, 
                    purchase_order_to_product, purchase_order_to_product_cutting, purchase_voucher_items, rack_for_raw_material,
                    raw_material_product_ref_items, raw_material_product_to_items, raw_material_product_wise_qty, raw_material_production_estimation, raw_material_production_total, sales_return_inward, sales_return_product, sales_return_voucher, sales_return_voucher_master, sales_voucher_master_outward_scan, sales_voucher_outward_scan,
                    set_prod_item_part_name, shade_godown_items,
                    shade_godown_items_temporary_table,purchase_order_for_raw_material_cutting_items,sales_voucher_finish_Goods,sales_voucher_master_finish_Goods)


from .forms import( Basepurchase_order_for_raw_material_cutting_items_form, ColorForm, 
                    CustomPProductaddFormSet, Finished_goods_Stock_TransferMaster_form, Outwardproductmasterform, PicklistProcessInOutwardFormsetupdate, Picklistvouchermasterform, ProductCreateSkuFormsetCreate,
                    ProductCreateSkuFormsetUpdate, Purchaseorderforpuchasevoucherrmform, Purchaseordermasterforpuchasevoucherrmform, SalesVoucherDeliveryChallanForm, SalesmaninfoForm, Salesvouchermasteroutwardscanform, SalesvoucheroutwardscanForm, bin_for_raw_material_form, cutting_room_form,
                    factory_employee_form, finished_goods_warehouse_racks_form, finished_goods_warehouse_zone_form, finished_product_warehouse_bin_form, 
                    labour_work_in_product_to_item_approval_formset, labour_work_in_product_to_item_form, labour_workin_master_form, labour_workout_child_form, 
                    labour_workout_cutting_items_form, ledger_types_form, product_purchase_voucher_master_form, purchase_order_for_raw_material_cutting_items_form, 
                    purchase_order_to_product_cutting_form, rack_for_raw_material_form, raw_material_product_estimation_items_form, raw_material_product_estimation_product_2_item_form, 
                    raw_material_production_estimation_form,raw_material_stock_trasfer_items_formset,
                    FabricFinishes_form, ItemFabricGroup, Itemform, LedgerForm,
                    OpeningShadeFormSetupdate, PProductAddForm, PProductCreateForm, ShadeFormSet,
                    StockItemForm, UnitName, account_sub_grp_form, PProductaddFormSet,
                    ProductImagesFormSet, ProductVideoFormSet, purchase_order_form,purchase_voucher_items_godown_formset_shade_change,
                    gst_form, item_purchase_voucher_master_form,
                    packaging_form, product_main_category_form,  Product2ItemFormsetExtraForm,Product2CommonItemFormSetExtraForm,
                    product_sub_category_form, purchase_voucher_items_formset,raw_material_product_estimation_formset_update,
                    purchase_voucher_items_godown_formset, purchase_voucher_items_formset_update, raw_material_stock_trasfer_master_form, sales_return_voucher_form, sales_return_voucher_master_form, salesreturninwardmasterform, salesvoucherfinishGoodsForm,
                    shade_godown_items_temporary_table_formset,shade_godown_items_temporary_table_formset_update,
                    Product2ItemFormset,Product2CommonItemFormSet,purchase_order_product_qty_formset,
                    purchase_order_raw_product_qty_formset,purchase_order_raw_product_qty_cutting_formset,product_purchase_voucher_items_formset_update,
                    purchase_order_cutting_approval_formset,product_purchase_voucher_items_formset,Finished_goods_transfer_records_formset_create,
                    purchase_order_raw_product_sheet_form,purchase_order_raw_material_cutting_form,
                    raw_material_product_estimation_formset, Finished_goods_transfer_records_formset_update,
                    stock_transfer_instance_formset_only_for_update,product_purchase_voucher_items_instance_formset_only_for_update, subcat_and_bin_form,
                    transfer_product_to_bin_formset, purchase_product_to_bin_formset,FinishedProductWarehouseBinFormSet,Purchaseorderforpuchasevoucherrmformset,Purchaseorderforpuchasevoucherrmformsetupdate,sub_category_and_bin_formset,picklistcreateformset,picklistcreateformsetupdate,salesvouchermasterfinishGoodsForm,salesvouchercreateformset,salesvoucherupdateformset,OutwardProductcreateFormSet,OutwardProductupdateFormSet,salesvoucherfromscanupdateformset,PicklistProcessInOutwardFormset,sales_return_product_formset,sales_return_product_formset_update,DeliveryChallanMasterForm,DeliveryChallanProductsCreateFormset,DeliveryChallanProductsUpdateFormset,SalesVoucherDeliveryChallanFormsetupdate)
    


logger = logging.getLogger('product_views')


"""
    logger.debug("This is a debug message")
    logger.info("This is an info message")
    logger.warning("This is a warning message")
    logger.error("This is an error message")
    logger.critical("This is a critical message")
    
"""

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)





def dashboard(request):
    return render(request,'misc/dashboard.html',{"page_name":"Dashboard"})





def edit_production_product(request,pk):
    
    user = request.user

    if not user.has_perm('product.view_product'):
        messages.error(request, "You do not have permission to view product")
        return redirect('pproductlist')
    
    gsts = gst.objects.all()
    pproduct = get_object_or_404(Product, Product_Refrence_ID=pk)
    
    product_skus = pproduct.productdetails.all()
    
    products_sku_counts = PProduct_Creation.objects.filter(Product__Product_Refrence_ID=pk).count()

    prod2cat_instance = Product2SubCategory.objects.filter(Product_id= pproduct.id)
    prod_main_cat_name = ''
    prod_main_cat_id = ''
    prod_sub_cat_dict = {}
    prod_sub_cat_dict_all = {}

    if prod2cat_instance.exists():
        prodmaincat = prod2cat_instance.first()
        
        prod_main_cat_name = prodmaincat.SubCategory_id.product_main_category.product_category_name
        prod_main_cat_id = prodmaincat.SubCategory_id.product_main_category.id

        for subcat in prod2cat_instance:
            prod_sub_cat_dict[subcat.SubCategory_id.id] = subcat.SubCategory_id.product_sub_category_name

        sub_categories = SubCategory.objects.filter(product_main_category = prod_main_cat_id)
        for sub_cat_all in sub_categories:
            prod_sub_cat_dict_all[sub_cat_all.id] = sub_cat_all.product_sub_category_name

    colors = Color.objects.all()
    main_categories = MainCategory.objects.all()

    form = PProductAddForm(instance=pproduct)
    formset = CustomPProductaddFormSet(instance=pproduct)

    if request.method == 'POST':

        if not pproduct and not user.has_perm('product.add_product'):
            messages.error(request, "You do not have permission to add product")
            return redirect('pproductlist')
    
        if pproduct and not user.has_perm('product.change_product'):
            messages.error(request, "You do not have permission to Update product")
            return redirect('pproductlist')
        
        product_ref_id = pk
        
        try:
            excel_file = request.FILES.get('excel_file')

            if excel_file:
                file_name = excel_file.name
                product_ref_id_file = file_name.split('_')[-1].split('.')[0].split(' ')[0]
                if not excel_file.name.endswith('.xlsx'):
                    messages.error(request, 'Invalid file format. Please upload a valid Excel file.')
                    logger.error('Invalid file format. Please upload a valid Excel file.')
                    return redirect('pproductlist')
                
                if int(product_ref_id_file) == product_ref_id:
                    with transaction.atomic():
                        wb = load_workbook(excel_file, data_only = True)
                        ws1 = wb['product_special_configs']
                        ws2 = wb['product_common_configs']

                        grand_total = 0
                        grand_total_combi = 0

                        for row in ws1.iter_rows(min_row=2,min_col=1):
                            id = row[0].value
                            item_name = row[1].value
                            product_sku = row[2].value
                            body_combi = row[7].value
                           
                            if id is not None:
                                if product_sku is not None and item_name is not None:
                                    part_name = row[3].value
                                    part_dimention = row[4].value
                                    dimention_total = row[5].value
                                    part_pieces = row[6].value
                                    body_combi = row[7].value
                                    
                                    if part_name is not None and part_dimention is not None and body_combi is not None:   

                                        if body_combi == 'body':
                                            grand_total = grand_total + float(dimention_total)  

                                        elif body_combi == 'combi':
                                            grand_total_combi = grand_total_combi + float(dimention_total)  

                                        p2i_config_instance = set_prod_item_part_name.objects.get(id=id)

                                        p2i_config_instance.producttoitem.grand_total = grand_total   
                                        p2i_config_instance.producttoitem.grand_total_combi = grand_total_combi   
                                        p2i_config_instance.part_name = part_name
                                        p2i_config_instance.part_dimentions = part_dimention
                                        p2i_config_instance.dimention_total = dimention_total
                                        p2i_config_instance.part_pieces = part_pieces
                                        p2i_config_instance.body_combi = body_combi
                                        p2i_config_instance.c_user = request.user
                                        p2i_config_instance.producttoitem.c_user = request.user
                                        p2i_config_instance.save()   
                                        p2i_config_instance.producttoitem.save()  

                                    else:
                                        p2i_config_instance = set_prod_item_part_name.objects.get(id = id)  
                                        p2i_config_instance.producttoitem.no_of_rows = p2i_config_instance.producttoitem.no_of_rows - 1   
                                        p2i_config_instance.producttoitem.c_user =  request.user
                                        p2i_config_instance.delete()
                                        p2i_config_instance.producttoitem.save()
                            else:
                                grand_total = 0
                                grand_total_combi = 0

                        
                        for product_c in PProduct_Creation.objects.filter(Product__Product_Refrence_ID = pk): 
                            product_sku = product_c.PProduct_SKU

                            row_no = 0  
                            grand_total = 0
                            grand_total_combi = 0

                            for row in ws2.iter_rows(min_row=2,min_col=1):  
                                
                                id = row[0].value
                                item_name = row[1].value
                                part_name = row[2].value
                                part_dimention = row[3].value
                                dimention_total = row[4].value
                                part_pieces = row[5].value
                                body_combi = row[6].value
                                
                                if id is not None and item_name is not None:  
                                    
                                    p2i_instances = product_2_item_through_table.objects.get(PProduct_pk = product_sku, Item_pk__item_name = item_name, common_unique = True)
                                    
                                    p2i_instances_configs = set_prod_item_part_name.objects.filter(producttoitem=p2i_instances).order_by('id')[row_no]

                                    if part_name is not None and body_combi is not None:  

                                        if body_combi == 'body':
                                            grand_total = grand_total + float(dimention_total)   
                                    
                                        elif body_combi == 'combi':
                                            grand_total_combi = grand_total_combi + float(dimention_total)  

                                        p2i_instances_configs.part_name = part_name
                                        p2i_instances_configs.part_dimentions = part_dimention
                                        p2i_instances_configs.part_pieces = part_pieces
                                        p2i_instances_configs.body_combi = body_combi
                                        p2i_instances_configs.dimention_total = dimention_total
                                        p2i_instances_configs.producttoitem.grand_total = grand_total 
                                        p2i_instances_configs.producttoitem.grand_total_combi = grand_total_combi   

                                        p2i_instances_configs.c_user = request.user
                                        p2i_instances_configs.producttoitem.c_user = request.user
                                        p2i_instances_configs.save()
                                        p2i_instances_configs.producttoitem.save() 
                                        row_no = row_no + 1  

                                    else:

                                        p2i_instances_configs.producttoitem.no_of_rows = p2i_instances_configs.producttoitem.no_of_rows - 1
                                        p2i_instances_configs.producttoitem.c_user = request.user
                                        p2i_instances_configs.delete()
                                        p2i_instances_configs.producttoitem.save()
                                        
                                else:
                                    row_no = 0
                                    grand_total = 0
                                    grand_total_combi = 0

                else:
                    messages.error(request, 'File with invalid Product Refrence Id uploaded')
                    logger.error("File with invalid Product Refrence Id uploaded")
                    return redirect('pproductlist')
            
        except Exception as e:
            logger.error(f"An error occured - {str(e)} - Product-Name - {pproduct}")
            messages.error(request, f'Error uploading Excel file: {str(e)}')
            return redirect('pproductlist')
        
        form = PProductAddForm(request.POST, request.FILES, instance = pproduct) 
        formset = CustomPProductaddFormSet(request.POST, request.FILES , instance=pproduct)
        
        formset.forms = [form for form in formset.forms if form.has_changed()]
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    form_instance = form.save(commit=False)
                    form_instance.c_user = request.user
                    
                    for form in formset.deleted_forms:
                        if form.instance.pk:
                            form.instance.delete()

                    for form in formset:
                        if form.is_valid():
                            if not form.cleaned_data.get('DELETE'):
                                formset_instances = form.save(commit=False)
                                formset_instances.c_user = request.user
                                formset_instances.save()

                    
                    p_id_pk = form_instance.pk
                    p_id = Product.objects.get(pk=p_id_pk)
        
                    
                    sub_category_ids = request.POST.getlist('Product_Sub_catagory')
            
                    
                    sub_cat_front_listcomp = [Product2SubCategory.objects.filter(Product_id=p_id,SubCategory_id=sub_cat_id).first() for sub_cat_id in sub_category_ids]
            
                    
                    sub_cat_backend = [x for x in Product2SubCategory.objects.filter(Product_id=p_id)]

                    
                    objects_to_delete = [obj for obj in sub_cat_backend if obj not in sub_cat_front_listcomp]
            
                    for obj in objects_to_delete:
                        obj.delete()

                    
                    for sub_cat_id in sub_category_ids:
                        sub_cat = SubCategory.objects.get(id = sub_cat_id)
                        p2c, created = Product2SubCategory.objects.get_or_create(Product_id=p_id, SubCategory_id=sub_cat)
                        p2c.c_user = request.user
                        p2c.save()

                    
                    form_instance.save()
                    logger.info(f"product-saved product")
                    logger.info(f"product-formset-saved product")
                    return redirect('pproductlist')

            except Exception as e:
                logger.error(f"An exception occured in Product - {e} - Product-name - {pproduct}")
                messages.error(request, f'An exception occured - {e}')
        
        else:

            logger.error(f"Productform not valid - {form.errors} - Product-name - {pproduct}")
            logger.error(f"Product formsets not valid- {formset.errors} - Product-name - {pproduct}")

            return render(request, 'product/edit_production_product.html', {'gsts':gsts,
                                                                            'form':form,
                                                                            'product_skus':product_skus,
                                                                            'formset':formset,
                                                                            'colors':colors,
                                                                            'products_sku_counts':products_sku_counts,
                                                                            'main_categories':main_categories,
                                                                            'prod_main_cat_name':prod_main_cat_name,
                                                                            'prod_main_cat_id':prod_main_cat_id,
                                                                            'prod_sub_cat_dict':prod_sub_cat_dict,
                                                                            'prod_sub_cat_dict_all':prod_sub_cat_dict_all,
                                                                            'page_name':'Edit Production'})


    return render(request, 'product/edit_production_product.html',{'gsts':gsts,
                                                                   'form': form,
                                                                   'product_skus':product_skus,
                                                                   'formset':formset,
                                                                   'colors':colors,
                                                                   'products_sku_counts':products_sku_counts,
                                                                   'main_categories':main_categories,
                                                                   'prod_main_cat_name':prod_main_cat_name,
                                                                    'prod_main_cat_id':prod_main_cat_id,
                                                                    'prod_sub_cat_dict':prod_sub_cat_dict,
                                                                    'prod_sub_cat_dict_all':prod_sub_cat_dict_all,
                                                                    'page_name':'Edit Production'})












def product_color_sku(request,ref_id = None):

    user = request.user

    logger.info(f"product_color_sku function run by {user}")

    if not user.has_perm('product.view_pproduct_creation'):
        messages.error(request, "You do not have permission to view product sku")
        if ref_id:
            return redirect(reverse('edit_production_product', args=[ref_id]))
        else:
            return redirect('pproductlist')
    
    color = Color.objects.all()

    if ref_id:
        instance = Product.objects.get(Product_Refrence_ID=ref_id) 
        ordered_product_details = instance.productdetails.all().order_by('created_date')
        formset = ProductCreateSkuFormsetUpdate(request.POST or None, request.FILES or None, instance=instance, c_user=request.user)
        
    else:
        instance = None
        formset = ProductCreateSkuFormsetCreate(request.POST or None, request.FILES or None, instance=instance, c_user=request.user)

    if request.method == 'POST':

        if not instance and not user.has_perm('product.add_pproduct_creation'):
            messages.error(request, "You do not have permission to add product")
            return redirect('pproductlist')

        if instance and not user.has_perm('product.change_pproduct_creation'):
            messages.error(request, "You do not have permission to Update product")
            return redirect('pproductlist')
        
        product_ref_id = request.POST.get('Product_Refrence_ID', None)
        
        formset.forms = [form for form in formset if form.has_changed()]

        if product_ref_id:
            if formset.is_valid():
                try:
                    for form in formset:
                        if form.is_valid():
                            form_instance = form.save(commit=False)  
                            obj, created =  Product.objects.get_or_create(Product_Refrence_ID=product_ref_id) 
                            obj.c_user = request.user
                            obj.save()
                            form_instance.Product = obj   
                            form_instance.save() 

                    return redirect(reverse('edit_production_product', args=[product_ref_id]))
                        
                except ValidationError as ve :
                    messages.error(request,f'{ve}')
                    logger.error(ve)

                except Exception as e :
                    messages.error(request,f'{e}')
                    logger.error(e)
            else:
                return render(request, 'product/product_color_sku.html', {'formset': formset, 'color': color,'ref_id': ref_id,'page_name':'Add Product'})
            
    return render(request, 'product/product_color_sku.html', {'formset': formset, 'color': color,'ref_id': ref_id,'page_name':'Add Product'})


def pproduct_list(request):

    user = request.user

    logger.info(f"pproduct_list function run by {user}")

    if not user.has_perm('product.view_product'):
        messages.error(request, "You do not have permission to view Products")
        return redirect('dashboard-main')
    
    queryset = Product.objects.all().order_by('Product_Name').select_related('Product_GST').prefetch_related('productdetails','productdetails__PProduct_color')

    product_search = request.GET.get('product_search','')
    
    if product_search != '':
        queryset = Product.objects.filter(
            Q(Product_Name__icontains=product_search)|
            Q(Model_Name__icontains=product_search)|
            Q(Product_Refrence_ID__icontains=product_search)|
            Q(productdetails__PProduct_SKU__icontains=product_search)
            ).distinct()

    # paginator = Paginator(queryset, 10)  
    
    
    # page_number = request.GET.get('page')
    
    # try:
    #     products = paginator.page(page_number)
    # except PageNotAnInteger:
        
    #     products = paginator.page(1)
    # except EmptyPage:
        
    #     products = paginator.page(paginator.num_pages)

    
    # index = products.number - 1
    # max_index = len(paginator.page_range)
    # start_index = index - 2 if index >= 2 else 0
    # end_index = index + 3 if index <= max_index - 3 else max_index
    # page_range = list(paginator.page_range)[start_index:end_index]

    context = {
        'products': queryset,
        # 'page_range': page_range,
        'product_search':product_search,
        'page_name':'Products List'
    }

    return render(request,'product/pproduct_list.html',context=context)


@permission_required('product.delete_product', raise_exception=True)
def pproduct_delete(request, pk):

    user = request.user

    logger.info(f"pproduct_delete function run by {user}")

    try:
        product = get_object_or_404(Product,Product_Refrence_ID=pk)
        product.delete()
        messages.success(request,f'Product {product.Product_Name} was deleted')
        logger.info(f"Product {product.Product_Name} was deleted by {user}")

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {product.Product_Name} because it is referenced by other objects.')
    return redirect('pproductlist')


def add_product_images(request, pk):

    user = request.user

    logger.info(f"add_product_images function run by {user}")

    if not (user.has_perm('product.add_productimage') or user.has_perm('product.change_productimage')):
        return JsonResponse({'error': 'You do not have permission to add or edit images.'})
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    if not user.has_perm('product.view_productimage'):
        messages.error(request, "You do not have permission to view images")
        return redirect('dashboard-main')


    product = PProduct_Creation.objects.get(pk=pk)   
    formset = ProductImagesFormSet(instance=product)  
    
    if request.method == 'POST':

        if product and not user.has_perm('product.change_productimage'):
            messages.error(request, "You do not have permission to update a images.")
            return redirect('dashboard-main')

        if product and not user.has_perm('product.delete_productimage'):
            messages.error(request, "You do not have permission to delete a images.")
            return redirect('dashboard-main')

        if not product and not user.has_perm('product.add_productimage'):
            messages.error(request, "You do not have permission to update a images.")
            return redirect('dashboard-main')


        formset = ProductImagesFormSet(request.POST, request.FILES, instance=product, c_user=request.user)
        if formset.is_valid():
            formset.save()

            if pk:
                messages.success(request,'Product images sucessfully update.')
                logger.info(f"Product images sucessfully update by {user}")
            else:
                messages.success(request,'Product images sucessfully added.')
                logger.info(f"Product images sucessfully added. by {user}")
            
            close_window_script = """
            <script>
            window.opener.location.reload(true);  // Reload parent window if needed
            window.close();  // Close current window
            </script>
            """
            return HttpResponse(close_window_script)
        else:
            return render(request, 'product/add_product_images.html', {'formset': formset, 'product': product})

    return render(request, 'product/add_product_images.html', {'formset': formset, 'product': product})


def add_product_video_url(request,pk):

    user = request.user

    logger.info(f"add_product_video_url function run by {user}")

    if not (user.has_perm('product.add_productvideourls') or user.has_perm('product.change_productvideourls')):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'You do not have permission to add or edit videos.'})
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    if not user.has_perm('product.view_productvideourls'):
        messages.error(request, "You do not have permission to view videos")
        return redirect('dashboard-main')
    
    product = PProduct_Creation.objects.get(pk=pk)   
    formset = ProductVideoFormSet(instance= product)  

    if request.method == 'POST':

        if product and not user.has_perm('product.change_productvideourls'):
            messages.error(request, "You do not have permission to update a videos.")
            return redirect('dashboard-main')

        if product and not user.has_perm('product.delete_productvideourls'):
            messages.error(request, "You do not have permission to delete a videos.")
            return redirect('dashboard-main')

        if not product and not user.has_perm('product.add_productvideourls'):
            messages.error(request, "You do not have permission to update a videos.")
            return redirect('dashboard-main')
        
        formset = ProductVideoFormSet(request.POST, instance=product, c_user=request.user)
        
        if formset.is_valid():
            formset.save()

            if pk:
                messages.success(request,'Product video url sucessfully update.')
                logger.info(f"Product video url sucessfully update by {user}")
            else:
                messages.success(request,'Product video url sucessfully added.')
                logger.info(f"Product video url sucessfully added. by {user}")
            
            close_window_script = """
            <script>
            window.opener.location.reload(true);  // Reload parent window if needed
            window.close();  // Close current window
            </script>
            """
            return HttpResponse(close_window_script)

    else:
            return render(request, 'product/add_product_videourl.html', {'formset': formset, 'product': product})

    return render(request, 'product/add_product_videourl.html', {'formset': formset, 'product': product})



def definemaincategoryproduct(request,pk=None):

    user = request.user

    logger.info(f"definemaincategoryproduct function run by {user}")

    if not user.has_perm('product.view_maincategory'):
        messages.error(request,"You do not have permission to view this page.")
        return redirect('dashboard-main')

    queryset = MainCategory.objects.all()

    main_cat_product_search = request.GET.get('main_cat_product_search','')

    if main_cat_product_search != '':
        queryset = MainCategory.objects.filter(product_category_name__icontains = main_cat_product_search)

    if pk:
        instance = MainCategory.objects.get(pk=pk)
        page_name = 'Edit Main Category'
    else:
        instance = None
        page_name = 'Create Main Category'
    
    form = product_main_category_form(instance=instance)
    
    if request.method == 'POST':

        if instance and not user.has_perm('product.change_maincategory'):
            messages.error(request, "You do not have permission to update a maincategory.")
            return redirect('define-main-category-product')

        if not instance and not user.has_perm('product.add_maincategory'):
            messages.error(request, "You do not have permission to add a maincategory.")
            return redirect('define-main-category-product')

        form = product_main_category_form(request.POST, instance= instance)

        if form.is_valid():
            form_instance  = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'Main Category created sucessfully')
                logger.info(f"Main Category created sucessfully")
            else:
                messages.success(request,'Main Category updated sucessfully')
                logger.info(f'Main Category updated sucessfully')
            return redirect('define-main-category-product')
        else:
            print(form.errors)
        
    return render(request,'product/definemaincategoryproduct.html',{'form':form,'main_cats':queryset,
                                                                    'main_cat_product_search':main_cat_product_search,
                                                                    'page_name':page_name})


@permission_required('product.delete_maincategory', raise_exception=True)
def definemaincategoryproductdelete(request,pk):

    user = request.user

    logger.info(f"definemaincategoryproductdelete function run by {user}")

    try:
        instance = MainCategory.objects.get(pk=pk)
        instance.delete()
        messages.success(request,f'{instance.product_category_name} Main Category Deleted Successfully.')
        logger.info(f"{instance.product_category_name} Main Category Deleted Successfully. by {user}")

    except Exception as e :
        messages.error(request,f'{e}')
    return redirect('define-main-category-product')


def definesubcategoryproduct(request, pk=None):

    user = request.user

    logger.info(f"define sub category product function run by {user}")

    if not user.has_perm('product.view_subcategory'):
        messages.error(request,"You do not have permission to view this page.")
        return redirect('dashboard-main')

    if pk:
        instance = SubCategory.objects.get(pk=pk)
        form = product_sub_category_form(request.POST or None, instance = instance)
        page_name = 'Edit Sub Category'
    else:
        instance = None
        form = product_sub_category_form()
        page_name = 'Add Sub Category'

    main_categories = MainCategory.objects.all()
    # sub_category = SubCategory.objects.all().select_related('product_main_category')

    if request.method == 'POST':
        print(request.POST)

        if instance and not user.has_perm('product.change_subcategory'):
            messages.error(request, "You do not have permission to update a subcategory.")
            return redirect('define-sub-category-product')

        if not instance and not user.has_perm('product.add_subcategory'):
            messages.error(request, "You do not have permission to add a subcategory.")
            return redirect('define-sub-category-product')
        
        try:
            form = product_sub_category_form(request.POST, instance = instance)

            if form.is_valid():
                
                form_instance = form.save(commit=False)

                form_instance.c_user = request.user

                form_instance.save()

                if pk:
                    messages.success(request,'Sub Category created sucessfully')
                    logger.info(f"Sub Category created sucessfully")
                else:
                    messages.success(request,'Sub Category updated sucessfully')
                    logger.info(f'Sub Category updated sucessfully')

                return redirect('define-sub-category-product')
        except Exception as e:
            print(e)
            messages.error(request,f'An Exception occoured - {e}')

    return render(request,'product/definesubcategoryproduct.html',{'main_categories':main_categories, 'form':form,'page_name':page_name,})


@permission_required('product.delete_subcategory', raise_exception=True)
def definesubcategoryproductdelete(request, pk):

    user = request.user

    logger.info(f"definesubcategoryproductdelete function run by {user}")

    try:
        instance = SubCategory.objects.get(pk=pk)
        instance.delete()
        messages.success(request,f'{instance.product_sub_category_name} Sub Category Deleted Successfully.')
        logger.info(f'{instance.product_sub_category_name} Sub Category Deleted Successfully by {user}')

    except Exception as e:
        messages.error(request,f'An Exception occoured - {e}')    
    return redirect('define-sub-category-product')



def assign_bin_to_product_ajax(request):

    formset = None
    rack_id = None
    post_data = None
    main_categories = MainCategory.objects.all()
    zones = finished_goods_warehouse_zone.objects.all()

    main_category = request.POST.get('mainProduct')
    product_main_category_id = request.POST.get('product_main_category')
    zone_id = request.POST.get('zone')
    rack_id = request.POST.get('rack')

    
    racks = finished_goods_warehouse_racks.objects.filter(zone_finished_name=zone_id).exclude(pk=rack_id)


    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        
        zonename = request.GET.get('mainZone')
        # rack_id = request.GET.get('rack')

        if zonename:
            racks = finished_goods_warehouse_racks.objects.filter(zone_finished_name=zonename)
            racks_data = [{'id': rack.id, 'name': rack.rack_name} for rack in racks]
            return JsonResponse({'racks_data': racks_data})

    if rack_id:
        bin_queryset = finished_product_warehouse_bin.objects.filter(rack_finished_name=rack_id).select_related('rack_finished_name','sub_catergory_id')
        formset = sub_category_and_bin_formset(queryset=bin_queryset,form_kwargs={'sub_cat_instance': main_category})
        # print(formset)

    if request.method == "POST":
        post_data = {
                    'product_main_category': MainCategory.objects.get(pk = product_main_category_id),
                    'zone': finished_goods_warehouse_zone.objects.get(pk = zone_id),
                    'rack': finished_goods_warehouse_racks.objects.get(pk = rack_id),
                }
        
    return render(request, 'product/assignbintoproduct.html', {
        'formset': formset,
        'main_categories': main_categories,
        'zones': zones,'post_data': post_data,
        'racks':racks
    })

def save_bin_to_subcategory(request,sub_id):

    print("in savebintosubcategory")
    
    if sub_id:
        instance = MainCategory.objects.get(pk=sub_id)
        title = 'Update'
        message = 'updated'
        page_name = 'Edit Sub Category'
    else:
        instance = None
        title = 'Create'
        message = 'created'
        page_name = 'Create Sub Category'

    if request.method == 'POST':
        try:
            print(request.POST)
            formset = sub_category_and_bin_formset(request.POST, form_kwargs={'sub_cat_instance': instance})
            
            if formset.is_valid():
                
                
                for form in formset:
                    
                    if form.cleaned_data.get('check_if_added_all') != False:
                        
                        if form.cleaned_data.get('check_if_added') == True:
                            formset_instance = form.save(commit=False)
                            formset_instance.sub_catergory_id = instance
                            formset_instance.save()

                        elif form.cleaned_data.get('check_if_added') == False :
                            formset_instance = form.save(commit=False)
                            formset_instance.sub_catergory_id = None
                            formset_instance.save()

                if message == 'created':
                    messages.success(request,'Sub-Category created sucessfully')

                if message == 'updated':
                    messages.success(request,'Sub-Category updated sucessfully')
            
                return redirect('subcategory-bin-list')
            else:
                print(formset.non_form_errors())
                
                logger.error(formset.non_form_errors)
        
        except Exception as e:
            print(formset.non_form_errors())
            print(e)
            messages.error(request,f'An Exception occoured - {e}')

    return render(request, 'product/assignbintoproduct.html')

def update_bin_to_subcategory(request,sub_id,r_id,bin_id=None):
    if sub_id:
        instance = MainCategory.objects.get(pk=sub_id)
        title = 'Update'
        message = 'updated'
        page_name = 'Edit Sub Category'
    else:
        instance = None
        title = 'Create'
        message = 'created'
        page_name = 'Create Sub Category'

    # main_categories = MainCategory.objects.all()
    zones = finished_goods_warehouse_zone.objects.all()
    
    # bins_queryset = finished_product_warehouse_bin.objects.filter(rack_finished_name = r_id)

    bins_queryset = finished_product_warehouse_bin.objects.filter(rack_finished_name=r_id)

    formset = sub_category_and_bin_formset(queryset=bins_queryset, form_kwargs={'sub_cat_instance': instance})

    main_cate = MainCategory.objects.get(pk = sub_id)

    rack = finished_goods_warehouse_racks.objects.get(pk = r_id)
   
    zone = rack.zone_finished_name

    racks = finished_goods_warehouse_racks.objects.filter(zone_finished_name = zone).exclude(pk = r_id)

    

    post_data = {
                'product_main_category': main_cate,
                'rack': rack,
                'zone': zone,
                }

    if request.method == 'POST':
        try:

            formset = sub_category_and_bin_formset(request.POST, form_kwargs={'sub_cat_instance': instance})
            
            if formset.is_valid():
                
                
                for form in formset:
                    
                    if form.cleaned_data.get('check_if_added_all') != False:
                        
                        if form.cleaned_data.get('check_if_added') == True:
                            formset_instance = form.save(commit=False)
                            formset_instance.sub_catergory_id = instance
                            formset_instance.save()

                        elif form.cleaned_data.get('check_if_added') == False :
                            formset_instance = form.save(commit=False)
                            formset_instance.sub_catergory_id = None
                            formset_instance.save()

                if message == 'created':
                    messages.success(request,'Sub-Category created sucessfully')

                if message == 'updated':
                    messages.success(request,'Sub-Category updated sucessfully')
            
                return redirect('subcategory-bin-list')
            
            else:
                print(formset.non_form_errors())
                
                logger.error(formset.non_form_errors)
        
        except Exception as e:
            print(formset.non_form_errors())
            print(e)
            messages.error(request,f'An Exception occoured - {e}')

    return render(request, 'product/assignbintoproduct.html',{'formset':formset,'post_data':post_data,'zones': zones,'editable_bin_id': bin_id,'racks':racks})

def subcategory_bin_list(request):
    
    subcategories = finished_product_warehouse_bin.objects.filter(sub_catergory_id__isnull=False).select_related( 
            'rack_finished_name',
            'rack_finished_name__zone_finished_name',
            'rack_finished_name__zone_finished_name__warehouse_finished_name',
            'sub_catergory_id'
        ).values(
            'id',
            'bin_name',
            'rack_finished_name__rack_name', 
            'rack_finished_name__id',
            'rack_finished_name__zone_finished_name__zone_name',
            'rack_finished_name__zone_finished_name__warehouse_finished_name__warehouse_name_finished',
            'sub_catergory_id__product_category_name',
            'product_size_in_bin',
            'sub_catergory_id__id'
        )

    return render(request, 'product/subcategorybinlist.html',{'subcategories': subcategories})



def product2subcategoryproductajax(request):
    selected_main_cat = request.GET.get('p_main_cat')
    sub_cats = SubCategory.objects.filter(product_main_category = selected_main_cat)
    
    sub_cat_dict = {}

    for sub_cat in sub_cats:
        sub_cat_dict[sub_cat.id] = sub_cat.product_sub_category_name 

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse({'sub_cat_dict':sub_cat_dict})


def product2subcategory(request):
    products = Product.objects.all()
    sub_category = SubCategory.objects.all()
    main_categories = MainCategory.objects.all()

    if request.method == 'POST':

        try:
            
            product_id_get = request.POST.get('product_name')
            
            
            sub_category_ids = request.POST.getlist('sub_category_name')
            
            
            p_id = get_object_or_404(Product, id = product_id_get)

            
            existing_instances =  Product2SubCategory.objects.filter(Product_id=p_id)

            updated_instances_front = []
            
            
            for sub_cat_id in sub_category_ids:
                
                s_c_id =  get_object_or_404(SubCategory, id = sub_cat_id)
                
                p_2_c_instance = Product2SubCategory.objects.filter(Product_id=p_id, SubCategory_id=s_c_id).first() 
                updated_instances_front.append(p_2_c_instance)

            updated_instance_pk = set(obj.pk for obj in updated_instances_front if obj is not None)
            
            instances_to_delete = [obj for obj in existing_instances if obj.pk not in updated_instance_pk]
            
            for obj in instances_to_delete:
                obj.delete()

            for sub_cat_id in sub_category_ids:
 
                s_c_id =  get_object_or_404(SubCategory, id = sub_cat_id)

                p2c, created = Product2SubCategory.objects.get_or_create(Product_id=p_id, SubCategory_id=s_c_id)
                p2c.c_user = request.user
                p2c.save()
            messages.success(request,f'Product sucessfully added to {s_c_id.product_sub_category_name}')
        
        except IntegrityError:
            messages.error(request, 'Product already present in Subcategory')

        except Exception as e:
            messages.error(request,f'An Exception occoured - {e}')

    return render(request,'product/product2subcategory.html',{'main_categories':main_categories,'products':products,'sub_category':sub_category
                                                              })


def product2subcategoryajax(request):

    productid = request.GET.get('selected_product_id')
    categoryforselectedproduct = Product2SubCategory.objects.filter(Product_id = productid)
    
    dict_result = {}

    for obj in categoryforselectedproduct:
        dict_result[obj.SubCategory_id.id] = obj.SubCategory_id.product_sub_category_name
    
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({'dict_result':dict_result})






# ======================== ITEM CREATION ===========================

def item_create(request):

    user = request.user

    logger.info(f"item_create function run by {user}")

    if not user.has_perm('product.add_item_creation'):
        messages.error(request, "You do not have permission to add a item.")
        return redirect('item-list')

    gsts = gst.objects.all()
    fab_grp = Fabric_Group_Model.objects.all()
    unit_name = Unit_Name_Create.objects.all()
    packaging_material_all = packaging.objects.all()
    fab_finishes = FabricFinishes.objects.all()
    items_to_clone = Item_Creation.objects.all()
    colors = Color.objects.all()
    form = Itemform()
    racks = rack_for_raw_material.objects.all()

    if request.path == '/itemcreatepopup/':
        template_name = 'product/item_create_popup.html'
    else:
        template_name = 'product/item_create_update.html'
        
    if request.method == 'POST':

        if not user.has_perm('product.add_item_creation'):
            messages.error(request, "You do not have permission to add a item.")
            return redirect('item-list')

        form = Itemform(request.POST, request.FILES)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()
            
            form_instance.bin.set(form.cleaned_data.get('bin', []))

            if request.path == '/itemcreatepopup/':
                return HttpResponse('item created', status = 200) 
            else:
                logger.info(f"{form_instance.item_name} Item Successfully Created by {user}")
                messages.success(request,'Item has been created')
                return redirect(reverse('item-edit', args=[form_instance.id]))
    
        else:
            logger.error(f"item form not valid{form.errors}")
            messages.error(request, f"item form not valid{form.errors}")
            return render(request,template_name, {'gsts':gsts,'fab_grp':fab_grp,'unit_name':unit_name,'colors':colors,'packaging_material_all':packaging_material_all,'fab_finishes':fab_finishes,'title''items_to_clone':items_to_clone,'page_name':'Create Raw Material'})
    
    
    return render(request,template_name,{'gsts':gsts,'fab_grp':fab_grp,'unit_name':unit_name,'colors':colors,'packaging_material_all':packaging_material_all,'fab_finishes':fab_finishes,'form':form,'items_to_clone':items_to_clone,'page_name':'Create raw material','racks':racks})

def item_edit(request,pk):

    user = request.user

    logger.info(f"item_edit function run by {user}")

    if not user.has_perm('product.change_item_creation'):
        messages.error(request, "You do not have permission to edit a item.")
        return redirect('item-list') 
    
    gsts = gst.objects.all()
    fab_grp = Fabric_Group_Model.objects.all()
    unit_name = Unit_Name_Create.objects.all()
    colors = Color.objects.all()
    packaging_material_all = packaging.objects.all()
    fab_finishes = FabricFinishes.objects.all()
    item_pk = get_object_or_404(Item_Creation ,pk = pk)
    racks = rack_for_raw_material.objects.all()

    form = Itemform(instance=item_pk)
    
    queryset = item_color_shade.objects.filter(items = pk).annotate(
        total_quantity=Sum('opening_shade_godown_quantity__opening_quantity'),
        total_value=Sum(F('opening_shade_godown_quantity__opening_quantity') * F('opening_shade_godown_quantity__opening_rate'),
        output_field=DecimalField(max_digits=10, decimal_places=2)))

    formset = ShadeFormSet(instance= item_pk, queryset=queryset)

    
    
    if request.method == 'POST':

        if not user.has_perm('product.change_item_creation'):
            messages.error(request, "You do not have permission to edit a item.")
            return redirect('item-list') 
    
        form = Itemform(request.POST, request.FILES , instance = item_pk)
        formset = ShadeFormSet(request.POST , request.FILES, instance = item_pk)
        print(request.POST)
        formset.forms = [form for form in formset if form.has_changed()]
        
        try:
            if form.is_valid() and formset.is_valid():
                form_instance = form.save(commit=False)  
                form_instance.c_user = request.user
                form_instance.save()

                form_instance.bin.set(form.cleaned_data.get('bin', []))

                for form in formset.deleted_forms: 
                    if form.instance.pk:
                        form.instance.delete()

                for form in formset: 
                    if form.is_valid():
                        if form.instance.pk: 
                            form_instance = form.save(commit=False)
                            form_instance.c_user = request.user
                            form_instance.save()

                        else:  
                            
                            if not form.cleaned_data.get('DELETE'):
                                shade_form_instance = form.save(commit=False) 
                                shade_form_instance.c_user = request.user
                                shade_form_instance.save() 

                                form_prefix_number = form.prefix[-1] 
                                opening_godown_quantity = request.POST.get(f'shades-{form_prefix_number}-openingValue') 

                                if opening_godown_quantity != '':
                                    opening_godown_quantity_dict = json.loads(opening_godown_quantity)
                                    opening_godown_qty_data = opening_godown_quantity_dict.get('newData')
                                    
                                    
                                    item_godown_formset_data = {}
                                    for key , value in opening_godown_qty_data.items():
                                        form_set_id = key.split('_')[-1]

                                        new_data_get = opening_godown_qty_data.get(key)
                                        
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-TOTAL_FORMS'] = str(len(opening_godown_qty_data))
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-INITIAL_FORMS'] =  str(0)
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-MIN_NUM_FORMS'] =  str(0)
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-MAX_NUM_FORMS'] =  str(1000)
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-{form_set_id}-opening_godown_id'] = new_data_get.get('godownData')
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-{form_set_id}-opening_quantity'] = new_data_get.get('qtyData')
                                        item_godown_formset_data[f'opening_shade_godown_quantity_set-{form_set_id}-opening_rate'] = new_data_get.get('rateData')
                                    
                                    
                                    new_godown_opening_formsets = OpeningShadeFormSetupdate(item_godown_formset_data, prefix='opening_shade_godown_quantity_set')

                                    for form in new_godown_opening_formsets:
                                        if form.is_valid():
                                            form_instance = form.save(commit = False)
                                            form_instance.opening_purchase_voucher_godown_item = shade_form_instance
                                            form_instance.save()
                logger.info(f"Item updated successfully by {user}")                     
                messages.success(request,'Item updated successfully')
                return redirect('item-list')
            
        except ProtectedError as e:
            messages.error(request,f"Cannot delete item_color_shade due to protected foreign keys: {e}")
            logger.error(f"Cannot delete item_color_shade due to protected foreign keys: {e}")
            print(f"Cannot delete item_color_shade due to protected foreign keys: {e}")

        except Exception as e:
             logger.error(f'An exception occured in item edit - {e}')
             return render(request,'product/item_create_update.html',{'gsts':gsts,'fab_grp':fab_grp,'unit_name':unit_name,'colors':colors,'packaging_material_all':packaging_material_all,'fab_finishes':fab_finishes,'form':form,'formset': formset,"page_name":"Edit raw material"})
        
    return render(request,'product/item_create_update.html',{'gsts':gsts,'fab_grp':fab_grp,'unit_name':unit_name,'colors':colors,'packaging_material_all':packaging_material_all,'fab_finishes':fab_finishes,'form':form,'formset': formset,"page_name":"Edit raw material",'racks':racks})

def item_clone_ajax(request):

    user = request.user

    if not user.has_perm('product.add_item_creation'):
        messages.error(request, "You do not have permission to add a item.")
        return redirect('item-list')
    
    selected_item_name_value = int(request.GET.get('itemValue'))
    
    if selected_item_name_value:
        selected_item = get_object_or_404(Item_Creation, pk=selected_item_name_value)

        response_data = {'fabric_group':{
            'fab_g_key':selected_item.Fabric_Group.id,
            'fab_g_value':selected_item.Fabric_Group.fab_grp_name},
            'color':{'color_key':selected_item.Item_Color.id,'color_value':selected_item.Item_Color.color_name}, 
            'material_code':selected_item.Material_code,'packing':{'packing_key':selected_item.Item_Packing.id ,'packing_value':selected_item.Item_Packing.packing_material},
            'unit_name':{'unit_name_key':selected_item.unit_name_item.id,'unit_name_value':selected_item.unit_name_item.unit_name},
            'panha':selected_item.Panha,'fab_non_fab':selected_item.Fabric_nonfabric,
            'fab_finishes':{'fab_finishes_key':selected_item.Item_Fabric_Finishes.id,'fab_finishes_value':selected_item.Item_Fabric_Finishes.fabric_finish},
            'gst':{'gst_key':selected_item.Item_Creation_GST.id,'gst_value':selected_item.Item_Creation_GST.gst_percentage},
            'hsn_code':selected_item.HSN_Code,'status':selected_item.status, 'unit_name_units':selected_item.unit_name_item.unit_value}
       
    return JsonResponse({'response_data':response_data})

def item_list(request):

    user = request.user

    logger.info(f"item_list function run by {user}")

    if not user.has_perm('product.view_item_creation'):
        messages.error(request, "You do not have permission to view items")
        return redirect('dashboard-main')
    
    g_search = request.GET.get('item_search','')

    queryset = Item_Creation.objects.all().annotate(total_quantity=Sum('shades__godown_shades__quantity'), shade_num = Count('shades', distinct=True),godown_num=Count('shades__godown_shades', distinct=True)).order_by('item_name').select_related('Item_Color','unit_name_item','Fabric_Group','Item_Creation_GST','Item_Fabric_Finishes','Item_Packing').prefetch_related('shades','shades__godown_shades','shades__godown_shades__godown_name','bin')
    
    bins = {}

    for item in queryset:
        item_name = item.item_name
        if item_name not in bins:
            bins[item_name] = {}

        for bin in item.bin.all():
            rack_name = bin.rack.rack_name

            if rack_name not in bins[item_name]:
                bins[item_name][rack_name] = []

            if bin.bin_name not in bins[item_name][rack_name]:
                bins[item_name][rack_name].append(bin.bin_name)

    if g_search != '':
        queryset = queryset.filter(Q(item_name__icontains = g_search)| Q(Item_Color__color_name__icontains = g_search)| Q(Fabric_Group__fab_grp_name__icontains = g_search)| Q(Material_code__icontains = g_search))
        
    sort_name = request.GET.get('sort_name')

    if sort_name == "item_name_sort_asc" :
        queryset = Item_Creation.objects.order_by('item_name')
    
    elif sort_name == "item_name_sort_dsc" :
        queryset = Item_Creation.objects.order_by('-item_name')

    elif sort_name == "fabgrp_sort_asc" :
        queryset = Item_Creation.objects.order_by('Fabric_Group__fab_grp_name')


    elif sort_name == "fabgrp_sort_dsc" :
        queryset = Item_Creation.objects.order_by('-Fabric_Group__fab_grp_name')

    elif sort_name == "item_color_sort_dsc" :
        queryset = Item_Creation.objects.order_by('Item_Color__color_name')
    
    elif sort_name == "item_color_sort_dsc" :
        queryset = Item_Creation.objects.order_by('-Item_Color__color_name')

    any_desc = request.GET.get('any_desc')
    exact_desc = request.GET.get('exact_desc')

    if any_desc:
        if any_desc != '' and any_desc is not None:
            queryset = Item_Creation.objects.filter(item_name__icontains=any_desc)
        
    if exact_desc:
        if exact_desc != '' and exact_desc is not None:
            queryset = Item_Creation.objects.filter(item_name__exact=exact_desc)

    return render(request,'product/list_item.html', {"items":queryset,"item_search":g_search,"page_name":"Raw Materials","bins": bins,})
    
def openingquantityformsetpopup(request,parent_row_id=None,primary_key=None):

    user = request.user

    if not user.has_perm('product.view_godown_raw_material'):
        messages.error(request, "You do not have permission to view qty")
        return redirect('dashboard-main')
    
    godowns =  Godown_raw_material.objects.all()

    formset = None
    shade_instance = None
    try:
        
        if parent_row_id is not None and primary_key is not None:
            shade_instance = get_object_or_404(item_color_shade,pk=primary_key)
            formset = OpeningShadeFormSetupdate(request.POST or None, instance = shade_instance, prefix = "opening_shade_godown_quantity_set")
            
        
        elif primary_key is None and parent_row_id is not None:

            decoded_data = False
            encoded_data = request.GET.get('data')
            
            
            if encoded_data:
                decoded_data = json.loads(urllib.parse.unquote(encoded_data))
                new_row_data = decoded_data.get('newData', {})
                initial_data_backend = []
                
                
                for key, value in new_row_data.items():
                    initial_data_backend.append({
                            "opening_godown_id": int(value['godownData']),
                            "opening_quantity": float(value['qtyData']),
                            "opening_rate": float(value['rateData'])})

                
                
                total_forms = len(initial_data_backend)
                opening_shade_godown_quantitycreateformset = modelformset_factory(opening_shade_godown_quantity, fields = ['opening_godown_id','opening_quantity','opening_rate'], extra=total_forms, can_delete=True)   
                formset = opening_shade_godown_quantitycreateformset(queryset=opening_shade_godown_quantity.objects.none(),initial=initial_data_backend, prefix = "opening_shade_godown_quantity_set")
                
            else:
                
                opening_shade_godown_quantitycreateformset = modelformset_factory(opening_shade_godown_quantity, fields = ['opening_rate','opening_quantity','opening_godown_id'], extra=1, can_delete=True)            
                formset = opening_shade_godown_quantitycreateformset(queryset=opening_shade_godown_quantity.objects.none(),prefix = "opening_shade_godown_quantity_set")
        
        if request.method == 'POST':
            
            if not user.has_perm('product.add_godown_raw_material'):
                messages.error(request, "You do not have permission to add qty")
                return redirect('dashboard-main')
            
            if not user.has_perm('product.change_godown_raw_material'):
                messages.error(request, "You do not have permission to update qty")
                return redirect('dashboard-main')
            
            if not user.has_perm('product.delete_godown_raw_material'):
                messages.error(request, "You do not have permission to delete qty")
                return redirect('dashboard-main')
            
            if primary_key is not None:

                for form in formset.deleted_forms: 
                    if form.instance.pk:
                        try:
                            form.instance.delete()
                        except Exception as e:

                            return JsonResponse({"error": str(e)}, status=400)

                formset.forms = [form for form in formset if form.has_changed()] 
                if formset.is_valid():
                    for form in formset:
                        if form.is_valid():
                            if not form.cleaned_data.get('DELETE'):
                                try:
                                    old_opening_instance = opening_shade_godown_quantity.objects.get(id= form.instance.id) 
                                    
                                    form_instance = form.save(commit=False)

                                    form_instance.old_opening_godown_id = old_opening_instance.opening_godown_id 
                                    form_instance.old_opening_g_quantity = old_opening_instance.opening_quantity 
                                    form_instance.save()

                                except opening_shade_godown_quantity.DoesNotExist:
                                    form_instance = form.save(commit=False)
                                    
                                    form_instance.old_opening_g_quantity = 0 
                                    form_instance.save()
                                
                                except Exception as e:
                                    
                                    return JsonResponse({"error": str(e)}, status=400)

                    return HttpResponse('<script>window.close();</script>') 
                
                else:
                    logging.error(f'Error in item opening godown formset{formset.error})')
                    return JsonResponse({"errors": formset.errors}, status=400)
                
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)     
    return render(request,'product/opening_godown_qty.html',{'formset':formset,'godowns':godowns ,"parent_row_id":parent_row_id, 'primary_key':primary_key,'shade_instance':shade_instance})


def openingquantityformsetpopupajax(request):
    itemValue_get = request.GET.get('itemValue')
    primary_key_id_get = request.GET.get('primary_key_id')        

    if itemValue_get is not None and primary_key_id_get != '':
        popup_url = reverse('opening-godown-qty-pk', args=[primary_key_id_get,itemValue_get])

    elif itemValue_get is not None:
        popup_url = reverse('opening-godown-qty', args=[itemValue_get])

    else:
        popup_url = None
    
    return JsonResponse({'popup_url':popup_url})


@permission_required('product.delete_item_creation', raise_exception=True)
def item_delete(request, pk):
    try:
        item_pk = get_object_or_404(Item_Creation,pk = pk)
        item_pk.delete()
        messages.success(request,f'Item {item_pk.item_name} was deleted')
        logger.info(f'Item {item_pk.item_name} was deleted')
    
    except Exception as e:
         messages.error(request, f'EXCEPTION-{e}')
         logger.error( f'EXCEPTION-{e}')
         print(e)
    return redirect('item-list')


# ======================== SUBCATEGORY FUNCTIONS ===========================


def color_create_update(request, pk=None):

    user = request.user

    
    
    if request.path == '/color_popup/':
        if not (user.has_perm('product.add_color') or user.has_perm('product.change_color')):
            return JsonResponse({'error': 'You do not have permission to add or edit color.'})

    if request.path == '/color_popup/':
        if not user.has_perm('product.view_color'):
            return JsonResponse({'error': 'You do not have permission to add color.'})

    if not user.has_perm('product.view_color'):
        messages.error(request, "You do not have permission to view Colors")
        return redirect('dashboard-main')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    queryset = Color.objects.all()
    color_search = request.GET.get('color_search','')

    if color_search != '':
        queryset = Color.objects.filter(color_name__icontains = color_search)

    if request.path == '/simple_colorcreate_update/':
        template_name = 'product/color_create_update.html'
        
    elif request.path == f'/simple_colorcreate_update/{pk}':
        template_name = 'product/color_create_update.html'
        
    elif request.path == '/colorcreate_update/':
        template_name = "product/create_color_modal.html"
        
    elif request.path == '/color_popup/':
        template_name = "product/color_popup.html"
        
    if pk:
        instance = get_object_or_404(Color, pk = pk)
        page_name='Edit Color'
    else:
        instance = None
        page_name='Add Color'

    form = ColorForm(instance=instance)

    if request.method == 'POST':
        
        if instance and not user.has_perm('product.change_color'):
            messages.error(request, "You do not have permission to update a color.")
            return redirect('simplecolorlist')
        
        if not instance and not user.has_perm('product.add_color'):
            messages.error(request, "You do not have permission to add a color.")
            return redirect('simplecolorlist')

        form = ColorForm(request.POST, instance=instance)

        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if request.path == '/simple_colorcreate_update/' or request.path == f'/simple_colorcreate_update/{pk}':
                if instance:
                    messages.success(request, 'Color updated successfully.')
                else:
                    messages.success(request, 'Color created successfully.')
                
                return redirect('simplecolorlist')
            
            elif template_name == "product/color_popup.html":
                
                color_all = Color.objects.all().values('id','color_name')
                messages.success(request, 'Color created successfully.')
                return JsonResponse({'color_all':list(color_all)}) 
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)
    
    return render(request, template_name , {'form': form, 'colors':queryset,'color_search':color_search,'page_name':page_name})


@permission_required('product.delete_color', raise_exception=True)
def color_delete(request, pk):
    product_color = get_object_or_404(Color,pk=pk)
    try:
        product_color.delete()
        messages.success(request,f'Color {product_color.color_name} was deleted')

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {product_color.color_name} because it is referenced by other objects.')
    return redirect('simplecolorlist')


def item_fabric_group_create_update(request, pk = None):

    user = request.user

    logger.info(f"item_fabric_group_create_update function run by {user}")
        
    if request.path == '/fabric_popup/':
        if not (user.has_perm('product.add_fabric_group_model') or user.has_perm('product.change_fabric_group_model')):
            return JsonResponse({'error': 'You do not have permission to add or edit fabric group.'})

        
    if request.path == '/fabric_popup/':
        if not user.has_perm('product.view_fabric_group_model'):
            return JsonResponse({'error': 'You do not have permission to view fabric group.'})


    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    if not user.has_perm('product.view_fabric_group_model'):
        messages.error(request, "You do not have permission to view fabric groups.")
        return redirect('dashboard-main')
    
    queryset = Fabric_Group_Model.objects.all()

    fabric_group_search = request.GET.get('fabric_group_search','')

    if fabric_group_search != '':
        queryset = Fabric_Group_Model.objects.filter(fab_grp_name__icontains = fabric_group_search)

    if pk:
        item_fabric_pk =  get_object_or_404(Fabric_Group_Model,pk=pk)
        instance = item_fabric_pk
        page_name = "Edit Fabric"
    else:
        form = ItemFabricGroup()
        instance = None
        page_name = "Add Fabric"

    if request.path == '/itemfabricgroupcreateupdate/':
        template_name = 'product/item_fabric_group_create_update.html'
        
    elif request.path == '/fabric_popup/':
        template_name = 'product/fabric_popup.html'
        
    elif request.path == f'/itemfabricgroupcreateupdate/{pk}':
        template_name = 'product/item_fabric_group_create_update.html'
        

    form = ItemFabricGroup(instance=instance)

    if request.method == 'POST':
        
        if instance and not user.has_perm('product.change_fabric_group_model'):
            messages.error(request, "You do not have permission to update a fabric.")
            return redirect('item-fabgroup-create-list')
        
        if not instance and not user.has_perm('product.add_fabric_group_model'):
            messages.error(request, "You do not have permission to add a fabric.")
            return redirect('item-fabgroup-create-list')

        form = ItemFabricGroup(request.POST, instance=instance)

        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'Fabric group updated sucessfully.')
                logger.info(f"fabric group {form.cleaned_data.get('fab_grp_name')} update by {user}")
            else:
                messages.success(request,'Fabric group created sucessfully.')
                logger.info(f"fabric group {form.cleaned_data.get('fab_grp_name')} created by {user}")

            if template_name == 'product/item_fabric_group_create_update.html':
                return redirect('item-fabgroup-create-list')

            elif template_name == 'product/fabric_popup.html':
                Fabric_Group_all = Fabric_Group_Model.objects.all().values('id','fab_grp_name')
                return JsonResponse({'Fabric_Group_all':list(Fabric_Group_all)})

        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,template_name,{"fab_group_all":queryset,"fabric_group_search":fabric_group_search,'form':form,'page_name':page_name})


@permission_required('product.delete_fabric_group_model', raise_exception=True)
def item_fabric_group_delete(request,pk):
    user = request.user

    logger.info(f"item_fabric_group_delete function run by {user}")

    try:
        item_fabric_pk = get_object_or_404(Fabric_Group_Model,pk=pk)
        item_fabric_pk.delete()

        messages.success(request,f'Fabric group {item_fabric_pk.fab_grp_name} was deleted')
        logger.info(f"Fabric group {item_fabric_pk.fab_grp_name} was deleted by {user}")

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {item_fabric_pk.fab_grp_name} because it is referenced by other objects.')
    
    return redirect('item-fabgroup-create-list')


def unit_name_create_update(request,pk=None):

    user = request.user

    logger.info(f"unit_name_create_update function run by {user}")

    if request.path == '/units_popup/':
        if not (user.has_perm('product.add_unit_name_create') or user.has_perm('product.change_unit_name_create')):
            return JsonResponse({'error': 'You do not have permission to add or edit unit.'})

    
    if request.path == '/units_popup/':
        if not user.has_perm('product.view_unit_name_create'):
            return JsonResponse({'error': 'You do not have permission to add or edit unit.'})

    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})

    if not user.has_perm('product.view_unit_name_create'):
        messages.error(request, "You do not have permission to view units")
        return redirect('dashboard-main')
    
    queryset = Unit_Name_Create.objects.all()

    unit_name_search =  request.GET.get('unit_name_search','')

    if unit_name_search != '':
        queryset = Unit_Name_Create.objects.filter(unit_name__icontains = unit_name_search)

    if pk:
        unit_name_pk = get_object_or_404(Unit_Name_Create,pk=pk)
        instance = unit_name_pk
        page_name = "Edit Unit"
    else:
        instance= None
        page_name = "Add Unit"

    form = UnitName(instance = instance)

    if request.path == '/unitnamecreate/':
        template_name = 'product/unit_name_create_update.html'
        
    elif request.path == '/units_popup/':
        template_name = 'product/units_popup.html'
        
    elif request.path == f'/unitnameupdate/{pk}':
        template_name = 'product/unit_name_create_update.html'
        
    if request.method == 'POST':

        if instance and not user.has_perm('product.change_unit_name_create'):
            messages.error(request, "You do not have permission to update a unit.")
            return redirect('unit_name-create_list')

        if not instance and not user.has_perm('product.add_unit_name_create'):
            messages.error(request, "You do not have permission to add a unit.")
            return redirect('unit_name-create_list')
        

        form = UnitName(request.POST, instance=instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'Unit updated sucessfully.')
                logger.info(f"{form_instance.unit_name} updated by {user}")
            else:
                messages.success(request,'Unit created sucessfully.')
                logger.info(f"{form_instance.unit_name} created by {user}")

            
            if  template_name == 'product/unit_name_create_update.html':
                return redirect('unit_name-create_list')

            elif template_name == 'product/units_popup.html':
                Unit_Name_all_values = Unit_Name_Create.objects.all().values('id','unit_name')
                return JsonResponse({'Unit_Name_all_values':list(Unit_Name_all_values)})

        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)
        
    else:
        return render(request, template_name, {'form':form,"unit_name_all":queryset,'unit_name_search':unit_name_search,'page_name':page_name})


def unit_name_units_ajax(request):
    
    user = request.user

    unit_name_pk = request.GET.get('unit_name_pk')

    if unit_name_pk is not None:

        if not user.has_perm('product.add_unit_name_create'):
            return JsonResponse({'error': 'You do not have permission to add a unit.', 'close': True}, status=403)
    
        unit_name_instance = get_object_or_404(Unit_Name_Create,pk=unit_name_pk)
        unit_name_units = unit_name_instance.unit_value

        logger.info(f"{unit_name_units} updated by {user}")

        return JsonResponse({'unit_name_units':unit_name_units})
    else:
        return JsonResponse({'error': 'Invalid Request'}, status=400)   


@permission_required('product.delete_unit_name_create', raise_exception=True)
def unit_name_delete(request,pk):

    user = request.user

    logger.info(f"unit_name_delete function run by {user}")

    try:
        unit_name_pk = get_object_or_404(Unit_Name_Create,pk=pk)
        unit_name_pk.delete()
        logger.info(f"Unit name {unit_name_pk.unit_name} deleted by {user}")
        messages.success(request,f'Unit name {unit_name_pk.unit_name} was deleted.')
    except IntegrityError as e:
        messages.error(request,f'Cannot delete {unit_name_pk.unit_name} because it is referenced by other objects.')
    return redirect('unit_name-create_list')


def packaging_create_update(request, pk = None):
    
    user = request.user

    logger.info(f"packaging_create_update function run by {user}")
        
    if request.path == '/packagingpop/':
        if not (user.has_perm('product.add_packaging') or user.has_perm('product.change_packaging')):
            return JsonResponse({'error': 'You do not have permission to add or edit Packing.'})
        
    if request.path == '/packagingpop/':
        if not user.has_perm('product.view_packaging'):
            return JsonResponse({'error': 'You do not have permission to view Packing.'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    if not user.has_perm('product.view_packaging'):
        messages.error(request, "You do not have permission to view Packing.")
        return redirect('dashboard-main')


    queryset =  packaging.objects.all()

    packaging_search = request.GET.get('packaging_search','')

    if packaging_search != '':
        queryset =  packaging.objects.filter(packing_material__icontains = packaging_search)

    if pk:
        packaging_instance = packaging.objects.get(pk=pk)
        page_name = "Edit Packaging"
    else:
        packaging_instance = None
        page_name = "Add Packaging"

    if request.path == '/packagingpop/':
        template_name = 'misc/packaging_popup.html'

    elif request.path == '/packaging_create/' or f'/packagingupdate/{pk}':
        template_name = 'misc/packaging_create_update.html'
        
    form = packaging_form(instance = packaging_instance)

    if request.method == 'POST':

        if packaging_instance and not user.has_perm('product.change_packaging'):
            messages.error(request, "You do not have permission to update a Packing.")
            return redirect('packaging-create-list')

        if not packaging_instance and not user.has_perm('product.add_packaging'):
            messages.error(request, "You do not have permission to add a Packing.")
            return redirect('packaging-create-list')


        form = packaging_form(request.POST ,instance=packaging_instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'packing updated sucessfully.')
                logger.info(f"{form_instance.packing_material} updated by {user}")
            else:
                messages.success(request,'packing created sucessfully.')
                logger.info(f"{form_instance.packing_material} created by {user}")

            if template_name == 'misc/packaging_create_update.html':
                return redirect('packaging-create-list')

            elif template_name == 'misc/packaging_popup.html':

                packaging_all_values = packaging.objects.all().values('id','packing_material')

                return JsonResponse({'packaging_all_values': list(packaging_all_values)})
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request, template_name ,{'form':form,'packaging_all':queryset,'page_name':page_name})


@permission_required('product.delete_packaging', raise_exception=True)
def packaging_delete(request,pk):

    user = request.user

    logger.info(f"packaging_delete function run by {user}")

    packaging_pk =  packaging.objects.get(pk=pk)
    packaging_pk.delete()

    logger.info(f"{packaging_pk.packing_material} delete by {user}")
    messages.success(request, 'Packing deleted.')

    return redirect('packaging-create-list')


def gst_create_update(request, pk = None):

    user = request.user

    logger.info(f"gst_create_update function run by {user}")

    if request.path == '/gstpopup/':
        if not (user.has_perm('product.add_gst') or user.has_perm('product.change_gst')):
            return JsonResponse({'error': 'You do not have permission to add or edit GST'})

    if request.path == '/gstpopup/':
        if not user.has_perm('product.view_gst'):
            return JsonResponse({'error': 'You do not have permission to view GST'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})

    if not user.has_perm('product.view_gst'):
        messages.error(request, "You do not have permission to view GST")
        return redirect('dashboard-main')
    
    queryset =  gst.objects.all()

    gst_search = request.GET.get('gst_search','')

    if gst_search != "":
        queryset =  gst.objects.filter(gst_percentage__contains = gst_search)

    if pk:
        instance = gst.objects.get(pk=pk)
        page_name = "Edit GST"

    else:
        instance = None
        page_name = "Add GST"


    if request.path == '/gstpopup/':
        template_name = 'accounts/gst_popup.html'

    elif request.path == '/gstcreate/':
        template_name = 'accounts/gst_create_update.html'
        
    elif request.path == f'/gstupdate/{pk}':
        template_name = 'accounts/gst_create_update.html'
        

    form = gst_form(instance = instance)
    if request.method == 'POST':

        if not instance and not user.has_perm('product.add_gst'):
            messages.error(request, "You do not have permission to add GST")
            return redirect('gst-create-list')

        if instance and not user.has_perm('product.change_gst'):
            messages.error(request, "You do not have permission to update GST")
            return redirect('gst-create-list')
    
        form = gst_form(request.POST, instance = instance)
        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()
            
            if pk:
                messages.success(request,'GST updated successfully.')
                logger.info(f"{form_instance.gst_percentage} updated by {user}")
            else:
                messages.success(request,'GST created successfully.')
                logger.info(f"{form_instance.gst_percentage} created by {user}")

            if template_name == 'accounts/gst_create_update.html':
                
                return redirect('gst-create-list')

            elif template_name == 'accounts/gst_popup.html':
                
                gst_updated = gst.objects.all().values('id', 'gst_percentage')
                
                return JsonResponse({"gst_updated": list(gst_updated)})
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,template_name,{'form':form,'gsts':queryset,'page_name':page_name})


@permission_required('product.delete_gst', raise_exception=True)
def gst_delete(request,pk):

    user = request.user

    logger.info(f"gst_delete function run by {user}")

    gst_pk = gst.objects.get(pk=pk)
    gst_pk.delete()

    logger.info(f"{gst_pk.gst_percentage} GST deleted {user}")
    messages.success(request,f"{gst_pk.gst_percentage}GST deleted")

    return redirect('gst-create-list')


def fabric_finishes_create_update(request, pk = None):

    user = request.user

    if request.path == '/fabricfinishespopup/':
        if not (user.has_perm('product.add_fabricfinishes') or user.has_perm('product.change_fabricfinishes')):
            return JsonResponse({'error': 'You do not have permission to add or edit fabric finishes.'})

    if request.path == '/fabricfinishespopup/':
        if not user.has_perm('product.view_fabricfinishes'):
            return JsonResponse({'error': 'You do not have permission to add fabric finishes.'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})

    
    if not user.has_perm('product.view_fabricfinishes'):
        messages.error(request, "You do not have permission to view fabric finishes.")
        return redirect('dashboard-main')
    
    queryset =  FabricFinishes.objects.all()

    fabric_finishes_search = request.GET.get('fabric_finishes_search','')

    if fabric_finishes_search != '':
        queryset =  FabricFinishes.objects.filter(fabric_finish__icontains = fabric_finishes_search)

    if pk:
        fabric_finishes_instance = FabricFinishes.objects.get(pk=pk)
        page_name = "Edit Fab Finish"
    else:
        fabric_finishes_instance = None
        page_name = "Add Fab Finish"

    if request.path == '/fabricfinishespopup/':
        template_name = 'misc/fabric_finishes_popup.html'

    elif request.path == '/fabricfinishesscreate/' or f'/fabricfinishesupdate/{pk}':
        template_name = 'misc/fabric_finishes_create_update.html'
    

    form = FabricFinishes_form(instance = fabric_finishes_instance)

    if request.method == 'POST':

        if fabric_finishes_instance and not user.has_perm('product.change_fabricfinishes'):
            messages.error(request, "You do not have permission to update a fabric finish.")
            return redirect('fabric-finishes-create-list')

        if not fabric_finishes_instance and not user.has_perm('product.add_fabricfinishes'):
            messages.error(request, "You do not have permission to add a fabric finish.")
            return redirect('fabric-finishes-create-list')
        
        
        form = FabricFinishes_form(request.POST,instance = fabric_finishes_instance)

        if form.is_valid():
            form_instance = form.save(commit=False)
            form_instance.c_user = request.user
            form_instance.save()

            if pk:
                messages.success(request,'fabric finish updated sucessfully')
            else:
                messages.success(request,'fabric finish created sucessfully')

            if template_name == 'misc/fabric_finishes_create_update.html':
                return redirect('fabric-finishes-create-list')
            
            elif template_name == 'misc/fabric_finishes_popup.html':
                fabric_finishes_all = FabricFinishes.objects.all().values('id', 'fabric_finish')
                
                return JsonResponse({"fabric_finishes_all": list(fabric_finishes_all)})
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,template_name,{'form':form,'fabricfinishes':queryset,'fabric_finishes_search':fabric_finishes_search,'page_name':page_name})


@permission_required('product.delete_fabricfinishes', raise_exception=True)
def fabric_finishes_delete(request,pk):

    user = request.user
    
    logger.info(f"fabric_finishes_delete function run by {user}")

    fabric_finish =  FabricFinishes.objects.get(pk=pk)
    fabric_finish.delete()

    logger.info(f"{fabric_finish.fabric_finish} fabric finish deleted by {user}")
    messages.success(request,'fabric finish deleted.')
    return redirect('fabric-finishes-create-list')


def cutting_room_create_update_list(request, pk=None):

    user = request.user
    
    logger.info(f"cutting_room_create_update_list function run by {user}")

    if not user.has_perm('product.view_cutting_room'):
        messages.error(request, "You do not have permission to view cutting rooms")
        return redirect('dashboard-main')
    
    cutting_rooms = cutting_room.objects.all()
   
    if pk:
        instance = cutting_room.objects.get(id = pk)
        page_name = "Edit Cutt Room"
    else:
        instance = None
        page_name = "Add Cutt Room"

    form = cutting_room_form(request.POST or None, instance = instance) 

    if request.method == 'POST':

        if instance and not user.has_perm('product.change_cutting_room'):
            messages.error(request, "You do not have permission to update a cutting room.")
            return redirect('cutting_room-create')

        if not instance and not user.has_perm('product.add_cutting_room'):
            messages.error(request, "You do not have permission to add a cutting room.")
            return redirect('cutting_room-create')
        
        if form.is_valid():
            form.save()

            if pk:
                messages.success(request,f"{form.cleaned_data['cutting_room_name']} cutting room created succesfully")
                logger.info(f"{form.cleaned_data['cutting_room_name']} cutting room by {user}")
            else:
                messages.success(request,f"{form.cleaned_data['cutting_room_name']} cutting room updated succesfully")
                logger.info(f"{form.cleaned_data['cutting_room_name']} cutting room by {user}")

            return redirect('cutting_room-create')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,'production/cuttingroomcreateupdatelist.html', {'form':form,'cutting_rooms':cutting_rooms,'page_name':page_name})


@permission_required('product.delete_cutting_room', raise_exception=True)
def cuttingroomdelete(request,pk):
    instance = cutting_room.objects.get(pk=pk)
    instance.delete()
    return redirect('cutting_room-create')


def factory_employee_create_update_list(request ,pk=None):

    user = request.user

    logger.info(f"factory_employee_create_update_list function run by {user}")

    if not user.has_perm('product.view_factory_employee'):
        messages.error(request, "You do not have permission to view factory employee")
        return redirect('dashboard-main')
        
    factory_employees = factory_employee.objects.all()
    cutting_rooms =  cutting_room.objects.all()

    if pk:
        instance = get_object_or_404(factory_employee,pk=pk)
        page_name = "Edit Factory Emp"
    else:
        instance = None
        page_name = "Add Factory Emp"
    
    form = factory_employee_form(request.POST or None, instance = instance)

    if request.method == 'POST':

        if instance and not user.has_perm('product.change_factory_employee'):
            messages.error(request, "You do not have permission to update a factory employee.")
            return redirect('factory-emp-create')

        if not instance and not user.has_perm('product.add_factory_employee'):
            messages.error(request, "You do not have permission to add a factory employee.")
            return redirect('factory-emp-create')
        
        if form.is_valid():
            form.save()

            messages.success(request,'Factory Employee created Successfully')
            logger.info(f"{form.factory_emp_name} Factory Employee created Successfully by {user}")
            return redirect('factory-emp-create')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,'production/factory_emp_create_update_list.html', {'form':form,'factory_employees':factory_employees,'cutting_rooms':cutting_rooms,'page_name':page_name})


@permission_required('product.delete_factory_employee', raise_exception=True)
def factoryempdelete(request,pk=None):

    user = request.user

    logger.info(f"factoryempdelete function run by {user}")

    try:
        instance = get_object_or_404(factory_employee,pk=pk)
        instance.delete()
        messages.success(request,f'Factory Employee {instance.factory_emp_name} was deleted')
        logger.info(f"Factory Employee {instance.factory_emp_name} was deleted by {user}")

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {instance.factory_emp_name} because it is referenced by other objects.')
    return redirect('factory-emp-create')


def salesman_create_update(request,e_id=None):

    user = request.user

    logger.info(f"salesman_create_update function run by {user}")

    if not user.has_perm('product.view_salesman_info'):
        messages.error(request, "You do not have permission to view salesman")
        return redirect('dashboard-main')

    salesman_list = Salesman_info.objects.all()

    if e_id:
        salesman_instance = Salesman_info.objects.get(id = e_id)
        form = SalesmaninfoForm(instance=salesman_instance)
        page_name = "Edit Salesman"
    else:
        salesman_instance = None
        form = SalesmaninfoForm()
        page_name = "Add Salesman"

    if request.method == 'POST':

        if salesman_instance and not user.has_perm('product.change_salesman_info'):
            messages.error(request, "You do not have permission to update a salesman.")
            return redirect('salesman-create')

        if not salesman_instance and not user.has_perm('product.add_salesman_info'):
            messages.error(request, "You do not have permission to add a salesman.")
            return redirect('salesman-create')
        
        form = SalesmaninfoForm(request.POST or None, instance=salesman_instance)

        if form.is_valid():
            salesman = form.save()
            messages.success(request, f"{salesman.salesman_name} added successfully")
            logger.info(f"{salesman.salesman_name} added successfully by {user}")
            return redirect('salesman-create')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,'production/salesman_create_update.html',{'form':form,'salesman_list':salesman_list,'page_name':page_name})


@permission_required('product.delete_salesman_info', raise_exception=True)
def delete_salesman(request,e_id):

    user = request.user

    logger.info(f"delete_salesman function run by {user}")

    try:
        instance = get_object_or_404(Salesman_info,pk=e_id)
        instance.delete()
        messages.success(request,f'salesman {instance.salesman_name} was deleted')
        logger.info(f"salesman {instance.salesman_name} was deleted by {user}")

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {instance.salesman_name} because it is referenced by other objects.')
    return redirect('salesman-create')


def create_update_rack_for_raw_material(request, r_id=None):

    user = request.user

    logger.info(f"create_update_rack_for_raw_material function run by {user}")

    try:

        if not user.has_perm('product.view_rack_for_raw_material'):
            messages.error(request, "You do not have permission to view rack")
            return redirect('dashboard-main')
        
        rack_list = rack_for_raw_material.objects.all()

        rack_instance = get_object_or_404(rack_for_raw_material, id=r_id) if r_id else None

        if rack_instance:
            page_name = "Edit Rack" 
        else:
            page_name = "Add Rack"

        form = rack_for_raw_material_form(request.POST or None, instance=rack_instance)

        if request.method == 'POST':

            if rack_instance and not user.has_perm('product.change_rack_for_raw_material'):
                messages.error(request, "You do not have permission to update a rack.")
                return redirect('create-rack-for-raw-material')

            if not rack_instance and not user.has_perm('product.add_rack_for_raw_material'):
                messages.error(request, "You do not have permission to add a rack.")
                return redirect('create-rack-for-raw-material')


            if form.is_valid():
                form.save()
                messages.success(request, f"{form.rack_name} rack saved successfully!")
                logger.info(f"{form.rack_name} rack saved successfully by {user}")
                return redirect('create-rack-for-raw-material')
            else:
                messages.error(request, "Please correct the errors below.")
 
    except bin_for_raw_material.DoesNotExist:
        messages.error(request, "The rack record does not exist.")
        return redirect('create-rack-for-raw-material')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        print(f"Error in create_update_rack_for_raw_material: {str(e)}")


    return render(request, 'product/create_update_rack_for_raw_material.html', {'form': form,'rack_list': rack_list,'page_name':page_name})


@permission_required('product.delete_rack_for_raw_material', raise_exception=True)
def delete_rack_for_raw_material(request,r_id):

    user = request.user

    logger.info(f"delete_rack_for_raw_material function run by {user}")

    try:
        rack_instance = get_object_or_404(rack_for_raw_material, id=r_id)
        
        rack_instance.delete()
        messages.success(request, f"{rack_instance.rack_name} rack deleted successfully!")
        logger.info(f"{rack_instance.rack_name} rack deleted successfully by {user}")

    except rack_for_raw_material.DoesNotExist:
        messages.error(request, "The rack does not exist.")

    except Exception as e:
        messages.error(request, f"An error occurred while deleting the rack: {str(e)}")
        print(f"Error in delete_rack_for_raw_material: {str(e)}")

    return redirect('create-rack-for-raw-material')


def create_update_bin_for_raw_material(request, r_id=None,b_id=None):

    user = request.user

    logger.info(f"create_update_bin_for_raw_material function run by {user}")

    try:

        if not user.has_perm('product.view_bin_for_raw_material'):
            messages.error(request, "You do not have permission to view bins")
            return redirect('create-rack-for-raw-material')
        
        bin_list = bin_for_raw_material.objects.filter(rack = r_id)

        bin_instance = get_object_or_404(bin_for_raw_material, id=b_id) if b_id else None

        if bin_instance:
            page_name = "Edit Bin"
        else:
            page_name = "Add Bin"

        form = bin_for_raw_material_form(request.POST or None, instance=bin_instance)

        rack_instance = get_object_or_404(rack_for_raw_material, id=r_id) if r_id else None

        if request.method == 'POST':

            if bin_instance and not user.has_perm('product.change_bin_for_raw_material'):
                messages.error(request, "You do not have permission to update a bin.")
                return redirect('create-bin-for-raw-material', r_id=r_id)

            if not bin_instance and not user.has_perm('product.add_bin_for_raw_material'):
                messages.error(request, "You do not have permission to add a bin.")
                return redirect('create-bin-for-raw-material', r_id=r_id)

            if form.is_valid():
                bin_form = form.save(commit=False)
                bin_form.rack = rack_instance
                bin_form.save()

                messages.success(request, f"{bin_form.bin_name} bin details saved successfully!")
                logger.info(f"{bin_form.bin_name} bin details saved successfully by {user}")
                return redirect('create-bin-for-raw-material', r_id=r_id)
            else:
                messages.error(request, "Please correct the errors below.")

    except bin_for_raw_material.DoesNotExist:
        messages.error(request, "The bin record does not exist.")
        return redirect('create-bin-for-raw-material', r_id=r_id)

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        print(f"Error in create_update_bin_for_raw_material: {str(e)}")

    return render(request, 'product/create_update_bin_for_raw_material.html', {'form': form,'bin_list':bin_list,'rack_instance':rack_instance,'page_name':page_name})


@permission_required('product.delete_bin_for_raw_material', raise_exception=True)
def delete_bin_for_raw_material(request,b_id):

    user = request.user

    logger.info(f"delete_bin_for_raw_material function run by {user}")

    try:
        bin_instance = get_object_or_404(bin_for_raw_material, id=b_id)
        rack_id = bin_instance.rack.id
        if Item_Creation.objects.filter(bin=bin_instance).exists():
            messages.error(request, "Cannot delete: This bin is assigned to one or more items.")
            return redirect('create-bin-for-raw-material', r_id=rack_id)
        
        bin_instance.delete()
        messages.success(request, f"{bin_instance.bin_name} bin deleted successfully!")
        logger.info(f"{bin_instance.bin_name} bin deleted successfully by {user}")
    except bin_for_raw_material.DoesNotExist:
        messages.error(request, "The bin does not exist.")

    except Exception as e:
        messages.error(request, f"An error occurred while deleting the bin: {str(e)}")
        print(f"Error in delete_bin_for_raw_material: {str(e)}")

    return redirect('create-bin-for-raw-material', r_id=rack_id)


# ================== ACCOUNTS FUNCTIONS =====================

def account_sub_group_create_update(request, pk=None):

    user = request.user

    logger.info(f"account_sub_group_create_update function run by {user}")

    if not user.has_perm('product.view_accountsubgroup'):
        messages.error(request, "You do not have permission to view account sub group")
        return redirect('dashboard-main')

    groups = AccountSubGroup.objects.select_related('acc_grp').all()

    if pk:
        instance = get_object_or_404(AccountSubGroup ,pk=pk)
        page_name='Edit Group'
    else:
        instance = None
        page_name='Add Group'

    main_grp = AccountGroup.objects.all()
    form = account_sub_grp_form(request.POST or None, instance=instance)

    if request.method == 'POST':

        if instance and not user.has_perm('product.change_accountsubgroup'):
            messages.error(request, "You do not have permission to update account sub group.")
            return redirect('account_sub_group-create')

        if not instance and not user.has_perm('product.add_accountsubgroup'):
            messages.error(request, "You do not have permission to add account sub group.")
            return redirect('account_sub_group-create')
        
        if form.is_valid():
            form.save()
            messages.success(request, f"{form.cleaned_data['account_sub_group']} Account sub-group created sucessfully")
            logger.info(f"{form.cleaned_data['account_sub_group']} Account sub-group created sucessfully by {user}")
            return redirect('account_sub_group-create')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)
    return render(request,'product/acc_sub_grp_create_update.html',{'main_grp':main_grp,'form':form,"groups":groups,'page_name':page_name})


@permission_required('product.delete_accountsubgroup', raise_exception=True)
def account_sub_group_delete(request, pk):

    user = request.user

    logger.info(f"account_sub_group_delete function run by {user}")

    try:
        group = get_object_or_404(AccountSubGroup ,pk=pk)
        group.delete()
        messages.success(request,f'Account Sub Group {group.account_sub_group} was deleted')
        logger.info(f"Account Sub Group {group.account_sub_group} was deleted by {user}")
    except IntegrityError as e:
        messages.error(request,f'Cannot delete {group.account_sub_group} because it is referenced by other objects.')
    return redirect('account_sub_group-create')


def stock_item_create_update(request,pk=None):

    user = request.user

    logger.info(f"stock_item_create_update function run by {user}")

    if not user.has_perm('product.view_stockitem'):
        messages.error(request, "You do not have permission to view stock item")
        return redirect('dashboard-main')
    
    if pk:
        instance = get_object_or_404(StockItem ,pk=pk)
        page_name = 'Edit Stock Item'
    else:
        instance = None
        page_name = 'Add Stock Item'

    if request.user.is_superuser:
        stocks = StockItem.objects.all()
    else:
        stocks = StockItem.objects.filter(company = request.user.company)
   
    accsubgrps = AccountSubGroup.objects.all()

    form = StockItemForm(instance = instance, user = request.user)

    if request.method == 'POST':

        if instance and not user.has_perm('product.change_stockitem'):
            messages.error(request, "You do not have permission to update a stock item.")
            return redirect('stock-item-create')

        if not instance and not user.has_perm('product.add_stockitem'):
            messages.error(request, "You do not have permission to add a stock item.")
            return redirect('stock-item-create')
        
        form = StockItemForm(request.POST ,instance=instance, user=request.user)
        if form.is_valid():
            try:
                form.save()

                if pk:
                    messages.success(request, f"{form.cleaned_data['stock_item_name']} Stock item updated sucessfully")
                    logger.info(f"{form.cleaned_data['stock_item_name']} Stock item updated sucessfully by {user}")
                else:
                    messages.success(request, f"{form.cleaned_data['stock_item_name']} Stock item created sucessfully")
                    logger.info(f"{form.cleaned_data['stock_item_name']} Stock item created sucessfully by {user}")
                return redirect('stock-item-create')
            
            except Exception as e:
                messages.error(request,f'{e}')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)
    
    return render(request,'product/stock_item_create_update.html', {'accsubgrps':accsubgrps,'form':form,'stocks':stocks,'page_name':page_name})


"""
When to Pass the User:
Role-Based Access: If you want to limit the choices in dropdowns or fields based on the user's role (e.g., superusers can see all companies, while regular users can only see their own), passing the user helps implement this.

Custom Logic: If you have logic in the form's __init__ method that depends on the user (like hiding fields or setting default values), you should pass the user.

Validation: If you need to perform validation based on the user’s company or permissions, having the user available can be useful.

When You Might Not Need to Pass the User:
Simple Forms: If your form doesn’t change based on the user and all users should see the same options, you might not need to pass the user at all.

Static Choices: If the form fields can be populated with static choices or querysets that do not depend on the user's identity, you can simply define those in the form without needing to know who the user is.
"""

@permission_required('product.delete_stockitem', raise_exception=True)
def stock_item_delete(request, pk):

    user = request.user

    logger.info(f"stock_item_delete function run by {user}")

    try:

        stock = get_object_or_404(StockItem ,pk=pk)
        stock.delete()
        messages.success(request,f'Stock Item {stock.stock_item_name} was deleted')
        logger.info(f"Stock Item {stock.stock_item_name} was deleted by {user}")

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {stock.stock_item_name} because it is referenced by other objects')        
    return redirect('stock-item-create')


@transaction.atomic
def ledgercreate(request):

    user = request.user

    logger.info(f"ledgercreate function run by {user}")

    if request.path == '/ledgerpopupcreate/':
        if not (user.has_perm('product.add_ledger') or user.has_perm('product.change_ledger')):
            return JsonResponse({'error': 'You do not have permission to add or edit ledger.'})
        
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})

    if not user.has_perm('product.add_ledger'):
        messages.error(request, "You do not have permission to add ledger")
        return redirect('ledger-list')


    if request.path == '/ledgerpopupcreate/':
        template_name = 'accounts/ledger_create_popup.html'
    else:
        template_name = 'accounts/ledger_create_update.html'

    under_groups = AccountSubGroup.objects.all()
    ledgerTypes_query = ledgerTypes.objects.all()

    form = LedgerForm()

    if request.method == 'POST':

        if not user.has_perm('product.add_color'):
            messages.error(request, "You do not have permission to add a ledger.")
            return redirect('ledger-list')

        form = LedgerForm(request.POST)
        
        if form.is_valid():
            ledger_instance = form.save(commit = False) 
            form.save()

            open_bal_value = form.cleaned_data['opening_balance']
            debit_credit_value = form.cleaned_data['Debit_Credit']

            if debit_credit_value == 'Debit':
                account_credit_debit_master_table.objects.create(ledger = ledger_instance, voucher_type = 'Ledger' ,debit= open_bal_value)

            elif debit_credit_value == 'Credit':
                account_credit_debit_master_table.objects.create(ledger = ledger_instance, voucher_type = 'Ledger',credit= open_bal_value)

            elif debit_credit_value == 'N/A':
                account_credit_debit_master_table.objects.create(ledger = ledger_instance, voucher_type = 'Ledger',credit= 0, debit= 0)

            messages.success(request,'Ledger Created successfully')
            logger.info(f"Ledger Created successfully by {user}")
            
            if request.path == '/ledgerpopupcreate/':
                ledger_labour = Ledger.objects.filter(under_group__account_sub_group='Sundry Creditors').values('id','name')         
                return JsonResponse({'ledger_labour':list(ledger_labour)})
            else:
                return redirect('ledger-list')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)
    
    return render(request, template_name,{'form':form,'under_groups':under_groups,'ledgerTypes_query':ledgerTypes_query,'page_name':'Add Leadger'})
    

@transaction.atomic
def ledgerupdate(request,pk):

    user = request.user

    logger.info(f"ledgerupdate function run by {user}")

    if not user.has_perm('product.change_ledger'):
        messages.error(request, "You do not have permission to edit ledger")
        return redirect('ledger-list')
    
    under_groups = AccountSubGroup.objects.all()
    
    Ledger_pk = get_object_or_404(Ledger,pk = pk)

    ledgers = Ledger_pk.transaction_entry.all() 

    Opening_ledger = ledgers.filter(voucher_type ='Ledger').first() 
    form = LedgerForm(instance = Ledger_pk)
    opening_balance = 0

    if form.instance.Debit_Credit == 'Debit':            
        opening_bal = Opening_ledger.debit               
        opening_balance = opening_balance + opening_bal  

    elif form.instance.Debit_Credit == 'Credit':         
        opening_bal = Opening_ledger.credit              
        opening_balance = opening_balance + opening_bal  

    elif form.instance.Debit_Credit == 'N/A': 
        opening_balance = opening_balance 

    else:
        messages.error(request,' Error with Credit Debit ')
    
    if request.method == 'POST':

        if Ledger_pk and not user.has_perm('product.change_ledger'):
            messages.error(request, "You do not have permission to edit ledger")
            return redirect('ledger-list')
        
        form = LedgerForm(request.POST, instance = Ledger_pk)
        name_for_message = request.POST['name']

        if form.is_valid():
            form.save()

            if request.POST['Debit_Credit'] == 'Debit':
                Opening_ledger.debit = request.POST['opening_balance']
                Opening_ledger.credit = 0
                Opening_ledger.save()

            if request.POST['Debit_Credit'] == 'Credit':
                Opening_ledger.credit = request.POST['opening_balance']
                Opening_ledger.debit = 0
                Opening_ledger.save()

            if request.POST['Debit_Credit'] == 'N/A':
                Opening_ledger.credit = 0
                Opening_ledger.debit = 0
                Opening_ledger.save()
            
            messages.success(request, f'Ledger of {name_for_message} Updated')
            logger.info(f"Ledger of {name_for_message} Updated by {user}")
            return redirect('ledger-list')
        else:
            messages.error(f"{form.error}")
            print(form.error)
    
    return render(request,'accounts/ledger_create_update.html',{'form':form,'under_groups':under_groups,'open_bal':opening_balance,'page_name':'Edit Leadger'})


def ledgerlist(request):

    user = request.user

    logger.info(f"ledgerlist function run by {user}")

    if not user.has_perm('product.view_ledger'):
        messages.error(request, "You do not have permission to view ledgers")
        return redirect('dashboard-main')
    
    ledgers = Ledger.objects.select_related('under_group').all()

    return render(request, 'accounts/ledger_list.html', {'ledgers':ledgers,'page_name':'Ledgers'})

@permission_required('product.delete_ledger', raise_exception=True)
def ledgerdelete(request, pk):

    user = request.user

    logger.info(f"ledgerdelete function run by {user}")

    try:
        Ledger_pk = get_object_or_404(Ledger ,pk=pk)
        Ledger_pk.delete()

        messages.success(request ,f'ledger of {Ledger_pk.name} was deleted')
        logger.info(f"ledger of {Ledger_pk.name} was deleted by {user}")

    except Exception as e:
        messages.error(request ,f'Cannot delete {Ledger_pk.name} because it is referenced by other objects.')
    return redirect('ledger-list')


def ledgerTypes_create_update(request,pk=None):

    user = request.user

    logger.info(f"ledgerTypes_create_update function run by {user}")

    if not user.has_perm('product.view_ledgertypes'):
        messages.error(request, "You do not have permission to view Ledger Types")
        return redirect('dashboard-main')
    
    if request.path == '/ledgertypecreatepopup/':
        if not (user.has_perm('product.add_ledgertypes') or user.has_perm('product.change_ledgertypes')):
            return JsonResponse({'error': 'You do not have permission to add or edit Ledger Types.'})

    if request.path == '/ledgertypecreatepopup/':
        if not user.has_perm('product.view_ledgertypes'):
            return JsonResponse({'error': 'You do not have permission to add Ledger Types.'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    

    ledger_types = ledgerTypes.objects.all()

    if request.path == '/ledgertypecreatepopup/':
        template_name = 'accounts/ledgertypecreatepopup.html'
    else:
        template_name = 'accounts/ledgerTypescreateupdate.html'

    if pk:
        type_instance = ledgerTypes.objects.get(pk=pk)
        page_name = 'Edit Ledger Type'
    else:
        type_instance = None
        page_name = 'Add Ledger Type'

    form = ledger_types_form(request.POST or None, instance = type_instance)

    if request.method == 'POST':

        if type_instance and not user.has_perm('product.change_ledgertypes'):
            messages.error(request, "You do not have permission to update a Ledger Type.")
            return redirect('ledger-Types-create')
        
        if not type_instance and not user.has_perm('product.add_ledgertypes'):
            messages.error(request, "You do not have permission to add a Ledger Type.")
            return redirect('ledger-Types-create')
        
        if form.is_valid():
            form.save()

            if request.path == '/ledgertypecreatepopup/':
                ledger_types = ledgerTypes.objects.all().values('id','type_name')
                return JsonResponse({'ledger_type':list(ledger_types)}, status = 200)
            else:
                messages.success(request,f"{form.cleaned_data['type_name']} created successfully")
                logger.info(f"{form.cleaned_data['type_name']} created successfully by {user}")
                return redirect('ledger-Types-create')
        else:
            messages.error(request,f"{form.errors}")
            print(form.errors)

    return render(request,template_name,{'form':form,'ledger_types':ledger_types,'page_name':page_name})


@permission_required('product.delete_ledgertypes', raise_exception=True)
def ledgerTypes_delete(request,pk):

    user = request.user

    logger.info(f"ledgerTypes_delete function run by {user}")

    type_instance = get_object_or_404(ledgerTypes,pk=pk)

    if type_instance:
        type_instance.delete()
        messages.success(request,f"{type_instance.type_name} deleted successfully")
        logger.info(f"{type_instance.type_name} deleted successfully by {user}")
        return redirect('ledger-Types-create')


# =================================== MAIN CATEGORY GODOWNS FUNCTIONS ================================
def godowncreate(request):

    user = request.user

    logger.info(f"godowncreate function run by {user}")

    if request.method == 'POST':
       
        godown_name = request.POST['godown_name']
        godown_type = request.POST['Godown_types']

        if godown_type == 'Raw Material':

            if not user.has_perm('product.add_godown_raw_material'):
                messages.error(request, "You do not have permission to add godown")
                return redirect('godown-list')
    
            try:
                godown_raw = Godown_raw_material(godown_name_raw=godown_name) 
                godown_raw.save() 
                messages.success(request,f"{godown_name} godown created for raw material")
                logger.info(f"{godown_name} godown created for raw material by {user}")

                if 'save' in request.POST:
                    return redirect('godown-list')
                elif 'save_and_add_another' in request.POST:
                    return redirect('godown-create')
                
            except ValidationError as ve:
                messages.error(request,f"{ve}")
                return redirect('godown-list')

            except Exception as e:
                messages.error(request,f"{e}")
                return redirect('godown-list')
            

        elif godown_type == 'Warehouse Finished Goods':

            if not user.has_perm('product.add_finished_goods_warehouse'):
                messages.error(request, "You do not have permission to add warehouse")
                return redirect('godown-list')

            try:
                warehouse = Finished_goods_warehouse(warehouse_name_finished=godown_name) 
                warehouse.save()  
                messages.success(request,f'{godown_name} warehouse created successfully')
                logger.info(f"{godown_name} warehouse created successfully by {user}")

                if 'save' in request.POST:
                    return redirect('godown-list')
                
                elif 'save_and_add_another' in request.POST:
                    return redirect('godown-create')
                
            except ValidationError as ve:
                messages.error(request,f"{ve}")
                return redirect('godown-list')

            except Exception as e:
                messages.error(request,f"{e}")
                return redirect('godown-list')

        elif godown_type == 'Finished Goods':
            
            if not user.has_perm('product.add_godown_finished_goods'):
                messages.error(request, "You do not have permission to add godown")
                return redirect('godown-list')
            
            try:
                godown_finished = Godown_finished_goods(godown_name_finished=godown_name) 

                godown_finished.save() 
                messages.success(request,f"{godown_name} Godown created for finished goods")
                logger.info(f"{godown_name} Godown created for finished goods material by {user}")

                if 'save' in request.POST:
                    return redirect('godown-list')
                elif 'save_and_add_another' in request.POST:
                    return redirect('godown-create')
            
            except ValidationError as ve:
                messages.error(request,f"{ve}")
                return redirect('godown-list')

            except Exception as e:
                messages.error(request,f"{e}")
                return redirect('godown-list')
            
        else:
            messages.error(request,'Error Selecting Godown.')
            return redirect('godown-list')
            
    return render(request,'misc/godown_create.html',{'page_name':"Add Gowdown & Warehouse"})


def godownupdate(request,str,pk):

    user = request.user

    logger.info(f"godownupdate function run by {user}")

    if str == 'finished':

        page_name = "Edit Godown"

        if not user.has_perm('product.change_godown_finished_goods'):
            messages.error(request, "You do not have permission to edit godown")
            return redirect('godown-list')
    
        godown_type = 'Finished Goods'
        finished_godown_pk = get_object_or_404(Godown_finished_goods, pk=pk)
        instance_data = finished_godown_pk.godown_name_finished

        if request.method == 'POST':
            godown_name =  request.POST['godown_name']
            finished_godown_pk.godown_name_finished = godown_name
            finished_godown_pk.save()

            messages.success(request,f'{godown_name} finished goods godown updated.')
            logger.info(f"{godown_name} finished goods godown updated")

            return redirect('godown-list')
        
    elif str == 'raw':

        page_name = "Edit Godown"

        if not user.has_perm('product.change_godown_raw_material'):
            messages.error(request, "You do not have permission to edit godown")
            return redirect('godown-list')
        
        godown_type = 'Raw Material'
        raw_godown_pk = get_object_or_404(Godown_raw_material , pk=pk)
        instance_data = raw_godown_pk.godown_name_raw

        if request.method == 'POST':
            godown_name =  request.POST['godown_name']
            raw_godown_pk.godown_name_raw = godown_name
            raw_godown_pk.save()
            
            messages.success(request,f'{godown_name} raw material godown updated.')
            logger.info(f"{godown_name} raw material godown updated")
            return redirect('godown-list')


    elif str == 'Warehouse':

        page_name = 'Edit Warehouse'

        if not user.has_perm('product.change_finished_goods_warehouse'):
            messages.error(request, "You do not have permission to edit warehouse")
            return redirect('godown-list')
        
        godown_type = 'Warehouse Finished Goods'
        warehouse_pk = get_object_or_404(Finished_goods_warehouse, pk=pk)
        instance_data = warehouse_pk.warehouse_name_finished

        if request.method == 'POST':
            godown_name =  request.POST['godown_name']
            warehouse_pk.warehouse_name_finished = godown_name
            warehouse_pk.save()
            
            messages.success(request,f'{godown_name} warehouse updated.')
            logger.info(f"{godown_name} warehouse updated")
            return redirect('godown-list')
    else:
        messages.error(request,'error in godownupdate str variable')
    
    context = {
        'instance_data': instance_data,
        'godown_type': godown_type,
        'page_name': page_name
    }
    
    return render(request,'misc/godown_update.html', context)


def godownlist(request):

    user = request.user

    logger.info(f"godownlist function run by {user}")

    if not user.has_perm('product.view_godown_raw_material'):
        messages.error(request, "You do not have permission to view godown")
        return redirect('dashboard-main')
    else:
        godowns_raw = Godown_raw_material.objects.all()

    if not user.has_perm('product.view_godown_finished_goods'):
        messages.error(request, "You do not have permission to view godown")
        return redirect('dashboard-main')
    else:
        godowns_finished = Godown_finished_goods.objects.all()

    if not user.has_perm('product.view_finished_goods_warehouse'):
        messages.error(request, "You do not have permission to view warehouse")
        return redirect('dashboard-main')
    else:
        warehouses = Finished_goods_warehouse.objects.all()

    return render(request,'misc/godown_list.html',{'godowns_raw':godowns_raw, 'godowns_finished':godowns_finished,'warehouses':warehouses,"page_name":'Godowns & Warehouse'})


def godowndelete(request,str,pk):

    user = request.user

    logger.info(f"godowndelete function run by {user}")

    if str == 'finished':

        if not user.has_perm('product.delete_godown_finished_goods'):
            messages.error(request, "You do not have permission to delete godown")
            return redirect('godown-list')
    
        try:
            finished_godown_pk = get_object_or_404(Godown_finished_goods, pk=pk)
            finished_godown_pk.delete()
            messages.success(request,f'Finished Goods Godown {finished_godown_pk.godown_name_finished} was deleted')
            logger.info(f"Finished Goods Godown {finished_godown_pk.godown_name_finished} was deleted by {user}")

        except Exception as e:
            messages.error(request,f'Cannot delete {finished_godown_pk.godown_name_finished} because it is referenced by other objects.')
            

    elif str == 'raw':

        if not user.has_perm('product.delete_godown_raw_material'):
            messages.error(request, "You do not have permission to delete godown")
            return redirect('godown-list')
        
        try:
            raw_godown_pk = get_object_or_404(Godown_raw_material, pk=pk)
            raw_godown_pk.delete()
            messages.success(request,f'Raw Material Godown - {raw_godown_pk.godown_name_raw} was deleted')
            logger.info(f"Raw Material Godown - {raw_godown_pk.godown_name_raw} was deleted by {user}")

        except Exception as e:
            messages.error(request,f'Cannot delete {raw_godown_pk.godown_name_raw} because it is referenced by other objects.')
            
    elif str == 'Warehouse':

        if not user.has_perm('product.delete_finished_goods_warehouse'):
            messages.error(request, "You do not have permission to delete warehouse")
            return redirect('godown-list')
        
        try:
            warehouse_pk = get_object_or_404(Finished_goods_warehouse, pk=pk)
            warehouse_pk.delete()
            messages.success(request,f'warehouse {warehouse_pk.warehouse_name_finished} was deleted')
            logger.info(f"{warehouse_pk.warehouse_name_finished} warehouse was deleted by {user}")

        except Exception as e:
            messages.error(request,f'Cannot delete {warehouse_pk.warehouse_name_finished} because it is referenced by other objects.')
    
    else:
        messages.error(request, f'Error Deleting Godowns')
    return redirect('godown-list')


def add_zone_in_warehouse(request,id,zone_id=None):

    user = request.user

    logger.info(f"add_zone_in_warehouse function run by {user}")

    if not user.has_perm('product.view_finished_goods_warehouse_zone'):
        messages.error(request, "You do not have permission to view zones")
        return redirect('dashboard-main')

    zones = finished_goods_warehouse_zone.objects.filter(warehouse_finished_name = id).select_related("warehouse_finished_name")

    warehouse_id = get_object_or_404(Finished_goods_warehouse, id=id)

    warehouses = Finished_goods_warehouse.objects.filter(id = warehouse_id.id).prefetch_related('warehouses','warehouses__zones','warehouses__zones__racks')

    if zone_id:
        zone_instanace = finished_goods_warehouse_zone.objects.get(pk = zone_id)
        form = finished_goods_warehouse_zone_form(request.POST or None, instance=zone_instanace)
        page_name = 'Edit Zone'
    else:
        zone_instanace = None
        form = finished_goods_warehouse_zone_form(request.POST or None, instance=zone_instanace)
        page_name = 'Add Zone'
        
    if request.method == "POST":

        if zone_instanace and not user.has_perm('product.change_finished_goods_warehouse_zone'):
            messages.error(request, "You do not have permission to update a zone.")
            return redirect('add-zone-in-warehouse' , id = warehouse_id.id)

        if not zone_instanace and not user.has_perm('product.add_finished_goods_warehouse_zone'):
            messages.error(request, "You do not have permission to add a zone.")
            return redirect('add-zone-in-warehouse' , id = warehouse_id.id)
        
        form = finished_goods_warehouse_zone_form(request.POST, instance=zone_instanace)
        if form.is_valid():
            zone = form.save(commit=False)
            zone.warehouse_finished_name = warehouse_id
            zone.save()

            if zone_id:
                messages.success(request,f"{zone.zone_name} has been updated")
                logger.info(f"{zone.zone_name} has been updated by {user}")
            else:
                messages.success(request,f"{zone.zone_name} has been created")
                logger.info(f"{zone.zone_name} has been created by {user}")

            return redirect(reverse('add-zone-in-warehouse', args=[id]))
    
    
    return render(request,'finished_product/add_zone_in_warehouse.html',{'form':form,'warehouse_id':warehouse_id,'warehouses':warehouses,'zones':zones,'page_name':page_name})


@permission_required('product.delete_finished_goods_warehouse_zone', raise_exception=True)
def delete_zone_in_warehouse(request,zone_id):

    user = request.user

    logger.info(f"delete_zone_in_warehouse function run by {user}")

    warehouse_id = finished_goods_warehouse_zone.objects.get(id = zone_id).warehouse_finished_name
    zone = finished_goods_warehouse_zone.objects.get(id = zone_id)
    zone.delete()

    messages.success(request,f"{zone.zone_name} has been deleted")
    logger.info(f"{zone.zone_name} has been deleted by {user}")

    return redirect('add-zone-in-warehouse',id=warehouse_id.id)



def add_rack_in_zone(request,zone_id,rack_id=None):

    user = request.user

    logger.info(f"add_rack_in_zone function run by {user}")

    if not user.has_perm('product.view_finished_goods_warehouse_racks'):
        messages.error(request, "You do not have permission to view racks")
        return redirect('dashboard-main')

    racks = finished_goods_warehouse_racks.objects.filter(zone_finished_name=zone_id).select_related("zone_finished_name")

    zone = finished_goods_warehouse_zone.objects.get(id = zone_id)

    if rack_id:
        rack_instance = finished_goods_warehouse_racks.objects.get(pk=rack_id)
        page_name = 'Edit Rack'
        form = finished_goods_warehouse_racks_form(request.POST or None, instance=rack_instance)
    else:
        rack_instance = None
        page_name = 'Add Rack'
        form = finished_goods_warehouse_racks_form()


    if request.method == "POST":

        if rack_instance and not user.has_perm('product.change_finished_goods_warehouse_racks'):
            messages.error(request, "You do not have permission to update a rack.")
            return redirect('add-rack-in-zone' , zone_id = zone_id)

        if not rack_instance and not user.has_perm('product.add_finished_goods_warehouse_racks'):
            messages.error(request, "You do not have permission to add a rack.")
            return redirect('add-rack-in-zone' , zone_id = zone_id)

        form = finished_goods_warehouse_racks_form(request.POST or None, instance=rack_instance)

        if form.is_valid():
            rack = form.save(commit=False)
            rack.zone_finished_name = zone
            rack.save()

            if rack_id:
                messages.success(request,f"{rack.rack_name} has been updated")
                logger.info(f"{rack.rack_name} has been updated by {user}")
            else:
                messages.success(request,f"{rack.rack_name} has been created")
                logger.info(f"{rack.rack_name} has been created by {user}")

            return redirect('add-rack-in-zone', zone_id=zone_id)
    
    return render(request,"finished_product/add_rack_in_zone.html",{'form':form,'racks':racks,'zone':zone,'page_name':page_name})



@permission_required('product.delete_finished_goods_warehouse_racks', raise_exception=True)
def delete_rack_in_zone(request,rack_id):

    user = request.user

    logger.info(f"delete_rack_in_zone function run by {user}")

    rack = finished_goods_warehouse_racks.objects.get(id=rack_id)
    zone = finished_goods_warehouse_racks.objects.get(id=rack_id).zone_finished_name
    rack.delete()

    messages.success(request,f"{rack.rack_name} has been deleted")
    logger.info(f"{rack.rack_name} has been deleted by {user}")

    return redirect('add-rack-in-zone', zone_id=zone.id)



def add_bin_in_rack(request,rack_id,bin_id=None):
    
    user = request.user

    logger.info(f"add_bin_in_rack function run by {user}")

    if not user.has_perm('product.view_finished_product_warehouse_bin'):
        messages.error(request, "You do not have permission to view bins")
        return redirect('dashboard-main')

    bins = finished_product_warehouse_bin.objects.filter(rack_finished_name = rack_id).select_related("rack_finished_name")

    rack = get_object_or_404(finished_goods_warehouse_racks,id=rack_id)

    if bin_id:
        bin_instance = finished_product_warehouse_bin.objects.get(id=bin_id)
        page_name = "Edit Bin"
        form = finished_product_warehouse_bin_form(request.POST or None, instance = bin_instance)
    else:
        bin_instance = None
        page_name = "Add Bin"
        form = finished_product_warehouse_bin_form()

    if request.method == "POST":

        if bin_instance and not user.has_perm('product.change_finished_product_warehouse_bin'):
            messages.error(request, "You do not have permission to update a bin.")
            return redirect('add-bin-in-rack' , rack_id = rack_id)

        if not bin_instance and not user.has_perm('product.add_finished_product_warehouse_bin'):
            messages.error(request, "You do not have permission to add a bin.")
            return redirect('add-bin-in-rack' , rack_id = rack_id)
        
        form = finished_product_warehouse_bin_form(request.POST or None, instance = bin_instance)

        if form.is_valid():
            bin = form.save(commit=False)
            bin.rack_finished_name = rack
            bin.save()

            if bin_id:
                messages.success(request,f"{bin.bin_name} has been updated")
                logger.info(f"{bin.bin_name} has been updated by {user}")
            else:
                messages.success(request,f"{bin.bin_name} has been created")
                logger.info(f"{bin.bin_name} has been created by {user}")

            return redirect('add-bin-in-rack', rack_id=rack_id)
    
    return render(request,"finished_product/add_bin_in_rack.html",{'form':form,'bins':bins,'rack':rack,'page_name':page_name})


@permission_required('product.delete_finished_product_warehouse_bin', raise_exception=True)
def delete_bin_in_rack(request,bin_id):

    user = request.user

    logger.info(f"delete_bin_in_rack function run by {user}")

    rack_id = finished_product_warehouse_bin.objects.get(id=bin_id).rack_finished_name
    bin = get_object_or_404(finished_product_warehouse_bin,id=bin_id)
    bin.delete()

    messages.success(request,f"{bin.bin_name} has been deleted")
    logger.info(f"{bin.bin_name} has been deleted by {user}")

    return redirect('add-bin-in-rack', rack_id = rack_id.id)
   








def stockTrasferRaw(request, pk=None):
    print(request.POST)
    godowns = Godown_raw_material.objects.all()
    source_godown_items_dict = None
    source_godown_items_dict_json = None

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        
        try:
            
            selected_source_godown_id = request.GET.get('selected_godown_id')
            
            
            destination_godowns_queryset = Godown_raw_material.objects.exclude(pk = selected_source_godown_id)
            
            destination_godowns = {}

            for records in destination_godowns_queryset:
                destination_godowns[records.id] = records.godown_name_raw
            
            selected_source_godown_items = item_godown_quantity_through_table.objects.filter(godown_name=selected_source_godown_id)

            
            items_in_godown = {}
            for items in selected_source_godown_items:
                item = items.Item_shade_name
                item_name =  item.items.item_name
                item_id = item.items.id
                items_in_godown[item_id] = item_name

            
            
            item_name_value = request.GET.get('item_value')
            
            
            item_shades_of_selected_item = item_color_shade.objects.filter(items=item_name_value)

            
            item_shades = {}

            
            items_shade_quantity_in_godown = {}

            
            for x in item_shades_of_selected_item:
                
                shades_of_item_in_selected_godown = item_godown_quantity_through_table.objects.filter(godown_name = selected_source_godown_id, Item_shade_name = x.id)
        
                
                
                for x in shades_of_item_in_selected_godown:

                    shade_name = x.Item_shade_name.item_shade_name
                    shade_id = x.Item_shade_name.id
                    item_shades[shade_id] = shade_name

                    
                    item_id = x.Item_shade_name.id
                    items_shade_quantity_in_godown[item_id] = x.quantity


            
            item_color = None
            item_per = None


            if item_name_value is not None:
                item_name_value = int(item_name_value)

                
                items =  get_object_or_404(Item_Creation ,id = item_name_value)
            
                item_color = items.Item_Color.color_name
                item_per = items.unit_name_item.unit_name


            
            shade_quantity = 0
            selected_shade = request.GET.get('selected_shade_id')
            
            if selected_shade is not None and selected_source_godown_id is not None:
                selected_shade = int(selected_shade)
                selected_source_godown_id = int(selected_source_godown_id)

                quantity_get = item_godown_quantity_through_table.objects.filter(Item_shade_name = selected_shade, godown_name = selected_source_godown_id).first()
                shade_quantity  = quantity_get.quantity
            
            
            return JsonResponse({'items_in_godown': items_in_godown, 'item_shades':item_shades,
                                    'item_color':item_color,'item_per':item_per,'items_shade_quantity_in_godown':items_shade_quantity_in_godown,
                                    'shade_quantity':shade_quantity,'destination_godowns':destination_godowns})
        

        except Exception as e:
            messages.error(request, f'An Error occoured {e}')
            logger.error(f'An Error occoured in stock transfer raw{e}')


    if pk:
        raw_transfer_instance = get_object_or_404(RawStockTransferMaster,voucher_no=pk)
        formset  = raw_material_stock_trasfer_items_formset(request.POST or None, instance = raw_transfer_instance)
        source_godown_items = item_godown_quantity_through_table.objects.filter(godown_name = raw_transfer_instance.source_godown.id)
        page_name = 'Edit Stock Transfer'

        source_godown_items_dict = {x.Item_shade_name.items.id : x.Item_shade_name.items.item_name for x in source_godown_items}
        
        source_godown_items_dict_json = json.dumps(source_godown_items_dict)

        # print(source_godown_items_dict)
        


    else:
        source_godown_items = None
        raw_transfer_instance = None
        formset  = raw_material_stock_trasfer_items_formset(request.POST or None,instance = raw_transfer_instance)
        page_name = 'Create Stock Transfer'

    masterstockform = raw_material_stock_trasfer_master_form(request.POST or None, instance = raw_transfer_instance)


    if request.method == 'POST':
        
        formset.forms = [form for form in formset if form.has_changed()]


        if masterstockform.is_valid() and formset.is_valid():
            masterforminstance = masterstockform.save(commit=False)
            masterforminstance.save()
            
            
            for form in formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()

            for form in formset:
                if form.is_valid():
                    if not form.cleaned_data.get('DELETE'):  
                        transfer_instance = form.save(commit=False)
                        transfer_instance.master_instance = masterforminstance 
                        transfer_instance.save()

            return redirect('stock-transfer-raw-list')


    context = {'masterstockform':masterstockform,'formset':formset,'godowns':godowns,'source_godown_items':source_godown_items,'source_godown_items_dict':source_godown_items_dict_json,'page_name':page_name}

    return render(request,'misc/stock_transfer_raw.html', context=context)




def stockTrasferRawList(request):

    stocktrasferall = RawStockTransferMaster.objects.all().order_by('created_date')

    return render(request,'misc/stock_transfer_raw_list.html',{'stocktrasferall':stocktrasferall,'page_name':'Stock Transfer List'})




def stockTrasferRawDelete(request,pk):
    stocktrasferinstance =  get_object_or_404(RawStockTransferMaster,pk = pk)
    stocktrasferinstance.delete()
    return redirect('stock-transfer-raw-list')





def purchase_voucher_rm_with_po_ajax(request):
    try:
        party_name = request.GET.get('partyName')
        open_po = request.GET.get('openPO')
        
        if open_po == "true":

            item_names_search = purchase_order_for_puchase_voucher_rm.objects.filter(master_instance__party_name=party_name).values('item_name','item_name__item_name','item_name__Material_code').annotate(total_qty=Sum('quantity'))

            if item_names_search:

                searched_item_name_dict = {}

                for i in item_names_search:
                    total_qty = str(i['total_qty'])
                    item_name_val = i['item_name__item_name'] + ' | ' + total_qty + ',' + i['item_name__Material_code'] + ',' + 'True'
                    item_id = i['item_name']
                    searched_item_name_dict[item_id] = item_name_val

                logger.info(f"searched result via itemdynamicsearchajax {searched_item_name_dict}")
                
                return JsonResponse({'searched_item_name_dict': searched_item_name_dict}, status=200)

            else:
                return JsonResponse({'error': 'No items found.'}, status=404)


    except Exception as e:
        print(e)





@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def purchasevouchercreateupdate(request, pk = None):
    
    item_name_searched = Item_Creation.objects.all()
    if request.META.get('HTTP_X_REQUESTED_WITH') != 'XMLHttpRequest':
        
        
        if pk:
            purchase_invoice_instance = get_object_or_404(item_purchase_voucher_master,pk=pk)
            item_formsets_change = purchase_voucher_items_formset_update(request.POST or None, instance=purchase_invoice_instance)
            page_name = 'Edit Purchase Invoice'
        
        else:
            purchase_invoice_instance = None
            item_formsets_change = purchase_voucher_items_formset(request.POST or None, instance=purchase_invoice_instance)
            page_name = 'Create Purchase Invoice'

        items_formset = item_formsets_change
        Purchase_gst = gst.objects.all()

        for forms in items_formset.forms:
            godown_items_formset = purchase_voucher_items_godown_formset()

        raw_material_godowns = Godown_raw_material.objects.all()

        master_form  = item_purchase_voucher_master_form(instance=purchase_invoice_instance)
        
        account_sub_grp = AccountSubGroup.objects.filter(account_sub_group__icontains = 'Sundry Creditors').first()
        
        if account_sub_grp is not None:
            
            party_names = Ledger.objects.filter(under_group=account_sub_grp.id)
        else:
            party_names = ''
    
    try:
        party_gst_no = ''
        selected_party_name = request.GET.get('selected_party_name')
        if selected_party_name is not None:
            ledger_instance = Ledger.objects.filter(id = selected_party_name).first()
            party_gst_no = party_gst_no + ledger_instance.Gst_no

        item_value = request.GET.get('item_value')
        
        item_color_out = ''
        item_per_out = ''
        item_gst_out = 0
        item_m_code_out = ''
        
    
        if item_value is not None: 
            item_value = int(item_value)
            item = Item_Creation.objects.get(id = item_value)

            item_color = item.Item_Color.color_name
            item_color_out =  item_color_out + item_color

            item_per = item.unit_name_item.unit_name
            item_per_out = item_per_out + item_per

            item_gst = item.Item_Creation_GST.gst_percentage
            item_gst_out = item_gst_out + item_gst

            item_material_code = item.Material_code
            item_m_code_out = item_m_code_out + item_material_code

        
        
        item_shades = item_color_shade.objects.filter(items = item_value)

        item_shades_dict = {}
        item_shades_total_quantity_dict = {}
        
        for shade in item_shades:
            item_shades_dict[shade.id] = shade.item_shade_name
            
            godown_shade_quantity = 0
            shade_godowns =  item_godown_quantity_through_table.objects.filter(Item_shade_name = shade)
            for godown in shade_godowns:
                godown_shade_quantity = godown_shade_quantity + godown.quantity
            item_shades_total_quantity_dict[shade.id] = godown_shade_quantity
        
        auto_popup_flag = False

        shade_count = len(item_shades_total_quantity_dict)

        if shade_count == 1:
            auto_popup_flag = True

        
    except Exception as e:
        print(f'exception occoured {e}')
    
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
             return JsonResponse({'item_color': item_color_out , 'item_shade': item_shades_dict,'auto_popup_flag':auto_popup_flag,
                                  "item_per":item_per_out, 'item_shades_total_quantity_dict':item_shades_total_quantity_dict,
                                  'item_gst_out':item_gst_out,'party_gst_no':party_gst_no,'item_m_code_out':item_m_code_out})

    if request.method == 'POST':
        try:
            with transaction.atomic(): 
                
                master_form = item_purchase_voucher_master_form(request.POST, instance = purchase_invoice_instance)

                items_formset = item_formsets_change

                godown_items_formset = purchase_voucher_items_godown_formset(request.POST, prefix='shade_godown_items_set')
                
                items_formset.forms = [form for form in items_formset.forms]


                if master_form.is_valid() and items_formset.is_valid():
                    
                    master_instance = master_form.save()
                    
                    # open_po = request.POST.get('opening_po')
                    # print('open_po = ',open_po)
                    
                    
                    for form in items_formset.deleted_forms:
                        
                        if form.instance.pk:
                            
                            form.instance.deleted_directly = True
                            form.instance.delete()

                    all_purchase_temp_data = []
                    
                    for form in items_formset:
                        
                        if form.is_valid():
                            # *****************************
                            from_open_po = form.cleaned_data.get('from_open_po', None)

                            if from_open_po:
                                item_name = form.instance.item_shade.items.item_name
                                qty = form.instance.quantity_total
                                
                                po_instances = purchase_order_for_puchase_voucher_rm.objects.filter(
                                    item_name__item_name=item_name,quantity__gt = 0
                                ).order_by('id')  # Ensure predictable ordering
                                
                                # Step 1: First look for an exact quantity match
                                exact_match_instance = po_instances.filter(quantity=qty).first()
                                if exact_match_instance:
                                    exact_match_instance.quantity -= qty
                                    exact_match_instance.save()
                                    continue  # Move to the next form if exact match is found and processed
                                
                                # Step 2: If no exact match, deduct from other instances
                                remaining_qty = qty
                                po_updated = False
                                
                                for po_instance in po_instances:
                                    if po_instance.quantity >= remaining_qty:
                                        # Deduct only the remaining quantity
                                        po_instance.quantity -= remaining_qty
                                        po_instance.save()
                                        po_updated = True
                                        break
                                    elif po_instance.quantity > 0:
                                        # Deduct what is possible and continue to the next instance
                                        remaining_qty -= po_instance.quantity
                                        po_instance.quantity = 0
                                        po_instance.save()
                                
                                if not po_updated:
                                    # If we couldn't deduct the required quantity, raise an error or create a new PO
                                    raise ValueError(
                                        f"Insufficient quantity available for {item_name}. Please create a new purchase order.",
                                        print(f"Insufficient quantity available for {item_name}. Please create a new purchase order.")
                                    )


                            
                            if not form.cleaned_data.get('DELETE'):
                                

                                items_instance = form.save(commit=False)
                                items_instance.item_purchase_master = master_instance
                                items_instance.save()

                                
                                
                                form_prefix_number = form.prefix[-1] 
                                # print(request.POST)
                                
                                unique_id_no = request.POST.get(f'item_unique_id_{form_prefix_number}')
                                primary_key = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-id')
                            
                                
                                if primary_key == '' or primary_key == None: 
                                    
                                    purchase_voucher_temp_data = shade_godown_items_temporary_table.objects.filter(unique_id=unique_id_no)
                                    
                                    for data in purchase_voucher_temp_data:
                                        all_purchase_temp_data.append(data)
                                    
                                    godown_temp_data = {}
                                    form_set_id = 0
                                    for temp_godown_row_data in purchase_voucher_temp_data:
                                        godown_temp_data[f'shade_godown_items_set-TOTAL_FORMS'] = str(len(purchase_voucher_temp_data))
                                        godown_temp_data[f'shade_godown_items_set-INITIAL_FORMS'] =  str(0)
                                        godown_temp_data[f'shade_godown_items_set-MIN_NUM_FORMS'] =  str(0)
                                        godown_temp_data[f'shade_godown_items_set-MAX_NUM_FORMS'] =  str(1000)
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-godown_id'] = temp_godown_row_data.godown_id
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-quantity'] = temp_godown_row_data.quantity
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-rate'] = temp_godown_row_data.rate
                                        godown_temp_data[f'shade_godown_items_set-{form_set_id}-amount'] = temp_godown_row_data.amount
                                        form_set_id =  form_set_id + 1
                                    
                                    godown_items_formset = purchase_voucher_items_godown_formset(godown_temp_data, prefix='shade_godown_items_set')
                                    
                                    saved_data_to_delete = 0
                                    for godown_form in godown_items_formset:
                                        if godown_form.is_valid():
                                            godown_instance = godown_form.save(commit = False)
                                            
                                            godown_instance.purchase_voucher_godown_item = items_instance 
                                            godown_instance.save()
                                            
                                            saved_data_to_delete = saved_data_to_delete + 1
                                            
                                        else:
                                           
                                            purchase_voucher_temp_data.delete()

                                    if saved_data_to_delete == form_set_id:
                                        purchase_voucher_temp_data.delete()


                                
                                godown_item_quantity = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-jsonDataInputquantity')
                               
                                if godown_item_quantity != '':
                                    voucher_row_godown_data = json.loads(godown_item_quantity)
                                    
                                    parent_row_prefix_id = voucher_row_godown_data.get('parent_row_prefix_id')

                                    if parent_row_prefix_id == form_prefix_number:

                                        new_row = voucher_row_godown_data.get('newRow')
                                        new_rate = float(voucher_row_godown_data.get('all_Rate'))
                                        row_item = items_instance.item_shade.id
                                        Item_instance =  item_color_shade.objects.get(id = row_item)

                                        for key, value in new_row.items():
                                            godown_id = int(value['gId'])    
                                            updated_quantity = value['jsonQty']   
                                            
                                            
                                            godown_old_id = value.get('popup_old_id', None)   
                                            if godown_old_id == '':
                                                godown_old_id = None

                                            if godown_old_id != '' and godown_old_id is not None:
                                                godown_old_id = int(godown_old_id)   

                                            popup_row_id = value.get('popup_row_id', None)  
                                            if popup_row_id == '':
                                                popup_row_id = None
                                        
                                            
                                            if godown_old_id == None or godown_old_id == godown_id:
                                                
                                                godown_instance = Godown_raw_material.objects.get(id = godown_id)
                                                Item, created = item_godown_quantity_through_table.objects.get_or_create(godown_name = godown_instance,Item_shade_name = Item_instance)
                                                
                                                if popup_row_id == None or popup_row_id == '' :
                                                    initial_quantity = 0

                                                else:
                                                    initial_quantity = shade_godown_items.objects.get(pk = popup_row_id)
                                                    initial_quantity = initial_quantity.quantity
                                                
                                                qty_to_update = updated_quantity - initial_quantity

                                                Item.quantity = Item.quantity + qty_to_update
                                                Item.item_rate = new_rate
                                                Item.save()
                                            
                                            
                                            if godown_old_id != None:
                                                
                                                godown_old_id = int(godown_old_id) 
                                                godown_instance_old = Godown_raw_material.objects.get(id = godown_old_id)

                                                godown_new_id = int(godown_id)
                                                godown_instance_new = Godown_raw_material.objects.get(id = godown_new_id)

                                                
                                                if godown_old_id != godown_new_id:
                                                    old_quantity_get = shade_godown_items.objects.get(pk = popup_row_id)
                                                    old_quantity = old_quantity_get.quantity

                                                    old_godown_through_row = item_godown_quantity_through_table.objects.get(godown_name = godown_instance_old,Item_shade_name=Item_instance)

                                                    old_godown_through_row.quantity = old_godown_through_row.quantity - old_quantity
                                                    old_godown_through_row.save()

                                                    new_godown_through_row, created  = item_godown_quantity_through_table.objects.get_or_create(godown_name = godown_instance_new,Item_shade_name=Item_instance)
                                                    
                                                    
                                                    if new_godown_through_row:
                                                        new_quantity_c = new_godown_through_row.quantity
                                                    else:
                                                        
                                                        new_quantity_c = 0

                                                    new_godown_through_row.quantity = new_quantity_c + updated_quantity
                                                    new_godown_through_row.save()


                                
                                popup_godowns_exists = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-popupData')
                                old_item_shade = request.POST.get(f'purchase_voucher_items_set-{form_prefix_number}-old_item_shade')
                                
                                if popup_godowns_exists != '':
                                    popup_godown_data = json.loads(popup_godowns_exists)
                                    
                                    row_prefix_id = popup_godown_data.get('prefix_id')

                                    if row_prefix_id == form_prefix_number:
                                        
                                        shade_id = int(popup_godown_data.get('shade_id'))
                                        prefix_id =  int(popup_godown_data.get('prefix_id'))
                                        primarykey = int(popup_godown_data.get('primary_id'))
                                        old_item_shade = int(old_item_shade)
                                        
                                        
                                        purchasevoucherpopupupdate(popup_godown_data,shade_id,prefix_id,primarykey,old_item_shade)
                                        
                    return redirect('purchase-voucher-list')
                else:
                    
                    if 'temp_data_exists' in request.session and 'temp_uuid' in request.session: 
                        temp_data_exists_bool = request.session['temp_data_exists']
                        temp_uuids = request.session['temp_uuid']
                        del request.session['temp_data_exists']
                        del request.session['temp_uuid']
                        for data in temp_uuids:
                            temp_uuids_data =  shade_godown_items_temporary_table.objects.filter(unique_id=data)
                            temp_uuids_data.delete()

                    return redirect('purchase-voucher-list')
            
        except Exception as e:
            print('an error occoured-test',e)
            messages.error(request,f'An error occoured{e} godown temporary data deleted')
            
        finally:
                
                
                
                if 'temp_data_exists' in request.session and 'temp_uuid' in request.session: 
                    temp_data_exists_bool = request.session['temp_data_exists']
                    temp_uuids = request.session['temp_uuid']
                    del request.session['temp_data_exists']
                    del request.session['temp_uuid']
                    for data in temp_uuids:
                        temp_uuids_data = shade_godown_items_temporary_table.objects.filter(unique_id=data)
                        temp_uuids_data.delete()

    context = {'master_form':master_form,
               'party_names':party_names,
               'items_formset':items_formset,
               'Purchase_gst':Purchase_gst,
               'godown_formsets':godown_items_formset,
               'item_godowns_raw':raw_material_godowns,
               'items':item_name_searched,'page_name' : page_name
               }

    return render(request,'accounts/purchase_invoice.html',context=context)







def purchasevoucherpopupupdate(popup_godown_data,shade_id,prefix_id,primarykey,old_item_shade):
        
        if primarykey is not None:
            voucher_item_instance = purchase_voucher_items.objects.get(id=primarykey)

            if old_item_shade != shade_id:
                all_godown_old_instances = shade_godown_items.objects.filter(purchase_voucher_godown_item = primarykey)
                if all_godown_old_instances:
                    for items in all_godown_old_instances:
                        items.deleted_directly = True

                        
                        items.extra_data_old_shade = old_item_shade
                        items.delete()

            formset = purchase_voucher_items_godown_formset(popup_godown_data, instance = voucher_item_instance ,prefix='shade_godown_items_set')
            
            if formset.is_valid():
                for form in formset.deleted_forms:
                    if form.instance.pk:
                        
                        form.instance.deleted_directly = True
                        form.instance.delete()
                formset.save()
            else:
                print('godown_errors',formset.errors)

                



def purchasevoucherpopup(request,shade_id,prefix_id,unique_id=None,primarykey=None,item_rate=None):
    
    if item_rate != None:
        item_rate_value = decimal.Decimal(item_rate)
    else:
        item_rate_value = None
    
   
    
    
    
    
    if unique_id is not None:
        
        temp_instances = shade_godown_items_temporary_table.objects.filter(unique_id=unique_id)
        
        if temp_instances:
            formsets = shade_godown_items_temporary_table_formset_update(request.POST or None, queryset = temp_instances,prefix='shade_godown_items_set')

        else:
            formsets = shade_godown_items_temporary_table_formset(request.POST or None, queryset = temp_instances,prefix='shade_godown_items_set')
    
    
    elif primarykey is not None:

        godowns_for_selected_shade = shade_godown_items.objects.filter(purchase_voucher_godown_item__item_shade = shade_id,purchase_voucher_godown_item = primarykey)
        
        voucher_item_instance = purchase_voucher_items.objects.get(id=primarykey)
        
        if godowns_for_selected_shade:
            formsets = purchase_voucher_items_godown_formset(instance = voucher_item_instance,prefix='shade_godown_items_set')
        else:
            formsets = purchase_voucher_items_godown_formset_shade_change(
            instance=voucher_item_instance,
            prefix='shade_godown_items_set',
            queryset=godowns_for_selected_shade)

    
    formset = formsets

    try:
        godowns = Godown_raw_material.objects.all()
        item = Item_Creation.objects.get(shades__id = shade_id) 
        item_shade = item_color_shade.objects.get(id = shade_id)

    except Exception as e:
        messages.error(request,'Error with Shades')

    
    if request.method == 'POST':
        formset = formsets

        if formset.is_valid():

            for form in formset.deleted_forms:
                if form.instance.id:
                    form.instance.delete()

            for form in formset:
                if not form.cleaned_data.get('DELETE'):
                    if form.is_valid():
                        form.save()
                    
            
            
            request.session['temp_data_exists'] = True
            temp_uuid = request.session.get('temp_uuid', [])
            temp_uuid.append(unique_id)
            request.session['temp_uuid'] = temp_uuid 

            return HttpResponse('<script>window.close();</script>') 

        else:
            print(formset.errors)
            print(formset.non_form_errors())
            context = {
                'godowns': godowns, 'item': item, 'item_shade': item_shade, 'formset': formset, 'item_rate_value':item_rate_value,
                'unique_id': unique_id, 'shade_id': shade_id, 'errors': formset.errors,'prefix_id':prefix_id, 'primary_key':primarykey
            }
            return render(request, 'accounts/purchase_popup.html', context)
        
    return render(request, 'accounts/purchase_popup.html' ,{'godowns':godowns,'item':item,'shade_id': shade_id,
                                                            'item_shade':item_shade,'formset':formset,'item_rate_value':item_rate_value,
                                                            'unique_id':unique_id,'prefix_id':prefix_id, 'primary_key':primarykey})





def purchasevouchercreategodownpopupurl(request):
    shade_id = request.GET.get('selected_shade')
    unique_id = request.GET.get('unique_invoice_row_id')
    primary_key = request.GET.get('purchase_id')
    prefix_id  = request.GET.get('prefix_id')
    item_instance = item_color_shade.objects.get(id=shade_id)
    fab_grp_instance = Fabric_Group_Model.objects.get(items__shades = shade_id)

    
    query_set_order = item_color_shade.objects.filter(items__Fabric_Group=fab_grp_instance).order_by('-modified_date_time').first()

    item_rate = query_set_order.rate



    
    if primary_key is not None:
        popup_url = reverse('purchase-voucher-popup-update', args=[shade_id,prefix_id,primary_key])
        
    elif unique_id is not None:
        popup_url = reverse('purchase-voucher-popup-create', args=[shade_id,prefix_id,item_rate,unique_id])
        
    else:
        popup_url = None
    

    return JsonResponse({'popup_url':popup_url})





def purchasevoucherlist(request):
    purchase_invoice_list = item_purchase_voucher_master.objects.all().order_by('created_date')
    return render(request,'accounts/purchase_invoice_list.html',{'purchase_invoice_list':purchase_invoice_list,'page_name':'Purchase List'})





def purchasevoucherdelete(request,pk):
    purchase_invoice_pk = get_object_or_404(item_purchase_voucher_master,pk=pk)
    purchase_invoice_pk.delete()
    return redirect('purchase-voucher-list')
                    



def product2item(request,product_refrence_id):

    user = request.user

    logger.info(f"product2item function run by {user}")
    
    if not user.has_perm('product.view_product_2_item_through_table'):
        messages.error(request, "You do not have permission to view set production")
        return redirect('dashboard-main')
        
    try:

        all_ref_ids = Product.objects.all()
        
        items = Item_Creation.objects.all().order_by('item_name')
        product_refrence_no = product_refrence_id
        model_name = Product.objects.get(Product_Refrence_ID=product_refrence_id)

        Products_all = PProduct_Creation.objects.filter(Product__Product_Refrence_ID=product_refrence_id).select_related('PProduct_color')

        if not Products_all.exists():
                raise ValueError("No products found for the given reference ID.")
        
        
        extraformspecial = True
        for product in Products_all:
            if product.product_2_item_through_table_set.filter(common_unique = False):
                extraformspecial = False

        
        extraformcommon = True
        for product in Products_all:
            if product.product_2_item_through_table_set.filter(common_unique = True):
                extraformcommon = False
        

        product2item_instances = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_refrence_id,
              common_unique = False).select_related('PProduct_pk','Item_pk','PProduct_pk__PProduct_color').order_by('row_number')
        
        if extraformspecial:
            formset_single = Product2ItemFormsetExtraForm(queryset = product2item_instances, prefix='product2itemuniqueformset')

        else:
            formset_single = Product2ItemFormset(queryset=product2item_instances , prefix = 'product2itemuniqueformset')

        distinct_product2item_commmon_instances = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_refrence_id,common_unique = True).order_by(
                'row_number','id').distinct('row_number').select_related('PProduct_pk','Item_pk')

        if extraformcommon:
            formset_common = Product2CommonItemFormSetExtraForm(queryset=distinct_product2item_commmon_instances,prefix='product2itemcommonformset')
        else:
            formset_common = Product2CommonItemFormSet(queryset=distinct_product2item_commmon_instances,prefix='product2itemcommonformset')


        clone_ajax_valid = False
        if extraformcommon and extraformspecial:
            clone_ajax_valid = True

        if request.method == 'POST':
            # print(request.POST)

            if not (user.has_perm('product.change_product_2_item_through_table') or user.has_perm('product.add_product_2_item_through_table') or user.has_perm('product.delete_product_2_item_through_table')):
                messages.error(request, "You do not have permission to Add or update a set production.")
                return redirect('dashboard-main')
            

            formset_single = Product2ItemFormset(request.POST, queryset=product2item_instances, prefix='product2itemuniqueformset')
            formset_common = Product2CommonItemFormSet(request.POST, queryset=distinct_product2item_commmon_instances, prefix='product2itemcommonformset') 
            
            formset_single_valid = False
            formset_common_valid = False
            
            if formset_single.is_valid():
                try:
                    for form in formset_single.deleted_forms:
                        if form.instance.pk:  
                            form.instance.delete()
                            
                    for form in formset_single:
                        if not form.cleaned_data.get('DELETE'): 
                            if form.cleaned_data.get('Item_pk'):                                  
                                if form.instance.pk:  
                                    existing_instance = product_2_item_through_table.objects.get(pk=form.instance.pk)  
                                    initial_rows = existing_instance.no_of_rows 
                                else:
                                    initial_rows = 0

                                p2i_instance = form.save(commit = False)
                                p2i_instance.c_user = request.user
                                p2i_instance.common_unique = False 
                                p2i_instance.save()

                                logger.info(f"Product to item created/updated special - {p2i_instance.id}")

                                no_of_rows_to_create = form.cleaned_data['no_of_rows'] - initial_rows   
                                p2i_instance.row_number = form.cleaned_data['row_number']

                                if no_of_rows_to_create > 0:
                                    for row in range(no_of_rows_to_create):
                                        logger.info(f" set prod item part name created of p2i instance - {p2i_instance.id}")
                                        set_prod_item_part_name.objects.create(producttoitem = p2i_instance, c_user = request.user)

                                p2i_instance.save()
                                formset_single_valid = True

                            else:
                                raise ValidationError('Please select existing Item Name or select from the dropdown')
                                
                except Exception as e:
                    logger.error(f'Error saving unique records - {e}')
                    messages.error(request, f'Error saving unique records - {e}')  
            
            else:
                logger.error(f'Error saving unique records - {formset_single.errors}')
                messages.error(request, f'Error saving unique records - {formset_single.errors}') 
                            
            if formset_common.is_valid():
                try:
                    for form in formset_common.deleted_forms:
                        if form.instance.id: 
                            deleted_item = form.instance.Item_pk  
                            
                            for product in Products_all: 
                                p2i_to_delete = product_2_item_through_table.objects.filter(PProduct_pk=product, Item_pk=deleted_item, common_unique=True)
                                logger.info(f"Deleted product to item instace of {product}, - {deleted_item}")
                                p2i_to_delete.delete()
                            
                            
                    for form in formset_common:
                        
                        if not form.cleaned_data.get('DELETE'): 

                            if form.cleaned_data.get('Item_pk'):  

                                for product in Products_all:

                                    old_item = form.initial.get('Item_pk')

                                    if old_item:
                                        item = old_item
                                        print("old",old_item)
                                    else:
                                        item = form.cleaned_data['Item_pk']
                                        print(item)

                                    obj, created = product_2_item_through_table.objects.get_or_create(PProduct_pk=product, Item_pk=item, common_unique=True)
                                    obj.c_user = request.user
                                    
                                    if created:
                                        initial_rows = 0

                                    if not created:
                                        initial_rows = obj.no_of_rows
                                        
                                    obj.Item_pk = form.cleaned_data['Item_pk']
                                    obj.no_of_rows = form.cleaned_data['no_of_rows']
                                    obj.Remark = form.cleaned_data['Remark']
                                    obj.row_number = form.cleaned_data['row_number']
                                    logger.info(f"Product to item created/updated common - {obj.id}")
                                    obj.save()
                                
                                    
                                    rows_to_create = form.cleaned_data['no_of_rows'] - initial_rows
                                    if rows_to_create > 0:
                                            for row in range(rows_to_create):
                                                set_prod_item_part_name.objects.create(producttoitem = obj,c_user = request.user)
                                                logger.info(f" set prod item part name created of - {obj.id}")

                                    formset_common_valid = True
                            else:
                                raise ValidationError('Please select existing Item Name or select from the dropdown')
                                    
                except Exception as e:
                    logger.error(f'Error saving common records - {e}')
                    messages.error(request, f'Error saving common records{e}.') 
            else:
                logger.error(f'Error saving unique records - {formset_common.errors}')
                messages.error(request, f'Error saving unique records - {formset_common.errors}')
        

            if formset_common_valid and formset_single_valid:

                messages.success(request,'Items to Product sucessfully added.')
                close_window_script = """<script>
                window.opener.location.reload(true);  // Reload parent window if needed
                window.close();  // Close current window
                </script>"""

                return HttpResponse(close_window_script)
            
            else:
                for form_errors in formset_common.errors:
                    if form_errors:
                        logger.error(f'Error with formset_common form - {product_refrence_id } - {form_errors}')
                        messages.error(request, f'{form_errors}')

                for form_errors in formset_single.errors:
                    if form_errors:
                        logger.error(f'Error with formset_common form - {product_refrence_id } - {form_errors}')
                        messages.error(request, f'{form_errors}')


                close_window_script = """
                            <script>
                                                window.opener.location.reload(true);  // Reload parent window if needed
                                                window.close();  // Close current window
                                                </script>
                                                """
                return HttpResponse(close_window_script)


        return render(request,'production/product2itemsetproduction.html', {'formset_single':formset_single,'formset_common':formset_common,
                                                                'Products_all':Products_all,'all_ref_ids':all_ref_ids,'clone_ajax_valid':clone_ajax_valid,
                                                                'items':items,'product_refrence_no': product_refrence_no,'model_name':model_name})

    except Exception as e:
        logger.error(f'Error with forms - {product_refrence_id } - {e}')
        messages.error(request, 'An unexpected error occurred. Please try again later.')
        return render(request, 'production/product2itemsetproduction.html', {
            'formset_single': formset_single,
            'formset_common': formset_common,
            'Products_all': Products_all,
            'items': items,
            'product_refrence_no': product_refrence_no,
            'model_name':model_name
        })



def export_Product2Item_excel(request,product_ref_id):
    
    user = request.user

    logger.info(f"export_Product2Item_excel function run by {user}")

    if not user.has_perm('product.view_product_2_item_through_table'):
        return JsonResponse({'error': 'You do not have permission to download set production.'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    try:
        
        products_in_i2p_special = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_ref_id, common_unique = False).order_by(
            'row_number')
    
        products_in_i2p_common = product_2_item_through_table.objects.filter(
            PProduct_pk__Product__Product_Refrence_ID=product_ref_id, common_unique = True).order_by(
            'row_number', 'id').distinct('row_number')

        if not products_in_i2p_special and not products_in_i2p_common:
            logger.error(f"Add Items to the respective Products first - {product_ref_id}")
            raise ObjectDoesNotExist("Add Items to the respective Products first")

        wb = Workbook()

        default_sheet = wb['Sheet']
        wb.remove(default_sheet)    

        wb.create_sheet('product_special_configs')
        wb.create_sheet('product_common_configs')

        sheet1 = wb.worksheets[0]
        sheet2 = wb.worksheets[1]

        column_widths = [10, 40, 20, 30, 20, 15, 12 ,12, 12, 12]  

        for i, column_width in enumerate(column_widths, start=1):  
            col_letter = get_column_letter(i)
            sheet1.column_dimensions[col_letter].width = column_width
        
        for i, column_width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(i)
            sheet2.column_dimensions[col_letter].width = column_width

        headers =  ['id','item name', 'product sku','part name', 'part dimention','dimention total','part pieces','body/combi','grand_total', 'combi_total']
        sheet1.append(headers)

        row_count_to_unlock_total = 1
        for product in products_in_i2p_special:
            grand_total_parent = product.grand_total
            grand_total_combi_parent = product.grand_total_combi

            rows_to_insert_s1 = []

            for product_configs in product.product_item_configs.all().order_by('id'):
                rows_to_insert_s1.append([
                product_configs.id,
                product_configs.producttoitem.Item_pk.item_name,
                product_configs.producttoitem.PProduct_pk.PProduct_SKU,
                product_configs.part_name,
                product_configs.part_dimentions,
                product_configs.dimention_total,
                product_configs.part_pieces,
                product_configs.body_combi
            ])

            row_count_to_unlock = 1

            for row in rows_to_insert_s1:
                sheet1.append(row)
                row_count_to_unlock = row_count_to_unlock + 1

            row_count_to_unlock_total =  row_count_to_unlock_total + row_count_to_unlock

            
            sheet1.append(['','','','','','','','', grand_total_parent , grand_total_combi_parent])
        
            rows_to_insert_s1.clear()

        
        for row in sheet1.iter_rows(min_row=2, max_row=row_count_to_unlock_total, min_col=4, max_col=8):
            for cell in row:
                cell.protection = Protection(locked = False)

        
        headers =  ['id','item name','part name', 'part dimention', 'dimention total','part pieces','body/combi','grand_total','combi_total']
        sheet2.append(headers)

        row_count_to_unlock_total_common = 1
        for product in products_in_i2p_common:
            grand_total_parent = product.grand_total
            grand_total_combi_parent = product.grand_total_combi

            rows_to_insert_s2 = []
            for product_configs in product.product_item_configs.all().order_by('id'):
                rows_to_insert_s2.append([
                product_configs.id,
                product_configs.producttoitem.Item_pk.item_name,
                product_configs.part_name,
                product_configs.part_dimentions,
                product_configs.dimention_total,
                product_configs.part_pieces,
                'body',
            ])

            row_count_to_unlock = 1

            for row in rows_to_insert_s2:
                sheet2.append(row)
                row_count_to_unlock = row_count_to_unlock + 1
            row_count_to_unlock_total_common = row_count_to_unlock_total_common + row_count_to_unlock

            
            sheet2.append(['','','','','','','',grand_total_parent, grand_total_combi_parent])

            rows_to_insert_s2.clear()

        
        for row in sheet2.iter_rows(min_row=2, max_row=row_count_to_unlock_total_common, min_col=3, max_col=7):
            for cell in row:
                cell.protection = Protection(locked = False)

        
        sheet1.protection.sheet = True
        sheet2.protection.sheet = True

        fileoutput = BytesIO()
        wb.save(fileoutput)
        
        
        response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name_with_pk = f'product_reference_id_{product_ref_id}'
        response['Content-Disposition'] = f'attachment; filename="{file_name_with_pk}.xlsx"'

        logger.info(f"{file_name_with_pk}.xlsx download by {user}")

        return response

        

    except product_2_item_through_table.DoesNotExist:
        messages.error(request, 'No data found for the given product reference ID.')
        logger.error(f"items for product Does not exists for refrence id {product_ref_id}")
        return redirect(reverse('edit_production_product', args=[product_ref_id]))
    
    
    except Exception as e:
        messages.error(request, f'An error occurred while exporting data {e}')
        logger.error(f"An exception occoured while exporting data - {e} for ref id {product_ref_id}")
        return redirect(reverse('edit_production_product', args=[product_ref_id]))




def viewproduct2items_configs(request, product_sku):

    user = request.user

    logger.info(f"viewproduct2items_configs function run by {user}")

    ref_id = PProduct_Creation.objects.get(PProduct_SKU = product_sku).Product.Product_Refrence_ID
    if not user.has_perm('product.view_product_2_item_through_table'):
        messages.error(request, "You do not have permission to view configs")
        return redirect(reverse('edit_production_product',args=[ref_id]))

    try:
        product2item_instances = product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU=product_sku).order_by('row_number')
        product2item_instances_first = product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU=product_sku).first()


        context = {
            'product2item_instances': product2item_instances,
            'product2item_instances_first': product2item_instances_first
        }

        return render(request, 'production/product2itemsconfigview.html', context)

    except product_2_item_through_table.DoesNotExist:
        return render(request, 'production/product2itemsconfigview.html', {
            'error_message': f'No items found for product SKU: {product_sku}'})

    except DatabaseError as e:
        return HttpResponseServerError(f'A database error occurred: {e}')

    except Exception as e:
        return HttpResponseServerError(f'An unexpected error occurred: {e}')
    


@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def purchaseordercreateupdate(request,pk=None):
    
    try:
        ledger_party_names = Ledger.objects.filter(under_group__account_sub_group = 'Sundry Debtors')
        products = Product.objects.all()

        raw_material_godowns = Godown_raw_material.objects.all()

        
        if pk:
            instance = get_object_or_404(purchase_order,pk=pk)
            model_name = instance.product_reference_number.Model_Name
            model_images = instance.product_reference_number.productdetails
            page_name = 'Edit Order'
        else:
            instance = None
            model_name = None
            model_images = None
            page_name = 'New Order'

        formset = purchase_order_product_qty_formset(instance=instance)
        form = purchase_order_form(instance=instance)

    except Exception as e:
        logger.error(f'An Exception Occoured {e}')

    
    if request.method == 'POST':

        
        
        if 'submit-form-1' in request.POST:
            form = purchase_order_form(request.POST, instance=instance)
        
            if form.is_valid():
                try:
                    form_instance = form.save(commit=False)
                    form_instance.balance_number_of_pieces = form.instance.number_of_pieces
                    form_instance.save()
                    logger.info(f'Purchase invoice created-updated-{form.instance.id}')
                    
                    return redirect(reverse('purchase-order-update', args=[form.instance.id]))
                
                except ValidationError as val_err:
                    logger.error(f'Validation error: {val_err} - {form.errors}')

                except DatabaseError as db_err:
                    logger.error(f'Database error during formset save: {db_err}')

                except Exception as e:
                    logger.error(f'Unexpected error during form save: {e}')
            else:
                logger.error(f'Purchase Order Quantities updated error-{form.instance.id} - {form.errors}')
            

        if 'submit-form-2' in request.POST:
            
            try:
                formset = purchase_order_product_qty_formset(request.POST, instance=instance)
                

                
                
                
                

                
                if formset.is_valid():
                    try:
                        formset.save()
                        
                        for form in formset:
                            if not form.cleaned_data.get('DELETE'):
                                p_o_instance = form.instance.purchase_order_id  
                                p_o_instance.purchase_order_to_product_saved = True
                                p_o_instance.save()
                                if p_o_instance.process_status == '1':  
                                    p_o_instance.process_status = '2'  
                                    p_o_instance.save()  
                            
                            
                        messages.success(request, 'Purchase Order Quantities updated successfully.')
                        logger.info(f'Purchase Order Quantities updated-{form.instance.id}')

                        
                        return redirect('purchase-order-raw-material-list')

                    except DatabaseError as db_err:

                        logger.error(f'Database error during form save: {db_err}')
                        messages.error(request, 'A database error occurred during form save.')

                    except Exception as e:
                        logger.error(f'Unexpected error during form save: {e}')
                        messages.error(request, 'An unexpected error occurred during form save.')
                else:
                    
                    print(formset.errors)
                    logger.error(f'Purchase Order Quantities update error - {formset.errors} {formset.non_form_errors()}')
                    messages.error(request, f'There were errors in the form. Please correct them and try again.')

            except ValidationError as val_err:
                logger.error(f'Validation error: {val_err}-{formset.errors}')
                messages.error(request, f'Validation error: {val_err}')

            except Exception as e:
                logger.error(f'Exception occurred: {e}')
                messages.error(request, f'An exception occurred: {e}') 
        

    return render(request,'production/purchaseordercreateupdate.html',{'form':form ,'formset':formset,'raw_material_godowns':raw_material_godowns,
                                                                          'ledger_party_names':ledger_party_names,
                                                                          "products":products,'model_name':model_name,'model_images':model_images,
                                                                          'page_name':page_name})




def purchaseorderlist(request):
    purchase_orders = purchase_order.objects.all().select_related('ledger_party_name','product_reference_number').prefetch_related('raw_materials').order_by('created_date')
    return render(request,'production/purchaseorderlist.html',{'purchase_orders': purchase_orders,'page_name':'Order List'})




def purchaseorderdelete(request,pk):

    try:
        instance = get_object_or_404(purchase_order, pk = pk)
        instance.delete()
        logger.info(f"Purchase Order with order no - {instance.purchase_order_number} was deleted")
        messages.success(request,f'Purchase order with order no - {instance.purchase_order_number} was deleted')
        
    except Exception as e:
        messages.error(request,f'Cannot delete {instance.purchase_order_number} - {e}.')
        logger.error(f"Cannot delete {instance.purchase_order_number} - {e}.")
    return redirect('purchase-order-list')
     

def excel_download_production(request, module_name, pk):

    print("excel_download_production")
    if module_name is not None and pk is not None:  

        file_name = None
        
        
        thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"))
        
        if module_name == 'purchase_order_raw':
            wb = Workbook()

            
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)  
            wb.create_sheet('purchase_order_raw')

            sheet = wb.worksheets[0]

            file_name = 'purchase_order_raw'

            column_widths = [13, 16, 10, 30, 13, 10, 10, 15, 15]  

            
            for i, column_width in enumerate(column_widths, start=1):  
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = column_width

            purchase_order_instance = purchase_order.objects.get(pk=pk)
            
            sheet.cell(row=1, column=1).value = f'PO No:  {purchase_order_instance.purchase_order_number}'
            
            sheet.cell(row=1, column=1).font = Font(bold=True)
            
            sheet.cell(row=1, column=1).border = thin_border
            sheet.cell(row=1, column=2).border = thin_border
            sheet.cell(row=1, column=3).border = thin_border
            sheet.cell(row=1, column=4).border = thin_border

            sheet.cell(row=2, column=1).value = f'Date:  {purchase_order_instance.created_date.replace(tzinfo=None).strftime("%d %B %Y")}'
            
            sheet.cell(row=2, column=1).font = Font(bold=True)

            sheet.cell(row=2, column=1).border = thin_border
            sheet.cell(row=2, column=2).border = thin_border
            sheet.cell(row=2, column=3).border = thin_border
            sheet.cell(row=2, column=4).border = thin_border

            sheet.cell(row=3, column=1).value = f'Ref No: { purchase_order_instance.product_reference_number.Product_Refrence_ID}'
            
            sheet.cell(row=3, column=1).font = Font(bold=True)

            sheet.cell(row=3, column=1).border = thin_border
            sheet.cell(row=3, column=2).border = thin_border
            sheet.cell(row=3, column=3).border = thin_border
            sheet.cell(row=3, column=4).border = thin_border

            sheet.cell(row=4, column=1).value = f'Name:  {purchase_order_instance.product_reference_number.Model_Name}'
            
            sheet.cell(row=4, column=1).font = Font(bold=True)
           
            sheet.cell(row=4, column=1).border = thin_border
            sheet.cell(row=4, column=2).border = thin_border
            sheet.cell(row=4, column=3).border = thin_border
            sheet.cell(row=4, column=4).border = thin_border

            sheet.cell(row=5, column=1).value = f'Party Name:  {purchase_order_instance.ledger_party_name.name}'
            
            sheet.cell(row=5, column=1).font = Font(bold=True)

            sheet.cell(row=5, column=1).border = thin_border
            sheet.cell(row=5, column=2).border = thin_border
            sheet.cell(row=5, column=3).border = thin_border
            sheet.cell(row=5, column=4).border = thin_border

            sheet.cell(row=6, column=1).value = f'Total PO Qty: {purchase_order_instance.number_of_pieces}'
           
            sheet.cell(row=6, column=1).font = Font(bold=True)

            sheet.cell(row=6, column=1).border = thin_border
            sheet.cell(row=6, column=2).border = thin_border
            sheet.cell(row=6, column=3).border = thin_border
            sheet.cell(row=6, column=4).border = thin_border

            sheet.cell(row=1, column=5).value = 'Product SKU'
            sheet.cell(row=1, column=6).value = 'Color'
            sheet.cell(row=1, column=8).value = 'Proc Qty'
            
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=8).font = Font(bold=True)
            
            sheet.cell(row=1, column=3).border = thin_border
            sheet.cell(row=1, column=4).border = thin_border
            sheet.cell(row=1, column=5).border = thin_border
            sheet.cell(row=1, column=6).border = thin_border
            sheet.cell(row=1, column=7).border = thin_border
            

            
            start_row = 2
            start_column = 5

            product_2_item_qs = purchase_order_instance.p_o_to_products.all()

            for index, instance in enumerate(product_2_item_qs.order_by('id'), start = start_row): 

                sheet.cell(row=index, column=start_column).value = instance.product_id.PProduct_SKU
                sheet.cell(row=index, column=start_column).border = thin_border
                sheet.cell(row=index, column=start_column + 1).value = instance.product_id.PProduct_color.color_name
                sheet.cell(row=index, column=start_column + 1).font = Font(bold=True)
                sheet.cell(row=index, column=start_column + 1).border = thin_border
                sheet.cell(row=index, column=start_column + 2).border = thin_border
                sheet.cell(row=index, column=start_column + 3).value = instance.order_quantity
                sheet.cell(row=index, column=start_column + 3).border = thin_border
                
                product_img_instance = instance.product_id.PProduct_image

                if product_img_instance:

                    product_img = product_img_instance.url

                    relative_path = str(product_img).lstrip('/media/') 
                
                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))

                
                
                
            
                
                if os.path.isfile(full_image_path):

                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    

                    
                    excel_img.width = 55  
                    excel_img.height = 50  

                    
                    

                    img_position = f'I{index}'

                    
                    sheet.add_image(excel_img, img_position)

                else:
                    sheet.cell(row=index, column=start_column + 3).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    

            length_queryset = len(product_2_item_qs)
            
            row_num = 2 + length_queryset

            sheet.cell(row=row_num, column = 6).value = 'Total'
            sheet.cell(row=row_num, column = 8).value = purchase_order_instance.number_of_pieces
            sheet.cell(row=row_num, column = 5).font = Font(bold=True)
            sheet.cell(row=row_num, column = 8).font = Font(bold=True)
            sheet.cell(row=row_num, column = 6).border = thin_border
            sheet.cell(row=row_num, column = 8).border = thin_border
            sheet.cell(row=row_num, column = 7).border = thin_border
            

            
            if length_queryset < 6:

                start_row_items = 9

            else:
                start_row_items = length_queryset + 4

            start_column_items = 1 

            header_row = start_row_items - 1
            
            headers = ["Body/Combi", "Product Color", "Pcs", "Material Name", "Rate", "Panha","Units", "Consmp","Combi-2-Consump","T-Consmp","Physical Stock","Bal Stock"]

            
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text=True)

            for index, instance in enumerate(purchase_order_instance.raw_materials.all().order_by('id'), start=start_row_items):
                sheet.cell(row=index, column = start_column_items).value = instance.Remark

                if instance.product_color != 'Common Item':
                    p_color = instance.product_color

                else:
                    p_color = ''

                sheet.cell(row=index, column=start_column_items + 1).value = p_color 
                sheet.cell(row=index, column=start_column_items + 2).value = instance.pcs
                sheet.cell(row=index, column=start_column_items + 3).value = instance.material_name
                sheet.cell(row=index, column=start_column_items + 4).value = instance.rate
                sheet.cell(row=index, column=start_column_items + 5).value = instance.panha
                sheet.cell(row=index, column=start_column_items + 6).value = instance.units
                sheet.cell(row=index, column=start_column_items + 7).value = instance.consumption
                sheet.cell(row=index, column=start_column_items + 8).value = instance.combi_consumption
                sheet.cell(row=index, column=start_column_items + 9).value = instance.total_comsumption
                sheet.cell(row=index, column=start_column_items + 10).value = instance.physical_stock
                sheet.cell(row=index, column=start_column_items + 11).value = instance.balance_physical_stock

            qs_length_list = len(purchase_order_instance.raw_materials.all())
            sheet.cell(row=qs_length_list + start_row_items, column=2).value =  'Narration'       
            sheet.cell(row=qs_length_list + start_row_items, column=3).value =  purchase_order_instance.note       
            sheet.cell(row=qs_length_list + start_row_items, column=2).font = Font(bold = True)
            sheet.cell(row=qs_length_list + start_row_items, column=3).font = Font(bold = True)




        elif module_name == 'purchase_order_cutting':
            wb = Workbook()

            
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)  
            wb.create_sheet('Cutting_Order')
            wb.create_sheet('Cutting_size')

            sheet = wb.worksheets[0]
            sheet1 = wb.worksheets[1]

            file_name = 'purchase_order_cutting'

            column_widths = [16, 20, 15, 15, 15, 15, 15, 15, 15]  

            column_widths_sheet_2 = [30, 25, 20, 20, 20, 20, 12, 12, 12]

            
            for i, column_width in enumerate(column_widths, start=1):  
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = column_width

            for i, column_width in enumerate(column_widths_sheet_2, start=1):  
                col_letter = get_column_letter(i)
                sheet1.column_dimensions[col_letter].width = column_width

        
            purchase_order_cutting_instance = purchase_order_raw_material_cutting.objects.get(raw_material_cutting_id=pk)
            
            sheet.cell(row=1, column=1).value = f'P O No:  {purchase_order_cutting_instance.purchase_order_id.purchase_order_number}'

            sheet.cell(row=2, column=1).value = f'Cutting No:  {purchase_order_cutting_instance.raw_material_cutting_id}'
            
            sheet.cell(row=3, column=1).value = f'Date:  {purchase_order_cutting_instance.created_date.replace(tzinfo=None).strftime("%d %B %Y")}'  
            
            sheet.cell(row=4, column=1).value = f'Cutter Name:  {purchase_order_cutting_instance.factory_employee_id.factory_emp_name}'
            
            sheet.cell(row=5, column=1).value = f'Prod Ref No:  {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Product_Refrence_ID}'

            sheet.cell(row=6, column=1).value = f'Model Name:  {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Model_Name}'
        
            sheet.cell(row=7, column=1).value = f'Processed Qty:  {purchase_order_cutting_instance.processed_qty}'
            
            sheet.cell(row=9, column=1).value = f'Party Name:  {purchase_order_cutting_instance.purchase_order_id.ledger_party_name.name}'
           
            sheet.cell(row=10, column=1).value = f'Total PO Qty:  {purchase_order_cutting_instance.purchase_order_id.number_of_pieces}'
            
            sheet.cell(row=11, column=1).value = f'Bal Qty:  {purchase_order_cutting_instance.balance_qty}'
            
            sheet.cell(row=12, column=1).value = f'Target Date:  {purchase_order_cutting_instance.purchase_order_id.target_date.strftime("%d %B %Y")}'



            for index in range(1,13):
            
                sheet.cell(row=index, column=1).font = Font(bold=True)
            
                sheet.cell(row=index, column=1).border = thin_border
                sheet.cell(row=index, column=2).border = thin_border
                sheet.cell(row=index, column=3).border = thin_border

                
            sheet.cell(row=1, column=4).value = 'Product SKU'
            sheet.cell(row=1, column=4).font = Font(bold=True)
            sheet.cell(row=1, column=4).border = thin_border

            sheet.cell(row=1, column=5).value = 'Color'
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=5).border = thin_border

            sheet.cell(row=1, column=6).value = 'Cutting Qty'
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=6).border = thin_border

            
            
            
            start_row = 2
            start_column = 4  

            for index, instance in enumerate(purchase_order_cutting_instance.purchase_order_to_product_cutting_set.all().order_by('id'), start=start_row):

                product_creation_instance = PProduct_Creation.objects.get(PProduct_SKU = instance.product_sku)

                product_img_instance = product_creation_instance.PProduct_image

                if product_img_instance:
                    product_img = product_img_instance.url
                    relative_path = str(product_img).lstrip('/media/') 
                
                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))


                
                if os.path.isfile(full_image_path):
                
                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    
                    
                    excel_img.width = 60  
                    excel_img.height = 55  

                    
                    
                    img_position = f'G{index}'

                    
                    sheet.add_image(excel_img, img_position)

                else:

                    sheet.cell(row=index, column=start_column + 3).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    


                sheet.cell(row=index, column=start_column).value = instance.product_sku
                sheet.cell(row=index, column=start_column).font = Font(bold=True)
                sheet.cell(row=index, column=start_column).border = thin_border


                sheet.cell(row=index, column=start_column + 1).value = instance.product_color
                sheet.cell(row=index, column=start_column + 1).font = Font(bold=True)
                sheet.cell(row=index, column=start_column + 1).border = thin_border


                sheet.cell(row=index, column=start_column + 2).value = instance.cutting_quantity
                sheet.cell(row=index, column=start_column + 2).font = Font(bold=True)
                sheet.cell(row=index, column=start_column + 2).border = thin_border



            qs_length = len(purchase_order_cutting_instance.purchase_order_to_product_cutting_set.all())

            sheet.cell(row=qs_length + 2, column = 5).value = 'Total'
            sheet.cell(row=qs_length + 2, column = 5).border = thin_border
            sheet.cell(row=qs_length + 2, column = 5).font = Font(bold=True)
            sheet.cell(row=qs_length + 2, column = 6).value = purchase_order_cutting_instance.processed_qty
            sheet.cell(row=qs_length + 2, column = 6).border = thin_border
            sheet.cell(row=qs_length + 2, column = 6).font = Font(bold=True)


            if  qs_length < 11:

                
                start_row_items = 14

            else:

                start_row_items = qs_length + 4

            start_column_items = 1 

            header_row = start_row_items - 1

            
            headers = ["Body/Combi", "Product Color","Pcs","Material Name", 'Shade Color',"Panha","Units","Consump","Combi-2-Consump","Total Consump","Physical Stock","Balance Stock"]

            
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text=True)


            cutting_items_qs = purchase_order_cutting_instance.purchase_order_for_raw_material_cutting_items_set.filter(material_color_shade__items__Fabric_nonfabric='Fabric')

            for index, instance in enumerate(cutting_items_qs.order_by('id'), start=start_row_items):
                sheet.cell(row=index, column=start_column_items).value = instance.Remark


                if instance.product_color != 'Common Item':
                    p_color = instance.product_color

                else:
                    p_color = ''

                sheet.cell(row=index, column=start_column_items + 1).value = p_color
                sheet.cell(row=index, column=start_column_items + 2).value = instance.pcs
                sheet.cell(row=index, column=start_column_items + 3).value = instance.material_name
                sheet.cell(row=index, column=start_column_items + 4).value = instance.material_color_shade.item_shade_name
                
                sheet.cell(row=index, column=start_column_items + 5).value = instance.panha
                sheet.cell(row=index, column=start_column_items + 6).value = instance.units
                sheet.cell(row=index, column=start_column_items + 7).value = instance.consumption
                sheet.cell(row=index, column=start_column_items + 8).value = instance.combi_consumption
                sheet.cell(row=index, column=start_column_items + 9).value = instance.total_comsumption
                sheet.cell(row=index, column=start_column_items + 10).value = instance.physical_stock
                sheet.cell(row=index, column=start_column_items + 11).value = instance.balance_physical_stock


            qs_length_list = len(cutting_items_qs)

            sheet.cell(row=qs_length_list + start_row_items, column=2).value =  'Narration :'       
            sheet.cell(row=qs_length_list + start_row_items, column=3).value =  purchase_order_cutting_instance.note
            sheet.cell(row=qs_length_list + start_row_items, column=2).font = Font(bold=True)
            sheet.cell(row=qs_length_list + start_row_items, column=3).font = Font(bold=True)


            

            sheet1.cell(row=1, column=1).value = f'Ref No - {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Product_Refrence_ID}'
            sheet1.cell(row=2, column=1).value = f'Model Name - {purchase_order_cutting_instance.purchase_order_id.product_reference_number.Model_Name}'

            sheet1.cell(row=1, column=1).font = Font(bold=True)
            sheet1.cell(row=1, column=1).border = thin_border

            sheet1.cell(row=2, column=1).font = Font(bold=True)
            sheet1.cell(row=2, column=1).border = thin_border
            
            sheet1.cell(row = 1, column=2).value = 'Product SKU'
            sheet1.cell(row = 1, column=2).font = Font(bold=True)
            sheet1.cell(row = 1, column=2).border = thin_border

            sheet1.cell(row = 1, column=3).value = 'color'
            sheet1.cell(row = 1, column=3).font = Font(bold=True)
            sheet1.cell(row = 1, column=3).border = thin_border

            sheet1.cell(row = 1, column=4).font = Font(bold=True)
            sheet1.cell(row = 1, column=4).border = thin_border

             
            start_row = 2
            start_column = 2

            purchase_order_cutting_p_2_item_qs = purchase_order_cutting_instance.purchase_order_to_product_cutting_set.all()

            for index, instance in enumerate(purchase_order_cutting_p_2_item_qs.order_by('id'), start=start_row):

                product_creation_instance = PProduct_Creation.objects.get(PProduct_SKU = instance.product_sku)

                product_img_instance = product_creation_instance.PProduct_image

                if product_img_instance:
                    product_img = product_img_instance.url
                    relative_path = str(product_img).lstrip('/media/') 
                
                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))


                
                if os.path.isfile(full_image_path):
                    
                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    
                    
                    excel_img.width = 60  
                    excel_img.height = 55  

                    
                    
                    img_position = f'D{index}'

                    
                    sheet1.add_image(excel_img, img_position)

                else:

                    sheet1.cell(row=index, column=start_column + 2).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    

                sheet1.cell(row=index, column=start_column).value = instance.product_sku
                sheet1.cell(row=index, column=start_column).font = Font(bold=True)
                sheet1.cell(row=index, column=start_column).border = thin_border

                sheet1.cell(row=index, column=start_column + 1).value = instance.product_color
                sheet1.cell(row=index, column=start_column + 1).font = Font(bold = True)
                sheet1.cell(row=index, column=start_column + 1).border = thin_border

                
                product_ref_no = purchase_order_cutting_instance.purchase_order_id.product_reference_number.Product_Refrence_ID

                items_qs = product_2_item_through_table.objects.filter(
                    PProduct_pk__Product__Product_Refrence_ID = product_ref_no , Remark = 'BODY')


                
                for index, instance in enumerate(items_qs, start = 1):
                    sheet1.cell(row=index , column=5).value = instance.Item_pk.item_name
                    sheet1.cell(row=index , column=5).font = Font(bold=True)

            
            if len(purchase_order_cutting_p_2_item_qs) < len(items_qs) :
                start_row_items = len(items_qs) + 2
            
            else:
                start_row_items = len(purchase_order_cutting_p_2_item_qs) + 3
                
            start_column_items = 4 

            header_row = start_row_items - 1

            
            headers = ["Item Name", "Part Name", "Part Dimentions", 'Dimention Total','Part Pcs','Body/Combi','Grand Total','Grand Total Combi']

            
            for col_num, header in enumerate(headers, start = 1):
                sheet1.cell(row = header_row, column = col_num).value = header
                sheet1.cell(row = header_row, column = col_num).font = Font(bold=True)
                sheet1.cell(row = header_row, column = col_num).alignment = Alignment(wrap_text = True)

            start_column_items = 1

            product_with_combi = [] 

            combi_found = False

            
            for product in product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID = product_ref_no):

                if product.product_item_configs.filter(body_combi='combi').exists():
                    product_with_combi.append(product)
                    combi_found = True
            
            
            if not combi_found:

                body_instance = product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID = product_ref_no).first()

                product_with_combi.append(body_instance)
                

            for instance in product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU = product_with_combi[0].PProduct_pk.PProduct_SKU).order_by('row_number'):
                list_to_append = []
                for record in instance.product_item_configs.all().order_by('id'):
                    list_1 = [
                        record.producttoitem.Item_pk.item_name,
                        record.part_name,
                        record.part_dimentions,
                        record.dimention_total,
                        record.part_pieces,
                        record.body_combi
                        ]
                    
                    list_to_append.append(list_1)

                for row in list_to_append:
                    sheet1.append(row)

                sheet1.append(['','','','','','', instance.grand_total, instance.grand_total_combi])
                    

        elif module_name == 'labour_workout':

            wb = Workbook()

            
            default_sheet = wb['Sheet']

            wb.remove(default_sheet)  

            wb.create_sheet('Labour_workout')
            
            sheet = wb.worksheets[0]

            file_name = 'labour_workout'

            
            column_widths = [35, 20, 12, 10, 15, 30, 15, 15, 15]  

            
            for i, column_width in enumerate(column_widths, start=1):  
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = column_width

            labour_workout_instance = labour_workout_childs.objects.get(id=pk)

            sheet.cell(row=1, column=1).value = f'PO No. : {labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number}'
            
            sheet.cell(row=2, column=1).value = f'Challan No : {labour_workout_instance.challan_no}, Date : {labour_workout_instance.created_date.replace(tzinfo=None).strftime("%d %B %Y")}'
            
            sheet.cell(row=3, column=1).value = f'Vendor Name : {labour_workout_instance.labour_name.name}'
            
            sheet.cell(row=4, column=1).value = f'Ref No : {labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID}'
            
            sheet.cell(row=5, column=1).value = f'Model Name : {labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name}'

            sheet.cell(row=6, column=1).value = f'No of Pcs : {labour_workout_instance.total_process_pcs}'
            

            for cell in range(1,7):
                sheet.cell(row=cell, column=1).border = thin_border
                sheet.cell(row=cell, column=2).border = thin_border
                sheet.cell(row=cell, column=3).border = thin_border

            sheet.cell(row=1, column=4).value = 'Product SKU'
            sheet.cell(row=1, column=4).font = Font(bold=True)
            sheet.cell(row=1, column=4).border = thin_border

            sheet.cell(row=1, column=5).value = 'Color'
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=5).border = thin_border

            sheet.cell(row=1, column=6).value = 'Process Qty'
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=6).border = thin_border

            
            start_row = 2 
            start_column = 4

            labour_workout_p2i_items_qs = labour_workout_instance.labour_workout_child_items.all()

            for index, instance in enumerate(labour_workout_p2i_items_qs.order_by('id'), start = start_row):
                sheet.cell(row=index, column=start_column).value = instance.product_sku
                
                sheet.cell(row=index, column=start_column).border = thin_border

                sheet.cell(row=index, column=start_column + 1).value = instance.product_color
                
                sheet.cell(row=index, column=start_column + 1).border = thin_border

                sheet.cell(row=index, column=start_column + 2).value = instance.processed_pcs
                
                sheet.cell(row=index, column=start_column + 2).border = thin_border

                product_creation_instance = PProduct_Creation.objects.get(PProduct_SKU = instance.product_sku)

                product_img_instance = product_creation_instance.PProduct_image

                if product_img_instance:

                    product_img = product_img_instance.url
                    relative_path = str(product_img).lstrip('/media/') 

                else:
                    relative_path = 'pproduct/images/Unknown_pic.png'

                
                
                full_image_path = os.path.join(settings.MEDIA_ROOT, str(relative_path))


                
                if os.path.isfile(full_image_path):
                    try:
                        img = PILImage.open(full_image_path)
                        img.verify()  

                    except Exception as e:
                        print(f"Error opening image: {e}")
                        exit()

                    
                    excel_img = Image(full_image_path)

                    
                    
                    excel_img.width = 60  
                    excel_img.height = 55  

                    
                    col_letter =  get_column_letter(index) 
                    img_position = f'G{index}'

                    
                    sheet.add_image(excel_img, img_position)

                else:
                    sheet.cell(row=index, column = start_column + 3).value = None
                    print(f"Image file does not exist at path: {full_image_path}")
                    

            qs_length = len(labour_workout_p2i_items_qs)

            sheet.cell(row=qs_length + 2, column = 4).value = 'Total'
            sheet.cell(row=qs_length + 2, column = 4).border = thin_border
            sheet.cell(row=qs_length + 2, column = 4).font = Font(bold=True)

            sheet.cell(row=qs_length + 2, column = 5).value = labour_workout_instance.total_process_pcs
            sheet.cell(row=qs_length + 2, column = 5).border = thin_border
            sheet.cell(row=qs_length + 2, column = 5).font = Font(bold=True)

            
            for index, instance in enumerate(product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.
                product_reference_number.Product_Refrence_ID, Remark='BODY'), start = 8):
                
                sheet.cell(row=index , column=1).value = instance.Item_pk.item_name
                
            len_items = len(product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.
                product_reference_number.Product_Refrence_ID, Remark='BODY'))

            
            if qs_length < len_items + 8:

                start_row_items = len_items + 9
            else:
                start_row_items = qs_length + 4
                
            header_row = start_row_items - 1

            
            
            headers = ["Item Name", "Part Name", "Total Part Pcs"]

            
            for col_num, header in enumerate(headers, start = 1):
                sheet.cell(row = header_row, column = col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text = True)

            product_with_combi = [] 

            combi_found = False

            
            for product in product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.
                purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID):
                
                if product.product_item_configs.filter(body_combi = 'combi').exists():
                    product_with_combi.append(product)
                    combi_found = True

            
            if not combi_found:
                body_instance = product_2_item_through_table.objects.filter(
                PProduct_pk__Product__Product_Refrence_ID=labour_workout_instance.labour_workout_master_instance.
                purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID).first()

                product_with_combi.append(body_instance)

            rows_iterated_outer = 0

            for record in product_2_item_through_table.objects.filter(PProduct_pk__PProduct_SKU = product_with_combi[0].PProduct_pk.PProduct_SKU,Item_pk__Fabric_nonfabric = 'Fabric').order_by('id'):
                list_to_append = []    

                rows_iterated_inner = 0
                item_name_last = ''
                for instance in record.product_item_configs.all().order_by('id'):

                    if instance.part_pieces is None or instance.part_pieces == 0 :

                        part_pcs = 0 
                    else:
                        part_pcs =  instance.part_pieces

                    if item_name_last == record.Item_pk.item_name:
                        item_name_present = ''

                    else:
                        item_name_present = record.Item_pk.item_name

                    list = [
                        item_name_present,
                        instance.part_name,
                        part_pcs * labour_workout_instance.total_process_pcs
                        ]
                    
                    rows_iterated_inner = rows_iterated_inner + 1

                    item_name_last = record.Item_pk.item_name

                    list_to_append.append(list)

                for x in list_to_append:
                    sheet.append(x)

                rows_iterated_outer = rows_iterated_outer + rows_iterated_inner + 1
                sheet.append(['', '' , ''])
            
            start_column_items = 4 

            
            headers = ["Body/Combi", "Pcs", "Material Name", "Total Consump"]

            
            for col_num, header in enumerate(headers, start=4):
                sheet.cell(row=header_row, column=col_num).value = header
                sheet.cell(row=header_row, column=col_num).font = Font(bold=True)
                sheet.cell(row=header_row, column=col_num).alignment = Alignment(wrap_text=True)


            for index, instance in enumerate(labour_workout_instance.labour_workout_cutting_items_set.all().order_by('id'), start = start_row_items):
                sheet.cell(row=index, column=start_column_items).value = instance.Remark
                sheet.cell(row=index, column=start_column_items + 1).value = instance.pcs
                sheet.cell(row=index, column=start_column_items + 2).value = instance.material_name
                sheet.cell(row=index, column=start_column_items + 3).value = instance.total_comsumption


            narration_row = rows_iterated_outer + len_items + 8
        
            sheet.cell(row=narration_row, column=2).value = f'Narration: {labour_workout_instance.note}'
            sheet.cell(row=narration_row, column=2).font = Font(bold=True)

            for cell in range(1,narration_row):
                sheet.cell(row=cell, column=1).font = Font(bold=True)

        fileoutput = BytesIO()
        wb.save(fileoutput)
            
        
        response = HttpResponse(fileoutput.getvalue(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name_with_pk = f'product_reference_id_{file_name}'
        response['Content-Disposition'] = f'attachment; filename="{file_name_with_pk}.xlsx"'

        return response
    
    else:

        return HttpResponse('INVALID ENTRY')





@login_required(login_url = 'login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def purchaseorderrawmaterial(request ,p_o_pk, prod_ref_no):
    
    purchase_order_instance = purchase_order.objects.get(pk=p_o_pk)

    form = purchase_order_form(instance = purchase_order_instance)
    
    product_refrence_no = prod_ref_no

    product_2_items_instances_unique = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no, common_unique = False).order_by(
                                'row_number')
    
    product_2_items_instances_common = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no, common_unique = True).order_by(
                                'row_number','id').distinct('row_number')
    

    
    product_2_items_instances = product_2_items_instances_unique.union(product_2_items_instances_common)
    
    
    model_name = purchase_order_instance.product_reference_number.Model_Name

    physical_stock_all_godowns = {}

    for item in product_2_items_instances:
        item_id = item.Item_pk
        item_name = item.Item_pk.item_name
        item_quantity = 0
        item_godowns = item_godown_quantity_through_table.objects.filter(Item_shade_name__items = item_id,
                                        godown_name = purchase_order_instance.temp_godown_select) 
        
        if item_godowns:
            for query in item_godowns:
                item_quantity = item_quantity + query.quantity
                physical_stock_all_godowns[item_name] = str(item_quantity)

        else:
            physical_stock_all_godowns[item_name] = str(item_quantity)

    purchase_order_raw_formset = purchase_order_raw_product_qty_formset(instance = purchase_order_instance)

    
    
    if not purchase_order_instance.raw_materials.all():
        
        physical_stock_all_godown_json = json.dumps(physical_stock_all_godowns) 

        initial_data = []
        for query in product_2_items_instances:
            rate_first = query.Item_pk.shades.order_by('id').first() 

            if query.common_unique == True:
                product_color_or_common_item = 'Common Item'
                product_sku_or_common_item = 'Common Item'

            else:
                product_color_or_common_item = query.PProduct_pk.PProduct_color
                product_sku_or_common_item = query.PProduct_pk.PProduct_SKU
            
            initial_data_dict = {'product_sku': product_sku_or_common_item,
                                'product_color' : product_color_or_common_item,
                                'material_name':query.Item_pk.item_name,
                                'rate':rate_first.rate,
                                'panha':query.Item_pk.Panha,
                                'units':query.Item_pk.Units,
                                'g_total':query.grand_total,
                                'g_total_combi':query.grand_total_combi,
                                'consumption':'0',
                                'combi_consumption':0,
                                'total_comsumption':'0',
                                'unit_value': query.Item_pk.unit_name_item.unit_name,
                                'physical_stock':'0',
                                'balance_physical_stock':'0',
                                'row_number':query.row_number,
                                'Remark':query.Remark, 
                                'pcs': '0' }
            
            initial_data.append(initial_data_dict)

        initial_sorted_data = sorted(initial_data, key=itemgetter('row_number'), reverse=False)

        
        purchase_order_raw_product_sheet_formset = inlineformset_factory(purchase_order, purchase_order_for_raw_material, form=purchase_order_raw_product_sheet_form, extra=len(initial_sorted_data) if initial_data else 0, can_delete=False)

        purchase_order_raw_sheet_formset = purchase_order_raw_product_sheet_formset(initial=initial_sorted_data, instance=purchase_order_instance)


    
    elif purchase_order_instance.raw_materials.all():

        physical_stock_all_godown_json = None

        purchase_order_raw_product_sheet_formset = inlineformset_factory(purchase_order, purchase_order_for_raw_material, form=purchase_order_raw_product_sheet_form, extra=0, can_delete=False)
        
        purchase_order_raw_sheet_formset = purchase_order_raw_product_sheet_formset(instance=purchase_order_instance)


    if request.method == 'POST':
        
        purchase_order_raw_formset = purchase_order_raw_product_qty_formset(request.POST)

        purchase_order_raw_sheet_formset = purchase_order_raw_product_sheet_formset(request.POST, instance=purchase_order_instance)

        try:
            if purchase_order_raw_formset.is_valid() and purchase_order_raw_sheet_formset.is_valid():
                
                try:
                    with transaction.atomic():
                        
                        purchase_order_raw_formset.save()
                        purchase_order_raw_sheet_formset.save()
                        
                        note_post = request.POST.get('note')

                        if note_post:
                            purchase_order_instance.note = note_post
                        else:
                            purchase_order_instance.note = None
                            
                        purchase_order_instance.save()


                        for form in purchase_order_raw_sheet_formset:
                            po_form_instance = form.instance.purchase_order_id  
                            if po_form_instance.process_status == '2':   
                                po_form_instance.process_status = '3'  
                                po_form_instance.save()  

                        
                        return redirect('purchase-order-raw-material-list')
                
                except ValueError as ve:
                    messages.error(request,f'Error Occured - {ve}')

                except Exception as e:
                    messages.error(request,f'Exception Occured - {e}')
                
                return render(request,'production/purchaseorderrawmaterial.html',{'form': form ,'model_name':model_name,
                                                                        'purchase_order_raw_formset':purchase_order_raw_formset,
                                                                        'purchase_order_raw_sheet_formset':purchase_order_raw_sheet_formset,
                                                                        'physical_stock_all_godown_json':physical_stock_all_godown_json,'page_name':'Purchase Order View'})

            
            else:
                print(purchase_order_raw_formset.errors,purchase_order_raw_sheet_formset.errors)
                raise ValidationError("No products found for the given reference ID.")
                
  
                
        except ValidationError as ve:
                messages.error(request,f' Please enter correct Procurement color wise QTY {ve}')
        
        except Exception as e:
            messages.error(request,f' An exception occoured {e}')



    return render(request,'production/purchaseorderrawmaterial.html',{'form':form, 'model_name':model_name,
                                                                      'purchase_order_raw_formset':purchase_order_raw_formset,
                                                                      'purchase_order_raw_sheet_formset':purchase_order_raw_sheet_formset,
                                                                      'physical_stock_all_godown_json':physical_stock_all_godown_json,'page_name':'Purchase Order View'})






def purchase_order_for_raw_material_list(request):
    
    purchase_orders_pending = purchase_order.objects.annotate(raw_material_count=Count('raw_materials')).filter(raw_material_count__lt=1, purchase_order_to_product_saved=True).order_by('created_date')

    purchase_orders_completed = purchase_order.objects.annotate(raw_material_count=Count('raw_materials')).filter(raw_material_count__gt=0).order_by('created_date')

    return render(request,'production/purchase_order_for_raw_material_list.html',{'purchase_orders_pending': purchase_orders_pending,'purchase_orders_completed':purchase_orders_completed,'page_name':'Purchase Order List'})




def purchase_order_for_raw_material_delete(request,pk):

    try:
        purchase_order_raw_instances = purchase_order_for_raw_material.objects.filter(purchase_order_id=pk)

        purchase_object = get_object_or_404(purchase_order,pk=pk)
        purchase_object.process_status = 2
        purchase_object.save()

        for instances in purchase_order_raw_instances:
            instances.delete()
        messages.success(request,'Deleted Successfully')
        return redirect('purchase-order-raw-material-list')
    
    except Exception as e:
        messages.error(request,f'Error While Deleting purchase Order {e}')
        return redirect('purchase-order-raw-material-list')

 










@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def purchaseordercuttingcreateupdate(request,p_o_pk,prod_ref_no,pk=None):

    if pk:
        purchase_order_cutting_instance = get_object_or_404(purchase_order_raw_material_cutting, pk=pk)
    else:
        purchase_order_cutting_instance = None

    labour_all = factory_employee.objects.all()

    
    purchase_order_instance = get_object_or_404(purchase_order, pk=p_o_pk)

    
    purchase_order_raw_instances = purchase_order_for_raw_material.objects.filter(purchase_order_id=p_o_pk).order_by('id')

    
    purchase_order_to_product_instances = purchase_order_to_product.objects.filter(purchase_order_id = p_o_pk)

    
    form = purchase_order_form(instance = purchase_order_instance)

    current_godown = form.instance.temp_godown_select

    
    purchase_order_cutting_form = purchase_order_raw_material_cutting_form(request.POST or None, instance=purchase_order_cutting_instance)

    
    if not pk:

        product_refrence_no = prod_ref_no

    
        product_2_items_instances_unique = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no,common_unique = False).order_by(
                                'row_number')
    
        product_2_items_instances_common = product_2_item_through_table.objects.filter(
                            PProduct_pk__Product__Product_Refrence_ID = product_refrence_no,common_unique = True).order_by(
                                'row_number','id').distinct('row_number')


        product_2_items_instances = product_2_items_instances_unique.union(product_2_items_instances_common)

        
        initial_data = []

        for purchase_items_raw in product_2_items_instances:
            
            current_godown_qty = item_color_shade.objects.filter(items=purchase_items_raw.Item_pk,godown_shades__godown_name=purchase_order_instance.temp_godown_select).annotate(total_godown_qty = Sum('godown_shades__quantity'))

            rate_first = purchase_items_raw.Item_pk.shades.order_by('id').first() 

            shade_single_list = []  
            
            
            for qs in purchase_items_raw.Item_pk.shades.all():
                try:
                    
                    shade_godown_qty = item_godown_quantity_through_table.objects.get(
                    godown_name=purchase_order_instance.temp_godown_select,
                    Item_shade_name=qs.id)

                    
                    shade_data = {
                    'shade_id': qs.id,
                    'shade_name': qs.item_shade_name,
                    'godown_quantity': shade_godown_qty.quantity}  
                        
                    shade_single_list.append(shade_data)

                except item_godown_quantity_through_table.DoesNotExist:
                    

                    shade_data = {
                        'shade_id': qs.id,
                        'shade_name': qs.item_shade_name,
                        'godown_quantity': 0 }
                    
                    shade_single_list.append(shade_data)


            if purchase_items_raw.common_unique == True:
                product_color_or_common_item = 'Common Item'
                product_sku_or_common_item = 'Common Item'

            else:
                product_color_or_common_item = purchase_items_raw.PProduct_pk.PProduct_color
                product_sku_or_common_item = purchase_items_raw.PProduct_pk.PProduct_SKU

            current_godown_qty_total = 0

            for qs in current_godown_qty:
                current_godown_qty_total = current_godown_qty_total + qs.total_godown_qty

            
            initial_data_dict = {
                'product_sku': product_sku_or_common_item,
                'product_color' : product_color_or_common_item,
                'material_name' : purchase_items_raw.Item_pk.item_name,
                'material_color_shade': shade_single_list, 
                'fabric_non_fab': purchase_items_raw.Item_pk.Fabric_nonfabric, 
                'rate' : rate_first.rate,
                'panha' : purchase_items_raw.Item_pk.Panha,
                'units' :  purchase_items_raw.Item_pk.Units,
                'g_total' : purchase_items_raw.grand_total,
                'g_total_combi': purchase_items_raw.grand_total_combi,
                'consumption' : '0',
                'total_comsumption' :'0',
                'combi_consumption' : '0',
                'unit_value' : purchase_items_raw.Item_pk.unit_name_item.unit_name,
                'physical_stock' : current_godown_qty_total,
                'balance_physical_stock': '0',
                'row_number': purchase_items_raw.row_number,
                'Remark': purchase_items_raw.Remark,
                'pcs': '0'}
            
            initial_data.append(initial_data_dict)

       
        initial_sorted_data = sorted(initial_data, key = itemgetter('row_number'), reverse=False)

        
        
        purchase_order_for_raw_material_cutting_items_formset = inlineformset_factory(purchase_order_raw_material_cutting, 
                                                                                      purchase_order_for_raw_material_cutting_items, 
                                                                                      form=purchase_order_for_raw_material_cutting_items_form,
                                                                                      formset=Basepurchase_order_for_raw_material_cutting_items_form, 
                                                                                      extra=len(initial_data), 
                                                                                      can_delete=False)


        
        purchase_order_for_raw_material_cutting_items_formset_form = purchase_order_for_raw_material_cutting_items_formset(initial=initial_sorted_data)


        
        initial_data_p_o_to_items = []

        for instances in purchase_order_to_product_instances:
            initial_data_dict = {
                'product_color': instances.product_id.PProduct_color,
                'product_sku': instances.product_id.PProduct_SKU,
                'order_quantity': instances.order_quantity,
                'process_quantity': instances.process_quantity,
                'cutting_quantity': '0',
            }

            initial_data_p_o_to_items.append(initial_data_dict)

        
        purchase_order_to_product_formset = inlineformset_factory(purchase_order_raw_material_cutting, 
                                                                  purchase_order_to_product_cutting,
                                                                    form=purchase_order_to_product_cutting_form,
                                                                      extra=len(initial_data_p_o_to_items),
                                                                      can_delete=False)

        purchase_order_to_product_formset_form = purchase_order_to_product_formset(initial= initial_data_p_o_to_items)

    
    elif pk:

        purchase_order_for_raw_material_cutting_items_formset = inlineformset_factory(purchase_order_raw_material_cutting,
                                                                                       purchase_order_for_raw_material_cutting_items,
                                                                                         form=purchase_order_for_raw_material_cutting_items_form, 
                                                                                         extra=0, can_delete=False)
        
        purchase_order_for_raw_material_cutting_items_formset_form = purchase_order_for_raw_material_cutting_items_formset(instance=purchase_order_cutting_instance)


        
        purchase_order_to_product_formset = inlineformset_factory(purchase_order_raw_material_cutting, purchase_order_to_product_cutting, 
                                                                  form=purchase_order_to_product_cutting_form, extra=0, can_delete=False)
        
        purchase_order_to_product_formset_form = purchase_order_to_product_formset(instance=purchase_order_cutting_instance)
        
    if request.method == 'POST':
        
        
        purchase_order_for_raw_material_cutting_items_formset_form = purchase_order_for_raw_material_cutting_items_formset(request.POST)

        
        purchase_order_to_product_formset_form = purchase_order_to_product_formset(request.POST)

        if purchase_order_cutting_form.is_valid() and purchase_order_to_product_formset_form.is_valid() and purchase_order_for_raw_material_cutting_items_formset_form.is_valid():
            try:
                with transaction.atomic():
                    
                    cutting_form_instance = purchase_order_cutting_form.save()
                    cutting_form_instance.purchase_order_id.cutting_total_processed_qty = cutting_form_instance.purchase_order_id.cutting_total_processed_qty + cutting_form_instance.processed_qty
                    cutting_form_instance.purchase_order_id.save()
                    
                    
                    if cutting_form_instance.purchase_order_id.process_status == '3':
                        cutting_form_instance.purchase_order_id.process_status = '4'
                        cutting_form_instance.purchase_order_id.save()
                    

                    
                    for form in purchase_order_to_product_formset_form:

                        if form.is_valid(): 
                            try:
                                p_o_to_order_form_instance = form.save(commit = False)
                                p_o_to_order_form_instance.purchase_order_cutting_id = cutting_form_instance
                                p_o_to_order_form_instance.save()

                                
                                product_sku = p_o_to_order_form_instance.product_sku
                                processed_qty = p_o_to_order_form_instance.cutting_quantity

                                p_o_id = p_o_to_order_form_instance.purchase_order_cutting_id.purchase_order_id
                            
                                purchase_order_products = purchase_order_to_product.objects.filter(purchase_order_id =p_o_id,product_id =product_sku).first()

                                if purchase_order_products:
                                    purchase_order_products.process_quantity =  purchase_order_products.process_quantity - processed_qty
                                    purchase_order_products.save()

                                else:
                                    logger.error(f'Product {product_sku} not found in purchase order {p_o_id}')
                                    messages.error(request, f'Product {product_sku} not found in the purchase order.')

                            except Exception as e:
                                logger.error(f'Error processing product form: {e}')
                                messages.error(request, f'An error occurred while processing the product form: {e}')


                    
                    for form in purchase_order_for_raw_material_cutting_items_formset_form:
                        if form.is_valid(): 
                            try:
                                form_instance = form.save(commit=False)
                                form_instance.purchase_order_cutting = cutting_form_instance
                                form_instance.total_comsumption_in_cutting = form_instance.total_comsumption
                                form_instance.entry_from_cutting_room = True
                                form_instance.save()
                            

                            except ValidationError as ve:
                                logger.error(f'Error with: {e}')

                            except Exception as e:
                                logger.error(f'Error processing raw material form: {e}')
                                messages.error(request, f'An error occurred while processing the raw material form: {e}')
                        
                        
                    
                    processed_quantity = int(request.POST['processed_qty'])
                    qty_to_process = cutting_form_instance.purchase_order_id.balance_number_of_pieces  
                    qty_to_process_minus_processed_qty = qty_to_process - processed_quantity  
                    cutting_form_instance.purchase_order_id.balance_number_of_pieces = qty_to_process_minus_processed_qty  
                    cutting_form_instance.purchase_order_id.save() 

                    messages.success(request, f'Cutting Order Created SuccessFully')
                    return(redirect(reverse('purchase-order-cutting-list', args = [cutting_form_instance.purchase_order_id.id, cutting_form_instance.purchase_order_id.product_reference_number.Product_Refrence_ID])))


            except ValidationError as val_err:
                logger.error(f'Validation error: {val_err}')
                messages.error(request, f'Validation error: {val_err}')


            except DatabaseError as db_err:
                logger.error(f'Database error: {db_err}')
                messages.error(request, f'Database error: {db_err}')


            except Exception as e:
                logger.error(f'Unexpected error: {e}')
                messages.error(request, f'An unexpected error occurred: {e}')

        else:
            
            
            

            
            


            return render(request,'production/purchase_order_cutting.html',{'form':form,'labour_all':labour_all,'purchase_order_cutting_form':purchase_order_cutting_form,'p_o_pk':p_o_pk,
                                                                    'purchase_order_to_product_formset_form':purchase_order_to_product_formset_form,
                                                                     'purchase_order_for_raw_material_cutting_items_formset_form':purchase_order_for_raw_material_cutting_items_formset_form,'page_name':'Cutting Order View'})


    return render(request,'production/purchase_order_cutting.html',{'form':form,'labour_all':labour_all,'purchase_order_cutting_form':purchase_order_cutting_form,'p_o_pk':p_o_pk,
                                                                    'purchase_order_to_product_formset_form':purchase_order_to_product_formset_form,
                                                                     'purchase_order_for_raw_material_cutting_items_formset_form':purchase_order_for_raw_material_cutting_items_formset_form,'page_name':'Cutting Order View'})





def purchaseordercuttinglistall(request):

    current_date = datetime.date.today

    raw_materials_exists = purchase_order_for_raw_material.objects.filter(purchase_order_id=OuterRef('pk')).values('pk')[:1] 

    purchase_orders_cutting_pending = (
        purchase_order.objects.annotate(
        raw_materials_exist=Exists(raw_materials_exists),
        total_approved_qty_sum=Sum('cutting_pos__approved_qty'),
        ).filter(
            balance_number_of_pieces__gt=0,
            raw_materials_exist=True  
        ).annotate(
            total_approved_balance=F('number_of_pieces') - F('total_approved_qty_sum')).order_by('created_date'))



    purchase_orders_cutting_completed = purchase_order.objects.filter(
        balance_number_of_pieces=0).annotate(
            total_processed_qty = Sum('cutting_pos__processed_qty'),
            total_approved_qty_sum = Sum('cutting_pos__approved_qty')).annotate(
                total_approved_balance = F('number_of_pieces') - F('total_approved_qty_sum')).order_by(
                    'created_date'
                    )


    purchase_order_cutting_all = purchase_order_raw_material_cutting.objects.all()

    vouchers_pending_count = purchase_order_raw_material_cutting.objects.filter(cutting_cancelled = False).filter(~Q(processed_qty=F('approved_qty'))).count()

    print('vouchers_pending_count = ', vouchers_pending_count)

    voucher_pending_quantity = purchase_order_raw_material_cutting.objects.filter(cutting_cancelled = False).exclude(processed_qty = F('approved_qty'))

    voucher_pending_quantity_total = 0

    for i in voucher_pending_quantity:

        processed_qty = i.processed_qty
        approved_qty = i.approved_qty

        balance = processed_qty - approved_qty

        voucher_pending_quantity_total += balance


    return render(request,'production/purchaseordercuttinglistall.html', {'purchase_orders_cutting_pending':purchase_orders_cutting_pending,'purchase_orders_cutting_completed':purchase_orders_cutting_completed,'page_name':'Cutting Order List','current_date':current_date,'purchase_order_cutting_all':purchase_order_cutting_all,'vouchers_pending_count':vouchers_pending_count,'voucher_pending_quantity_total':voucher_pending_quantity_total})




def purchaseordercuttinglist(request,p_o_pk,prod_ref_no):
    p_o_cutting_order_all = purchase_order_raw_material_cutting.objects.filter(
        purchase_order_id = p_o_pk).select_related('purchase_order_id__ledger_party_name',
        'factory_employee_id').order_by('created_date')
    
    Purchase_order_no = purchase_order.objects.get(id=p_o_pk)

    return render(request,'production/purchaseordercuttinglist.html', {'p_o_cutting_order_all':p_o_cutting_order_all, 'p_o_number':Purchase_order_no, 'prod_ref_no':prod_ref_no, 'p_o_pk':p_o_pk,'page_name':'Edit Cutting Order'})




def purchaseordercuttingapprovalcheckajax(request):
    
    cutting_pk_no = request.GET.get('cutting_pk_no')

    if not cutting_pk_no:
        return JsonResponse({'status': 'error', 'message': 'Cutting ID not provided'}, status=400)

    try:
        
        purchase_order_master_instances = labour_workout_master.objects.filter(purchase_order_cutting_master__raw_material_cutting_id=cutting_pk_no).annotate(total_processed_qty = Sum('labour_workout_childs__total_process_pcs')).order_by('created_date').values('created_date', 'total_approved_pcs', 'total_pending_pcs','total_processed_qty')

        if not purchase_order_master_instances.exists():
            return JsonResponse({'status': 'error', 'message': 'No records found for the given Cutting ID'}, status=404)

        approval_data = []

        for x in purchase_order_master_instances:
            dict_to_append = {
                'Approved_Date': x['created_date'],
                'Approved_Name': 'approved name',  
                'Approved_Qty': x['total_approved_pcs'],
                'pending_Qty': x['total_pending_pcs'],
                'total_processed_qty': x['total_processed_qty']
                }
            approval_data.append(dict_to_append)

        print('approval_data = ',approval_data)

        return JsonResponse({'status': 'success','approval_data': approval_data,'cutting_pk_no':cutting_pk_no})

    except ObjectDoesNotExist as e:
        logger.error(f'Record not found: {e}')
        return JsonResponse({'status': 'error', 'message': 'Record not found'}, status=404)

    except ValidationError as e:
        logger.error(f'Invalid data: {e}')
        return JsonResponse({'status': 'error', 'message': 'Invalid input data'}, status=400)

    except Exception as e:
        logger.error(f'Unexpected error: {e}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'An unexpected error occurred'}, status = 500)






def pendingapprovall(request):
    pending_approval_query = purchase_order_raw_material_cutting.objects.exclude(processed_qty = F('approved_qty')).order_by('created_date') 
    return render(request,'production/cuttingapprovallistall.html', {'pending_approval_query': pending_approval_query,'page_name':'Cutting Appr Pending List'})




@login_required(login_url = 'login')
def purchaseordercuttingpopup(request, cutting_id):

    if cutting_id:
        cutting_order_instance = purchase_order_raw_material_cutting.objects.get(raw_material_cutting_id = cutting_id)
    
    else:
        cutting_order_instance = None

    formset = purchase_order_cutting_approval_formset(request.POST or None, instance = cutting_order_instance)




    if request.method == 'POST':
        if formset.is_valid():
            
            if any(form.has_changed() for form in formset): 
                formset_instance = formset.save(commit=False)

                
                raw_material_cutting_instance = purchase_order_raw_material_cutting.objects.get(raw_material_cutting_id=cutting_id)

                
                old_total_approved_qty_total = raw_material_cutting_instance.approved_qty

                
                labour_workout_master_instance = labour_workout_master.objects.create(purchase_order_cutting_master=raw_material_cutting_instance)

                total_approved_pcs = 0

                for form in formset_instance:
                    p_o_to_cutting_instance = purchase_order_to_product_cutting.objects.get(id = form.id) 
                    old_total_approved_qty_diffrence  =  form.approved_pcs - p_o_to_cutting_instance.approved_pcs  
                    form.approved_pcs_diffrence = old_total_approved_qty_diffrence  
                    old_total_approved_qty_total = old_total_approved_qty_total + old_total_approved_qty_diffrence 
                    form.save() 
                    
                    total_approved_pcs = total_approved_pcs + old_total_approved_qty_diffrence

                    
                    product_to_item_labour_workout.objects.create(labour_workout = labour_workout_master_instance,
                                                                product_color=form.product_color,product_sku=form.product_sku,
                                                                pending_pcs = old_total_approved_qty_diffrence, processed_pcs = old_total_approved_qty_diffrence)

                raw_material_cutting_instance.approved_qty = old_total_approved_qty_total 
                raw_material_cutting_instance.save() 

                labour_workout_master_instance.total_approved_pcs = total_approved_pcs 
                labour_workout_master_instance.total_pending_pcs = total_approved_pcs 
                labour_workout_master_instance.save() 

            
            close_window_script = """
            <script>
            window.opener.location.reload(true);  // Reload parent window if needed
            window.close();  // Close current window
            </script>
            """

            return HttpResponse(close_window_script)
        else:
            return render(request,'production/purchaseordercuttingpopup.html', {'formset':formset})
            
    return render(request,'production/purchaseordercuttingpopup.html', {'formset':formset,'cutting_order_instance':cutting_order_instance})





def purchaseordercuttingmastercancelajax(request):

    if request.method == 'POST':
        try:
            cutting_key = request.POST.get('cuttingId')
            
            cutting_instance = get_object_or_404(purchase_order_raw_material_cutting,pk=cutting_key)

            with transaction.atomic():
                if cutting_instance:
                    cutting_instance.cutting_cancelled = True
                    cutting_instance.save()

                    if cutting_instance.approved_qty == 0:
                        processed_qty_to_revert = cutting_instance.processed_qty

                        
                        cutting_instance.purchase_order_id.cutting_total_processed_qty = cutting_instance.purchase_order_id.cutting_total_processed_qty - processed_qty_to_revert
                        cutting_instance.purchase_order_id.balance_number_of_pieces = cutting_instance.purchase_order_id.balance_number_of_pieces + processed_qty_to_revert
                        cutting_instance.purchase_order_id.save()
                        
                        
                        for record in cutting_instance.purchase_order_to_product_cutting_set.all():
                            purchase_order_to_product_instance = purchase_order_to_product.objects.get(purchase_order_id=cutting_instance.purchase_order_id.id,product_id__PProduct_SKU=record.product_sku)
                            purchase_order_to_product_instance.process_quantity = purchase_order_to_product_instance.process_quantity + record.cutting_quantity 
                            purchase_order_to_product_instance.save()

                        
                        for cutting_items in cutting_instance.purchase_order_for_raw_material_cutting_items_set.all():

                            item_godown_instance, created = item_godown_quantity_through_table.objects.get_or_create(Item_shade_name=cutting_items.material_color_shade,godown_name=cutting_items.purchase_order_cutting.purchase_order_id.temp_godown_select)

                            if created:
                                item_godown_instance_qty = 0 
                            else:
                                item_godown_instance_qty = item_godown_instance.quantity

                            item_godown_instance.quantity = item_godown_instance_qty + cutting_items.total_comsumption
                            item_godown_instance.save()
                            cutting_items.cutting_room_status = 'cutting_room_cancelled'
                            cutting_items.total_comsumption_in_cutting = 0
                            cutting_items.entry_from_cutting_room = True
                            cutting_items.save()

                        messages.success(request, 'Cutting Order cancelled successfully')
                        return JsonResponse({'status' : 'success'}, status=200)
                
                    else:
                        return JsonResponse({'status':'Cutting Already Approved'}, status=404)
                else:
                    return JsonResponse({'status':'Instance not found'}, status=404)
            
        except ObjectDoesNotExist as ne:
            logger.error(f'Instance not found -{ne}')
            messages.error(request, f'Error with labour workout: {ne}')
            return JsonResponse({'status':f'Instance not found -{ne}'}, status=404)
        

        except IntegrityError as ie:
            messages.error(request, 'Database integrity error occurred. Please try again.')
            logger.error(f'Database integrity error - {ie}')
            return JsonResponse({'status': 'Database integrity error occurred.'}, status=500)
        

        except Exception as e:
            logger.error(f'Instance not found -{e}')
            messages.error(request, f'Error with labour workout: {e}')
            return JsonResponse({'status':f'Instance not found -{e}'}, status=404)
        
    else:
        return JsonResponse({'status': 'Invalid request method.'}, status=405)



def labourworkoutlistall(request):
    labour_workout_pending = labour_workout_master.objects.all().annotate(total_processed_qty = Sum('labour_workout_childs__total_process_pcs')).filter(total_pending_pcs__gt=0).order_by('created_date')
    
    labour_workout_completed = labour_workout_master.objects.all().annotate(total_processed_qty = Sum('labour_workout_childs__total_process_pcs')).filter(total_pending_pcs__lt=1).order_by('created_date')
    
    labour_workout_instance_all = labour_workout_childs.objects.all().annotate(total_labour_workin_qty=Sum('labour_work_in_master__total_return_pcs')).order_by('created_date')

    current_date = datetime.date.today

    return render(request,'production/labourworkoutlistall.html', {'labour_workout_pending':labour_workout_pending, 'labour_workout_completed':labour_workout_completed, 'current_date':current_date, 'page_name':'Labour WorkOut List', 'labour_workout_instance_all':labour_workout_instance_all})




@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def labourworkoutsingle(request, labour_workout_child_pk=None, pk=None):

    try:
        ledger_labour_instances = Ledger.objects.filter(under_group__account_sub_group = 'Job charges(Exp of Mfg)')

        godown_id = None

        
        if pk is not None:
            
            labourworkoutinstance = get_object_or_404(labour_workout_master,id = pk)

            if labourworkoutinstance:
                godown_id = labourworkoutinstance.purchase_order_cutting_master.purchase_order_id.temp_godown_select.id

            else:
                
                raise ObjectDoesNotExist

            godown_instance = get_object_or_404(Godown_raw_material,id=godown_id)


            child_master_intial_data = {'total_approved_pcs' : labourworkoutinstance.total_approved_pcs,
                                        'total_balance_pcs' : labourworkoutinstance.total_pending_pcs}

            
            labour_work_out_child_form = labour_workout_child_form(initial=child_master_intial_data)
            
            
            product_to_item_instances = product_to_item_labour_workout.objects.filter(labour_workout = labourworkoutinstance)
        
            initial_items_data_dict = []

            for instance in product_to_item_instances:
                data_dict = {
                    'product_sku':instance.product_sku,
                    'product_color':instance.product_color,
                    'pending_pcs': instance.processed_pcs, 
                    'balance_pcs': instance.pending_pcs, 
                    'processed_pcs': 0
                    }

                initial_items_data_dict.append(data_dict)
            
            
            labour_workout_child_product_to_items_formset = inlineformset_factory(
                                    labour_workout_childs,product_to_item_labour_child_workout,fields=['product_sku',
                                                            'product_color','processed_pcs',
                                                            'pending_pcs','balance_pcs'], can_delete=False,extra=len(initial_items_data_dict))
            
            
            product_to_item_formset = labour_workout_child_product_to_items_formset(initial=initial_items_data_dict)


            
            raw_material_cutting_items_instances = purchase_order_for_raw_material_cutting_items.objects.filter(purchase_order_cutting = labourworkoutinstance.purchase_order_cutting_master).order_by('id')

            initial_data_dict = []
        
            for instance in raw_material_cutting_items_instances:

                item_instance = item_color_shade.objects.filter(items__item_name= instance.material_name,item_shade_name=instance.material_color_shade).first()

    
                if item_instance:
                    
                    current_balance = item_godown_quantity_through_table.objects.filter(godown_name = godown_instance,Item_shade_name = item_instance).first()
                    
                    if current_balance:
                        total_current_balance = current_balance.quantity
                    else:
                        total_current_balance = 0

                data = {
                    'product_sku': instance.product_sku,
                    'product_color': instance.product_color,
                    'material_name': instance.material_name,
                    'material_color_shade': instance.material_color_shade,
                    'rate': instance.rate,
                    'panha': instance.panha,
                    'units': instance.units,
                    'g_total': instance.g_total,
                    'g_total_combi' : instance.g_total_combi,
                    'consumption' : instance.consumption,
                    'combi_consumption' :instance.combi_consumption,
                    'total_comsumption': 0,
                    'unit_value': instance.unit_value,
                    'physical_stock': total_current_balance,
                    'balance_physical_stock': total_current_balance,
                    'fab_non_fab': instance.material_color_shade.items.Fabric_nonfabric,
                    'Remark': instance.Remark,
                    'pcs' : instance.pcs
                    }
                
                initial_data_dict.append(data)

            labour_workout_cutting_items_form_formset = inlineformset_factory(labour_workout_childs,labour_workout_cutting_items,
                                                                            form = labour_workout_cutting_items_form,               
                                                                            extra=len(initial_data_dict))
            
            
            labour_workout_cutting_items_formset_form =  labour_workout_cutting_items_form_formset(initial = initial_data_dict) 


        
        elif pk is None:

            labour_workout_child_instance = labour_workout_childs.objects.get(id = labour_workout_child_pk)
            

            
            labour_work_out_child_form = labour_workout_child_form(instance = labour_workout_child_instance)

            

            
            labour_workout_child_product_to_items_formset = inlineformset_factory(
                                    labour_workout_childs,product_to_item_labour_child_workout,fields=['product_sku',
                                                            'product_color','processed_pcs',
                                                            'pending_pcs','balance_pcs'], can_delete=False,extra=0)
            
            
            product_to_item_formset = labour_workout_child_product_to_items_formset(instance = labour_workout_child_instance)
            

            labour_workout_cutting_items_form_formset = inlineformset_factory(labour_workout_childs,labour_workout_cutting_items,
                                                                            form = labour_workout_cutting_items_form, 
                                                                            can_delete=False,
                                                                            extra=0)
            
            labour_workout_cutting_items_formset_form =  labour_workout_cutting_items_form_formset(instance = labour_workout_child_instance)

    except Exception as e:
        logger.error(f'exception occured at labour workout create GET {e}')



    if request.method == 'POST':
        

        
        labour_work_out_child_form = labour_workout_child_form(request.POST)

        
        product_to_item_formset = labour_workout_child_product_to_items_formset(request.POST)

        
        labour_workout_cutting_items_formset_form =  labour_workout_cutting_items_form_formset(request.POST) 


        if labour_work_out_child_form.is_valid() and product_to_item_formset.is_valid() and labour_workout_cutting_items_formset_form.is_valid():
            try:
                with transaction.atomic():
                    labour_workout_form_instance = labour_work_out_child_form.save(commit=False)
                    labour_workout_form_instance.labour_workout_master_instance = labourworkoutinstance
                    processed_qty = labour_workout_form_instance.total_process_pcs
                    labour_workout_form_instance.labour_workin_pending_pcs = processed_qty
                    labour_workout_form_instance.save()
                    
                    labourworkoutinstance.total_pending_pcs = labourworkoutinstance.total_pending_pcs - processed_qty
                
                    labourworkoutinstance.save()
                    
                    for form in product_to_item_formset:
                        if form.is_valid():
                            product_to_item_form = form.save(commit=False)
                            product_to_item_form.labour_workout = labour_workout_form_instance
                            product_to_item_form.labour_w_in_pending = product_to_item_form.processed_pcs
                            product_to_item_form.save()

                            
                            single_p2i_instance = product_to_item_instances.get(product_sku=product_to_item_form.product_sku,product_color=product_to_item_form.product_color)

                            single_p2i_instance.pending_pcs = single_p2i_instance.pending_pcs - product_to_item_form.processed_pcs
                            single_p2i_instance.save()
                    
                    
                    for form in labour_workout_cutting_items_formset_form:
                        if form.is_valid():
                            formset_form = form.save(commit=False)
                            formset_form.labour_workout_child_instance = labour_workout_form_instance
                            material_name = formset_form.material_name
                            material_color_shade = formset_form.material_color_shade

                            item_shade_instance = item_color_shade.objects.get(items__item_name=material_name,item_shade_name = material_color_shade)

                            if item_shade_instance.items.Fabric_nonfabric == 'Non Fabric':
                                total_comsumption = formset_form.total_comsumption
                                obj, created = item_godown_quantity_through_table.objects.get_or_create(godown_name=godown_instance,Item_shade_name=item_shade_instance)

                                if created:
                                    qty_to_deduct = 0
                        
                                elif not created:
                                    qty_to_deduct = obj.quantity

                                obj.quantity = qty_to_deduct - total_comsumption
                                obj.save()

                            elif item_shade_instance.items.Fabric_nonfabric == 'Fabric':
                                purchase_order_cutting_items = purchase_order_for_raw_material_cutting_items.objects.get(
                                    purchase_order_cutting__raw_material_cutting_id=labour_workout_form_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,
                                    product_sku = formset_form.product_sku, material_color_shade = item_shade_instance)
                                
                                purchase_order_cutting_items.total_comsumption_in_cutting = purchase_order_cutting_items.total_comsumption_in_cutting - formset_form.total_comsumption
                                purchase_order_cutting_items.entry_from_cutting_room = False
                                purchase_order_cutting_items.save()

                            formset_form.save()

                    return redirect(reverse('labour-workout-child-list', args=[pk]))
                
                
            except Exception as e:
                logger.error(f'An exception occoured in labour workout create {e}')

        else:
            logger.error(f'labour_workout_cutting_items_formset_form{labour_workout_cutting_items_formset_form.errors}')
            logger.error(f'product_to_item_formset {product_to_item_formset.non_form_errors()}')
            logger.error(f'labour_workout_cutting_items_formset_form - {labour_workout_cutting_items_formset_form.non_form_errors()}')

            return render(request,'production/labourworkoutsingle.html',
                  {'product_to_item_formset':product_to_item_formset,'labour_work_out_child_form':labour_work_out_child_form,
                   'labour_workout_cutting_items_formset_form':labour_workout_cutting_items_formset_form,
                   'ledger_labour_instances':ledger_labour_instances,'godown_id':godown_id,'page_name':'Labour WorkOut Create'})


            

    return render(request,'production/labourworkoutsingle.html',
                  {'product_to_item_formset':product_to_item_formset,'labour_work_out_child_form':labour_work_out_child_form,
                   'labour_workout_cutting_items_formset_form':labour_workout_cutting_items_formset_form,
                   'ledger_labour_instances':ledger_labour_instances,'godown_id':godown_id,'page_name':'Labour WorkOut Create'})





def labour_workout_child_list(request, labour_master_pk):
    labour_work_out_master = labour_workout_master.objects.get(id=labour_master_pk)
    labour_workout_child_instances = labour_workout_childs.objects.filter(labour_workout_master_instance = labour_master_pk).order_by('created_date')
    return render(request,'production/labourworkoutchilds.html', {'labour_master_pk':labour_master_pk,
                                                                  'labour_workout_child_instances':labour_workout_child_instances,
                                                                  'labour_work_out_master':labour_work_out_master,
                                                                  'page_name':'Labour Workout Create'})






def labourworkoutsingledeleteajax(request):
    
    if request.method == 'POST':

        labour_workout_child_pk =  request.POST.get('labour_workout_child_pk')
        try:
            with transaction.atomic():
                
                labour_workout_child_instance = labour_workout_childs.objects.get(id=labour_workout_child_pk)
                
                labour_master_instance = labour_workout_child_instance.labour_workout_master_instance
                labour_master_instance.total_pending_pcs = labour_master_instance.total_pending_pcs + labour_workout_child_instance.total_process_pcs

                labour_master_instance.save()
                
                for product_2_Item_child_instancs in labour_workout_child_instance.labour_workout_child_items.all():
                    product_2_item_master_instance = product_to_item_labour_workout.objects.get(labour_workout = labour_master_instance,product_sku=product_2_Item_child_instancs.product_sku,product_color=product_2_Item_child_instancs.product_color)
                    product_2_item_master_instance.pending_pcs = product_2_item_master_instance.pending_pcs + product_2_Item_child_instancs.processed_pcs
                    product_2_item_master_instance.save()


                for labour_workout_child_items in labour_workout_child_instance.labour_workout_cutting_items_set.all():
                    item_in_row = item_color_shade.objects.get(items__item_name=labour_workout_child_items.material_name,item_shade_name=labour_workout_child_items.material_color_shade)
                    
                    if item_in_row.items.Fabric_nonfabric == 'Fabric':
                        
                        purchase_order_cutting_item = purchase_order_for_raw_material_cutting_items.objects.get(purchase_order_cutting=labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id, material_color_shade=item_in_row, product_sku=labour_workout_child_items.product_sku)
                        
                        purchase_order_cutting_item.total_comsumption_in_cutting = purchase_order_cutting_item.total_comsumption_in_cutting + labour_workout_child_items.total_comsumption
                        purchase_order_cutting_item.entry_from_cutting_room = False
                        purchase_order_cutting_item.save()

                    elif item_in_row.items.Fabric_nonfabric == 'Non Fabric':

                        obj, created = item_godown_quantity_through_table.objects.get_or_create(Item_shade_name=item_in_row,godown_name=labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.temp_godown_select)

                        if created:
                            qty_to_add = 0
                        else:
                            qty_to_add = obj.quantity

                        obj.quantity = qty_to_add + labour_workout_child_items.total_comsumption
                        obj.save()

                labour_workout_child_instance.delete()

                messages.success(request,'labour workout deleted successfully')
                return JsonResponse({'status':'success'}, status=200) 
             

        except ObjectDoesNotExist as ne:
            messages.error(request, f'Error with labour workout: {ne}')
            logger.error(f'Instance not found - {ne}')
            return JsonResponse({'status': f'Instance not found - {ne}'}, status=404)
        

        except IntegrityError as ie:
            messages.error(request, 'Database integrity error occurred. Please try again.')
            logger.error(f'Database integrity error - {ie}')
            return JsonResponse({'status': 'Database integrity error occurred.'}, status=500)
        

        except Exception as e:
            logger.error(f'An unexpected error occurred - {e}')
            messages.error(request, f'Error with labour workout: {e}')
            return JsonResponse({'status': f'An unexpected error occurred - {e}'}, status=500)
        
    else:
        return JsonResponse({'status': 'Invalid request method.'}, status=405)





def cuttingroomqty(request):
    cutting_room_items = purchase_order_for_raw_material_cutting_items.objects.filter(total_comsumption_in_cutting__gt=0)

    return render(request,'production/cuttingroomqty.html',{'cutting_room_items':cutting_room_items,'page_name':'Cutting Room Qty'})





def labourworkincreatelist(request,l_w_o_id):

    user = request.user

    logger.info(f"labour workin create list function run by {user}.")

    labour_workout_child_instance = labour_workout_childs.objects.get(id=l_w_o_id)

    labour_workin_instances = list(
        labour_work_in_master.objects.filter(
            labour_voucher_number=labour_workout_child_instance
        ).annotate(
            approved_Qty_total=Sum('l_w_in_products__approved_qty'),
            total_approved_pcs=Sum('l_w_in_products__approved_qty'),
            pending_for_approval_pcs=Sum('l_w_in_products__pending_for_approval'),
            total_balance_qty=Sum('l_w_in_products__dummy_balance_qty')
        ).order_by('created_date')
    )


    return render(request,'production/labour_work_in_list.html',{'labour_workout_child_instance':labour_workout_child_instance,'labour_workin_instances':labour_workin_instances,'page_name':'Challan Wise GRN'})







@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def labourworkincreate(request, l_w_o_id = None, pk = None, approved=False):
    
    template_name = 'production/labourworkincreate.html'
    
    labour_workin_all = None

    approval_check = approved

    last_entry_for_submit_button = None
    
    
    if l_w_o_id is None:

        template_name = 'production/labourworkincreateraw.html'
        labour_workin_master_instance = None
        
        master_form = labour_workin_master_form()

        product_to_item_formset = None

        labour_work_in_product_to_item_formset = inlineformset_factory(labour_work_in_master,labour_work_in_product_to_item, form = labour_work_in_product_to_item_form, extra = 0, can_delete = False)

        labour_workout_child_instance = None


        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            try:
                vendor_name_value = request.GET.get('nameValue')

                vendor_name_dict = None
    
                if vendor_name_value:
                    selected_vendor_name = Ledger.objects.filter(under_group__account_sub_group='Job charges(Exp of Mfg)',name__icontains=vendor_name_value)
                    vendor_name_dict = {}
                    for record in selected_vendor_name:
                        vendor_name_dict[record.id] = record.name

                choosed_vendor_name = request.GET.get('itemValue')       

                labour_workout_instance_dict = []

                if choosed_vendor_name:
                    labour_workout_instances = labour_workout_childs.objects.filter(labour_name=choosed_vendor_name).exclude(labour_workin_pending_pcs=0).exclude(labour_workin_pending_pcs__isnull=True)
                    
                    for instance in labour_workout_instances:
                        dict_to_append = {
                            'Challan_No': instance.challan_no,
                            'PO_No':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
                            'PO_Total_QTY':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.number_of_pieces,
                            'cutting_vch_no' : instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,
                            'Ref_No':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID,
                            'Model_Name':instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name,
                            'Issued_QTY':instance.total_process_pcs,
                            'Rec_QTY': instance.labour_workin_pcs,
                            'Balance_QTY': instance.labour_workin_pending_pcs,
                            'labour_workout_id': instance.id,
                            'created_date' : instance.created_date}

                        labour_workout_instance_dict.append(dict_to_append)



                labour_work_out_id = request.GET.get('labourWorkOutId')
               
                master_initial_data = None

                formset_initial_data = None

                labour_workout_child_instance_id = None

                if labour_work_out_id:
                    labour_workout_child_instance = labour_workout_childs.objects.get(id = labour_work_out_id)

                    labour_workin_all_qd = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).annotate(total_approved_quantity = Sum('l_w_in_products__approved_qty')).values('voucher_number','total_return_pcs','total_approved_quantity','created_date')
                    labour_workin_all = list(labour_workin_all_qd)

                    last_entry = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).order_by('-created_date').first()

                    last_labour_charges = last_entry.labour_charges if last_entry else labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.labour_charges

                    labour_workout_child_instance_id = labour_workout_child_instance.id
                    
                    master_initial_data = {
                        'labour_name': labour_workout_child_instance.labour_name.name,
                        'challan_no' : labour_workout_child_instance.challan_no,
                        'challan_no_id' : labour_workout_child_instance.id,
                        'purchase_order_no' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
                        'purchase_order_no_id':labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.id,
                        'cutting_vch_no' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,
                        'cutting_vch_no_id' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,
                        'refrence_number' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID,
                        'model_name': labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name,
                        'total_p_o_qty' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.number_of_pieces,
                        'labour_workout_qty' : labour_workout_child_instance.total_process_pcs,
                        'labour_charges': last_labour_charges,
                        'total_balance_pcs' : labour_workout_child_instance.labour_workin_pending_pcs,
                        }

                    print('master_initial_data = ',master_initial_data)


                    product_to_item_l_w_in_instance = product_to_item_labour_child_workout.objects.filter(labour_workout=labour_workout_child_instance)
 
                    formset_initial_data = []

                    for instances in product_to_item_l_w_in_instance:

                        initial_data_dict = { 
                            'product_sku': instances.product_sku,
                            'product_color': instances.product_color,
                            'L_work_out_pcs': instances.processed_pcs,
                            'pending_to_return_pcs': instances.labour_w_in_pending,
                            'total_recieved_qty' : instances.processed_pcs - instances.labour_w_in_pending,
                            'return_pcs' : '0',
                            'qty_to_compare':  instances.labour_w_in_pending,
                            'cur_bal_plus_return_qty': instances.labour_w_in_pending,
                            'dummy_balance_qty': instances.processed_pcs - instances.labour_w_in_pending,
                        }

                        formset_initial_data.append(initial_data_dict)
                

                return JsonResponse({'vendor_name_dict':vendor_name_dict,
                                     'labour_workout_instance_dict':labour_workout_instance_dict,
                                     'master_initial_data':master_initial_data,
                                     'formset_initial_data':formset_initial_data,
                                     'labour_workout_child_instance_id': labour_workout_child_instance_id ,
                                     'labour_workin_all':labour_workin_all}
                                    )

            except ValueError as ve:
                    messages.error(request,f'Error Occured - {ve}')
                    return JsonResponse({'status': f'Error with ajax request - {ve}'}, status=404)
        
            except Exception as e:
                messages.error(request,f'Exception Occured - {e}')
                return JsonResponse({'status': f'Error with ajax request - {e}'}, status=404)


    
    elif l_w_o_id is not None and pk is None:
        
        print("********* IN CREATE MODE **********")

        labour_workin_master_instance = None

        labour_workout_child_instance = labour_workout_childs.objects.get(id=l_w_o_id)

        labour_workin_all = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).annotate(total_approved_quantity = Sum('l_w_in_products__approved_qty'))

        last_entry = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).order_by('-created_date').first()
        

        last_labour_charges = last_entry.labour_charges if last_entry else labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.labour_charges

        initial_data = {
            'labour_name': labour_workout_child_instance.labour_name.name,

            'challan_no' : labour_workout_child_instance.challan_no,

            'challan_no_id' : labour_workout_child_instance.id,

            'purchase_order_no' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,

            'purchase_order_no_id':labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.id,

            'cutting_vch_no' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,

            'cutting_vch_no_id' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.raw_material_cutting_id,

            'refrence_number' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID,

            'model_name': labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.product_reference_number.Model_Name,

            'total_p_o_qty' : labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.number_of_pieces,

            'labour_workout_qty' : labour_workout_child_instance.total_process_pcs,

            'labour_charges': last_labour_charges,

            'total_balance_pcs' :  labour_workout_child_instance.labour_workin_pending_pcs,  
        }

        master_form = labour_workin_master_form(initial=initial_data)

        product_to_item_l_w_in = product_to_item_labour_child_workout.objects.filter(labour_workout=labour_workout_child_instance)

        formset_initial_data = []

        for instances in product_to_item_l_w_in:

            initial_data_dict = { 
                'product_sku': instances.product_sku,
                'product_color': instances.product_color,
                'L_work_out_pcs': instances.processed_pcs,
                'pending_to_return_pcs': instances.labour_w_in_pending,
                'total_recieved_qty' : instances.processed_pcs - instances.labour_w_in_pending,
                'return_pcs' : '0',
                'qty_to_compare':  instances.labour_w_in_pending,
                'cur_bal_plus_return_qty': instances.labour_w_in_pending,
                'dummy_balance_qty': instances.processed_pcs - instances.labour_w_in_pending,
                }
            
            formset_initial_data.append(initial_data_dict)


        labour_work_in_product_to_item_formset = inlineformset_factory(labour_work_in_master,labour_work_in_product_to_item,form=labour_work_in_product_to_item_form, extra=len(formset_initial_data), can_delete=False)

        product_to_item_formset = labour_work_in_product_to_item_formset(initial=formset_initial_data)


    
    elif l_w_o_id is not None and pk is not None:
        
        print("********* IN EDIT MODE **********")

        labour_workout_child_instance = labour_workout_childs.objects.get(id = l_w_o_id)

        last_entry_for_submit_button = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).order_by('-created_date').first()

        labour_workin_all = labour_work_in_master.objects.filter(labour_voucher_number=labour_workout_child_instance).annotate(total_approved_quantity = Sum('l_w_in_products__approved_qty'))

        product_to_item_l_w_in = product_to_item_labour_child_workout.objects.filter(labour_workout=labour_workout_child_instance)

        labour_workin_master_instance = labour_work_in_master.objects.get(pk=pk)
    
        master_form = labour_workin_master_form(instance = labour_workin_master_instance)

        labour_work_in_product_to_item_formset = inlineformset_factory(labour_work_in_master,labour_work_in_product_to_item, 
            form = labour_work_in_product_to_item_form, extra = 0, can_delete = False)
        
        product_to_item_formset = labour_work_in_product_to_item_formset(instance = labour_workin_master_instance) 
        
        
        for form, instance in zip(product_to_item_formset.forms, product_to_item_l_w_in):  

            if instance:
                form.initial['qty_to_compare'] = instance.labour_w_in_pending
                form.initial['cur_bal_plus_return_qty'] = instance.labour_w_in_pending  + form.instance.return_pcs
                form.initial['total_recieved_qty'] = instance.processed_pcs - instance.labour_w_in_pending

    if request.method == 'POST':
        
        print(request.POST)

        labour_workout_child_i = request.POST.get('labour_workout_child_instance_id')

        print(labour_workout_child_i)

        if labour_workout_child_i:
            labour_workout_child_instance = labour_workout_childs.objects.get(id=labour_workout_child_i)

        master_form = labour_workin_master_form(request.POST, instance = labour_workin_master_instance)

        old_return_qty =  master_form.instance.total_return_pcs

        product_to_item_formset = labour_work_in_product_to_item_formset(request.POST,instance = labour_workin_master_instance)

        try:
            with transaction.atomic():
                if master_form.is_valid() and product_to_item_formset.is_valid():
                    
                    #save master form 
                    parent_form = master_form.save(commit = False)

                    master_form.instance.labour_charges = master_form.cleaned_data['labour_charges']

                    parent_form.labour_voucher_number = labour_workout_child_instance
                    
                    if master_form.instance.id:
                        labour_workin_qty = parent_form.total_return_pcs - old_return_qty 
                    else:
                        labour_workin_qty = parent_form.total_return_pcs
                    
                    labour_workout_child_instance.labour_workin_pcs = labour_workout_child_instance.labour_workin_pcs + labour_workin_qty
                    parent_form.labour_voucher_number.labour_workin_pending_pcs = parent_form.total_balance_pcs

                    labour_workout_child_instance.save()
                    
                    parent_form.save()




                    for form in product_to_item_formset:

                        if form.is_valid():
                            product_to_item_form = form.save(commit= False)

                            product_to_item_form.labour_workin_instance = parent_form

                            product_to_item_form.pending_for_approval = product_to_item_form.return_pcs
                            
                            l_w_o_instance = product_to_item_labour_child_workout.objects.get(labour_workout=labour_workout_child_instance,product_sku=product_to_item_form.product_sku,product_color=product_to_item_form.product_color)
                            
                            if product_to_item_form.pk:
                                labour_workin_product2item = labour_work_in_product_to_item.objects.get(pk = product_to_item_form.pk)
                                qty_to_change = product_to_item_form.return_pcs - labour_workin_product2item.return_pcs
                
                            else:
                                qty_to_change = product_to_item_form.return_pcs

                            l_w_o_instance.labour_w_in_pending = l_w_o_instance.labour_w_in_pending - qty_to_change
                            l_w_o_instance.save()

                            product_to_item_form.save()

                    return redirect(reverse('labour-workin-list-create', args=[labour_workout_child_instance.id]) )

                else:
                    
                    return redirect(reverse('labour-workin-list-create', args=[labour_workout_child_instance.id]) )
                    
                
                
        except ValidationError as ve:
            messages.error(request,f'Validation error {ve}')

        except Exception as e:
            messages.error(request,f'Other exceptions {e}')
    
    return render(request,template_name,{'master_form':master_form,'labour_work_in_product_to_item_formset':product_to_item_formset,'approval_check':approval_check,'page_name':'Labour Workin Create','labour_workin_all':labour_workin_all,'last_entry_id': last_entry_for_submit_button.id if last_entry_for_submit_button else None,'current_entry_id': master_form.instance.id if master_form.instance else None,})










def labourworkinlistall(request):

    current_date = datetime.date.today


    approved_qty_subquery = labour_work_in_product_to_item.objects.filter(
    labour_workin_instance__labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id=OuterRef('pk')
    ).values('labour_workin_instance__labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id') \
    .annotate(
        total_approved_qty=Sum('approved_qty')
    ).values('total_approved_qty')



    pending_approved_qty_subquery = labour_work_in_product_to_item.objects.filter(
    labour_workin_instance__labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id=OuterRef('pk')
    ).values('labour_workin_instance__labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id') \
    .annotate(
        total_pending_approved_qty=Sum('pending_for_approval')
    ).values('total_pending_approved_qty')



    labour_workout_childs_exists = labour_workout_childs.objects.filter(
    labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id = OuterRef('pk')
    ).values('pk')[:1]


    #for all 
    purchase_orders_with_labour_workout_childs = purchase_order.objects.annotate(
    has_labour_workout_childs=Exists(labour_workout_childs_exists)
    ).filter(
        has_labour_workout_childs=True
    ).annotate(
        total_lwo_pcs=Sum('cutting_pos__labourworkouts__labour_workout_childs__total_process_pcs'),
        total_labour_workin_pcs=Sum('cutting_pos__labourworkouts__labour_workout_childs__labour_workin_pcs'),
        total_labour_workin_pending=Sum('cutting_pos__labourworkouts__labour_workout_childs__labour_workin_pending_pcs'),
        total_approved_qty=Subquery(approved_qty_subquery),
        total_pending_approved_qty=Subquery(pending_approved_qty_subquery)
    ).order_by('created_date')

    #vendor name wise
    labour_workout_child_instances_all = labour_workout_childs.objects.all().annotate(total_approved_qty = Sum('labour_work_in_master__l_w_in_products__approved_qty'),total_pending_approved_qty = Sum('labour_work_in_master__l_w_in_products__pending_for_approval')).order_by('created_date')

    # labour workin grn no wise
    labour_workin_child_instances_all = labour_work_in_master.objects.all().order_by('created_date')


    # Calculate approve pending quantity
    labour_workin_pending_quantity = labour_work_in_master.objects.all().aggregate(pending_for_approval_pcs = Sum('l_w_in_products__pending_for_approval'))['pending_for_approval_pcs']


    # Calculate approve pending count
    labour_workin_pending_count = labour_work_in_master.objects.annotate(
        total_pending_for_approval=Sum('l_w_in_products__pending_for_approval'),
        total_approved_qty_sum=Sum('l_w_in_products__approved_qty')
    ).filter(
        Q(total_pending_for_approval__gt=F('total_approved_qty_sum'))
    ).count()


    return render(request,'production/labour_workin_listall.html',
                  {'labour_workout_child_instances_all':labour_workout_child_instances_all,
                   'purchase_order_instances': purchase_orders_with_labour_workout_childs,'current_date':current_date,'page_name':'Labour WorkIn List','labour_workin_child_instances_all':labour_workin_child_instances_all,'labour_workin_pending_count':labour_workin_pending_count,'labour_workin_pending_quantity':labour_workin_pending_quantity})







def labourworkinpurchaseorderlist(request ,p_o_no):

    user = request.user

    logger.info(f"labour workin purchase order list function run by {user}.")

    purchase_order_instance = purchase_order.objects.get(id = p_o_no)

    labour_workin_purchase_order_list = labour_workout_childs.objects.filter(labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__id = p_o_no).annotate(total_approved_qty=Sum('labour_work_in_master__l_w_in_products__approved_qty'), total_pending_approved_qty=Sum('labour_work_in_master__l_w_in_products__pending_for_approval')).order_by('created_date')

    lab_workout_master = labour_workout_childs.objects.filter(labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__id = p_o_no)
    
    total_lab_qty = 0
    for instance in lab_workout_master:
        total_lab_qty = total_lab_qty + instance.total_process_pcs

    lab_work_in_qty = labour_work_in_master.objects.filter(labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__id = p_o_no)

    total_lab_workin_qty = 0
    for instance in lab_work_in_qty:
        total_lab_workin_qty = total_lab_workin_qty + instance.total_return_pcs

    balance_qty = total_lab_qty - total_lab_workin_qty

    return render(request,'production/labour_workin_purchase_order_list.html',{'labour_workin_purchase_order_list':labour_workin_purchase_order_list,'purchase_order_instance':purchase_order_instance, 'total_lab_qty':total_lab_qty, 'total_lab_workin_qty':total_lab_workin_qty, 'balance_qty':balance_qty,'page_name':'PO Wise Challan'})




def labourworkinsingledeleteajax(request):
    
    if request.method == 'POST':
        labour_workin_id = request.POST.get('labour_workin_pk')

        if labour_workin_id:
            try:
                with transaction.atomic():
                    labour_workin_instance = labour_work_in_master.objects.get(pk=labour_workin_id)
                    
                    labour_workin_instance.labour_voucher_number.labour_workin_pending_pcs = labour_workin_instance.labour_voucher_number.labour_workin_pending_pcs + labour_workin_instance.total_return_pcs
                    labour_workin_instance.labour_voucher_number.labour_workin_pcs = labour_workin_instance.labour_voucher_number.labour_workin_pcs - labour_workin_instance.total_return_pcs
                    labour_workin_instance.labour_voucher_number.save()


                    for instances in labour_workin_instance.l_w_in_products.all():
                        product_2_item_child_instances =  product_to_item_labour_child_workout.objects.get(labour_workout=labour_workin_instance.labour_voucher_number,product_sku=instances.product_sku,product_color=instances.product_color)
                        product_2_item_child_instances.labour_w_in_pending = product_2_item_child_instances.labour_w_in_pending + instances.return_pcs

                        product_2_item_child_instances.save()

                    labour_workin_instance.delete()

                    return redirect(reverse('labour-workin-list-create', args=[labour_workin_instance.labour_voucher_number.id]) )
                    

            except ObjectDoesNotExist as ne:
                messages.error(request, f'Error with labour workout: {ne}')
                logger.error(f'Instance not found - {ne}')
                return JsonResponse({'status': f'Instance not found - {ne}'}, status=404)
        

            except IntegrityError as ie:
                messages.error(request, 'Database integrity error occurred. Please try again.')
                logger.error(f'Database integrity error - {ie}')
                return JsonResponse({'status': 'Database integrity error occurred.'}, status=500)
        

            except Exception as e:
                logger.error(f'An unexpected error occurred - {e}')
                messages.error(request, f'Error with labour workout: {e}')
                return JsonResponse({'status': f'An unexpected error occurred - {e}'}, status=500)
        
    else:
        return JsonResponse({'status': 'Invalid request method.'}, status=405)



def goods_return_pending_list(request):
    current_date = datetime.date.today
    labour_workin_instances = labour_work_in_master.objects.annotate(total_approved_pcs=Sum('l_w_in_products__approved_qty'),pending_for_approval_pcs = Sum('l_w_in_products__pending_for_approval')).filter(pending_for_approval_pcs__gt=0)
    return render(request,'production/goodsreturnpendinglist.html',{'labour_workin_instances':labour_workin_instances, 'current_date':current_date,'page_name':'Goods Return List'})





def goods_return_popup(request,pk):

    if pk:
        finished_goods_godowns = Godown_finished_goods.objects.all()
        labour_workin_instance = labour_work_in_master.objects.get(pk=pk)
        formset = labour_work_in_product_to_item_approval_formset(request.POST or None, instance=labour_workin_instance)

        if request.method == 'POST':
            
            godown_name_post = request.POST.get('godown_name_post')
            formset.forms = [form for form in formset if form.has_changed()]
            if formset.is_valid():
                try:
                    with transaction.atomic():

                        unique_id1 = uuid.uuid4()

                        for form in formset:
                            form_instance = form.save(commit=False)
            
                            godown_instance = Godown_finished_goods.objects.get(id = godown_name_post)
                            selected_product = PProduct_Creation.objects.get(PProduct_SKU = form_instance.product_sku)
                            obj, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = godown_instance, product_color_name = selected_product)

                            quantity_to_add = obj.quantity if not created else 0
                            
                            if form_instance.pk:
                                labour_workin_instance = labour_work_in_product_to_item.objects.get(pk = form_instance.pk)

                                approved_qty_differ = form_instance.approved_qty - labour_workin_instance.approved_qty 
                            else:
                                approved_qty_differ = form_instance.approved_qty

                            obj.quantity = quantity_to_add + approved_qty_differ  
                            
                            obj.save()

                            form_instance.save()

                            P2I_instance = labour_work_in_product_to_item.objects.get(id = form_instance.pk)

                            labour_workin_approval_report.objects.create(labour_w_i_p_2_i = P2I_instance, difference_qty = approved_qty_differ, unique_id = unique_id1)

                        messages.success(request,'Products Successfully recieved')

                except IntegrityError as e:
                    print(f"Caught IntegrityError:{e}")
                    messages.error(request,f'Error with Formset - {e}')
                    
                except Exception as e:
                    messages.error(request,f'Error with Formset - {e}')
                    
                
                close_window_script = """
                <script>
                window.opener.location.reload(true);  // Reload parent window if needed
                window.close();  // Close current window
                </script>
                """
                return HttpResponse(close_window_script)

            else:
                messages.error(request,f'Error with Formset {formset.errors}')

                
                close_window_script = """
                <script>
                window.opener.location.reload(true);  // Reload parent window if needed
                window.close();  // Close current window
                </script>
                """
                return HttpResponse(close_window_script)

    return render(request,'production/goods_return_popup.html',{'formset':formset,
                                                                'finished_goods_godowns':finished_goods_godowns,
                                                                'labour_workin_instance' : labour_workin_instance})


def purchaseorderlabourworkinapprovecheckajax(request):

    labour_workin_key = request.GET.get('labour_workin_key')
    
    if labour_workin_key:
        try:
            lab_workin_app_report = labour_workin_approval_report.objects.filter(labour_w_i_p_2_i__labour_workin_instance__id = labour_workin_key).order_by('creation_date')

            if not lab_workin_app_report.exists():
                return JsonResponse({'status': 'error', 'message': 'No records found for the given Cutting ID'}, status=404)

            
            grouped_reports = defaultdict(list)

            for report in lab_workin_app_report:
                grouped_reports[report.unique_id].append({
                "creation_date": report.creation_date,
                "difference_qty": report.difference_qty,
                "approved_sku": report.labour_w_i_p_2_i.product_sku,
                "color":report.labour_w_i_p_2_i.product_color ,
            })
            
            
            result = [{"unique_id": unique_id, "records": records} for unique_id, records in grouped_reports.items()]
                

            return JsonResponse({'status': 'success', 'result': result})

        except ObjectDoesNotExist as e:
            logger.error(f'Record not found: {e}')
            return JsonResponse({'status': 'error', 'message': 'Record not found'}, status=404)

        except ValidationError as e:
            logger.error(f'Invalid data: {e}')
            return JsonResponse({'status': 'error', 'message': 'Invalid input data'}, status=400)

        except Exception as e:
            logger.error(f'Unexpected error: {e}', exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'An unexpected error occurred'}, status = 500)




def finished_goods_godown_wise_report(request, g_id):
    
    
    all_godown_stock = product_godown_quantity_through_table.objects.filter(
        product_color_name__Product_id = OuterRef('pk')).values('product_color_name__Product_id'
        ).annotate(sum_all_g_qty=Sum('quantity')).values('sum_all_g_qty')

    
    product_quantity = Product.objects.filter(
        productdetails__godown_colors__godown_name__id=g_id).annotate(
        total_quantity_product=Sum('productdetails__godown_colors__quantity'),  
        all_godown_qty = Subquery(all_godown_stock))  

    return render(request,'production/godown_product_qty.html', {'product_quantity' : product_quantity})





def finished_goods_godown_wise_report_all(request):
    
    
    purchase_order_instances = purchase_order.objects.filter(
        product_reference_number__Product_Refrence_ID = OuterRef('Product_Refrence_ID')).values(
            'product_reference_number__Product_Refrence_ID').annotate(pending_approval_total = Sum(
        'cutting_pos__labourworkouts__labour_workout_childs__labour_work_in_master__l_w_in_products__pending_for_approval')).values('pending_approval_total')

    purchase_order_bal_vendor = purchase_order.objects.filter(
        product_reference_number__Product_Refrence_ID = OuterRef('Product_Refrence_ID')).values(
            'product_reference_number__Product_Refrence_ID').annotate(vendor_bal_qty = Sum(
        'cutting_pos__labourworkouts__labour_workout_childs__labour_workout_child_items__labour_w_in_pending')).values('vendor_bal_qty')


    product_quantity = Product.objects.annotate(total_quantity_product=Sum(
        'productdetails__godown_colors__quantity'), total_qty_pending = Subquery(purchase_order_instances),
        total_bal_vendor = Subquery(purchase_order_bal_vendor))
    
    
    return render(request,'production/labour_workout_report.html', {'product_quantity' : product_quantity})


""" Coalesce is a SQL function provided by Django through the django.db.models.functions module.
    It is used to handle NULL (or None in Python) values by replacing them with a specified fallback value.
    This is especially useful in aggregate queries, where calculations like Sum, Count, or Avg might return
    None if there are no matching rows or if the data contains only NULL values.

    Sum('productdetails__godown_colors__quantity'):

    Calculates the sum of quantity from the related model.
    If no rows match or all quantity values are NULL, it returns None.
    Coalesce(Sum(...), 0):

    If the Sum result is None, it replaces it with 0.
    Result:

    The query will always return a number (e.g., 0 if no data is found or None values exist).
    
    """




def finished_goods_godown_product_ref_wise_report(request, ref_no):

    if ref_no:
        product_instance = Product.objects.get(Product_Refrence_ID = ref_no)

        
        pending_approval_subquery = labour_workout_childs.objects.filter(
            id=OuterRef('id')).annotate(total_pending = Sum('labour_work_in_master__l_w_in_products__pending_for_approval')
                                        ).values('total_pending')


        approved_subquery = labour_workout_childs.objects.filter(
            id=OuterRef('id')).annotate(total_approved = Sum('labour_work_in_master__l_w_in_products__approved_qty')
                                        ).values('total_approved')

        
        
        total_return_subquery = labour_workout_childs.objects.filter(
            id=OuterRef('id')).annotate(total_return_pcs = Sum('labour_work_in_master__total_return_pcs')
                                        ).values('total_return_pcs')


        purchase_instances = labour_workout_childs.objects.filter(
            labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID=ref_no
            ).annotate(
            total_balance_to_vendor=Sum('labour_workout_child_items__labour_w_in_pending'),
            total_return_pcs_1=Subquery(total_return_subquery),
            total_pending_to_approval_qty = Subquery(pending_approval_subquery), 
            total_approved_qty = Subquery(approved_subquery)
            )

        
        
        
        product_quantity = Product.objects.filter(Product_Refrence_ID=ref_no).aggregate(
            total_quantity=Coalesce(Sum('productdetails__godown_colors__quantity'), 0))
        
        
        purchase_order_instance = Product.objects.filter(Product_Refrence_ID = ref_no).aggregate(
            total_approval_pending = Sum('purchase_order__cutting_pos__labourworkouts__labour_workout_childs__labour_work_in_master__l_w_in_products__pending_for_approval'))
        
        

    return render(request,'production/godown_model_wise.html', {'purchase_instances': purchase_instances,
                                                                'product_instance' : product_instance,'product_quantity':product_quantity,
                                                                'purchase_order_instance':purchase_order_instance})


def finished_goods_vendor_model_wise_report(request, ref_no, challan_no):

    if ref_no is not None and challan_no is not None:
        
        labour_workin_instances = labour_work_in_master.objects.filter(labour_voucher_number__challan_no=challan_no)

        total_labour_workout = labour_workout_childs.objects.get(challan_no=challan_no)

        product_instance = get_object_or_404(Product, Product_Refrence_ID=ref_no)
        reference_no = ref_no

        SKU_List = []

        for sku_instance in product_instance.productdetails.all().order_by('PProduct_SKU'):
            SKU_List.append(f'{sku_instance.PProduct_SKU}-{sku_instance.PProduct_color}')

        queryset_list=[]
                
        labour_workout_p_2_i={}

        sku_to_processed_pcs = dict(total_labour_workout.labour_workout_child_items.all().values_list('product_sku', 'processed_pcs'))
        
        for x in SKU_List:

            sku = str(x.split('-')[0])  

            if sku in sku_to_processed_pcs: 

                labour_workout_p_2_i[sku] = sku_to_processed_pcs[sku] 
            
            else:
                labour_workout_p_2_i[sku] = 0


        for instance in labour_workin_instances:
            dict_to_append = {
                'GRN_No': instance.voucher_number,
                'date':instance.created_date,
                'description' : f'LWI = {instance.total_return_pcs}'}
        
            product_qty = {}
            
            sku_to_processed_pcs = dict(instance.l_w_in_products.values_list('product_sku', 'return_pcs'))
            
            for x in SKU_List:

                sku = int(x.split('-')[0])  
                
                if sku in sku_to_processed_pcs: 

                    str_sku =  str(sku)  
                    
                    product_qty[str_sku] = sku_to_processed_pcs[sku] 

                else:
                    str_sku =  str(sku)                    
                    product_qty[str_sku] = 0

            dict_to_append['sku_qty'] = product_qty

            queryset_list.append(dict_to_append)

        report_data_sorted = sorted(queryset_list, key = itemgetter('date'), reverse=False)
        

        total_sku_qty = {}
       
        for i in report_data_sorted:
            for key, value in i['sku_qty'].items():
                if key in total_sku_qty:
                    total_sku_qty[key] += value
                else:
                    total_sku_qty[key] = value
        
        final_dict = {key : labour_workout_p_2_i.get(key, 0) - total_sku_qty.get(key, 0) for key in total_sku_qty.keys()}

        lwo_total_qty = total_labour_workout.total_process_pcs,
        lwo_date = total_labour_workout.created_date 
        lwo_id = challan_no

        return render(request,'production/finishedgoodsvendormodelwisereport.html',{'report_data_sorted':report_data_sorted,'reference_no':reference_no,'product_instance':product_instance,'total_labour_workout':total_labour_workout,'labour_workout_p_2_i':labour_workout_p_2_i, 'SKU_List':SKU_List,'result_dict':final_dict,'lwo_total_qty':lwo_total_qty, 'lwo_date':lwo_date , 'lwo_id': lwo_id, 'total_sku_qty':total_sku_qty})




def finished_goods_color_wise_report(request):
    lwo_processed = product_to_item_labour_child_workout.objects.values(
        'product_sku', 'product_color'
    ).annotate(total_process=Sum('processed_pcs')).order_by('product_sku')
    
    lwi_processed = labour_work_in_product_to_item.objects.values(
        'product_sku', 'product_color'
    ).annotate(total_process=Sum('return_pcs')).order_by('product_sku')
    

    Products = PProduct_Creation.objects.all().values('PProduct_image','PProduct_color__color_name','Product__Product_Refrence_ID','PProduct_SKU','Product__Model_Name')

    first_merge = []

    for lwo in lwo_processed:
        for product in Products:
            sku = product['PProduct_SKU']
            color = product['PProduct_color__color_name']

            if sku == int(lwo['product_sku']) and color == lwo['product_color']:

                total_process_lwi = 0
            
                for lwi in lwi_processed:
                    if int(lwi['product_sku']) == product['PProduct_SKU'] and lwi['product_color'] == color:
                        total_process_lwi = lwi['total_process']
                        break

                first_merge.append({
                    'product_image': product['PProduct_image'] if product else '',
                    'product_color': product['PProduct_color__color_name'] if product else '',
                    'product_sku': sku,
                    'reference_id': product['Product__Product_Refrence_ID'] if product else '',
                    'model_name': product['Product__Model_Name'] if product else '',
                    'total_process': lwo['total_process'],
                    'total_process_lwi': total_process_lwi,
                    'total_balance':lwo['total_process'] - total_process_lwi
                })

    return render(request,'production/finished_goods_color_wise_report.html',{'final_report':first_merge,'MEDIA_URL': settings.MEDIA_URL})


def finished_goods_color_challan_and_grn_vise_report(request,ref_no):

    lwo_instances = labour_workout_childs.objects.filter(labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_no)

    lwo_total = labour_workout_childs.objects.filter(labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_no).aggregate(total_lwo = Sum('labour_workout_child_items__processed_pcs'))['total_lwo'] or 0
    
    lwo_processed = []

    for i in lwo_instances:

        challan_no = i.challan_no
        created_date = i.created_date
        vendor = i.labour_name.name

        v_dict = {}

        v_dict["voucher_data"] = {}

        v_dict["voucher_data"]["created_date"] = created_date
        v_dict["voucher_data"]["vendor"] = vendor
        v_dict["voucher_data"]["challan_no"] = challan_no

        v_dict["sku_data"] = {}

        for item in i.labour_workout_child_items.all():
            sku = int(item.product_sku)
            v_dict["sku_data"][sku] = item.processed_pcs
    
        lwo_processed.append(v_dict)
    
    # print("lwo_processed = ",lwo_processed)


    lwi_instance = labour_work_in_master.objects.filter(labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_no)

    lwi_processed = []

    for i in lwi_instance:

        grn_no = i.voucher_number
        created_date = i.created_date
        vendor = i.labour_voucher_number.labour_name.name

        g_dict = {}

        g_dict["voucher_data"] = {}

        g_dict["voucher_data"]["created_date"] = created_date
        g_dict["voucher_data"]["vendor"] = vendor
        g_dict["voucher_data"]["grn_no"] = grn_no

        g_dict["sku_data"] = {}

        for item in i.l_w_in_products.all():
            sku = int(item.product_sku)
            g_dict["sku_data"][sku] = item.return_pcs

        lwi_processed.append(g_dict)

    # print("lwi_processed = ", lwi_processed)

    skus = PProduct_Creation.objects.filter(Product__Product_Refrence_ID = ref_no)

    final_processed = lwo_processed + lwi_processed
    final_processed = sorted(final_processed, key=lambda x: x['voucher_data']['created_date'])

    print("final_processed = ", final_processed)

    

    return render(request,"production/finished_goods_color_challan_and_grn_vise_report.html",{'skus':skus,'final_processed':final_processed,'lwo_total':lwo_total})









def rawmaterialestimationlist(request):
    product_estimation_instances  = raw_material_production_estimation.objects.all()
    return render(request,'reports/rawmaterialestimationlist.html',{'product_estimation_instances':product_estimation_instances})



def rawmaterialestimationcreateupdate(request,pk=None):

    godown_id = Godown_raw_material.objects.all()
    product_all = Product.objects.all()

    if pk:
        raw_material_production_estimation_instance = raw_material_production_estimation.objects.get(pk=pk)
        product_estimation_form = raw_material_production_estimation_form(instance=raw_material_production_estimation_instance)
        product_estimation_formset  = raw_material_product_estimation_formset_update(instance=raw_material_production_estimation_instance)

        instance_exists_check = raw_material_product_ref_items.objects.filter(raw_material_estimation_master = raw_material_production_estimation_instance).first().raw_material_product_ref_itemss_p_2_i.exists()

    else:
        raw_material_production_estimation_instance = None
        product_estimation_form = raw_material_production_estimation_form()
        product_estimation_formset  = raw_material_product_estimation_formset()
        instance_exists_check = False
        


    if request.method == 'POST':

        product_estimation_form = raw_material_production_estimation_form(request.POST, instance=raw_material_production_estimation_instance)

        product_estimation_formset = raw_material_product_estimation_formset(request.POST, instance=raw_material_production_estimation_instance)

        product_estimation_formset.forms = [form for form in product_estimation_formset if form.has_changed()]
        
        if product_estimation_form.is_valid() and product_estimation_formset.is_valid():
            
            product_estimation_form_instance = product_estimation_form.save()

            for form in product_estimation_formset.deleted_forms:
                    if form.instance.pk:
                        form.instance.delete()

            for form in product_estimation_formset:
                if form.is_valid():
                    if not form.cleaned_data.get('DELETE'):
                        product_e_form = form.save(commit=False)
                        product_e_form.raw_material_estimation_master = product_estimation_form_instance
                        product_e_form.save()
            
            return redirect(reverse('rawmaterial-estimation-update', args=[product_estimation_form_instance.pk]))

    return render(request,'reports/rawmaterialestimationcreate.html',{
                  'product_estimation_formset': product_estimation_formset,
                  'product_estimation_form': product_estimation_form, 
                  'product_all':product_all, 'godown_id':godown_id,
                  'instance_exists_check':instance_exists_check})




def raw_material_estimation_popup(request, pk=None):

    if pk:
        try:
            product_ref_items_instance = raw_material_product_ref_items.objects.get(pk=pk) 

            product_creation_instances = PProduct_Creation.objects.filter(Product=product_ref_items_instance.product_id) 

            initial_p_2_i_dict = []
            for instance in product_creation_instances:
                dict_to_append = {
                    'product_sku' :instance.PProduct_SKU,
                    'product_color':instance.PProduct_color.color_name,
                    'estimate_qty' : '0' 
                    }

                initial_p_2_i_dict.append(dict_to_append)
            
            print(initial_p_2_i_dict)

            product_2_items_instances_unique = product_2_item_through_table.objects.filter(
                                PProduct_pk__Product__Product_Refrence_ID = product_ref_items_instance.product_id.Product_Refrence_ID, common_unique = False).order_by(
                                    'row_number')
        
            product_2_items_instances_common = product_2_item_through_table.objects.filter(
                                PProduct_pk__Product__Product_Refrence_ID = product_ref_items_instance.product_id.Product_Refrence_ID, common_unique = True).order_by(
                                    'row_number','id').distinct('row_number')
        
            
            
            product_2_items_instances = product_2_items_instances_unique.union(product_2_items_instances_common)

            
            initial_p_2_i_items_dict = []
            for query in product_2_items_instances:
                rate_first = query.Item_pk.shades.order_by('id').first() 
                
                
                if query.common_unique == True:
                    product_color_or_common_item = 'Common Item'
                    product_sku_or_common_item = 'Common Item'

                else:
                    product_color_or_common_item = query.PProduct_pk.PProduct_color
                    product_sku_or_common_item = query.PProduct_pk.PProduct_SKU
                
                initial_data_dict = {'product_sku': product_sku_or_common_item,
                                    'product_color' : product_color_or_common_item,
                                    'material_name':query.Item_pk.item_name,
                                    'rate':rate_first.rate,
                                    'panha':query.Item_pk.Panha,
                                    'units':query.Item_pk.Units,
                                    'g_total':query.grand_total,
                                    'g_total_combi':query.grand_total_combi,
                                    'consumption':'0',
                                    'combi_consumption':0,
                                    'total_comsumption':'0',
                                    'unit_value': query.Item_pk.unit_name_item.unit_name,
                                    'physical_stock':'0',
                                    'balance_physical_stock':'0',
                                    'row_number':query.row_number,
                                    'Remark':query.Remark, 
                                    'pcs': '0' }
                
                initial_p_2_i_items_dict.append(initial_data_dict)

            initial_sorted_data = sorted(initial_p_2_i_items_dict, key = itemgetter('row_number'), reverse=False)

            if product_ref_items_instance.raw_material_product_ref_itemss.exists():
                
                raw_material_product_estimation_product_2_item_formset = inlineformset_factory(raw_material_product_ref_items, raw_material_product_wise_qty, 
                                                                    form=raw_material_product_estimation_product_2_item_form, extra = 0, can_delete = False)
                

                raw_material_product_estimation_items_formset = inlineformset_factory(raw_material_product_ref_items, raw_material_product_to_items, 
                                                    form=raw_material_product_estimation_items_form , extra = 0, can_delete = False)
            else:
                
                raw_material_product_estimation_product_2_item_formset = inlineformset_factory(raw_material_product_ref_items, raw_material_product_wise_qty, 
                                                                    form=raw_material_product_estimation_product_2_item_form, extra = len(initial_p_2_i_dict), can_delete = False)
                

                raw_material_product_estimation_items_formset = inlineformset_factory(raw_material_product_ref_items, raw_material_product_to_items, 
                                                    form=raw_material_product_estimation_items_form , extra = len(initial_sorted_data), can_delete = False)


            product_2_item_formset = raw_material_product_estimation_product_2_item_formset(request.POST or None, initial = initial_p_2_i_dict, instance=product_ref_items_instance)

            
            material_product_estimation_items_formset = raw_material_product_estimation_items_formset(request.POST or None, initial=initial_sorted_data, instance=product_ref_items_instance)


        except ObjectDoesNotExist as e:
            logger.error(f"Product reference item with pk={pk} does not exist: {e}")
            return render(request, '404.html')  
        
        except Exception as e:
            logger.exception("An unexpected error occurred while setting up the formsets.")
            return render(request, '500.html')  


    if request.method == 'POST':

        if product_2_item_formset.is_valid() and material_product_estimation_items_formset.is_valid():
            try:
                for form in product_2_item_formset:
                    
                    p_2_i_form = form.save(commit= False)
                    p_2_i_form.raw_material_ref_id = product_ref_items_instance
                    p_2_i_form.save()
                    logger.info(f"Saved product_2_item form instance with id={p_2_i_form.id}")

                for form in material_product_estimation_items_formset:
                    
                    p_2_i_items_form = form.save(commit=False)
                    p_2_i_items_form.raw_material_ref_id = product_ref_items_instance
                    p_2_i_items_form.save()
                    logger.info(f"Saved material_product_estimation_items form instance with id={p_2_i_items_form.id}")
                
                close_window_script = """
                <script>
                window.opener.location.reload(true);  // Reload parent window if needed
                window.close();  // Close current window
                </script>
                """

                return HttpResponse(close_window_script)
            
            except Exception as e:
                logger.exception("Error saving form data in the POST request.")
                return render(request, '500.html')

        else:
            logger.warning("Form validation failed in POST request.")
            logger.debug(f"product_2_item_formset errors: {product_2_item_formset.errors}")
            logger.debug(f"material_product_estimation_items_formset errors: {material_product_estimation_items_formset.errors}")

    return render(request,'reports/raw_material_estimation_popup.html',{
                  'product_2_item_formset': product_2_item_formset,
                  'raw_material_product_estimation_items_formset':material_product_estimation_items_formset,
                  'product_ref_items_instance':product_ref_items_instance})




def labour_workin_approval_split(request,ref_id):

    product_info = Product.objects.get(Product_Refrence_ID = ref_id)
    product_images = PProduct_Creation.objects.filter(Product=product_info)

    queryset = labour_work_in_master.objects.filter(labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_id)

    sku_list = [f'{sku.PProduct_SKU}-{sku.PProduct_color.color_name}' for sku in PProduct_Creation.objects.filter(Product__Product_Refrence_ID=ref_id)]

    
    list_to_send = []
    for query in queryset:
        dict_to_append = {
            'vendor_name' : query.labour_voucher_number.labour_name.name,
            "grn_no" : query.voucher_number,
            'challan_no' : query.labour_voucher_number.challan_no,
            'approve_qty' : [],  
        }
        qty_dict = dict(query.l_w_in_products.all().values_list('product_sku','approved_qty'))

        
        qty = {}

        for sku in sku_list:

            sku_name = int(sku.split('-')[0])

            if sku_name in qty_dict:
                qty[sku_name] = qty_dict[sku_name]

            else:
                qty[sku_name] = 0
                
            dict_to_append['approve_qty'] = qty
        
        list_to_send.append(dict_to_append)

    return render(request,'finished_product/labourworkinapprovalsplit.html',{'list_to_send':list_to_send,'sku_list':sku_list,'product_info':product_info,'product_images':product_images})




def labour_workin_pending_split(request,ref_id):

    product_info = Product.objects.get(Product_Refrence_ID = ref_id)
    product_images = PProduct_Creation.objects.filter(Product=product_info)

    queryset = labour_work_in_master.objects.filter(labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID = ref_id)
    
    sku_list = [f'{sku.PProduct_SKU}-{sku.PProduct_color.color_name}' for sku in PProduct_Creation.objects.filter(Product__Product_Refrence_ID=ref_id)]

    list_to_send = []

    for query in queryset:
        dict_to_append = {
            'vendor_name' : query.labour_voucher_number.labour_name.name,
            "grn_no" : query.voucher_number,
            'challan_no' : query.labour_voucher_number.challan_no,
            'pending_qty' : [],  
        }

        qty_dict = dict(query.l_w_in_products.all().values_list('product_sku','pending_for_approval'))

        qty = {}

        for sku in sku_list:

            sku_name = int(sku.split('-')[0])

            if sku_name in qty_dict:
                qty[sku_name] = qty_dict[sku_name]

            else:
                qty[sku_name] = 0
                
            dict_to_append['pending_qty'] = qty
        
        list_to_send.append(dict_to_append)

    return render(request,'finished_product/labourworkinpendingsplit.html',{'list_to_send':list_to_send,'sku_list':sku_list,'product_info':product_info,'product_images':product_images})










def raw_material_estimation_calculate(request,u_id):
    
    if u_id:

        try:
            estimation_master_instance = get_object_or_404(raw_material_production_estimation,pk = u_id)

        except ObjectDoesNotExist as e:
            print(e)

        for instance in estimation_master_instance.raw_material_production_estimations_total.all():
            instance.delete()
        
        if estimation_master_instance:
            final_total_list = {}
            for master_instance in estimation_master_instance.raw_material_production_estimations.all():
                
                primary_list = []
                
                for master_items in master_instance.raw_material_product_ref_itemss_p_2_i.values('material_name','total_comsumption'):
                    primary_list.append(master_items)

                    
                for x in primary_list:
                    material_name = x['material_name']
                    qty = x['total_comsumption']
                    
                    material_in_dict  = final_total_list.get(material_name)

                    if material_in_dict:
                        final_total_list[material_name] = qty + material_in_dict

                    else:
                        final_total_list[material_name] = qty

            for key, value in final_total_list.items():
                    
                difference_quantity_in_p_o_stage = purchase_order_for_raw_material.objects.filter(material_name=key).aggregate(total_po_qty=Sum('total_comsumption')) 

                total_po_qty = difference_quantity_in_p_o_stage['total_po_qty'] if difference_quantity_in_p_o_stage['total_po_qty'] else 0

                difference_quantity_in_cutting_stage = purchase_order_for_raw_material_cutting_items.objects.filter(material_name=key).aggregate(total_cutting_qty = Sum('total_comsumption'))

                total_cutting_qty = difference_quantity_in_cutting_stage['total_cutting_qty'] if difference_quantity_in_cutting_stage['total_cutting_qty'] else 0

                diffrence_qty = total_po_qty - total_cutting_qty

                
                godown_item_instances_new = item_godown_quantity_through_table.objects.filter(Item_shade_name__items__item_name=key,godown_name=estimation_master_instance.raw_material_godown_id)

                if godown_item_instances_new.exists():
                    godown_qty = godown_item_instances_new.aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
                else:
                    godown_qty = 0

                total_godown_stock = godown_qty
                
                total_balance_stock = godown_qty - value
                
                raw_material_production_total.objects.create(raw_material_estination_master = estimation_master_instance,item_name=key,total_consump=value,godown_stock = total_godown_stock,balance_stock = total_balance_stock)

        response_dict = raw_material_production_total.objects.filter(raw_material_estination_master = estimation_master_instance).values('item_name','total_consump','godown_stock','balance_stock')
        

        ref_id = []

        product_ref_ids = estimation_master_instance.raw_material_production_estimations.values('product_id__Product_Refrence_ID')
    
        for product_ref_id in product_ref_ids:
            for key,val in product_ref_id.items():
                ref_id.append(val)
                

        material_for_ref_id_queryset = product_2_item_through_table.objects.filter(PProduct_pk__Product__Product_Refrence_ID__in = ref_id).select_related('PProduct_pk__Product', 'Item_pk__Item_Color')
        
        material_for_ref_id_list = []

        for material in material_for_ref_id_queryset:
            reference_id = material.PProduct_pk.Product.Product_Refrence_ID
            material_name = material.Item_pk.item_name
            color = material.Item_pk.Item_Color.color_name
            pro_sku = material.PProduct_pk.PProduct_SKU
            pro_fab_grp = material.Item_pk.Fabric_nonfabric
            panha_val = material.Item_pk.Panha
            unit_val = material.Item_pk.Units
            grand_total = material.grand_total
            grand_total_combi = material.grand_total_combi
            consumption = round(grand_total / (panha_val * unit_val),3)
            consumtionCombi = round(grand_total_combi / (panha_val * unit_val),3)
            common_unique = material.common_unique

            set_production_data_dict = {
                'ref_id': reference_id,
                'material_name': material_name,
                'color':color,
                'pro_sku': pro_sku,
                'pro_fab_grp': pro_fab_grp,
                'panha_val':panha_val,
                'unit_val':unit_val,
                'grand_total':grand_total,
                'consumption': consumption,
                'consumtionCombi': consumtionCombi,
                'common_unique': common_unique
                }
            material_for_ref_id_list.append(set_production_data_dict)


        # The goal is to organize materials into a dictionary, where Each key is a ref_id,Each value is a list of dictionaries, with each dictionary containing the material details for that ref_id.
        grouped_by_ref_id_unique = {}

        for material in material_for_ref_id_list:

            ref_no = material['ref_id']

            if ref_no not in grouped_by_ref_id_unique:
                grouped_by_ref_id_unique[ref_no] = []

            if material['common_unique']:

                if not any(m['material_name'] == material['material_name'] for m in grouped_by_ref_id_unique[ref_no]):
                    grouped_by_ref_id_unique[ref_no].append(material)
            else:
                grouped_by_ref_id_unique[ref_no].append(material)


        


        # when only cutting pending at stage of cutting and balance cutting
        purchase_orders = purchase_order.objects.filter(product_reference_number__Product_Refrence_ID__in = ref_id,process_status__gt = 2)

        sku_total_qty_cutting = {}

        if purchase_orders:  
            for order in purchase_orders:
                for x in order.p_o_to_products.all():

                    product_reference_number = x.purchase_order_id.product_reference_number.Product_Refrence_ID
                    sku = x.product_id.PProduct_SKU
                    process_quantity = x.process_quantity
                    
                    if product_reference_number not in sku_total_qty_cutting:
                        sku_total_qty_cutting[product_reference_number] = {}

                    if sku not in sku_total_qty_cutting[product_reference_number]:
                        sku_total_qty_cutting[product_reference_number][sku] = process_quantity
                    else:
                        sku_total_qty_cutting[product_reference_number][sku] += process_quantity

            for product_reference_number, skus in sku_total_qty_cutting.items():
                total_quantity = sum(skus.values())
                skus['total'] = total_quantity


        print('sku_total_qty_cutting = ', sku_total_qty_cutting)




        # when cutting approval pending at stage of cutting inside
        cutting_pending_for_lwo_approval = purchase_order_to_product_cutting.objects.filter(purchase_order_cutting_id__purchase_order_id__product_reference_number__Product_Refrence_ID__in=ref_id).filter(~Q(cutting_quantity=F('approved_pcs'))).filter(purchase_order_cutting_id__cutting_cancelled = False)

        sku_total_qty_cutting_lwo_aprv_pending = {}

        if cutting_pending_for_lwo_approval:
            for x in cutting_pending_for_lwo_approval:
                product_reference_number = x.purchase_order_cutting_id.purchase_order_id.product_reference_number.Product_Refrence_ID
                sku = x.product_sku
                process_quantity = x.cutting_quantity if x.balance_pcs == 0 else x.balance_pcs

                if product_reference_number not in sku_total_qty_cutting_lwo_aprv_pending:
                    sku_total_qty_cutting_lwo_aprv_pending[product_reference_number] = {}

                if sku not in sku_total_qty_cutting_lwo_aprv_pending[product_reference_number]:
                    sku_total_qty_cutting_lwo_aprv_pending[product_reference_number][sku] = process_quantity
                else:
                    sku_total_qty_cutting_lwo_aprv_pending[product_reference_number][sku] += process_quantity

            for product_reference_number, skus in sku_total_qty_cutting_lwo_aprv_pending.items():
                total_quantity = sum(skus.values())
                skus['total'] = total_quantity

        print('sku_total_qty_cutting_lwo_aprv_pending',sku_total_qty_cutting_lwo_aprv_pending)



       
        # when only lWO pending at stage of lwo 
        lwo_pending = product_to_item_labour_workout.objects.filter(labour_workout__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID__in = ref_id)

        sku_total_qty_lwo_pending = {}

        if lwo_pending:
            for x in lwo_pending:
                product_reference_number = x.labour_workout.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID
                sku = x.product_sku
                process_quantity = x.pending_pcs

                if product_reference_number not in sku_total_qty_lwo_pending:
                    sku_total_qty_lwo_pending[product_reference_number] = {}
                    
                if sku not in sku_total_qty_lwo_pending[product_reference_number]:
                    sku_total_qty_lwo_pending[product_reference_number][sku] = process_quantity
                    
                else:
                    sku_total_qty_lwo_pending[product_reference_number][sku] += process_quantity

            for product_reference_number, skus in sku_total_qty_lwo_pending.items():
                total_quantity = sum(skus.values())
                skus['total'] = total_quantity

        print('sku_total_qty_lwo_pending = ', sku_total_qty_lwo_pending)



       


        dataset_to_send = []

        list_to_send_for_cutting = []

        list_to_send_for_cutting_aprv_pending = []

        list_to_send_for_lwo = []

        if sku_total_qty_cutting and sku_total_qty_lwo_pending and sku_total_qty_cutting_lwo_aprv_pending:

            print("******* IN ALL *******")

            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }
                                    
                                    # if we have multiple ref_id then it is possible that the materials are duplicate 
                                    # this code is for this condition when the material is duplicate then sum this material cutting_consumption
                                    existing_item = next((item for item in list_to_send_for_cutting if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        list_to_send_for_cutting.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    # if we have multiple ref_id then it is possible that the materials are duplicate 
                                    # this code is for this condition when the material is duplicate then sum this material cutting_consumption
                                    existing_item = next((item for item in list_to_send_for_cutting if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        list_to_send_for_cutting.append(dict_append)

            # this code goal is if any ref_id is not in the pending stage but its still in the rm calculation then add this ref id material in list_to_send_for_cutting with cutting_consumption 0 
            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting):

                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_consumption': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting.append(dict_append)
            


            for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


            for key, value in sku_total_qty_cutting_lwo_aprv_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if data['pro_fab_grp'] == 'Non Fabric':

                                if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting_aprv_pending):
                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_con_aprv_pen': 0,
                                        'common_unique': data['common_unique']
                                    }

                                    list_to_send_for_cutting_aprv_pending.append(dict_append)
          






            for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


            for key, value in sku_total_qty_lwo_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if data['pro_fab_grp'] == 'Non Fabric':

                                if all(data['material_name'] != item['material_name'] for item in list_to_send_for_lwo):
                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'lwo_consumption': 0,
                                        'common_unique': data['common_unique']
                                    }

                                    list_to_send_for_lwo.append(dict_append)


        

            merge_two_list = []

            for x in list_to_send_for_cutting:
                material_name = x['material_name']
                match_found = False 
                
                for y in list_to_send_for_lwo:
                    if material_name == y['material_name']:
                        dict_to_append = {
                            'material_name': x['material_name'],
                            'cutting_consumption': x['cutting_consumption'],
                            'lwo_consumption': y['lwo_consumption'],
                            'color': x['color'],
                        }
                        merge_two_list.append(dict_to_append)
                        match_found = True
                        break
                
                if not match_found:
                    dict_to_append = {
                        'material_name': x['material_name'],
                        'cutting_consumption': x['cutting_consumption'],
                        'lwo_consumption': 0,
                        'color': x['color'],
                    }
                    merge_two_list.append(dict_to_append)

            merge_three_data = []
            for x in merge_two_list:
                material_name = x['material_name']
                match_found = False

                for y in list_to_send_for_cutting_aprv_pending:
                    if material_name == y['material_name']:
                        dict_to_append = {
                            'material_name': x['material_name'],
                            'cutting_consumption': x['cutting_consumption'],
                            'lwo_consumption': x['lwo_consumption'],
                            'cutting_con_aprv_pen': y['cutting_con_aprv_pen'],
                            'color': x['color'],
                        }
                        merge_three_data.append(dict_to_append)
                        match_found = True
                        break
                
                if not match_found:
                    dict_to_append = {
                        'material_name': x['material_name'],
                        'cutting_consumption': x['cutting_consumption'],
                        'lwo_consumption': x['lwo_consumption'],
                        'cutting_con_aprv_pen': 0,
                        'color': x['color'],
                    }
                    merge_three_data.append(dict_to_append)



            for i in response_dict:
                    material_name = i['item_name']
                    match_found = False
                    
                    item = get_object_or_404(Item_Creation, item_name = i['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']

                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names:
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None


                    for j in merge_three_data:
                        if material_name == j['material_name']:

                            dict_to_append = {
                                'item_id':item.id,
                                'material_name':i['item_name'],
                                'color': j['color'],
                                'cutting_consumption': 0 if j['cutting_consumption'] == 0.000 else j['cutting_consumption'],
                                'lwo_consumption': 0 if j['lwo_consumption'] == 0.000 else  j['lwo_consumption'],
                                'cutting_con_aprv_pen': 0 if j['cutting_con_aprv_pen'] == 0.000 else j['cutting_con_aprv_pen'],
                                'total_consump':i['total_consump'],
                                'godown_stock': i['godown_stock'],
                                'po_qty':po_qty,
                                'balance_stock': (i['godown_stock']) - ((i['total_consump']) + (j['cutting_consumption']) + (j['cutting_con_aprv_pen']) + (j['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                                'party_name' : p_name,
                                'mobile_no' : mobile,
                            }
                            dataset_to_send.append(dict_to_append)
                            match_found = True
                            break
                    if not match_found:
                        dict_to_append = {
                                'item_id':item.id,
                                'material_name':i['item_name'],
                                'color': j['color'],
                                'cutting_consumption': 0,
                                'lwo_consumption': 0,
                                'cutting_con_aprv_pen': 0,
                                'total_consump':i['total_consump'],
                                'godown_stock': i['godown_stock'],
                                'po_qty':po_qty,
                                'balance_stock': (i['godown_stock']) - ((i['total_consump']) + (0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                                'party_name' : p_name,
                                'mobile_no' : mobile,
                            }
                        dataset_to_send.append(dict_to_append)


# *******************************************************************************************
              
        elif sku_total_qty_cutting and sku_total_qty_cutting_lwo_aprv_pending:
            
            print("IN CUTTING PENDING AND CUTTING APPROVE PENDING")

            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }
                                    
                                    # if we have multiple ref_id then it is possible that the materials are duplicate 
                                    # this code is for this condition when the material is duplicate then sum this material cutting_consumption
                                    existing_item = next((item for item in list_to_send_for_cutting if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        list_to_send_for_cutting.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    # if we have multiple ref_id then it is possible that the materials are duplicate 
                                    # this code is for this condition when the material is duplicate then sum this material cutting_consumption
                                    existing_item = next((item for item in list_to_send_for_cutting if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        list_to_send_for_cutting.append(dict_append)

            # this code goal is if any ref_id is not in the pending stage but its still in the rm calculation then add this ref id material in list_to_send_for_cutting with cutting_consumption 0 
            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting):

                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_consumption': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting.append(dict_append)


            for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen': total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in list_to_send_for_cutting_aprv_pending
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen': total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in list_to_send_for_cutting_aprv_pending
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

            for key, value in sku_total_qty_cutting_lwo_aprv_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting_aprv_pending):
                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_con_aprv_pen': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting_aprv_pending.append(dict_append)
            
            merge_two_list = []

            for x in list_to_send_for_cutting:
                material_name = x['material_name']
                match_found = False 
                
                for y in list_to_send_for_cutting_aprv_pending:
                    if material_name == y['material_name']:
                        dict_to_append = {
                            'material_name': x['material_name'],
                            'cutting_consumption': x['cutting_consumption'],
                            'cutting_con_aprv_pen': y['cutting_con_aprv_pen'],
                            'color': x['color'],
                        }
                        merge_two_list.append(dict_to_append)
                        match_found = True
                        break
                
                if not match_found:
                    dict_to_append = {
                        'material_name': x['material_name'],
                        'cutting_consumption': x['cutting_consumption'],
                        'cutting_con_aprv_pen': 0,
                        'color': x['color'],
                    }
                    merge_two_list.append(dict_to_append)


            for x in response_dict:
                material_name = x['item_name']
                match_found = False
                for y in merge_two_list:
                    
                    item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']
                    
                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names: 
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None

                    if material_name == y['material_name']:
                        dict_to_append = {
                            'item_id':item.id,
                            'material_name':x['item_name'],
                            'color': y['color'],
                            'total_consump':x['total_consump'],
                            'cutting_consumption': 0 if y['cutting_consumption'] == 0.000 else y['cutting_consumption'],
                            'cutting_con_aprv_pen': 0 if y['cutting_con_aprv_pen'] == 0.000 else y['cutting_con_aprv_pen'],
                            'lwo_consumption': 0,
                            'godown_stock': x['godown_stock'],
                            'po_qty':po_qty,
                            'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (y['cutting_consumption']) + (y['cutting_con_aprv_pen']) + (0) - (po_qty if po_qty is not None else 0)),
                            'party_name' : p_name,
                            'mobile_no' : mobile,
                        }
                        dataset_to_send.append(dict_to_append)
                        match_found = True
                        break
                # if not match_found:
                #     dict_to_append = {
                #             'item_id':item.id,
                #             'material_name':x['item_name'],
                #             'total_consump':x['total_consump'],
                #             'cutting_consumption': 0,
                #             'lwo_consumption': 0,
                #             'godown_stock': x['godown_stock'],
                #             'balance_stock': x['balance_stock'],
                #             'party_name' : p_name,
                #             'mobile_no' : mobile,
                #         }
                #     dataset_to_send.append(dict_to_append)


# *******************************************************************************************
             
        elif sku_total_qty_cutting and sku_total_qty_lwo_pending:
            
            print("IN CUTTING PENDING AND LWO PENDING")

            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }
                                    
                                    # if we have multiple ref_id then it is possible that the materials are duplicate 
                                    # this code is for this condition when the material is duplicate then sum this material cutting_consumption
                                    existing_item = next((item for item in list_to_send_for_cutting if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        list_to_send_for_cutting.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    # if we have multiple ref_id then it is possible that the materials are duplicate 
                                    # this code is for this condition when the material is duplicate then sum this material cutting_consumption
                                    existing_item = next((item for item in list_to_send_for_cutting if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        list_to_send_for_cutting.append(dict_append)

            # this code goal is if any ref_id is not in the pending stage but its still in the rm calculation then add this ref id material in list_to_send_for_cutting with cutting_consumption 0 
            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting):

                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_consumption': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting.append(dict_append)



            for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


            for key, value in sku_total_qty_lwo_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if data['pro_fab_grp'] == 'Non Fabric':

                                if all(data['material_name'] != item['material_name'] for item in list_to_send_for_lwo):
                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'lwo_consumption': 0,
                                        'common_unique': data['common_unique']
                                    }

                                    list_to_send_for_lwo.append(dict_append)


            merge_two_list = []

            for x in list_to_send_for_cutting:
                material_name = x['material_name']
                match_found = False 
                
                for y in list_to_send_for_lwo:
                    if material_name == y['material_name']:
                        dict_to_append = {
                            'material_name': x['material_name'],
                            'cutting_consumption': x['cutting_consumption'],
                            'lwo_consumption': y['lwo_consumption'],
                            'color': x['color'],
                        }
                        merge_two_list.append(dict_to_append)
                        match_found = True
                        break
                
                if not match_found:
                    dict_to_append = {
                        'material_name': x['material_name'],
                        'cutting_consumption': x['cutting_consumption'],
                        'lwo_consumption': 0,
                        'color': x['color'],
                    }
                    merge_two_list.append(dict_to_append)


            for x in response_dict:
                material_name = x['item_name']
                match_found = False
                for y in merge_two_list:
                    
                    item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']
                    
                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names: 
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None

                    if material_name == y['material_name']:
                        dict_to_append = {
                            'item_id':item.id,
                            'material_name':x['item_name'],
                            'color': y['color'],
                            'total_consump':x['total_consump'],
                            'cutting_consumption': 0 if y['cutting_consumption'] == 0.000 else y['cutting_consumption'],
                            'cutting_con_aprv_pen': 0,
                            'lwo_consumption': 0 if y['lwo_consumption'] == 0.000 else y['lwo_consumption'],
                            'godown_stock': x['godown_stock'],
                            'po_qty':po_qty,
                            'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (y['cutting_consumption']) + (0) + (y['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                            'party_name' : p_name,
                            'mobile_no' : mobile,
                        }
                        dataset_to_send.append(dict_to_append)
                        match_found = True
                        break

                # if not match_found:
                #     dict_to_append = {
                #             'item_id':item.id,
                #             'material_name':x['item_name'],
                #             'total_consump':x['total_consump'],
                #             'cutting_consumption': 0,
                #             'lwo_consumption': 0,
                #             'godown_stock': x['godown_stock'],
                #             'balance_stock': x['balance_stock'],
                #             'party_name' : p_name,
                #             'mobile_no' : mobile,
                #         }
                #     dataset_to_send.append(dict_to_append)


#**************************************************************************************************

        elif sku_total_qty_cutting_lwo_aprv_pending and sku_total_qty_lwo_pending:

            print("IN CUTTING APPROVE PENDING AND LWO PENDING")

            for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen': total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in list_to_send_for_cutting_aprv_pending
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen': total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in list_to_send_for_cutting_aprv_pending
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


            for key, value in sku_total_qty_cutting_lwo_aprv_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting_aprv_pending):
                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_con_aprv_pen': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting_aprv_pending.append(dict_append)


            for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


            for key, value in sku_total_qty_lwo_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if data['pro_fab_grp'] == 'Non Fabric':

                                if all(data['material_name'] != item['material_name'] for item in list_to_send_for_lwo):
                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'lwo_consumption': 0,
                                        'common_unique': data['common_unique']
                                    }

                                    list_to_send_for_lwo.append(dict_append)

            merge_two_list = []

            for x in list_to_send_for_cutting_aprv_pending:
                material_name = x['material_name']
                match_found = False 
                
                for y in list_to_send_for_lwo:
                    if material_name == y['material_name']:
                        dict_to_append = {
                            'material_name': x['material_name'],
                            'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                            'lwo_consumption': y['lwo_consumption'],
                            'color': x['color'],
                        }
                        merge_two_list.append(dict_to_append)
                        match_found = True
                        break
                
                if not match_found:
                    dict_to_append = {
                        'material_name': x['material_name'],
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': 0,
                        'color': x['color'],
                    }
                    merge_two_list.append(dict_to_append)


            for x in response_dict:
                material_name = x['item_name']
                match_found = False
                for y in merge_two_list:
                    
                    item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']
                    
                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names: 
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None

                    if material_name == y['material_name']:
                        dict_to_append = {
                            'item_id':item.id,
                            'material_name':x['item_name'],
                            'color': y['color'],
                            'total_consump':x['total_consump'],
                            'cutting_consumption': 0,
                            'cutting_con_aprv_pen': 0 if y['cutting_con_aprv_pen'] == 0.000 else y['cutting_con_aprv_pen'],
                            'lwo_consumption': 0 if y['lwo_consumption'] == 0.000 else y['lwo_consumption'],
                            'godown_stock': x['godown_stock'],
                            'po_qty':po_qty,
                            'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (0) + (y['cutting_con_aprv_pen']) + (y['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                            'party_name' : p_name,
                            'mobile_no' : mobile,

                        }
                        dataset_to_send.append(dict_to_append)
                        match_found = True
                        break

                # if not match_found:
                #     dict_to_append = {
                #             'item_id':item.id,
                #             'material_name':x['item_name'],
                #             'total_consump':x['total_consump'],
                #             'cutting_consumption': 0,
                #             'lwo_consumption': 0,
                #             'godown_stock': x['godown_stock'],
                #             'balance_stock': x['balance_stock'],
                #             'party_name' : p_name,
                #             'mobile_no' : mobile,
                #         }
                #     dataset_to_send.append(dict_to_append)


        elif sku_total_qty_cutting:

            print("in only cutting")

            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next(
                                        (item for item in list_to_send_for_cutting
                                        if item['material_name'] == data['material_name'] and
                                        item['common_unique'] == data['common_unique']),
                                        None
                                    )

                                    if existing_item:
                                        
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        
                                        list_to_send_for_cutting.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next(
                                        (item for item in list_to_send_for_cutting
                                        if item['material_name'] == data['material_name'] and
                                        item['common_unique'] == data['common_unique']),
                                        None
                                    )

                                    if existing_item:
                                        
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        
                                        list_to_send_for_cutting.append(dict_append)


            for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting):

                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_consumption': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting.append(dict_append)



            for x in response_dict:
                material_name = x['item_name']
                match_found = False
                for y in list_to_send_for_cutting:
                    
                    item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']

                    
                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names: 
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None

                    if material_name == y['material_name']:
                        dict_to_append = {
                            'item_id':item.id,
                            'material_name':x['item_name'],
                            'product_sku':y['product_sku'],
                            'color':y['color'],
                            'total_consump':x['total_consump'],
                            'cutting_consumption': 0 if y['cutting_consumption'] == 0.000 else y['cutting_consumption'],
                            'lwo_consumption': 0,
                            'cutting_con_aprv_pen': 0,
                            'godown_stock': x['godown_stock'],
                            'po_qty':po_qty,
                            'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (y['cutting_consumption']) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                            'party_name' : p_name,
                            'mobile_no' : mobile,
                        }
                        dataset_to_send.append(dict_to_append)
                        match_found = True
                        break
                # if not match_found:
                #     dict_to_append = {
                #             'item_id':item.id,
                #             'material_name':x['item_name'],
                #             'total_consump':x['total_consump'],
                #             'cutting_consumption': 0,
                #             'lwo_consumption': 0,
                #             'cutting_con_aprv_pen': 0,
                #             'godown_stock': x['godown_stock'],
                #             'balance_stock': x['balance_stock'],
                #             'party_name' : p_name,
                #             'mobile_no' : mobile,
                #             'po_exist':po_exist
                #         }
                #     dataset_to_send.append(dict_to_append)


        elif sku_total_qty_cutting_lwo_aprv_pending:

            print("IN ONLY CUTTING APPROVE PENDING")

            for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen': total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in list_to_send_for_cutting_aprv_pending
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen': total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in list_to_send_for_cutting_aprv_pending
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


            for key, value in sku_total_qty_cutting_lwo_aprv_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if all(data['material_name'] != item['material_name'] for item in list_to_send_for_cutting_aprv_pending):
                                dict_append = {
                                    'ref_id': data['ref_id'],
                                    'product_sku': data['pro_sku'],
                                    'color': data['color'],
                                    'material_name': data['material_name'],
                                    'product_fabric_grp': data['pro_fab_grp'],
                                    'cutting_con_aprv_pen': 0,
                                    'common_unique': data['common_unique']
                                }

                                list_to_send_for_cutting_aprv_pending.append(dict_append)

            for x in response_dict:
                material_name = x['item_name']
                match_found = False
                for y in list_to_send_for_cutting_aprv_pending:
                    
                    item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']
                    
                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names: 
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None

                    if material_name == y['material_name']:
                        dict_to_append = {
                            'item_id':item.id,
                            'material_name':x['item_name'],
                            'product_sku':y['product_sku'],
                            'color':y['color'],
                            'total_consump':x['total_consump'],
                            'cutting_consumption': 0,
                            'lwo_consumption': 0,
                            'cutting_con_aprv_pen': 0 if y['cutting_con_aprv_pen'] == 0.000 else y['cutting_con_aprv_pen'],
                            'godown_stock': x['godown_stock'],
                            'po_qty':po_qty,
                            'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (0) + (y['cutting_con_aprv_pen']) + (0) - (po_qty if po_qty is not None else 0)),
                            'party_name' : p_name,
                            'mobile_no' : mobile,
                        }
                        dataset_to_send.append(dict_to_append)
                        match_found = True
                        break
                # if not match_found:
                #     dict_to_append = {
                #             'item_id':item.id,
                #             'material_name':x['item_name'],
                #             'total_consump':x['total_consump'],
                #             'cutting_consumption': 0,
                #             'lwo_consumption': 0,
                #             'cutting_con_aprv_pen': 0,
                #             'godown_stock': x['godown_stock'],
                #             'balance_stock': x['balance_stock'],
                #             'party_name' : p_name,
                #             'mobile_no' : mobile,
                #             'po_exist':po_exist
                #         }
                #     dataset_to_send.append(dict_to_append)
        

        elif sku_total_qty_lwo_pending:

            print("IN ONLY LWO PENDING")

            for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_lwo if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            list_to_send_for_lwo.append(dict_append)


            for key, value in sku_total_qty_lwo_pending.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] != key:

                            if data['pro_fab_grp'] == 'Non Fabric':

                                if all(data['material_name'] != item['material_name'] for item in list_to_send_for_lwo):
                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'lwo_consumption': 0,
                                        'common_unique': data['common_unique']
                                    }

                                    list_to_send_for_lwo.append(dict_append)

            for x in response_dict:
                material_name = x['item_name']
                match_found = False
                for y in list_to_send_for_lwo:
                    
                    item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                    po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']
                    
                    party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                    if party_names: 
                        p_name = party_names.item_purchase_master.party_name.name
                        mobile = party_names.item_purchase_master.party_name.mobile_no
                    else:
                        p_name = None
                        mobile = None

                    if material_name == y['material_name']:
                        dict_to_append = {
                            'item_id':item.id,
                            'material_name':x['item_name'],
                            'product_sku':y['product_sku'],
                            'color':y['color'],
                            'total_consump':x['total_consump'],
                            'cutting_consumption': 0,
                            'lwo_consumption': 0 if y['lwo_consumption'] == 0.000 else y['lwo_consumption'],
                            'cutting_con_aprv_pen': 0,
                            'godown_stock': x['godown_stock'],
                            'po_qty':po_qty,
                            'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (0) + (0) + (y['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                            'party_name' : p_name,
                            'mobile_no' : mobile,

                        }
                        dataset_to_send.append(dict_to_append)
                        match_found = True
                        break
                # if not match_found:
                #     dict_to_append = {
                #             'item_id':item.id,
                #             'material_name':x['item_name'],
                #             'total_consump':x['total_consump'],
                #             'cutting_consumption': 0,
                #             'lwo_consumption': 0,
                #             'cutting_con_aprv_pen': 0,
                #             'godown_stock': x['godown_stock'],
                #             'balance_stock': x['balance_stock'],
                #             'party_name' : p_name,
                #             'mobile_no' : mobile,
                #             'po_exist':po_exist
                #         }
                #     dataset_to_send.append(dict_to_append)


        else:
            print("******* IN NOTHING *******")

            for x in response_dict:
                              
                item = get_object_or_404(Item_Creation, item_name=x['item_name'])

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=x['item_name']).aggregate(total_qty=Sum('quantity'))['total_qty']
                
                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = x['item_name']).select_related('item_purchase_master__party_name').order_by('-created_date').first()
                    
                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None
                    

                dict_to_apnd = {
                    # 'id':dataset['id'],
                    'item_id' : item.id,
                    'material_name':item.item_name,
                    'color':item.Item_Color.color_name,
                    'total_consump':x['total_consump'],
                    'cutting_consumption': 0,
                    'lwo_consumption': 0,
                    'cutting_con_aprv_pen':0,
                    'godown_stock': x['godown_stock'],
                    'po_qty':po_qty,
                    'balance_stock': (x['godown_stock']) - ((x['total_consump']) + (0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
                    'mobile_no' : mobile,

                }

                dataset_to_send.append(dict_to_apnd)

        sorted_list = sorted(dataset_to_send,key = itemgetter('material_name'))
        for_excel = json.dumps(dataset_to_send, default=custom_serializer)

        
    return render(request,'reports/raw_material_estimation_calculation_pop_up.html',{'final_data': sorted_list,'for_excel':for_excel})





   
def custom_serializer(obj):
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    if obj is None:
        return None
    raise TypeError("Type not serializable")





def raw_material_estimate_delete(request,pk):

    try:
        raw_material_estimation_instance = get_object_or_404(raw_material_production_estimation,pk = pk)
        raw_material_estimation_instance.delete()

        messages.success(request,f'Account Sub Group {raw_material_estimation_instance.id} was deleted')

    except IntegrityError as e:
        messages.error(request,f'Cannot delete {raw_material_estimation_instance.id} because it is referenced by other objects. {e}')
    return redirect('rawmaterial-estimation-list')
                






def raw_material_estimation_calculate_excel_download(request):

    final_dict = request.GET.get('finalDic')
    
    data = json.loads(final_dict)
    
    wb = Workbook()
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    wb.create_sheet('Product Estimation')
    sheet1 = wb.worksheets[0]
    headers = ['Material Name', 'Total Consumption', 'Cutting Consumption','Cutting Aprv Pend Qty','LWO Consumption', 'Godown Stock', 'Balance Stock', 'Party Name', 'Mobile No']
    sheet1.append(headers)

    column_widths = [40, 20, 20, 20, 20, 15, 15, 25, 15]
    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)  
        sheet1.column_dimensions[column_letter].width = width

    for item in data:
        sheet1.append([item.get('material_name'), item.get('total_consump'),
                       item.get('cutting_consumption'), item.get('cutting_con_aprv_pen'), item.get('lwo_consumption'), item.get('godown_stock'),
                       item.get('balance_stock'), item.get('party_name'), item.get('mobile_no')])
    
    
    file_output = BytesIO()
    wb.save(file_output)
    file_output.seek(0)  # Rewind to the beginning of the file

    # Return a binary response
    response = HttpResponse(file_output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Product_Estimation.xlsx"'
    
    return response

        
















def product_purchase_voucher_create_update(request, pk=None):

    products = PProduct_Creation.objects.all()
    warehouses = Finished_goods_warehouse.objects.all()
    party_names = Ledger.objects.filter(under_group__account_sub_group = 'Sundry Creditors')
    
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':

        sku = request.GET.get('productName')
        print(sku)
        pro = PProduct_Creation.objects.get(PProduct_SKU = sku)
        color = pro.PProduct_color.color_name
        gst = pro.Product.Product_GST.gst_percentage
        return JsonResponse({'color':color, 'gst':gst})
    
    
    if pk:
        product_pur_vouch_instance = product_purchase_voucher_master.objects.filter(pk=pk).annotate(
            total_qc_qty=Sum('product_purchase_voucher_items__qc_recieved_qty')).first()
        
        product_pur_vouch_form = product_purchase_voucher_master_form(instance=product_pur_vouch_instance)
        product_purchase_voucher_items_formset_instance = product_purchase_voucher_items_formset_update(instance=product_pur_vouch_instance)
        
    else:
        product_pur_vouch_instance = None
        product_pur_vouch_form = product_purchase_voucher_master_form()
        product_purchase_voucher_items_formset_instance = product_purchase_voucher_items_formset()
    

    if request.method == 'POST':
        print(request.POST)
        product_pur_vouch_form = product_purchase_voucher_master_form(request.POST,instance=product_pur_vouch_instance)

        product_purchase_voucher_items_formset_instance = product_purchase_voucher_items_formset(request.POST,instance=product_pur_vouch_instance)

        product_purchase_voucher_items_formset_instance.forms = [form for form in product_purchase_voucher_items_formset_instance.forms if form.has_changed()]

        if product_pur_vouch_form.is_valid() and product_purchase_voucher_items_formset_instance.is_valid():
           
            try:
                with transaction.atomic():
                    product_pur_vouch_form_instance = product_pur_vouch_form.save()
                    
                    selected_warehouse = product_pur_vouch_form_instance.finished_godowns

                    for form in product_purchase_voucher_items_formset_instance.deleted_forms:
                        if form.instance.pk:
                            form.instance.delete()

                    for form in product_purchase_voucher_items_formset_instance:

                        if form.instance.pk and not form.cleaned_data.get('DELETE'): 
                            
                            
                            product_purchase_voucher_items_form = form.save(commit=False)
                            product_purchase_voucher_items_form.product_purchase_master = product_pur_vouch_form_instance
                            product_purchase_voucher_items_form.diffrence_qty = product_purchase_voucher_items_form.quantity_total

                            old_quantity = form.initial.get('quantity_total')  

                            
                            if form.has_changed() and 'product_name' in form.changed_data: 

                                old_product_name = form.initial.get('product_name') 
                                
                                
                                if old_product_name and old_quantity:
                                    
                                    obj1, created = Product_warehouse_quantity_through_table.objects.get_or_create(
                                    warehouse= selected_warehouse, 
                                    product = old_product_name )
                                    obj1.quantity = obj1.quantity - old_quantity 
                                    obj1.save()

                                    obj2, created = Product_warehouse_quantity_through_table.objects.get_or_create(
                                    warehouse= selected_warehouse , 
                                    product = product_purchase_voucher_items_form.product_name)

                                    obj2.quantity = obj1.quantity + product_purchase_voucher_items_form.quantity_total 
                                    obj2.save()

                            else:  

                                diff_qty_to_deduct = product_purchase_voucher_items_form.quantity_total - old_quantity 
                                
                                obj, created = Product_warehouse_quantity_through_table.objects.get_or_create(
                                    warehouse= selected_warehouse, 
                                    product = product_purchase_voucher_items_form.product_name)
                                
                                if created:
                                    obj.quantity = diff_qty_to_deduct
                                
                                else:
                                    obj.quantity = obj.quantity + diff_qty_to_deduct 

                                obj.save()

                            product_purchase_voucher_items_form.save()

                        elif not form.instance.pk: 
                            if not form.cleaned_data.get('DELETE'):

                                product_purchase_voucher_items_form = form.save(commit=False)
                                product_purchase_voucher_items_form.product_purchase_master = product_pur_vouch_form_instance
                                product_purchase_voucher_items_form.diffrence_qty = product_purchase_voucher_items_form.quantity_total
                                
                                obj, created = Product_warehouse_quantity_through_table.objects.get_or_create(
                                    warehouse= selected_warehouse, 
                                    product = product_purchase_voucher_items_form.product_name)
                            
                                if created:
                                    obj.quantity = product_purchase_voucher_items_form.quantity_total
                                
                                else:
                                    obj.quantity = obj.quantity + product_purchase_voucher_items_form.quantity_total 

                                obj.save()

                                product_purchase_voucher_items_form.save()

                    return redirect('product-purchase-voucher-list')
                
            except Exception as e:
                print(e)

        else:
            print(product_pur_vouch_form.errors)
            print(product_purchase_voucher_items_formset_instance.errors)
            print(product_purchase_voucher_items_formset_instance.non_form_errors())

    return render(request,'finished_product/product_purchase_voucher_create_update.html',{'product_pur_vouch_form':product_pur_vouch_form,'product_pur_vouch_instance':product_pur_vouch_instance,
            'product_purchase_voucher_items_formset_instance' : product_purchase_voucher_items_formset_instance,'products':products,'warehouses':warehouses,'party_names':party_names})







def product_purchase_voucher_list(request):

    product_purchase_voucher_all = product_purchase_voucher_master.objects.all().annotate(check_diff_qty = Sum('product_purchase_voucher_items__qc_recieved_qty'),total_qty = Sum('product_purchase_voucher_items__quantity_total')).order_by('created_date')

    return render(request,'finished_product/product_purchase_voucher_list.html',{'product_purchase_voucher_all':product_purchase_voucher_all})





def product_purchase_voucher_delete(request,pk):
    if pk:
        product_purchase_voucher_instance = get_object_or_404(product_purchase_voucher_master,pk=pk)
        product_purchase_voucher_instance.delete()
        return redirect('product-purchase-voucher-list')
    


@login_required(login_url = 'login')
def warehouse_product_transfer_create_and_update(request,pk=None):
    
    products = PProduct_Creation.objects.all()
    godowns = Godown_finished_goods.objects.all()
    warehouses =Finished_goods_warehouse.objects.all()

    dict_to_send = None

    if pk:
        voucher_instance = Finished_goods_Stock_TransferMaster.objects.filter(pk=pk).annotate(qc_all_qty = Sum('finished_goods_transfer_records__qc_recieved_qty')).first()
        form = Finished_goods_Stock_TransferMaster_form(instance=voucher_instance)
        formset = Finished_goods_transfer_records_formset_update(instance=voucher_instance)
        godown_id = voucher_instance.source_warehouse.id

        filtered_product = list(product_godown_quantity_through_table.objects.filter(
            godown_name__id = godown_id).values('product_color_name__Product__Product_Name','product_color_name__PProduct_SKU','product_color_name__PProduct_color__color_name','quantity','product_color_name__Product__Model_Name','product_color_name__Product__Product_Refrence_ID','product_color_name__Product__Product_UOM','product_color_name__Product__Product_GST__gst_percentage'))
        
        if filtered_product:
            
            dict_to_send = {}

            for query in filtered_product:
                ref_no = query.get('product_color_name__Product__Product_Refrence_ID')
                p_sku = query.get('product_color_name__PProduct_SKU')
                product_name = query.get('product_color_name__Product__Product_Name')
                product_model_name = query.get('product_color_name__Product__Model_Name')
                color = query.get('product_color_name__PProduct_color__color_name')
                uom = query.get('product_color_name__Product__Product_UOM')
                qty = query.get('quantity')
                gst = query.get('product_color_name__Product__Product_GST__gst_percentage')
                
                dict_to_send[p_sku] = [product_name,color,qty,product_model_name,ref_no,uom,gst]

    else:

        voucher_instance = None
        form = Finished_goods_Stock_TransferMaster_form()
        formset = Finished_goods_transfer_records_formset_create()

    
    if request.method == 'POST':

        form = Finished_goods_Stock_TransferMaster_form(request.POST,instance=voucher_instance)

        if pk:
            formset = Finished_goods_transfer_records_formset_update(request.POST,instance=voucher_instance)
        else:
            formset = Finished_goods_transfer_records_formset_create(request.POST,instance=voucher_instance)

        formset.forms = [form for form in formset.forms if form.has_changed()]

        if not form.is_valid():
            print("Form Errors:", form.errors)

        if not formset.is_valid():
            print("Formset Errors:", formset.errors)

        if form.is_valid() and formset.is_valid():

            try:

                with transaction.atomic():

                    first_form_instance = form.save(commit=False)
                    first_form_instance.save()

                    selected_godown = first_form_instance.source_warehouse
                    selected_warehouse = first_form_instance.destination_warehouse

                    for form in formset.deleted_forms:

                        if form.instance.pk:
            
                            old_product = form.instance.product
                            old_quantity = form.instance.product_quantity_transfer

                            godown_qty_revert_obj, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name = old_product)

                            if godown_qty_revert_obj:
                                godown_qty_revert_obj.quantity += old_quantity
                                godown_qty_revert_obj.save()

                            warehouse_obj,created = Product_warehouse_quantity_through_table.objects.get_or_create(warehouse = selected_warehouse, product = old_product)

                            if warehouse_obj:
                                warehouse_obj.quantity = max(0,warehouse_obj.quantity - old_quantity)
                                warehouse_obj.save()

                            form.instance.delete()

                    
                    for form in formset:
                        if not form.cleaned_data.get('DELETE'):
                            form_instance = form.save(commit=False)
                            form_instance.Finished_goods_Stock_TransferMasterinstance = first_form_instance
                            form_instance.diffrence_qty = form_instance.product_quantity_transfer
                            

                            old_product_name = form.initial.get('product')
                            old_product_quantity = form.initial.get('product_quantity_transfer')
                            
                            #if both has change
                            if pk and form.has_changed() and 'product' and 'product_quantity_transfer' in form.changed_data:

                                print(" IN BOTH ")

                                print("old godown")
                                old_godown_qty, _ = product_godown_quantity_through_table.objects.get_or_create(godown_name=selected_godown, product_color_name=old_product_name)

                                old_godown_qty.quantity += old_product_quantity
                                old_godown_qty.save()

                                print("new godown")
                                new_godown_qty, _ = product_godown_quantity_through_table.objects.get_or_create(godown_name=selected_godown, product_color_name=form_instance.product)

                                new_godown_qty.quantity -= form_instance.product_quantity_transfer
                                new_godown_qty.save()

                                
                                print("old warehouse")
                                old_warehouse_qty, _ = Product_warehouse_quantity_through_table.objects.get_or_create(warehouse=selected_warehouse, product = old_product_name)

                                old_warehouse_qty.quantity = max(0, old_warehouse_qty.quantity - old_product_quantity)
                                old_warehouse_qty.save()

                                
                                print("new warehouse")
                                new_warehouse_qty, _ = Product_warehouse_quantity_through_table.objects.get_or_create(warehouse=selected_warehouse, product=form_instance.product)

                                new_warehouse_qty.quantity += form_instance.product_quantity_transfer
                                new_warehouse_qty.save()
                            
                            
                            #if any quantity has change
                            elif pk and form.has_changed() and 'product_quantity_transfer' in form.changed_data:
                                print(" ONLY QTY ")

                                warehouse_obj , created = Product_warehouse_quantity_through_table.objects.get_or_create(warehouse=selected_warehouse,product=form_instance.product)

                                difference =  form_instance.product_quantity_transfer - old_product_quantity
                                warehouse_obj.quantity =  warehouse_obj.quantity + difference
                                warehouse_obj.save()

                                godown_qty_update_obj_through_quantity, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name = old_product_name)
                            
                                godown_qty_update_obj_through_quantity.quantity -= difference
                                godown_qty_update_obj_through_quantity.save()

                            

                            #if any product has change
                            elif pk and form.has_changed() and 'product' in form.changed_data:

                                print(" ONLY PRODUCT ")

                                print("old godown")
                                godown_qty_update_obj_through_old_product, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name = old_product_name)

                                godown_qty_update_obj_through_old_product.quantity += old_product_quantity
                                godown_qty_update_obj_through_old_product.save()

                                print("new godown")
                                godown_qty_update_obj_through_new_product, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name = form_instance.product)

                                godown_qty_update_obj_through_new_product.quantity -= form_instance.product_quantity_transfer
                                godown_qty_update_obj_through_new_product.save()
                                
                                print("old warehouse")
                                warehouse_qty_update_obj_through_old_product, created = Product_warehouse_quantity_through_table.objects.get_or_create(warehouse = selected_warehouse,product = old_product_name)

                                warehouse_qty_update_obj_through_old_product.quantity -= old_product_quantity
                                warehouse_qty_update_obj_through_old_product.save()
                                
                                print("new warehouse")
                                warehouse_qty_update_obj_through_new_product, created = Product_warehouse_quantity_through_table.objects.get_or_create(warehouse=selected_warehouse, product=form_instance.product)

                                warehouse_qty_update_obj_through_new_product.quantity += old_product_quantity
                                warehouse_qty_update_obj_through_new_product.save()

                            form_instance.save()

                    return redirect('all-product-transfer-to-warehouse')
            
            except Exception as e:
                print(e)

    return render(request,'finished_product/product_transfer_to_warehouse.html',{'form':form,'formset':formset,'godowns':godowns,'voucher_instance':voucher_instance,'warehouses':warehouses,'dict_to_send':dict_to_send})












def product_transfer_to_warehouse_list(request):

    warehouse_product_transfer_list = Finished_goods_Stock_TransferMaster.objects.all().annotate(all_qc_qty=Sum('finished_goods_transfer_records__qc_recieved_qty'),total_recieved_qty=Sum('finished_goods_transfer_records__product_quantity_transfer')).order_by('created_date')

    return render(request,'finished_product/product_transfer_to_warehouse_list.html',{'warehouse_product_transfer_list':warehouse_product_transfer_list})








def product_transfer_to_warehouse_delete(request):
    id = request.POST.get('ProductId')
    
    master_instance = Finished_goods_Stock_TransferMaster.objects.get(id=id)
    if master_instance:
        master_instance.transnfer_cancelled = True
        master_instance.save()

        transfer_records = Finished_goods_transfer_records.objects.filter(Finished_goods_Stock_TransferMasterinstance = id)
        for record in transfer_records:
            if not record.transnfer_cancelled_records:
                godown_transfer_records_quantity_revert = product_godown_quantity_through_table.objects.get(product_color_name = record.product, godown_name = record.Finished_goods_Stock_TransferMasterinstance.source_warehouse)
                godown_transfer_records_quantity_revert.quantity = godown_transfer_records_quantity_revert.quantity + record.product_quantity_transfer
                godown_transfer_records_quantity_revert.save()
                
                warehouse_transfer_records_quantity_revert = Product_warehouse_quantity_through_table.objects.get(product = record.product)
                warehouse_transfer_records_quantity_revert.quantity = warehouse_transfer_records_quantity_revert.quantity - record.product_quantity_transfer
                warehouse_transfer_records_quantity_revert.save()
                transfer_records.update(transnfer_cancelled_records=True)
        
    return redirect('all-product-transfer-to-warehouse')






    







def stock_transfer_instance_list_and_recieve(request,id,voucher_type):
    
    try:
        if voucher_type == 'transfer':
            # print("in transfer")
            stock_transfer_instance = Finished_goods_Stock_TransferMaster.objects.get(pk=id)
            
            finished_goods_transfer_items_instances =  Finished_goods_transfer_records.objects.filter(Finished_goods_Stock_TransferMasterinstance = stock_transfer_instance, transnfer_cancelled_records=False)

            purchase_number = stock_transfer_instance.voucher_no

            formset = stock_transfer_instance_formset_only_for_update(request.POST or None, queryset=finished_goods_transfer_items_instances, instance=stock_transfer_instance)

            completed_qs = Finished_goods_transfer_records.objects.filter(Finished_goods_Stock_TransferMasterinstance=stock_transfer_instance, qc_recieved_qty__gt = 0, transnfer_cancelled_records=False)

            completed_formset = stock_transfer_instance_formset_only_for_update(queryset = completed_qs, instance=stock_transfer_instance)

            entries = finishedgoodsbinallocation.objects.filter(related_transfer_record__Finished_goods_Stock_TransferMasterinstance__voucher_no = purchase_number).select_related('related_transfer_record','related_transfer_record__Finished_goods_Stock_TransferMasterinstance','product','bin_number').order_by('created_date')

            
            # print(entries)


        elif voucher_type == 'purchase':
            # print("in purchase")
            
            product_purchase_voucher_items_instance = product_purchase_voucher_master.objects.get(pk=id)

            purchase_number = product_purchase_voucher_items_instance.purchase_number

            formset = product_purchase_voucher_items_instance_formset_only_for_update(request.POST or None, instance=product_purchase_voucher_items_instance)

            completed_qs = product_purchase_voucher_items.objects.filter(product_purchase_master=product_purchase_voucher_items_instance, qc_recieved_qty__gt = 0)

            completed_formset = product_purchase_voucher_items_instance_formset_only_for_update(queryset=completed_qs, instance=product_purchase_voucher_items_instance)

            entries = finishedgoodsbinallocation.objects.filter(related_purchase_item__product_purchase_master__purchase_number = purchase_number).select_related('related_purchase_item','related_purchase_item__product_purchase_master','product','bin_number').order_by('created_date')
            # print(entries)

    except Exception as e:
        print(e)

    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                manual_serial_number = request.POST.get('manual_serial_number')
                scanned_sku = int(request.POST.get('product_sku'))
                scanned_serialnumber = request.POST.get('scanned_serial_number')
                selected_product_bin = request.POST.get('product_bin')
                selected_voucher_type = request.POST.get('voucher_type')
                
                if scanned_serialnumber:
                    prathamesh = scanned_serialnumber
                else:
                    prathamesh = manual_serial_number
                
                if prathamesh and scanned_sku and selected_product_bin and selected_voucher_type:

                    if formset.is_valid():
                        form_present = False 

                        if selected_voucher_type == 'purchase':
                            
                            for form in formset:
                                if form.instance.product_name.PProduct_SKU == scanned_sku and form.instance.quantity_total > form.instance.qc_recieved_qty and form.instance.diffrence_qty != 0:
                                    
                                    try:
                                        product_pur_items_instance = get_object_or_404(product_purchase_voucher_items,pk=form.instance.id)

                                        product_instance = get_object_or_404(PProduct_Creation, pk=scanned_sku)

                                        bin_instance = get_object_or_404(finished_product_warehouse_bin,pk=selected_product_bin)

                                        request.session['last_selected_bin_id'] = bin_instance.id

                                        bin_instance.products_in_bin += 1
                                        bin_instance.save()
                                        
                                        finishedgoodsbinallocation.objects.create(related_purchase_item = product_pur_items_instance, 
                                                unique_serial_no = prathamesh,product = product_instance, 
                                                bin_number = bin_instance ,source_type = 'purchase')
                                        
                                        form.instance.qc_recieved_qty = form.instance.qc_recieved_qty + 1
                                        form.instance.diffrence_qty = form.instance.diffrence_qty - 1

                                        bin_qty_instance, created = Product_bin_quantity_through_table.objects.get_or_create(bin=bin_instance,product=product_instance,defaults={'product_quantity': 1})

                                        if not created:
                                            bin_qty_instance.product_quantity += 1
                                            bin_qty_instance.save()

                                    except ValueError as ve:
                                        messages.error(request, f'{ve}')
                                        raise

                                    except Exception as e:
                                        messages.error(request, f'{e}')
                                        raise

                                    form.save()
                                    form_present = True
                                    messages.success(request,f'SerialNumber - {prathamesh} added to Bin - {bin_instance.bin_name} Sucessfully')
                                    break            
                                else:
                                    form_present = False
                                    
                                    if form.instance.product_name.PProduct_SKU == scanned_sku:
                                        if form.instance.quantity_total >= form.instance.qc_recieved_qty:
                                            messages.error(request, f' Scaning Completed ( All Product Scan Sucessfully Thank you)')
                                            return redirect(reverse('stock-transfer-instance-list-popup', args=[id, selected_voucher_type]))
                                
                            
                        elif selected_voucher_type == 'transfer':
                            for form in formset:
                                if form.instance.product.PProduct_SKU == scanned_sku and form.instance.product_quantity_transfer > form.instance.qc_recieved_qty and form.instance.diffrence_qty != 0:
                                    
                                    try:

                                        product_transfer_items_instance = get_object_or_404(Finished_goods_transfer_records,pk=form.instance.id)

                                        product_instance = get_object_or_404(PProduct_Creation,pk=scanned_sku)

                                        bin_instance = get_object_or_404(finished_product_warehouse_bin,pk=selected_product_bin)

                                        request.session['last_selected_bin_id'] = bin_instance.id

                                        bin_instance.products_in_bin += 1
                                        bin_instance.save()

                                        finishedgoodsbinallocation.objects.create(related_transfer_record = product_transfer_items_instance, 
                                                    unique_serial_no = prathamesh, product = product_instance, 
                                                    bin_number = bin_instance ,source_type='transfer')
                                        
                                        form.instance.qc_recieved_qty = form.instance.qc_recieved_qty + 1
                                        form.instance.diffrence_qty = form.instance.diffrence_qty - 1

                                        bin_qty_instance, created = Product_bin_quantity_through_table.objects.get_or_create(bin=bin_instance,product=product_instance,defaults={'product_quantity': 1})
                                        
                                        if not created:
                                            bin_qty_instance.product_quantity += 1
                                            bin_qty_instance.save()

                                    except ValueError as ve:
                                        messages.error(request, f'{ve}')
                                        raise

                                    except Exception as e:
                                        messages.error(request, f'{e}')
                                        raise

                                    form.save()
                                    form_present = True
                                    messages.success(request,f'SerialNumber - {scanned_serialnumber} added to Bin - {bin_instance.bin_name} Sucessfully')
                                    break
                                else:
                                    form_present = False

                                    if form.instance.product.PProduct_SKU == scanned_sku:
                                        if form.instance.product_quantity_transfer >= form.instance.qc_recieved_qty:
                                            messages.error(request, f' Scaning Completed ( All Product Scan Sucessfully Thank you)')
                                            return redirect(reverse('stock-transfer-instance-list-popup', args=[id, selected_voucher_type]))
                                    
                        if form_present == False:
                            messages.error(request,f'No Product SKU found in the scanned Voucher')
                            return redirect(reverse('stock-transfer-instance-list-popup', args=[id,selected_voucher_type]))

                    else:
                        messages.error(request,f'Error with formset validation - {formset.errors}')
                        return redirect(reverse('stock-transfer-instance-list-popup', args=[id,selected_voucher_type]))

                else:
                    raise ValidationError('Name , SKU, Color or Bin not selected')

        except IntegrityError:
            messages.error(request,f'The scanned Serial Number already Exists')
            return redirect(reverse('stock-transfer-instance-list-popup', args=[id,selected_voucher_type]))

        except ValidationError as ve:
            messages.error(request,f'Required Data not filled {ve} ')
            return redirect(reverse('stock-transfer-instance-list-popup', args=[id,selected_voucher_type]))

        except Exception as e:
            messages.error(request,f'Exception Occoured - {e}')
            return redirect(reverse('stock-transfer-instance-list-popup', args=[id,selected_voucher_type]))


    
    return render(request, 'finished_product/stock_transfer_instance_list_popup.html',{'formset': formset,
                                                'purchase_number':purchase_number,'voucher_type': voucher_type,
                                                'completed_formset':completed_formset,'single_entries':entries})









def delete_sigle_entries(request, e_id, voucher_type):
    try:
        delete_instance = finishedgoodsbinallocation.objects.get(pk=e_id)
        
        if voucher_type == "purchase":
            related_purchase_item = delete_instance.related_purchase_item
            if related_purchase_item:  # Check if related_purchase_item exists
                parent_id = related_purchase_item.product_purchase_master.id
            else:
                raise ValueError("No related purchase item found.")
        else:
            related_transfer_record = delete_instance.related_transfer_record
            if related_transfer_record:  # Check if related_transfer_record exists
                parent_id = related_transfer_record.Finished_goods_Stock_TransferMasterinstance.id
            else:
                raise ValueError("No related transfer record found.")

        delete_instance.delete()
        return redirect(reverse('stock-transfer-instance-list-popup', args=[parent_id, voucher_type]))

    except finishedgoodsbinallocation.DoesNotExist:
        return redirect(reverse('error-page'))  # Redirect to an error page or handle it appropriately
    except Exception as e:
        print(f"Error in delete_single_entries: {e}")
        return redirect(reverse('error-page'))





def scan_product_qty_list(request):

    product_purchase_voucher = product_purchase_voucher_items.objects.filter(qc_recieved_qty__gt = 0).select_related( 'product_purchase_master', 'product_name').values('id','product_purchase_master__ledger_type','product_name__PProduct_SKU','product_name__Product__Product_Refrence_ID','product_name__Product__Model_Name','product_name__PProduct_color__color_name','product_name__PProduct_image','quantity_total','qc_recieved_qty','diffrence_qty','created_date')

    stock_transfer_voucher = Finished_goods_transfer_records.objects.filter(qc_recieved_qty__gt = 0).select_related( 'Finished_goods_Stock_TransferMasterinstance', 'product').values('id','Finished_goods_Stock_TransferMasterinstance','product__PProduct_SKU','product__Product__Product_Refrence_ID','product__Product__Model_Name','product__PProduct_color__color_name','product__PProduct_image','product_quantity_transfer','qc_recieved_qty','diffrence_qty','created_date')

    merged_queryset = chain(product_purchase_voucher, stock_transfer_voucher)

    merged_list = sorted(list(merged_queryset), key=lambda x: x['created_date'])

    return render(request,'finished_product/scan_product_qty_list.html',{'merged_list':merged_list,'MEDIA_URL': settings.MEDIA_URL})





from django.utils import timezone

def scan_product_list(request,pk,v_type):

    instance_entries_all = []
    instance_entries = []

    if v_type == "purchase":

        instance_entries = finishedgoodsbinallocation.objects.filter(related_purchase_item = pk).select_related('related_purchase_item','product','bin_number').values(
            'related_purchase_item__product_purchase_master__id',
            'related_purchase_item__product_purchase_master__purchase_number',
            'related_purchase_item__product_purchase_master__ledger_type',
            'product__Product__Product_Refrence_ID',
            'product__PProduct_image',
            'product__Product__Model_Name',
            'product__PProduct_color__color_name',
            'product__PProduct_SKU',
            'unique_serial_no',
            'created_date',
            'bin_number__rack_finished_name__zone_finished_name__zone_name',
            'bin_number__rack_finished_name__rack_name',
            'bin_number__bin_name'
            ).order_by('created_date')

    elif v_type == "transfer":

        instance_entries = finishedgoodsbinallocation.objects.filter(related_transfer_record = pk).select_related('related_transfer_record','product','bin_number').values(
            'related_transfer_record__Finished_goods_Stock_TransferMasterinstance__id',
            'related_transfer_record__Finished_goods_Stock_TransferMasterinstance__voucher_no',
            'product__Product__Product_Refrence_ID',
            'product__PProduct_image',
            'product__Product__Model_Name',
            'product__PProduct_color__color_name',
            'product__PProduct_SKU',
            'unique_serial_no',
            'created_date',
            'bin_number__rack_finished_name__zone_finished_name__zone_name',
            'bin_number__rack_finished_name__rack_name',
            'bin_number__bin_name'
            ).order_by('created_date')
    else:
        instance_entries_all = finishedgoodsbinallocation.objects.all().select_related('related_transfer_record','product','bin_number').values(
            'related_purchase_item',
            'related_transfer_record',
            'related_purchase_item__product_purchase_master__id',
            'related_purchase_item__product_purchase_master__purchase_number',
            'related_purchase_item__product_purchase_master__ledger_type',
            'related_transfer_record__Finished_goods_Stock_TransferMasterinstance__id',
            'related_transfer_record__Finished_goods_Stock_TransferMasterinstance__voucher_no',
            'product__Product__Product_Refrence_ID',
            'product__PProduct_image',
            'product__Product__Model_Name',
            'product__PProduct_color__color_name',
            'product__PProduct_SKU',
            'unique_serial_no',
            'created_date',
            'bin_number__rack_finished_name__zone_finished_name__zone_name',
            'bin_number__rack_finished_name__rack_name',
            'bin_number__bin_name'
            ).annotate(total_qty = Sum('related_purchase_item__qc_recieved_qty'),total_qty_tr = Sum('related_transfer_record__qc_recieved_qty')).order_by('created_date')


    current_date = timezone.now()
    for entry in instance_entries:
        created_date = entry['created_date']
        age_in_days = (current_date - created_date).days
        entry['age_in_days'] = age_in_days


    for entry in instance_entries_all:
        created_date = entry['created_date']
        age_in_days = (current_date - created_date).days
        entry['age_in_days'] = age_in_days


    return render(request,'finished_product/scan_product_list.html',{'instance_entries':instance_entries,'MEDIA_URL':settings.MEDIA_URL,'instance_entries_all':instance_entries_all})







def warehouse_stock(request):

    purchase_sales_quantity_subquery = sales_voucher_outward_scan.objects.filter(product_name__PProduct_SKU=OuterRef('product_name__PProduct_SKU')).values('product_name__PProduct_SKU').annotate(sales_quantity=Sum('quantity')).values('sales_quantity')

    purchase_sales_return_quantity_subquery = sales_return_voucher.objects.filter(product_name__PProduct_SKU=OuterRef('product_name__PProduct_SKU')).values('product_name__PProduct_SKU').annotate(sales_retutn_quantity=Sum('quantity')).values('sales_retutn_quantity')
    
    product_purchase_voucher = (product_purchase_voucher_items.objects.all().values('product_name__PProduct_SKU','product_name__Product__Product_Refrence_ID','product_name__Product__Model_Name','product_name__PProduct_color__color_name','product_name__PProduct_image').annotate(total_quantity=Sum('quantity_total'),total_sale=Subquery(purchase_sales_quantity_subquery),total_inward=Sum('qc_recieved_qty'),total_balance = Sum('diffrence_qty'),total_sales_return = Subquery(purchase_sales_return_quantity_subquery)))

    transfer_sales_quantity_subquery = sales_voucher_outward_scan.objects.filter(product_name__PProduct_SKU=OuterRef('product__PProduct_SKU')).values('product_name__PProduct_SKU').annotate(sales_quantity=Sum('quantity')).values('sales_quantity')

    transfer_sales_return_quantity_subquery = sales_return_voucher.objects.filter(product_name__PProduct_SKU=OuterRef('product__PProduct_SKU')).values('product_name__PProduct_SKU').annotate(sales_retutn_quantity=Sum('quantity')).values('sales_retutn_quantity')

    stock_transfer_voucher = Finished_goods_transfer_records.objects.filter(transnfer_cancelled_records = False).values('product__PProduct_SKU','product__Product__Product_Refrence_ID','product__Product__Model_Name','product__PProduct_color__color_name','product__PProduct_image').annotate(total_quantity=Sum('product_quantity_transfer'),total_sale=Subquery(transfer_sales_quantity_subquery),total_inward=Sum('qc_recieved_qty'),total_balance = Sum('diffrence_qty'),total_sales_return = Subquery(transfer_sales_return_quantity_subquery))

    

    product_dict = {
        x['product_name__PProduct_SKU']: {
            'ref_id': x['product_name__Product__Product_Refrence_ID'],
            'model_name': x['product_name__Product__Model_Name'],
            'color': x['product_name__PProduct_color__color_name'],
            'product_sku': x['product_name__PProduct_SKU'],
            'total_quantity': x.get('total_quantity', 0) or 0,
            'total_sale': x.get('total_sale', 0) or 0,
            'total_inward': x.get('total_inward', 0) or 0,
            'total_balance': x.get('total_balance', 0) or 0,
            'total_sale_return': x.get('total_sales_return', 0) or 0,
            'img': x['product_name__PProduct_image'],
            'total': x.get('total_quantity', 0) or 0,
        }
        for x in product_purchase_voucher
    }

    for y in stock_transfer_voucher:
        sku = y['product__PProduct_SKU']
        if sku in product_dict:
            
            product_dict[sku]['total_quantity'] += y.get('total_quantity', 0) or 0
            product_dict[sku]['total_inward'] += y.get('total_inward', 0) or 0
            product_dict[sku]['total_balance'] += y.get('total_balance', 0) or 0
            product_dict[sku]['total'] += y.get('total_quantity', 0) or 0
        else:
            # Add new SKU from stock transfer
            product_dict[sku] = {
                'ref_id': y['product__Product__Product_Refrence_ID'],
                'model_name': y['product__Product__Model_Name'],
                'color': y['product__PProduct_color__color_name'],
                'product_sku': sku,
                'total_quantity': y.get('total_quantity', 0) or 0,
                'total_sale': y.get('total_sale', 0) or 0,
                'total_inward': y.get('total_inward', 0) or 0,
                'total_balance': y.get('total_balance', 0) or 0,
                'total_sale_return': y.get('total_sales_return', 0) or 0,
                'img': y['product__PProduct_image'],
                'total': y.get('total_quantity', 0) or 0,
            }

    merged_list = []
    for data in product_dict.values():
        data['inward_minus_sales'] = (data['total_inward'] + data['total_sale_return']) - data['total_sale']
        merged_list.append(data)


    return render(request,'finished_product/warehouse_stock.html',{'merged_list':merged_list,'MEDIA_URL': settings.MEDIA_URL})





def scan_single_product_list(request,sku):
    instance_entries = finishedgoodsbinallocation.objects.filter(product__PProduct_SKU = sku).values(
        'related_purchase_item__product_purchase_master__purchase_number' or None,
        'related_purchase_item__product_purchase_master__id' or None,
        'related_transfer_record__Finished_goods_Stock_TransferMasterinstance__voucher_no' or None,
        'related_transfer_record__Finished_goods_Stock_TransferMasterinstance__id' or None,
        'product__Product__Product_Refrence_ID',
        'product__PProduct_image',
        'product__Product__Model_Name',
        'product__PProduct_color__color_name',
        'product__PProduct_SKU',
        'unique_serial_no',
        'created_date',
        'bin_number__rack_finished_name__zone_finished_name__zone_name',
        'bin_number__rack_finished_name__rack_name',
        'bin_number__bin_name'
        )
    
    current_date = timezone.now()
    for entry in instance_entries:
        created_date = entry['created_date']
        age_in_days = (current_date - created_date).days
        entry['age_in_days'] = age_in_days

    return render(request,'finished_product/scan_single_product_list.html',{'instance_entries':instance_entries,'MEDIA_URL':settings.MEDIA_URL})




def model_name_wise_purchase_transfer_sales_report(request,sku):
    purchase_instance = product_purchase_voucher_items.objects.filter(product_name__PProduct_SKU = sku)

    transfer_instance = Finished_goods_transfer_records.objects.filter(product__PProduct_SKU = sku)    

    sales_instance = sales_voucher_finish_Goods.objects.filter(product_name__PProduct_SKU = sku)
    
    merge_list = []

    for item in purchase_instance:

        model_name = item.product_name.Product.Model_Name
        ref_no = item.product_name.Product.Product_Refrence_ID


        dict_to_append = {
            'instance_id':item.product_purchase_master.id,
            'date':item.created_date,
            'voucher_no':item.product_purchase_master.purchase_number,
            'voucher_type':"purchase(product)",
            'ship_from':item.product_purchase_master.party_name.name,
            'sale_no':"",
            'party_name':"",
            
            'color':item.product_name.PProduct_color.color_name,
            'pro_sku':item.product_name.PProduct_SKU,
            'inward':item.quantity_total,
            'outward':'',
            'balance':''
        }
        print(dict_to_append)
        merge_list.append(dict_to_append)

    
    for item in transfer_instance:

        model_name = item.product.Product.Model_Name
        ref_no = item.product.Product.Product_Refrence_ID

        dict_to_append = {
            'instance_id':item.Finished_goods_Stock_TransferMasterinstance.id,
            'date':item.created_date,
            'voucher_no':item.Finished_goods_Stock_TransferMasterinstance.voucher_no,
            'voucher_type':"Transfer",
            'ship_from':item.Finished_goods_Stock_TransferMasterinstance.source_warehouse.godown_name_finished,
            'sale_no':"",
            'party_name':"",
            
            'color':item.product.PProduct_color.color_name,
            'pro_sku':item.product.PProduct_SKU,
            'inward':item.product_quantity_transfer,
            'outward':'',
            'balance':''
        }

        merge_list.append(dict_to_append)


    for item in sales_instance:

        model_name = item.product_name.Product.Model_Name
        ref_no = item.product_name.Product.Product_Refrence_ID

        dict_to_append = {
            'instance_id':item.sales_voucher_master.id,
            'date':"",
            'voucher_no':"",
            'voucher_type':"Sales",
            'ship_from':"",
            'sale_no':item.sales_voucher_master.sales_no,
            'party_name':item.sales_voucher_master.party_name.name,
            'color':item.product_name.PProduct_color.color_name,
            'pro_sku':item.product_name.PProduct_SKU,
            'inward':0,
            'outward':item.quantity,
            'balance':'',
            'selected_godown':item.sales_voucher_master.selected_godown
        }
        print(dict_to_append)
        merge_list.append(dict_to_append)

    # list_to_send = sorted(merge_list, key = itemgetter('date'))

    return render(request,'finished_product/modelnamewisepurchasetransfersalesreport.html',{'merge_list':merge_list,'model_name':model_name,'ref_no':ref_no})




def product_wise_sales_report(request,sku):
    sales_voucher_queryset = sales_voucher_outward_scan.objects.filter(product_name = sku).select_related(
        'sales_voucher_master'
        ).values(
            'created_date',
            'product_name__PProduct_image',
            'sales_voucher_master__sale_no',
            'sales_voucher_master__outward_no__outward_no',
            'sales_voucher_master__salesman__salesman_name',
            'sales_voucher_master__party_name__name',
            'product_name__Product__Product_Refrence_ID',
            'product_name__Product__Model_Name',
            'product_name__PProduct_color__color_name',
            'product_name__PProduct_SKU',
            'quantity',
            'sales_voucher_master__ledger_type', 
        )
    return render(request,'finished_product/product_wise_sales_report.html',{'MEDIA_URL':settings.MEDIA_URL,'sales_voucher_queryset':sales_voucher_queryset})




def product_wise_sales_return_report(request,sku):
    sales_voucher_queryset = sales_return_voucher.objects.filter(product_name = sku).select_related(
        'sales_return_master'
        ).values(
            'created_date',
            'product_name__PProduct_image',
            'sales_return_master__sales_voucher_master__sale_no',
            'sales_return_master__sales_return_inward_instance__sales_return_no',
            'sales_return_master__salesman__salesman_name',
            'sales_return_master__party_name__name',
            'product_name__Product__Product_Refrence_ID',
            'product_name__Product__Model_Name',
            'product_name__PProduct_color__color_name',
            'product_name__PProduct_SKU',
            'quantity',
            'sales_return_master__ledger_type', 
        )
    return render(request,'finished_product/product_wise_sales_return_report.html',{'MEDIA_URL':settings.MEDIA_URL,'sales_voucher_queryset':sales_voucher_queryset})





def model_name_wise_purchase_transfer_report(request,sku):

    purchase_instance = product_purchase_voucher_items.objects.filter(product_name__PProduct_SKU = sku).select_related('product_purchase_master','product_name')

    transfer_instance = Finished_goods_transfer_records.objects.filter(product__PProduct_SKU = sku)    

    product_data = PProduct_Creation.objects.filter(PProduct_SKU = sku).select_related('Product').values('Product__Product_Refrence_ID','Product__Model_Name','PProduct_color__color_name','PProduct_image')

    merge_list = []

    for item in purchase_instance:

        model_name = item.product_name.Product.Model_Name
        ref_no = item.product_name.Product.Product_Refrence_ID

        dict_to_append = {
            'instance_id':item.product_purchase_master.id,
            'date':item.created_date,
            'voucher_no':item.product_purchase_master.purchase_number,
            'voucher_type':item.product_purchase_master.ledger_type,
            'party_name':item.product_purchase_master.party_name.name,
            'model_name':item.product_name.Product.Model_Name,
            'color':item.product_name.PProduct_color.color_name,
            'pro_sku':item.product_name.PProduct_SKU,
            'total_qty':item.quantity_total,
        }
        
        merge_list.append(dict_to_append)

    
    for item in transfer_instance:

        model_name = item.product.Product.Model_Name
        ref_no = item.product.Product.Product_Refrence_ID

        dict_to_append = {
            'instance_id':item.Finished_goods_Stock_TransferMasterinstance.id,
            'date':item.created_date,
            'voucher_no':item.Finished_goods_Stock_TransferMasterinstance.voucher_no,
            'voucher_type':"Transfer",
            'party_name':item.Finished_goods_Stock_TransferMasterinstance.source_warehouse.godown_name_finished,
            'model_name':item.product.Product.Model_Name,
            'color':item.product.PProduct_color.color_name,
            'pro_sku':item.product.PProduct_SKU,
            'total_qty':item.product_quantity_transfer,
        }

        merge_list.append(dict_to_append)

    list_to_send = sorted(merge_list, key = itemgetter('date'))

    return render(request,'finished_product/modelnamewisepurchasetransferreport.html',{'purchase_instance':purchase_instance,'list_to_send':list_to_send,'product_data':product_data,'MEDIA_URL': settings.MEDIA_URL})




def process_serial_no(request):
    print("in scan def")
    
    if request.method == 'POST':

        serial_no = request.POST.get('serialNo')
        voucher_type = request.POST.get('instance_type_post')
        voucher_no = request.POST.get('instance_number_post')

        try:
            if voucher_type == "purchase":
                exists = product_purchase_voucher_items.objects.filter(product_purchase_master__purchase_number=voucher_no).aggregate(total_sum=Sum('diffrence_qty'))['total_sum'] or 0

                if exists == 0:
                    print("Scan complete")
                    return JsonResponse({'error':'Scaning Completed (All Product Scan Sucessfully Thank you).'}, status=400)

            if voucher_type == "transfer":
                exists = Finished_goods_transfer_records.objects.filter(Finished_goods_Stock_TransferMasterinstance__voucher_no=voucher_no).aggregate(total_sum=Sum('diffrence_qty'))['total_sum'] or 0
                            
                if exists == 0:
                    print("Scan complete")
                    return JsonResponse({'error':'Scaning Completed (All Product Scan Sucessfully Thank you).'}, status=400)

        except Exception:
            return JsonResponse({'message': 'Failed to fetch data from external API.'}, status=500)



        if serial_no:
            try:
                # Call external API
                url = f'https://www.cosmusbags.com/cosmus/qrcode.php?wc={serial_no}'
                response_post = requests.get(url)
                response_data = response_post.json()
                print(response_data)

                if response_data.get('response_code') == 200 and response_data.get('response_desc') == 'success':

                    product_scanned_sku = response_data['sku']

                    try:
                        if voucher_type == "purchase":
                            exists = product_purchase_voucher_items.objects.filter(product_purchase_master__purchase_number=voucher_no,product_name__PProduct_SKU=product_scanned_sku).aggregate(total_sum=Sum('diffrence_qty'))['total_sum']

                            if exists == 0:
                                print("Scan complete")
                                return JsonResponse({'error': 'This Product Scanning Completed'}, status=400)

                        if voucher_type == "transfer":
                            exists = Finished_goods_transfer_records.objects.filter(Finished_goods_Stock_TransferMasterinstance__voucher_no=voucher_no,product__PProduct_SKU=product_scanned_sku).aggregate(total_sum=Sum('diffrence_qty'))['total_sum']         
                            
                            if exists == 0:
                                print("Scan complete")
                                return JsonResponse({'error':'This Product Scaning Completed'}, status=400)

                    except Exception:
                        return JsonResponse({'message': 'Failed to fetch data from external API.'}, status=500)
                    
            
                    try:
                        product_instance = PProduct_Creation.objects.get(PProduct_SKU=product_scanned_sku)
                    except ObjectDoesNotExist:
                        return JsonResponse({'error': 'SKU PRODUCT DOES NOT EXIST IN VOUCHER'}, status=400)
                    

                    try:
                        # If Serial Number has already been processed.
                        if finishedgoodsbinallocation.objects.only('id').filter(unique_serial_no=serial_no).exists():

                            return JsonResponse({'error': 'This Serial Number has already been processed.'}, status=400)
                            
                    except IntegrityError:
                        return JsonResponse({'error': 'This serial number has already been processed'}, status=400)

                    # Fetch product details
                    model_name = product_instance.Product.Model_Name if product_instance.Product.Model_Name else None
                    product_name = product_instance.Product.Product_Name if product_instance.Product.Product_Name else None
                    product_sku = product_instance.PProduct_SKU
                    product_color = product_instance.PProduct_color.color_name if product_instance.PProduct_color else None
                    product_image = product_instance.PProduct_image.url if product_instance.PProduct_image else None

                    
                    # Get product main categories
                    product_main_cats = product_instance.Product.product_cats.all()
                    main_cats_all = [x.SubCategory_id.product_main_category for x in product_main_cats]

                    bins_related_to_product = []

                    # Fetch all bins related to the product
                    for bin_obj in finished_product_warehouse_bin.objects.filter(sub_catergory_id__in=main_cats_all):

                        product_count = finishedgoodsbinallocation.objects.filter(bin_number=bin_obj, outward_done=False).count()

                        bins_related_to_product.append({
                            'bin_id': bin_obj.id,
                            'bin_name': bin_obj.bin_name,
                            'bin_size': bin_obj.product_size_in_bin,
                            'products_in_bin': product_count
                        })

                    # Retrieve last selected bin from table
                    last_selected_bin_id = finishedgoodsbinallocation.objects.filter(Q(related_purchase_item__product_purchase_master__purchase_number=voucher_no) |Q(related_transfer_record__Finished_goods_Stock_TransferMasterinstance__voucher_no=voucher_no)).order_by('-created_date').first()

                    # Sort bins: Show last selected bin first
                    if last_selected_bin_id:
                        bins_related_to_product.sort(key=lambda x: x['bin_id'] != last_selected_bin_id.bin_number.id)

                    return JsonResponse({
                        'model_name': model_name,
                        'product_name': product_name,
                        'product_sku': product_sku,
                        'bin_to_dict': bins_related_to_product,
                        'product_color': product_color,
                        'product_image': product_image,
                        'message': f'Serial No {serial_no} processed successfully.'
                    })

                else:
                    return JsonResponse({'message': 'Invalid response from external API.'}, status=400)

            except ObjectDoesNotExist:
                return JsonResponse({'message': 'Product SKU does not exist.'}, status=404)

            except requests.RequestException:
                return JsonResponse({'message': 'Failed to fetch data from external API.'}, status=500)

            except Exception:
                return JsonResponse({'message': 'Failed to fetch data from external API.'}, status=500)

        else:
            return JsonResponse({'message': 'Invalid Serial No.'}, status=400)

    return JsonResponse({'message': 'Invalid Request Method.'}, status=405)














def purchase_order_for_puchase_voucher_rm_create_update(request,p_id=None):
    party_names = Ledger.objects.filter(under_group__account_sub_group = 'Sundry Creditors')
    selected_item_id = request.GET.get('selectedItemId')

    if p_id:

        Purchaseorderforpuchasevoucherrmformset = inlineformset_factory(purchase_order_master_for_puchase_voucher_rm, purchase_order_for_puchase_voucher_rm, form=Purchaseorderforpuchasevoucherrmform, extra=0, can_delete=True)

        order_instance = purchase_order_master_for_puchase_voucher_rm.objects.get(id=p_id)
        master_form = Purchaseordermasterforpuchasevoucherrmform(instance=order_instance)
        formset = Purchaseorderforpuchasevoucherrmformsetupdate(instance=order_instance)

    else:
        
        Purchaseorderforpuchasevoucherrmformset = inlineformset_factory(purchase_order_master_for_puchase_voucher_rm, purchase_order_for_puchase_voucher_rm, form=Purchaseorderforpuchasevoucherrmform, extra=1, can_delete=True)
        if selected_item_id:
            selected_item = json.loads(selected_item_id)

            converted_str = [value for key,value in selected_item.items()]

            list_with_data = []
            for x in converted_str:

                item_name_queryset = Item_Creation.objects.get(pk = x)
                initial_data_dict = {
                            'item_id' : item_name_queryset.id,
                            'item_name': item_name_queryset.item_name,
                            'Material_code':item_name_queryset.Material_code,
                            'item_color':item_name_queryset.Item_Color.color_name,
                            'item_unit':item_name_queryset.unit_name_item.unit_name,
                            'item_gst':item_name_queryset.Item_Creation_GST.gst_percentage
                        }
                list_with_data.append(initial_data_dict)


            Purchaseorderforpuchasevoucherrmformset = inlineformset_factory(purchase_order_master_for_puchase_voucher_rm, purchase_order_for_puchase_voucher_rm, form=Purchaseorderforpuchasevoucherrmform, extra=len(list_with_data), can_delete=True)

            order_instance = None
            master_form = Purchaseordermasterforpuchasevoucherrmform()
            formset = Purchaseorderforpuchasevoucherrmformset(initial=list_with_data) 
        else:
            order_instance = None
            master_form = Purchaseordermasterforpuchasevoucherrmform()
            formset = Purchaseorderforpuchasevoucherrmformset()

    if request.method == 'POST':
        print(request.POST)
        master_form = Purchaseordermasterforpuchasevoucherrmform(request.POST, instance=order_instance)
        formset = Purchaseorderforpuchasevoucherrmformset(request.POST, instance=order_instance)

        if master_form.is_valid() and formset.is_valid():
            master_form_instance = master_form.save(commit=False)
            master_form_instance.save()


            for form in formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()

            for form in formset:
                if not form.cleaned_data.get('DELETE'):
                    form_instance = form.save(commit=False)
                    form_instance.master_instance = master_form_instance
                    qty = form.instance.quantity
                    form_instance.demo_quantity = qty
                    form_instance.save()

        else:
            print(master_form.errors)
            print(formset.errors)
        return redirect('purchase-order-for-puchase-voucher-rm-list')
    return render(request,'accounts/purchaseorderforpuchasevoucherrmcreateupdate.html',{'master_form':master_form,'formset':formset,'party_names':party_names})






def purchase_order_for_puchase_voucher_rm_list(request):
  
    order_list = purchase_order_master_for_puchase_voucher_rm.objects.annotate(total_qty=Sum('purchase_order_for_puchase_voucher_rm__quantity')).filter(total_qty__gt=0).order_by('created_date')
    

    order_list_complete = purchase_order_master_for_puchase_voucher_rm.objects.annotate(total_qty=Sum('purchase_order_for_puchase_voucher_rm__quantity'),total_demo_qty = Sum('purchase_order_for_puchase_voucher_rm__demo_quantity')).filter(total_qty = 0).order_by('created_date')

    return render(request,'accounts/purchaseorderforpuchasevoucherrmlist.html',{'order_list':order_list ,'order_list_complete':order_list_complete})




def negetive_stock(request):
    negetive_stock_sellerwise = Ledger.objects.filter(under_group__account_sub_group = 'Sundry Creditors')

    selected_vendor = negetive_stock_sellerwise.get(id=2)

    voucher_master = item_purchase_voucher_master.objects.filter(party_name = selected_vendor)

    filter_name  = request.GET.get('sort_name')
    selected_fabric_grp = request.GET.get('Fabric_Group')
    less_Number = request.GET.get('less_Number')


    negetive_stock_report =  Item_Creation.objects.all().annotate(total_qty = Sum('shades__godown_shades__quantity')).order_by('item_name').select_related('Item_Color','Fabric_Group')

    negetive_stock_report_list = list(negetive_stock_report)

    # print('negetive_stock_report_list --- ', negetive_stock_report_list)

    
    
    pending_ref_no = purchase_order_to_product.objects.filter(purchase_order_id__process_status__gt = 2,process_quantity__gt=0)

    sku_total_qty_cutting = {}

    if pending_ref_no:
        for item in pending_ref_no:
            product_reference_number = item.purchase_order_id.product_reference_number.Product_Refrence_ID
            sku = item.product_id.PProduct_SKU
            process_quantity = item.process_quantity

            if product_reference_number not in sku_total_qty_cutting:
                sku_total_qty_cutting[product_reference_number] = {}

            if sku not in sku_total_qty_cutting[product_reference_number]:
                sku_total_qty_cutting[product_reference_number][sku] = process_quantity
            else:
                sku_total_qty_cutting[product_reference_number][sku] += process_quantity

        for product_reference_number, skus in sku_total_qty_cutting.items():
            total_quantity = sum(skus.values())
            skus['total'] = total_quantity

    # print('sku_total_qty_cutting = ', sku_total_qty_cutting)



    cutting_pending_for_lwo_approval = purchase_order_to_product_cutting.objects.filter(~Q(cutting_quantity=F('approved_pcs'))).filter(purchase_order_cutting_id__cutting_cancelled = False)
    
    sku_total_qty_cutting_lwo_aprv_pending = {}

    if cutting_pending_for_lwo_approval:
        for x in cutting_pending_for_lwo_approval:
            product_reference_number = x.purchase_order_cutting_id.purchase_order_id.product_reference_number.Product_Refrence_ID
            sku = x.product_sku
            process_quantity = x.cutting_quantity if x.balance_pcs == 0 else x.balance_pcs

            if product_reference_number not in sku_total_qty_cutting_lwo_aprv_pending:
                sku_total_qty_cutting_lwo_aprv_pending[product_reference_number] = {}

            if sku not in sku_total_qty_cutting_lwo_aprv_pending[product_reference_number]:
                sku_total_qty_cutting_lwo_aprv_pending[product_reference_number][sku] = process_quantity
            else:
                sku_total_qty_cutting_lwo_aprv_pending[product_reference_number][sku] += process_quantity

        for product_reference_number, skus in sku_total_qty_cutting_lwo_aprv_pending.items():
            total_quantity = sum(skus.values())
            skus['total'] = total_quantity

    # print('sku_total_qty_cutting_lwo_aprv_pending = ',sku_total_qty_cutting_lwo_aprv_pending)


    
    lwo_pending = product_to_item_labour_workout.objects.all()

    sku_total_qty_lwo_pending = {}

    if lwo_pending:
        for x in lwo_pending:
            product_reference_number = x.labour_workout.purchase_order_cutting_master.purchase_order_id.product_reference_number.Product_Refrence_ID
            sku = x.product_sku
            process_quantity = x.pending_pcs

            if product_reference_number not in sku_total_qty_lwo_pending:
                sku_total_qty_lwo_pending[product_reference_number] = {}
                
            if sku not in sku_total_qty_lwo_pending[product_reference_number]:
                sku_total_qty_lwo_pending[product_reference_number][sku] = process_quantity
                
            else:
                sku_total_qty_lwo_pending[product_reference_number][sku] += process_quantity

        for product_reference_number, skus in sku_total_qty_lwo_pending.items():
            total_quantity = sum(skus.values())
            skus['total'] = total_quantity

    # print('sku_total_qty_lwo_pending = ', sku_total_qty_lwo_pending)





    material_for_ref_id_list = []

    if sku_total_qty_cutting:

        material_for_ref_id_queryset = product_2_item_through_table.objects.filter(PProduct_pk__Product__Product_Refrence_ID__in = sku_total_qty_cutting).select_related('PProduct_pk__Product', 'Item_pk__Item_Color')
        
        for material in material_for_ref_id_queryset:
            reference_id = material.PProduct_pk.Product.Product_Refrence_ID
            material_name = material.Item_pk.item_name
            color = material.Item_pk.Item_Color.color_name
            pro_sku = material.PProduct_pk.PProduct_SKU
            pro_fab_grp = material.Item_pk.Fabric_nonfabric
            panha_val = material.Item_pk.Panha
            unit_val = material.Item_pk.Units
            grand_total = material.grand_total
            grand_total_combi = material.grand_total_combi
            consumption = round(grand_total / (panha_val * unit_val),3)
            consumtionCombi = round(grand_total_combi / (panha_val * unit_val),3)
            common_unique = material.common_unique

            set_production_data_dict = {
                'ref_id': reference_id,
                'material_name': material_name,
                'color':color,
                'pro_sku': pro_sku,
                'pro_fab_grp': pro_fab_grp,
                'panha_val':panha_val,
                'unit_val':unit_val,
                'grand_total':grand_total,
                'consumption': consumption,
                'consumtionCombi': consumtionCombi,
                'common_unique': common_unique
            }
            material_for_ref_id_list.append(set_production_data_dict)
        
        grouped_by_ref_id_unique = {}

        for material in material_for_ref_id_list:

            ref_no = material['ref_id']

            if ref_no not in grouped_by_ref_id_unique:
                grouped_by_ref_id_unique[ref_no] = []

            if material['common_unique']:

                if not any(m['material_name'] == material['material_name'] for m in grouped_by_ref_id_unique[ref_no]):
                    grouped_by_ref_id_unique[ref_no].append(material)
            else:
                grouped_by_ref_id_unique[ref_no].append(material)

        # print('grouped_by_ref_id_unique = ', grouped_by_ref_id_unique)


    material_for_ref_id_list_for_cutting_aprv_pending = []

    if sku_total_qty_cutting_lwo_aprv_pending:

        material_for_ref_id_queryset_for_cutting_aprv_pending = product_2_item_through_table.objects.filter(PProduct_pk__Product__Product_Refrence_ID__in = sku_total_qty_cutting_lwo_aprv_pending).select_related('PProduct_pk__Product', 'Item_pk__Item_Color')

        for material in material_for_ref_id_queryset_for_cutting_aprv_pending:
            reference_id = material.PProduct_pk.Product.Product_Refrence_ID
            material_name = material.Item_pk.item_name
            color = material.Item_pk.Item_Color.color_name
            pro_sku = material.PProduct_pk.PProduct_SKU
            pro_fab_grp = material.Item_pk.Fabric_nonfabric
            panha_val = material.Item_pk.Panha
            unit_val = material.Item_pk.Units
            grand_total = material.grand_total
            grand_total_combi = material.grand_total_combi
            consumption = round(grand_total / (panha_val * unit_val),3)
            consumtionCombi = round(grand_total_combi / (panha_val * unit_val),3)
            common_unique = material.common_unique

            set_production_data_dict = {
                'ref_id': reference_id,
                'material_name': material_name,
                'color':color,
                'pro_sku': pro_sku,
                'pro_fab_grp': pro_fab_grp,
                'panha_val':panha_val,
                'unit_val':unit_val,
                'grand_total':grand_total,
                'consumption': consumption,
                'consumtionCombi': consumtionCombi,
                'common_unique': common_unique
            }
            material_for_ref_id_list_for_cutting_aprv_pending.append(set_production_data_dict)
        


        grouped_by_ref_id_unique_cutting_aprv_pending = {}

        for material in material_for_ref_id_list_for_cutting_aprv_pending:

            ref_no = material['ref_id']

            if ref_no not in grouped_by_ref_id_unique_cutting_aprv_pending:
                grouped_by_ref_id_unique_cutting_aprv_pending[ref_no] = []

            if material['common_unique']:
                if not any(m['material_name'] == material['material_name'] for m in grouped_by_ref_id_unique_cutting_aprv_pending[ref_no]):
                    grouped_by_ref_id_unique_cutting_aprv_pending[ref_no].append(material)
            else:
                grouped_by_ref_id_unique_cutting_aprv_pending[ref_no].append(material)

        # print("grouped_by_ref_id_unique_cutting_aprv_pending = ", grouped_by_ref_id_unique_cutting_aprv_pending)
    

    material_for_ref_id_list_for_lwo = []

    if sku_total_qty_lwo_pending:

        material_for_ref_id_queryset_for_low = product_2_item_through_table.objects.filter(PProduct_pk__Product__Product_Refrence_ID__in = sku_total_qty_lwo_pending).select_related('PProduct_pk__Product', 'Item_pk__Item_Color')

        for material in material_for_ref_id_queryset_for_low:
            reference_id = material.PProduct_pk.Product.Product_Refrence_ID
            material_name = material.Item_pk.item_name
            color = material.Item_pk.Item_Color.color_name
            pro_sku = material.PProduct_pk.PProduct_SKU
            pro_fab_grp = material.Item_pk.Fabric_nonfabric
            panha_val = material.Item_pk.Panha
            unit_val = material.Item_pk.Units
            grand_total = material.grand_total
            grand_total_combi = material.grand_total_combi
            consumption = round(grand_total / (panha_val * unit_val),3)
            consumtionCombi = round(grand_total_combi / (panha_val * unit_val),3)
            common_unique = material.common_unique

            set_production_data_dict = {
                'ref_id': reference_id,
                'material_name': material_name,
                'color':color,
                'pro_sku': pro_sku,
                'pro_fab_grp': pro_fab_grp,
                'panha_val':panha_val,
                'unit_val':unit_val,
                'grand_total':grand_total,
                'consumption': consumption,
                'consumtionCombi': consumtionCombi,
                'common_unique': common_unique
            }
            material_for_ref_id_list_for_lwo.append(set_production_data_dict)
        
        grouped_by_ref_id_unique_lwo = {}

        for material in material_for_ref_id_list_for_lwo:

            ref_no = material['ref_id']

            if ref_no not in grouped_by_ref_id_unique_lwo:
                grouped_by_ref_id_unique_lwo[ref_no] = []

            
            if material['common_unique']:

                if not any(m['material_name'] == material['material_name'] for m in grouped_by_ref_id_unique_lwo[ref_no]):
                    grouped_by_ref_id_unique_lwo[ref_no].append(material)
            else:
                grouped_by_ref_id_unique_lwo[ref_no].append(material)


        # print('grouped_by_ref_id_unique_lwo = ', grouped_by_ref_id_unique_lwo)




    merged_data = []

    if sku_total_qty_cutting and sku_total_qty_cutting_lwo_aprv_pending and sku_total_qty_lwo_pending:

        print("IN ALL")

        cutting_consumption_list = []

        for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)

        # print('cutting_consumption_list = ',cutting_consumption_list)
        # print(len(cutting_consumption_list))

# #******************************************************************************************************************************
        

        list_to_send_for_cutting_aprv_pending = []

        for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_cutting_aprv_pending.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

        # print('list_to_send_for_cutting_aprv_pending --- ', list_to_send_for_cutting_aprv_pending)
        # print(len(list_to_send_for_cutting_aprv_pending))

# #******************************************************************************************************************************

        

        lwo_consumption_list = []

        for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_lwo.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():
                                
                                sku = str(data['pro_sku'])

                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:

                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

        # print('lwo_consumption_list = ', lwo_consumption_list)
        # print(len(lwo_consumption_list))


        merge_aprv_pend_and_lwo_pend = []

        for i in list_to_send_for_cutting_aprv_pending:
            item_name = i['material_name']
            

            for j in lwo_consumption_list:
                if item_name == j['material_name']:
                    dict_to_append = {
                        'material_name': item_name,
                        'cutting_con_aprv_pen': i['cutting_con_aprv_pen'],
                        'lwo_consumption': j['lwo_consumption'], 
                    }
                    if dict_to_append not in merge_aprv_pend_and_lwo_pend:  
                        merge_aprv_pend_and_lwo_pend.append(dict_to_append)
                    break

        for x in list_to_send_for_cutting_aprv_pending:
            if not any(d['material_name'] == x['material_name'] for d in merge_aprv_pend_and_lwo_pend):
                merge_aprv_pend_and_lwo_pend.append(
                    {
                        'material_name': x['material_name'],
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': 0, 
                    }
                )

        for x in lwo_consumption_list:
            if not any(d['material_name'] == x['material_name'] for d in merge_aprv_pend_and_lwo_pend):
                merge_aprv_pend_and_lwo_pend.append(
                    {
                        'material_name': x['material_name'],
                        'cutting_con_aprv_pen':0,
                        'lwo_consumption':  x['lwo_consumption'], 
                    }
                )
        
        # print('merge_aprv_pend_and_lwo_pend = ', merge_aprv_pend_and_lwo_pend)
        # print(len(merge_aprv_pend_and_lwo_pend))

        data_list = []

        for x in merge_aprv_pend_and_lwo_pend:
            item_name = x['material_name']
  
            for i in cutting_consumption_list:
                if item_name == i['material_name']:
                    
                    dict_to_append = {
                        'material_name': i['material_name'],
                        'cutting_consumption': i['cutting_consumption'],
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': x['lwo_consumption'], 
                    }

                    if dict_to_append not in data_list:  
                        data_list.append(dict_to_append)
                    break


        for i in cutting_consumption_list:
            if not any(d['material_name'] == i['material_name'] for d in data_list):
                data_list.append({
                    'material_name': i['material_name'],
                    'cutting_consumption': i['cutting_consumption'],
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                })


        for x in data_list:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None

                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': x['cutting_consumption'],
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': x['lwo_consumption'],
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'po_qty':po_qty,
                        'balance' : (i.total_qty if i.total_qty is not None else 0)-((x['cutting_consumption']) + (x['cutting_con_aprv_pen']) + (x['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
					    'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break  
 
        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'po_qty':po_qty,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
					'mobile_no' : mobile,
                })


    elif sku_total_qty_cutting and sku_total_qty_cutting_lwo_aprv_pending:

        print("IN CUTTING PENDING AND CUTTING APPROVE PENDING")

        cutting_consumption_list = []

        for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)



        list_to_send_for_cutting_aprv_pending = []

        for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_cutting_aprv_pending.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

        merge_two_list = []

        for i in cutting_consumption_list:
            item_name = i['material_name']

            for j in list_to_send_for_cutting_aprv_pending:
                if item_name == j['material_name']:
                    dict_to_append = {
                        'material_name': item_name,
                        'cutting_consumption': i['cutting_consumption'],
                        'cutting_con_aprv_pen': j['cutting_con_aprv_pen'],
                    }
                    if dict_to_append not in merge_two_list:  
                        merge_two_list.append(dict_to_append)
                    break

        for i in cutting_consumption_list:
            if not any(d['material_name'] == i['material_name'] for d in merge_two_list):
                merge_two_list.append({
                    'material_name': i['material_name'],
                    'cutting_consumption': i['cutting_consumption'],
                    'cutting_con_aprv_pen': 0,
                })

        for i in list_to_send_for_cutting_aprv_pending:
            if not any(d['material_name'] == i['material_name'] for d in merge_two_list):
                merge_two_list.append({
                    'material_name': i['material_name'],
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': i['cutting_con_aprv_pen'],
                })               



        for x in merge_two_list:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None


                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': x['cutting_consumption'],
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': 0,
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'po_qty':po_qty,
                        'balance' : (i.total_qty if i.total_qty is not None else 0)-((x['cutting_consumption']) + (x['cutting_con_aprv_pen']) + (0) - (po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
					    'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break  

        
        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None


            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'po_qty':po_qty,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
					'mobile_no' : mobile,
                })


    elif sku_total_qty_cutting_lwo_aprv_pending and sku_total_qty_lwo_pending:

        print("ONLY IF CUTTING APPROVE PENDING AND LWO PENDING")
        
        list_to_send_for_cutting_aprv_pending = []

        for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_cutting_aprv_pending.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)

        lwo_consumption_list = []

        for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_lwo.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():
                                
                                sku = str(data['pro_sku'])

                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:

                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

        merge_two_list = []

        for i in lwo_consumption_list:
            item_name = i['material_name']

            for j in list_to_send_for_cutting_aprv_pending:
                if item_name == j['material_name']:
                    dict_to_append = {
                        'material_name': item_name,
                        'lwo_consumption': i['lwo_consumption'],
                        'cutting_con_aprv_pen': j['cutting_con_aprv_pen'],
                    }
                    if dict_to_append not in merge_two_list:  
                        merge_two_list.append(dict_to_append)
                    break

        for i in lwo_consumption_list:
            if not any(d['material_name'] == i['material_name'] for d in merge_two_list):
                merge_two_list.append({
                    'material_name': i['material_name'],
                    'lwo_consumption': i['lwo_consumption'],
                    'cutting_con_aprv_pen': 0,
                })

        for i in list_to_send_for_cutting_aprv_pending:
            if not any(d['material_name'] == i['material_name'] for d in merge_two_list):
                merge_two_list.append({
                    'material_name': i['material_name'],
                    'lwo_consumption': 0,
                    'cutting_con_aprv_pen': i['cutting_con_aprv_pen'],
                }) 
        


        for x in merge_two_list:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None


                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': 0,
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': x['lwo_consumption'],
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'po_qty':po_qty,
                        'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (x['cutting_con_aprv_pen']) + (x['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
					    'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break  

        
        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'po_qty':po_qty,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
					'mobile_no' : mobile,
                })


    elif sku_total_qty_cutting and sku_total_qty_lwo_pending:
        
        print("ONLY IF CUTTING PENDING AND LWO PENDING")

        cutting_consumption_list = []

        for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)


        lwo_consumption_list = []

        for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_lwo.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():
                                
                                sku = str(data['pro_sku'])

                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:

                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

        merge_two_list = []

        for i in cutting_consumption_list:
            item_name = i['material_name']

            for j in lwo_consumption_list:
                if item_name == j['material_name']:
                    dict_to_append = {
                        'material_name': item_name,
                        'cutting_consumption': i['cutting_consumption'],
                        'lwo_consumption': j['lwo_consumption'],
                    }
                    if dict_to_append not in merge_two_list:  
                        merge_two_list.append(dict_to_append)
                    break

        for i in cutting_consumption_list:
            if not any(d['material_name'] == i['material_name'] for d in merge_two_list):
                merge_two_list.append({
                    'material_name': i['material_name'],
                    'cutting_consumption': i['cutting_consumption'],
                    'lwo_consumption': 0,
                })

        for i in lwo_consumption_list:
            if not any(d['material_name'] == i['material_name'] for d in merge_two_list):
                merge_two_list.append({
                    'material_name': i['material_name'],
                    'cutting_consumption': 0,
                    'lwo_consumption': i['lwo_consumption'],
                })

        
        for x in merge_two_list:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None

                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': x['cutting_consumption'],
                        'cutting_con_aprv_pen': 0,
                        'lwo_consumption': x['lwo_consumption'],
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'po_qty':po_qty,
                        'balance' : (i.total_qty if i.total_qty is not None else 0)-((x['cutting_consumption']) + x['lwo_consumption'] + (0) - (po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
					    'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break 

        
        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'po_qty':po_qty,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
					'mobile_no' : mobile,
                })
    

    elif sku_total_qty_cutting:

        print("ONLY IF CUTTING PENDING")

        cutting_consumption_list = []

        for key, value in sku_total_qty_cutting.items():

                for refno, value_list in grouped_by_ref_id_unique.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k, v in value.items():

                                if data['pro_sku'] == k and data['common_unique'] == False:

                                    total_consumption_value = v * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)
                                        

                                elif data['pro_sku'] == k and data['common_unique'] == True:

                                    total_consumption_value = value['total'] * data['consumption']

                                    total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                    total_consumption = total_consumption_value + total_combi_consumption_value

                                    dict_append = {
                                        'ref_id': data['ref_id'],
                                        'product_sku': data['pro_sku'],
                                        'color': data['color'],
                                        'material_name': data['material_name'],
                                        'product_fabric_grp': data['pro_fab_grp'],
                                        'cutting_consumption': total_consumption,
                                        'common_unique': data['common_unique']
                                    }

                                    existing_item = next( (item for item in cutting_consumption_list if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']), None)

                                    if existing_item:
                                        existing_item['cutting_consumption'] += total_consumption
                                    else:
                                        cutting_consumption_list.append(dict_append)


        for x in cutting_consumption_list:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None

                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': x['cutting_consumption'],
                        'cutting_con_aprv_pen': 0,
                        'lwo_consumption': 0,
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'po_qty':po_qty,
                        'balance' : (i.total_qty if i.total_qty is not None else 0)-((x['cutting_consumption']) + (0) + (0)-(po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
                        'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break  

        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'po_qty':po_qty,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (0) + (0)-(po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
		            'mobile_no' : mobile,
                })


    elif sku_total_qty_cutting_lwo_aprv_pending:

        print(" ONLY CUTTING APPROVED PENDING")
  
        list_to_send_for_cutting_aprv_pending = []

        for key,value in sku_total_qty_cutting_lwo_aprv_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_cutting_aprv_pending.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():

                                sku = str(data['pro_sku'])
                                
                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                    
                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'cutting_con_aprv_pen' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next((item for item in list_to_send_for_cutting_aprv_pending if item['material_name'] == data['material_name'] and item['common_unique'] == data['common_unique']),None)

                                        if existing_item:
                                            existing_item['cutting_con_aprv_pen'] += total_consumption
                                        else:
                                            list_to_send_for_cutting_aprv_pending.append(dict_append)


        for x in list_to_send_for_cutting_aprv_pending:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None

                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': 0,
                        'cutting_con_aprv_pen': x['cutting_con_aprv_pen'],
                        'lwo_consumption': 0,
                        'po_qty':po_qty,
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (x['cutting_con_aprv_pen']) + (0) - (po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
                        'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break  

        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'po_qty':po_qty,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
                    'mobile_no' : mobile,
                })
    
    elif sku_total_qty_lwo_pending:

        print("ONLY LWO PENDING")

        lwo_consumption_list = []

        for key,value in sku_total_qty_lwo_pending.items():
                    
                for refno, value_list in grouped_by_ref_id_unique_lwo.items():

                    for data in value_list:

                        if data['ref_id'] == key:

                            for k,v in value.items():
                                
                                sku = str(data['pro_sku'])

                                if sku == k and data['common_unique'] == False:
                                    
                                    if data['pro_fab_grp'] == 'Non Fabric':

                                        total_consumption_value = v * data['consumption']

                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value


                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:

                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

                                elif sku == k and data['common_unique'] == True:

                                    if data['pro_fab_grp'] == 'Non Fabric':
                                    
                                        total_consumption_value = value['total'] * data['consumption']
                                        
                                        total_combi_consumption_value = value['total'] * data['consumtionCombi']

                                        total_consumption = total_consumption_value + total_combi_consumption_value

                                        dict_append = {
                                            'ref_id' : data['ref_id'],
                                            'product_sku' : data['pro_sku'],
                                            'color': data['color'],
                                            'material_name':data['material_name'],
                                            'product_fabric_grp': data['pro_fab_grp'],
                                            'lwo_consumption' : total_consumption,
                                            'common_unique': data['common_unique']
                                        }

                                        existing_item = next(
                                            (item for item in lwo_consumption_list
                                            if item['material_name'] == data['material_name'] and
                                            item['common_unique'] == data['common_unique']),
                                            None
                                        )

                                        if existing_item:
                                            
                                            existing_item['lwo_consumption'] += total_consumption
                                        else:
                                            lwo_consumption_list.append(dict_append)

        for x in lwo_consumption_list:
            item_name = x['material_name']
            matched = False  
            
            for i in negetive_stock_report_list:

                po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

                party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

                if party_names: 
                    p_name = party_names.item_purchase_master.party_name.name
                    mobile = party_names.item_purchase_master.party_name.mobile_no
                else:
                    p_name = None
                    mobile = None

                
                
                if item_name == i.item_name:
                    matched = True  
                    
                    dict_to_append = {
                        'item_id' : i.id,
                        'Material_code' : i.Material_code,
                        'Item_Color':i.Item_Color.color_name,
                        'Fabric_Group':i.Fabric_Group.fab_grp_name,
                        'unit_name' : i.unit_name_item.unit_name,
                        'material_name': i.item_name,
                        'cutting_consumption': 0,
                        'cutting_con_aprv_pen': 0,
                        'lwo_consumption': x['lwo_consumption'],
                        'total_qty': i.total_qty if i.total_qty is not None else 0,
                        'po_qty':po_qty,
                        'balance' : (i.total_qty if i.total_qty is not None else 0) - ((0) + (0) + (x['lwo_consumption']) - (po_qty if po_qty is not None else 0)),
                        'party_name' : p_name,
                        'mobile_no' : mobile,
                    }

                    if dict_to_append not in merged_data:  
                        merged_data.append(dict_to_append)
                    break  

        for i in negetive_stock_report_list:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            if not any(d['material_name'] == i.item_name for d in merged_data):

                merged_data.append({
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'po_qty':po_qty,
                    'balance' : (i.total_qty if i.total_qty is not None else 0) - ((0) + (0) + (0) - (po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
                    'mobile_no' : mobile,
                })


    else:
        negetive_stock_report =  Item_Creation.objects.all().annotate(total_qty = Sum('shades__godown_shades__quantity')).order_by('item_name').select_related('Item_Color','Fabric_Group')

        for i in negetive_stock_report:

            po_qty = purchase_order_for_puchase_voucher_rm.objects.filter(item_name__item_name=i.item_name).aggregate(total_qty=Sum('quantity'))['total_qty']

            party_names = purchase_voucher_items.objects.filter(item_shade__items__item_name = i.item_name).select_related('item_purchase_master__party_name').order_by('-created_date').first()

            if party_names: 
                p_name = party_names.item_purchase_master.party_name.name
                mobile = party_names.item_purchase_master.party_name.mobile_no
            else:
                p_name = None
                mobile = None

            dict_to_append = {
                    'item_id' : i.id,
                    'Material_code' : i.Material_code,
                    'Item_Color':i.Item_Color.color_name,
                    'Fabric_Group':i.Fabric_Group.fab_grp_name,
                    'unit_name' : i.unit_name_item.unit_name,
                    'material_name': i.item_name,
                    'cutting_consumption': 0,
                    'cutting_con_aprv_pen': 0,
                    'lwo_consumption': 0,
                    'po_qty':po_qty,
                    'total_qty': i.total_qty if i.total_qty is not None else 0,
                    'balance' : (i.total_qty if i.total_qty is not None else 0)-((0)+(0)+(0)-(po_qty if po_qty is not None else 0)),
                    'party_name' : p_name,
                    'mobile_no' : mobile,
                }

            merged_data.append(dict_to_append)


    data_for_frontend = sorted(merged_data , key=itemgetter('total_qty'))

    if filter_name == 'highest' and filter_name != '':
        data_for_frontend = sorted(merged_data, key=lambda x: x["total_qty"], reverse=True)

    if filter_name == 'lowest' and filter_name != '':
        data_for_frontend = sorted(merged_data, key=lambda x: x["total_qty"])
    
    if selected_fabric_grp:
        selected_fabric_grp_lower = selected_fabric_grp.lower()
        data_for_frontend = [item for item in merged_data if item["Fabric_Group"].lower() == selected_fabric_grp_lower]

    if less_Number:
        less_Number = decimal.Decimal(less_Number)
        data_for_frontend = [item for item in merged_data if item["total_qty"] < less_Number]

    if selected_fabric_grp and less_Number:
        selected_fabric_grp_lower = selected_fabric_grp.lower()
        less_Number = decimal.Decimal(less_Number)
        data_for_frontend = [item for item in merged_data if item["Fabric_Group"].lower() == selected_fabric_grp_lower and item["total_qty"] < less_Number]

    return render(request,'accounts/negetive_stock.html',{'negetive_stock_report':negetive_stock_report ,'negetive_stock_sellerwise':negetive_stock_sellerwise,'selected_fabric_grp':selected_fabric_grp, 'less_Number':less_Number,'merged_data':data_for_frontend,})










def purchase_order_for_purchase_rm_delete(request,pk):

    if pk:
        purchase_order_instance = get_object_or_404(purchase_order_master_for_puchase_voucher_rm,pk=pk)
        purchase_order_instance.delete()
        return redirect('purchase-order-for-puchase-voucher-rm-list')





def excel_download_for_purchase_order(request,p_id):

    purchase_order_child = purchase_order_for_puchase_voucher_rm.objects.filter(master_instance = p_id)
    party_details = purchase_order_master_for_puchase_voucher_rm.objects.get(id=p_id)
    wb = Workbook()
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    ws = wb.create_sheet('Purchase Order')

    logo_path = "static/images/PO_for_RM_logo.jpg"
    img = Image(logo_path)
    img.width = 400
    img.height = 100
    ws.add_image(img, "E1")

    thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),

    )

    ws.merge_cells('A6:C6')

    ws['A6'] = "To"
    ws['A7'] = party_details.party_name.name
    ws['A8'] = "EMAIL ID - "
    ws['A9'] = f"Contact - {party_details.party_name.mobile_no}"
    ws['A10'] = f"GSTIN - {party_details.party_name.Gst_no}"
    ws.merge_cells('D6:K8')


    ws['D6'] = f"{party_details.party_name.address} {party_details.party_name.city} {party_details.party_name.state} {party_details.party_name.country} {party_details.party_name.pincode}"
    ws['D6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['E9'] = "CIN - "
    ws['E10'] = f"GSTIN - "
    ws['N6'] = "Date - "
    ws['O6'] = datetime.date.today()
    ws['N7'] = "PO NO "
    ws['O7'] = party_details.po_no
    ws['N8'] = "Payment Term"
    ws['N9'] = "Remark - "

    headers = ['No','M code/Color Code','Item Name', 'Color', 'Quantity', 'Units', 'Rate', '1st Delivery', '2st Delivery', '3st Delivery', 'Tax', 'Gst%', 'CGST', 'SGST/IGST', 'Total Amount']

    start_row = 11
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)


    ws.row_dimensions[start_row].height = 40

    column_widths = [5, 7, 30, 7, 9, 7, 7, 9, 9, 9, 7, 7, 7, 11, 15]
    for col_num, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width
    data_start_row = 12
    max_rows = 35
    rows_filled = 0

    for row_num, child in enumerate(purchase_order_child, start=data_start_row):
        ws.cell(row=row_num, column=1, value=row_num - data_start_row + 1)
        ws.cell(row=row_num, column=2, value=child.item_name.Material_code)
        ws.cell(row=row_num, column=3, value=child.item_name.item_name)
        ws.cell(row=row_num, column=4, value=child.item_name.Item_Color.color_name)
        ws.cell(row=row_num, column=5, value=child.quantity)
        ws.cell(row=row_num, column=6, value=child.item_name.unit_name_item.unit_name)
        ws.cell(row=row_num, column=7, value=child.rate)
        ws.cell(row=row_num, column=8, )
        ws.cell(row=row_num, column=9, )
        ws.cell(row=row_num, column=10, )
        ws.cell(row=row_num, column=11, )
        ws.cell(row=row_num, column=12, value=child.item_name.Item_Creation_GST.gst_percentage)
        ws.cell(row=row_num, column=13, value=child.item_name.Item_Creation_GST.gst_percentage)
        ws.cell(row=row_num, column=14, )
        ws.cell(row=row_num, column=15, )

        rows_filled += 1
        total_quantity = purchase_order_for_puchase_voucher_rm.objects.filter(master_instance=p_id).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

        if rows_filled > 35:
            ws.cell(row = rows_filled + 1, column = 13, value=f"Total Taxable Value")
            ws.cell(row = rows_filled + 2, column = 13, value = "GST")
            ws.cell(row = rows_filled + 2, column = 14, value = child.item_name.Item_Creation_GST.gst_percentage)


            ws.cell(row = rows_filled + 3, column = 13, value = "Freight")
            ws.cell(row = rows_filled + 4, column = 13, value = "Round off")
            ws.cell(row = rows_filled + 5, column = 13, value = "Total Value")

            ws.cell(row=rows_filled + 4, column=5, value=total_quantity)
            ws.cell(row=rows_filled + 4, column=6, value=child.item_name.unit_name_item.unit_name)
            ws.merge_cells(start_row=rows_filled + 5, start_column=1, end_row=rows_filled + 5, end_column=12)
            ws.cell(row=rows_filled + 5, column = 1, value="Rupees")

            ws.merge_cells(start_row=rows_filled + 6, start_column=1, end_row=rows_filled + 6, end_column=15)
            ws.cell(row=rows_filled + 6, column = 1, value="THANK YOU FOR YOUR BUSINESS").alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=rows_filled + 8, column = 1, value="Order by")
            ws.cell(row=rows_filled + 8, column = 4, value="Prepared by")
            ws.cell(row=rows_filled + 8, column = 12, value="Authorised Signature")
            ws.cell(row=rows_filled + 10, column = 12, value="For Cosmus Lifestyle Pvt Ltd")

            for row in ws.iter_rows(min_row=6, max_row=rows_filled + 5, min_col=1, max_col=15):
                for cell in row:
                    cell.border = thin_border

        else:
            ws.cell(row = max_rows + 1, column = 13, value=f"Total Taxable Value")
            ws.cell(row = max_rows + 2, column = 13, value = "GST")
            ws.cell(row = max_rows + 2, column = 14, value = f"{child.item_name.Item_Creation_GST.gst_percentage} %")
            ws.cell(row = max_rows + 3, column = 13, value = "Freight")
            ws.cell(row = max_rows + 4, column = 13, value = "Round off")
            ws.cell(row = max_rows + 5, column = 13, value = "Total Value")

            ws.cell(row=max_rows + 4, column=5, value=total_quantity)
            ws.cell(row=max_rows + 4, column=6, value=child.item_name.unit_name_item.unit_name)
            ws.merge_cells(start_row=max_rows + 5, start_column=1, end_row=max_rows + 5, end_column=12)
            ws.cell(row=max_rows + 5, column = 1, value="Rupees")
            ws.merge_cells(start_row=max_rows + 6, start_column=1, end_row=max_rows + 6, end_column=15)
            ws.cell(row=max_rows + 6, column = 1, value="THANK YOU FOR YOUR BUSINESS").alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=max_rows + 8, column = 1, value="Order by")
            ws.cell(row=max_rows + 8, column = 4, value="Prepared by")
            ws.cell(row=max_rows + 8, column = 12, value="Authorised Signature")
            ws.cell(row=max_rows + 10, column = 12, value="For Cosmus Lifestyle Pvt Ltd")

            for row in ws.iter_rows(min_row=6, max_row=max_rows + 5, min_col=1, max_col=15):
                for cell in row:
                    cell.border = thin_border

    fileoutput = BytesIO()
    wb.save(fileoutput)
    response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="purchase_order.xlsx"'
    return response










def itemdynamicsearchajax(request):
    
    try:
        
        item_name_typed = request.GET.get('nameValue')

        if not item_name_typed:
            raise ValidationError("partial name provided.")
        
        logger.info(f"searched keyword via itemdynamicsearchajax {item_name_typed}")
        
        item_name_searched = Item_Creation.objects.filter(Q(item_name__icontains=item_name_typed) | Q(Material_code__contains=item_name_typed)).annotate(total_qty= Sum('shades__godown_shades__quantity'))
        

        if item_name_searched:
            

            searched_item_name_dict = {}
            for queryset in item_name_searched:

                if queryset.total_qty is not None:
                    total_qty = str(queryset.total_qty)

                else:
                    total_qty = '0'

                item_name = queryset.item_name + ' | ' + total_qty + ',' + queryset.Material_code
                item_id = queryset.id
                searched_item_name_dict[item_id] = item_name
            
            logger.info(f"searched result via itemdynamicsearchajax {searched_item_name_dict}")
            print(searched_item_name_dict)
            return JsonResponse({'item_name_typed': item_name_typed, 'searched_item_name_dict': searched_item_name_dict}, status=200)
        
        else:
            return JsonResponse({'error': 'No items found.'}, status=404)

    except ValidationError as ve:
        error_message = str(ve)
        logger.error(f"Validaton errorin itemdynamicsearchajax - {ve}")
        return JsonResponse({'error': error_message}, status = 400)
    
    except Exception as e:
        logger.error(f"Exception in itemdynamicsearchajax - {ve}")
        error_message = f"An error occurred:{str(e)}"
        return JsonResponse({'error': error_message}, status = 500)






def productdynamicsearchajax(request):
    try:
        product_name_typed = request.GET.get('productnamevalue')

        if not product_name_typed:
            return JsonResponse({'error': 'Please enter a search term.'}, status=400)
        
        logger.info(f"Search initiated by {request.user}: {product_name_typed}")
        
        products = PProduct_Creation.objects.filter(
            Q(PProduct_SKU__icontains=product_name_typed) |
            Q(PProduct_color__color_name__icontains=product_name_typed) |
            Q(Product__Product_Name__icontains=product_name_typed) |
            Q(Product__Product_Refrence_ID__icontains=product_name_typed)
        ).annotate(quantity=Subquery(product_godown_quantity_through_table.objects.filter(product_color_name=OuterRef('pk')
        ).values('quantity'))).distinct().values('PProduct_SKU', 'PProduct_color__color_name', 'Product__Model_Name', 'Product__Product_GST__gst_percentage','Product__Product_MRP','Product__Product_SalePrice_CustomerPrice','quantity','Product__Product_Refrence_ID')
        
        if products.exists():
            product_searched_dict = {
                product['PProduct_SKU']: [
                    product.get('Product__Model_Name', ''),
                    product.get('PProduct_color__color_name', ''),
                    product.get('Product__Product_GST__gst_percentage', ''),
                    product.get('Product__Product_MRP', ''),
                    product.get('Product__Product_SalePrice_CustomerPrice', ''),
                    product.get('quantity', '') if product.get('quantity', '') else 0,
                    product.get('Product__Product_Refrence_ID', ''),
                    product.get('Product', '')

                ] for product in products
            }
            
            
            logger.info(f"Search results for {product_name_typed}: {product_searched_dict}")
            return JsonResponse({'typed': product_name_typed, 
                                 'products': product_searched_dict}, status=200)
        
        return JsonResponse({'error': 'No items found.'}, status=404)

    except Exception as e:
        logger.error(f"Error during search by {request.user}: {e}")
        return JsonResponse({'error': f"An error occurred: {str(e)}"}, status=500)






def CheckUniqueFieldDuplicate(model_name, searched_value, col_name):
    
    if searched_value:
        validation_flag = False
        try:
            
            lookup = {f"{col_name}__iexact": searched_value}
            check_instance_valid = model_name.objects.get(**lookup)
            
            validation_flag = True

        except model_name.DoesNotExist:
            validation_flag = False
            
        except Exception as e:
            return JsonResponse({f'Status':'Exception Occoured - {e}'}, status=404)
        
        return JsonResponse({'validation_flag':validation_flag})
    else:
        return JsonResponse({f'Status':'No data recieved - {e}'}, status=404)







def UniqueValidCheckAjax(request):
    searched_from = request.GET.keys()
    
    if 'purchase_number' in searched_from:
        searched_value = request.GET.get('purchase_number').strip()
        
        model_name = item_purchase_voucher_master
        col_name = 'purchase_number'

    elif 'new_order_number' in searched_from:
        searched_value = request.GET.get('new_order_number').strip()
        model_name = purchase_order
        col_name = 'purchase_order_number'

    elif 'cutting_order_number' in searched_from:
        searched_value = request.GET.get('cutting_order_number').strip()
        model_name = purchase_order_raw_material_cutting
        col_name = 'raw_material_cutting_id'

    elif 'labour_workout_challan_no' in searched_from:
        searched_value = request.GET.get('labour_workout_challan_no').strip()
        model_name = labour_workout_childs
        col_name = 'challan_no'

    elif 'item_name' in searched_from:
        searched_value = request.GET.get('item_name').strip()
        model_name = Item_Creation
        col_name = 'item_name'

    elif 'item_material_code' in searched_from:
        searched_value = request.GET.get('item_material_code').strip()
        model_name = Item_Creation
        col_name = 'Material_code'
    
    elif 'voucher_number' in searched_from:
        searched_value = request.GET.get('voucher_number').strip()
        model_name = labour_work_in_master
        col_name = 'voucher_number'

    elif 'purchase_no' in searched_from:
        searched_value = request.GET.get('purchase_no').strip()
        model_name = product_purchase_voucher_master
        col_name = 'purchase_number'

    elif 'voucher_no' in searched_from:
        searched_value = request.GET.get('voucher_no').strip()
        model_name = Finished_goods_Stock_TransferMaster
        col_name = 'voucher_no'

    elif 'sales_no' in searched_from:
        searched_value = request.GET.get('sales_no').strip()
        model_name = sales_voucher_master_finish_Goods
        col_name = 'sales_no'

    elif 'sales_no' in searched_from:
        searched_value = request.GET.get('sales_no').strip()
        model_name = sales_voucher_master_outward_scan
        col_name = 'sales_no'
    
    else:
        model_name = None
        searched_value = None
        col_name = None


    
    if model_name and searched_value and col_name:
        return CheckUniqueFieldDuplicate(model_name, searched_value, col_name)
    else:
        return JsonResponse({'Status': 'Invalid data received'}, status=400)









def session_data_test(request):
        
    session_data = request.session
    
    for key, value in session_data.items():
        print(f"Key: {key}, Value: {value}")

    context = {}
    return render(request,'misc/session_test.html', context=context)






def finished_goods_stock_all(request,pk=None):
    wareshouse_all = Finished_goods_warehouse.objects.all()
    
    button_value = request.POST.get('All_stock')
    
    if pk:
        warehouse_data = Product_warehouse_quantity_through_table.objects.filter(product__PProduct_SKU = OuterRef('pk'),                     
                         warehouse__id = pk).values('quantity')     

        finished_godown_all = PProduct_Creation.objects.annotate(total_warehouse_stock = Subquery(
            warehouse_data)).order_by('Product__Product_Name').select_related('PProduct_color')

    elif button_value:

        
        finished_godown_all = PProduct_Creation.objects.annotate(total_warehouse_stock = Sum( 
            'product_warehouse_quantity_through_table__quantity')).order_by('Product__Product_Name').select_related('PProduct_color') 
    else:

        finished_godown_all = PProduct_Creation.objects.annotate(total_warehouse_stock = Sum( 
            'product_warehouse_quantity_through_table__quantity')).order_by('Product__Product_Name').select_related('PProduct_color')

    return render(request,'finished_product/finishedgoodsstockall.html',{'finished_godown_all':finished_godown_all, 'wareshouse_all':wareshouse_all})






def product_2_item_ajax(request):

    product_ref_id_get = request.GET.get('product_ref_id')
    if product_ref_id_get is not None or product_ref_id_get != '':
        try:

            product_ref_id = get_object_or_404(Product, pk=product_ref_id_get)

            product_2_items_instances_unique = list(product_2_item_through_table.objects.filter(
                                PProduct_pk__Product__Product_Refrence_ID = product_ref_id.Product_Refrence_ID, common_unique = False).order_by(
                                    'row_number').values('Item_pk','Item_pk__item_name','row_number','no_of_rows','Remark'))
        
            product_2_items_instances_common = list(product_2_item_through_table.objects.filter(
                                PProduct_pk__Product__Product_Refrence_ID = product_ref_id.Product_Refrence_ID, common_unique = True).order_by(
                                    'row_number','id').distinct('row_number').values('Item_pk','Item_pk__item_name','row_number','no_of_rows','Remark'))
            

            return JsonResponse({'status': 'success', 'product_2_items_instances_unique':product_2_items_instances_unique,
                                     'product_2_items_instances_common':product_2_items_instances_common})
        
        except ObjectDoesNotExist as e:
            logger.error(f'Record not found: {e}')
            return JsonResponse({'status': 'error', 'message': 'Record not found'}, status=404)

        except Exception as e:
            logger.error(f'Invalid data: {e}')
            return JsonResponse({'status': 'error', 'message': 'Invalid input data'}, status=400)

    else:
        logger.error(f'Invalid data: {e}')
        return JsonResponse({'status': 'error', 'message': 'Invalid input data else'}, status=400)








def creditdebitreport(request):
    all_reports = account_credit_debit_master_table.objects.all()

    return render(request,'misc/credit_debit_master_report.html',{'all_reports':all_reports,'page_name':'Credit/Debit Report'})







def godown_stock_raw_material_report_fab_grp(request,g_id,fab_id=None):
    
    
    items_in_godown = item_godown_quantity_through_table.objects.filter(godown_name=g_id)
    
    
    Fabric_grp_name = None
    querylist = None

    
    if fab_id:
        page_id = 'item_page'

    else:
        page_id = 'fabric_page'
    
    if page_id == 'fabric_page':
        
        fabric_in_godown = items_in_godown.distinct('Item_shade_name__items__Fabric_Group')
        
        list_fab_grp = []

        
        for fab in fabric_in_godown:
            list_fab_grp.append(fab.Item_shade_name.items.Fabric_Group.id)

        queryset = []

        
        for items in list_fab_grp:
            values = Fabric_Group_Model.objects.filter(id=
                    items).filter(items__shades__godown_shades__godown_name=g_id).annotate(total_qty = 
                    Round(Sum('items__shades__godown_shades__quantity'), 2),
                    avg_rate=Round(Avg('items__shades__rate'),2)).first()

                    

            queryset.append(values)

    elif page_id == 'item_page':

        

        Fabric_grp_name = Fabric_Group_Model.objects.get(id=fab_id)

        items_in_fab_grp = Item_Creation.objects.filter(Fabric_Group=fab_id).filter(
            shades__godown_shades__godown_name__id=g_id).annotate(
                total_qty =Round(Sum('shades__godown_shades__quantity')))
                

        querylist = []
        
        
        for query in items_in_fab_grp: 
            item_dict = {}
            item_dict['item_name'] = query.item_name
            item_dict['total_qty'] = query.total_qty

            shades_list = []
            for shade in query.shades.filter(godown_shades__godown_name__id=g_id): 
                shade_dict = {}
                shade_dict['rate'] = shade.rate
                shades_list.append(shade_dict)

                shade_godown_list = []
                for godown_items in shade.godown_shades.filter(godown_name__id=g_id): 
                    godown_shade_dict = {}
                    shade_dict['item_shade'] = godown_items.Item_shade_name.item_shade_name
                    shade_dict['item_shade_id'] = godown_items.Item_shade_name.id
                    shade_dict['quantity'] = godown_items.quantity
                    shade_godown_list.append(godown_shade_dict)

            item_dict['shades'] = shades_list
            querylist.append(item_dict)
        queryset = items_in_fab_grp

    godown_name = Godown_raw_material.objects.get(id = g_id)
    
    
    return render(request,'reports/godownstockrawmaterialreportfabgrp.html',{'page_id':page_id,
                                                                             'godown_id':g_id,
                                                                             'godown_name':godown_name,
                                                                             'Fabric_grp_name':Fabric_grp_name,
                                                                             'queryset':queryset,
                                                                             'querylist':querylist})








def godown_item_report(request, shade_id,g_id=None):

    """
    scheme = request.scheme
    host = request.get_host()
    path = request.get_full_path()
    request.build_absolute_uri()    
    """
    scheme = request.scheme
    host = request.get_host()

    root_url = f'{scheme}://{host}/'

    shade_name = item_color_shade.objects.get(id=shade_id)
    godown_name = 'All Stock'

    report_data = []

    opening_godown_data = []

    if g_id is not None:
        godown_name = Godown_raw_material.objects.get(id=g_id)


        
        opening_godown_qty = opening_shade_godown_quantity.objects.filter(
            opening_purchase_voucher_godown_item=shade_name, opening_godown_id=godown_name)
        

        
        
        purchase_voucher_godown_qty = item_purchase_voucher_master.objects.filter(
            purchase_voucher_items__item_shade = shade_name , purchase_voucher_items__shade_godown_items__godown_id = godown_name).annotate(
                godown_qty_total=Sum('purchase_voucher_items__shade_godown_items__quantity'), item_rate=Round(Avg(
                    'purchase_voucher_items__rate')), filter=Q(purchase_voucher_items__shade_godown_items__godown_id = godown_name))
        

        
        purchase_order_cutting_room_qty = godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = False, godown_id = g_id)
        
        
    
        
        purchase_order_cutting_room_qty_cancelled =  godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = True, godown_id = g_id)
        

        labour_workout_report = labour_workout_cutting_items.objects.filter(material_name = shade_name.items.item_name,
            material_color_shade = shade_name.item_shade_name,labour_workout_child_instance__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__temp_godown_select = g_id)
        


        
        Raw_Stock_Transfer_Master_object_qty_outward = RawStockTrasferRecords.objects.filter(master_instance__source_godown__godown_name_raw = godown_name, item_shade_transfer = shade_name) 

        Raw_Stock_Transfer_Master_object_qty_inward = RawStockTrasferRecords.objects.filter(master_instance__destination_godown__godown_name_raw = godown_name, item_shade_transfer = shade_name) 

        Raw_Stock_Transfer_Master_object_qty_both_with_shade_inward = []

    else:
        
        Raw_Stock_Transfer_Master_object_qty_inward = []

        Raw_Stock_Transfer_Master_object_qty_outward = []

        
        opening_godown_qty = opening_shade_godown_quantity.objects.filter(
            opening_purchase_voucher_godown_item = shade_name)
        


        
        
        purchase_voucher_godown_qty = item_purchase_voucher_master.objects.filter(
            purchase_voucher_items__item_shade = shade_name).annotate(
                godown_qty_total=Sum('purchase_voucher_items__shade_godown_items__quantity'), item_rate=Round(Avg(
                    'purchase_voucher_items__rate')))
        

        
        purchase_order_cutting_room_qty = godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = False)
        
        
    
        
        purchase_order_cutting_room_qty_cancelled =  godown_item_report_for_cutting_room.objects.filter(
            material_color_shade = shade_id, inward = True)
        

        labour_workout_report = labour_workout_cutting_items.objects.filter(material_name = shade_name.items.item_name,
            material_color_shade = shade_name.item_shade_name)



        
        Raw_Stock_Transfer_Master_object_qty_both_with_shade_inward = RawStockTrasferRecords.objects.filter(item_shade_transfer = shade_name)
        
    for godown_qty in opening_godown_qty:

        opening_godown_data.append({
            'date': godown_qty.created_date,
            'particular': 'Opening Balance',
            'voucher_type': '',
            'vch_no': '',
            'inward_quantity': f"{godown_qty.opening_quantity}",
            'inward_value': godown_qty.opening_rate * godown_qty.opening_quantity,
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity': 0,
            'closing_value': 0,
            'rate':godown_qty.opening_rate
        })

    for purchase_voucher_item_qty in purchase_voucher_godown_qty:

        report_data.append({
            'date': purchase_voucher_item_qty.created_date,
            'particular': purchase_voucher_item_qty.party_name.name,
            'voucher_type': purchase_voucher_item_qty.ledger_type,
            'vch_no': purchase_voucher_item_qty.purchase_number,
            'inward_quantity': f"{purchase_voucher_item_qty.godown_qty_total}",
            'inward_value': purchase_voucher_item_qty.item_rate * purchase_voucher_item_qty.godown_qty_total,
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity':0,
            'closing_value': 0,
            'rate': purchase_voucher_item_qty.item_rate,
            'embedded_url' : f'{root_url}purchasevoucherupdate/{purchase_voucher_item_qty.id}'
            })
        

    for fabric_cutting_items in purchase_order_cutting_room_qty:
        outward_value = round(fabric_cutting_items.total_comsumption * fabric_cutting_items.rate , 2)
        report_data.append({
            'date': fabric_cutting_items.creation_date,
            'particular': fabric_cutting_items.particular,
            'voucher_type': fabric_cutting_items.voucher_type,
            'vch_no': fabric_cutting_items.voucher_number,
            'inward_quantity': '',
            'inward_value': '',
            'outward_quantity': f"{fabric_cutting_items.total_comsumption}",
            'outward_value': outward_value,
            'closing_quantity': 0,
            'closing_value': 0,
            'rate': fabric_cutting_items.rate,
            'embedded_url' : f'{root_url}purchaseordercuttingupdate/{fabric_cutting_items.p_o_id}/{fabric_cutting_items.product_ref_no}/{fabric_cutting_items.cutting_pk}'
             })
    

    for fabric_cutting_cancelled_items in purchase_order_cutting_room_qty_cancelled:
        inward_value = round(fabric_cutting_cancelled_items.total_comsumption * fabric_cutting_items.rate, 2)
        report_data.append({
            'date': fabric_cutting_cancelled_items.creation_date,
            'particular': fabric_cutting_cancelled_items.particular,
            'voucher_type': fabric_cutting_cancelled_items.voucher_type,
            'vch_no': fabric_cutting_cancelled_items.voucher_number,
            'inward_quantity': f"{fabric_cutting_cancelled_items.total_comsumption}",
            'inward_value': inward_value,
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity': 0,
            'closing_value': 0,
            'rate': fabric_cutting_cancelled_items.rate,
            'embedded_url':f'{root_url}purchaseordercuttingupdate/{fabric_cutting_items.p_o_id}/{fabric_cutting_items.product_ref_no}/{fabric_cutting_items.cutting_pk}'
            })
        
    for record in labour_workout_report:
        item_instance = item_color_shade.objects.get(items__item_name=record.material_name,item_shade_name=record.material_color_shade)
        
        if item_instance.items.Fabric_nonfabric == 'Non Fabric':
            outward_value = round(record.total_comsumption * record.rate , 2)
            
            report_data.append({
                'date': record.created_date,
                'particular': record.labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.ledger_party_name.name,
                'voucher_type': f'Labour workout/{record.labour_workout_child_instance.labour_name.name}',
                'vch_no': record.labour_workout_child_instance.labour_workout_master_instance.purchase_order_cutting_master.purchase_order_id.purchase_order_number,
                'inward_quantity': '',
                'inward_value': '',
                'outward_quantity': f"{record.total_comsumption}",
                'outward_value': outward_value,
                'closing_quantity': 0,
                'closing_value': 0,
                'rate': record.rate,
                'embedded_url' : f'{root_url}labourworkoutsingleview/{record.labour_workout_child_instance.id}'
                })
    
    for stock_qty in Raw_Stock_Transfer_Master_object_qty_outward:

        report_data.append({
            'date': stock_qty.created_date,
            'particular': 'Stock Transfer',
            'voucher_type': 'Stock Transfer',
            'vch_no': stock_qty.master_instance.voucher_no,
            'inward_quantity': "",
            'inward_value': "",
            'outward_quantity': stock_qty.item_quantity_transfer,
            'outward_value': round(stock_qty.item_shade_transfer.rate * stock_qty.item_quantity_transfer),
            'closing_quantity': 0,
            'closing_value': 0,
            'rate': stock_qty.item_shade_transfer.rate,
            'embedded_url' : f'{root_url}stocktransferrawupdate/{stock_qty.master_instance.voucher_no}'
        })

    for stock_qty in Raw_Stock_Transfer_Master_object_qty_inward:

        report_data.append({
            'date': stock_qty.created_date,
            'particular': 'Stock Transfer',
            'voucher_type': 'Stock Transfer',
            'vch_no': stock_qty.master_instance.voucher_no,
            'inward_quantity': stock_qty.item_quantity_transfer,
            'inward_value': round(stock_qty.item_shade_transfer.rate * stock_qty.item_quantity_transfer),
            'outward_quantity': '',
            'outward_value': '',
            'closing_quantity': 0,
            'closing_value': 0,
            'rate':stock_qty.item_shade_transfer.rate,
            'embedded_url' : f'{root_url}stocktransferrawupdate/{stock_qty.master_instance.voucher_no}'
        })



    for stock_qty in Raw_Stock_Transfer_Master_object_qty_both_with_shade_inward:

            report_data.append({
                'date': stock_qty.created_date,
                'particular': 'Stock Transfer',
                'source_gowdown':stock_qty.master_instance.source_godown,
                'destination_godown':stock_qty.master_instance.destination_godown,
                'voucher_type': 'Stock Transfer',
                'vch_no': stock_qty.master_instance.voucher_no,
                'inward_quantity': stock_qty.item_quantity_transfer,
                'inward_value': round(stock_qty.item_quantity_transfer * stock_qty.item_shade_transfer.rate),
                'outward_quantity': stock_qty.item_quantity_transfer,
                'outward_value': round(stock_qty.item_quantity_transfer * stock_qty.item_shade_transfer.rate),
                'closing_quantity': 0,
                'closing_value': 0,
                'rate':stock_qty.item_shade_transfer.rate
            })


    report_data_sorted = sorted(report_data, key = itemgetter('date'), reverse=False)

    
    """
    The sorted() function is used to sort the list. It returns a new list that is sorted 
    according to the specified key.

    The key parameter specifies a function that is used to extract the sorting key from each
    dictionary. Here, itemgetter('date') is used to get the value associated with the date
    key in each dictionary.

    If you set reverse=False, it would sort the list in ascending order.

    """
    
    return render(request,'reports/godownstockrawmaterialreportsingle.html',{'godoown_name':godown_name,
                                                                             'shade_name':shade_name,'report_data':report_data_sorted,
                                                                             'page_name':'Raw Material Report','opening_godown_data':opening_godown_data})





def allrawmaterialstockreport(request):
    queryset = item_color_shade.objects.all().annotate(total_qty = Sum(
        'godown_shades__quantity')).order_by('items__item_name').prefetch_related(
            'godown_shades','godown_shades__godown_name').select_related('items')
    
    return render(request,'reports/allrawmaterialstockreport.html',{'queryset':queryset})







def allfinishedgoodsstockreport(request,action=None):

    if action == 'All_Record':
        product_queryset_subquery = labour_work_in_product_to_item.objects.filter(
            product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(
            total_labour_workin_qty_sum = Coalesce(Sum('return_pcs'), 0)).values('total_labour_workin_qty_sum')


        product_pending_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(
                total_labour_workin_pen_qty_sum = Coalesce(Sum('pending_for_approval'), 0)).values('total_labour_workin_pen_qty_sum')
        

        product_approve_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(total_labour_workin_aprv_qty_sum = Coalesce(Sum('approved_qty'), 0)).values('total_labour_workin_aprv_qty_sum')


        product_sales_queryset_subquery = sales_voucher_finish_Goods.objects.filter(product_name__PProduct_SKU = OuterRef('PProduct_SKU')).values('product_name__PProduct_SKU').annotate(total_sales_qty_sum = Sum('quantity')).values('total_sales_qty_sum')


        product_queryset = PProduct_Creation.objects.all().annotate(total_qty = Sum(
            'godown_colors__quantity'),total_labour_workin_qty = Subquery(
                product_queryset_subquery),total_labour_workin_pending_qty = Subquery(product_pending_queryset_subquery),total_labour_workin_approve_qty = Subquery(product_approve_queryset_subquery),total_sales_qty = Subquery(product_sales_queryset_subquery)).order_by('Product__Model_Name').select_related('Product','PProduct_color')

    else:
        product_queryset_subquery = labour_work_in_product_to_item.objects.filter(
        product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(
            total_labour_workin_qty_sum = Coalesce(Sum('return_pcs'), 0)).values('total_labour_workin_qty_sum')


        product_pending_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(
                total_labour_workin_pen_qty_sum = Coalesce(Sum('pending_for_approval'), 0)).values('total_labour_workin_pen_qty_sum')
        

        product_approve_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(total_labour_workin_aprv_qty_sum = Coalesce(Sum('approved_qty'), 0)).values('total_labour_workin_aprv_qty_sum')


        product_sales_queryset_subquery = sales_voucher_finish_Goods.objects.filter(product_name__PProduct_SKU = OuterRef('PProduct_SKU')).values('product_name__PProduct_SKU').annotate(total_sales_qty_sum = Sum('quantity')).values('total_sales_qty_sum')


        product_queryset = PProduct_Creation.objects.all().annotate(total_qty = Sum(
            'godown_colors__quantity'),total_labour_workin_qty = Subquery(
                product_queryset_subquery),total_labour_workin_pending_qty = Subquery(product_pending_queryset_subquery),total_labour_workin_approve_qty = Subquery(product_approve_queryset_subquery),total_sales_qty = Subquery(product_sales_queryset_subquery)).filter(Q(total_qty__gt=0) | Q(total_labour_workin_qty__gt=0) | Q(total_labour_workin_pending_qty__gt=0) | Q(total_labour_workin_approve_qty__gt=0) | Q(total_sales_qty__gt=0)).order_by('Product__Model_Name').select_related('Product','PProduct_color')
    
    return render(request,'reports/allfinishedgoodsstockreport.html',{'product_queryset':product_queryset})




def qc_approved_model_wise_report(request,ref_id):
    
    if ref_id:
        product_instance = get_object_or_404(Product,Product_Refrence_ID=ref_id)

        all_sku = [f'{sku.PProduct_SKU}-{sku.PProduct_color}' for sku in product_instance.productdetails.all().order_by('PProduct_SKU')]

        
        
        l_w_in_instances = labour_work_in_master.objects.filter(
            labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID=ref_id)
        
        all_l_w_in_dict = []

        approved_qty_all = []
        for record in l_w_in_instances:

            dict_to_append = {
                'GRN_No':record.voucher_number,
                'Date' : record.created_date,
                'Description' : f'Description /{record.labour_voucher_number.labour_name.name}',
                'challan_no' : record.labour_voucher_number.challan_no,
                'total_lwo' : record.labour_voucher_number.total_process_pcs,
                'desc':'LWI',
                'id' : record.id,
                'l_w_o_id' :record.labour_voucher_number.id,
            }

            lwi_p_2_i = dict(record.l_w_in_products.all().values_list('product_sku','return_pcs')) 

            lwi_p_2_i_approved = dict(record.l_w_in_products.all().values_list('product_sku','approved_qty'))


            skus = {}
            
            approved_qty = {}
            for sku in all_sku:   
                
                only_sku = int(sku.split('-')[0])
                
                if only_sku in lwi_p_2_i:
                    skus[only_sku] = lwi_p_2_i[only_sku]

                else:
                    
                    skus[only_sku] = 0

                value = approved_qty.get(only_sku, 0)
                approved_qty[only_sku] = value + lwi_p_2_i_approved.get(only_sku,0)

            dict_to_append['SKUS'] = skus

            all_l_w_in_dict.append(dict_to_append)

            approved_qty_all.append(approved_qty)
        

        consolidated_approval_dict = {}

        for dicts in approved_qty_all:
            
            for key,value in dicts.items():
                if key in consolidated_approval_dict:
                    consolidated_approval_dict[key] = consolidated_approval_dict[key] + value
                else:
                     consolidated_approval_dict[key] = value

        sorted_data = sorted(all_l_w_in_dict, key = itemgetter('Date'), reverse=False)

    return render(request, 'reports/qcapprovedmodelwisereport.html',{'product_instance':product_instance,
                                                                     'sorted_data':sorted_data,'all_sku':all_sku,
                                                                     'consolidated_approval_dict':consolidated_approval_dict,'ref_id':ref_id})






def raw_material_excel_download(request):

    wb = Workbook()

    default_sheet = wb['Sheet']
    wb.remove(default_sheet) 

    wb.create_sheet('raw_material_create')
    sheet1 = wb.worksheets[0]
    headers =  ['Raw Material Name', 'Material Code','Color','Packing','Unit Name','Units','Panha', 
                'Fabric or Non Fabric','Fabric Finishes','Fabric Group','GST','HSN Code','Status']
    
    sheet1.append(headers)

    fileoutput = BytesIO()
    wb.save(fileoutput)
        
    
    response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = 'raw_material_create'
    response['Content-Disposition'] = f'attachment; filename = "{file_name}.xlsx"'

    return response




def raw_material_excel_upload(request):
    
    if request.method == "POST":
        try:
            excel_file = request.FILES.get('excel_file')
        
            if excel_file:
                file_name = excel_file.name.split('.')[0].split(' ')[0]

            
                if file_name == 'raw_material_create' and excel_file.name.endswith('.xlsx'):
                    
                    df = pd.read_excel(excel_file)
                    
                    
                    required_columns = {
                        'Raw Material Name':'item_name',
                        'Material Code':'Material_code',
                        'Color':'Item_Color',
                        'Packing':'Item_Packing',
                        'Unit Name':'unit_name_item',
                        'Panha':'Panha', 
                        'Units': 'Units',
                        'Fabric or Non Fabric':'Fabric_nonfabric',
                        'Fabric Finishes':'Item_Fabric_Finishes',
                        'Fabric Group':'Fabric_Group',
                        'GST':'Item_Creation_GST',
                        'HSN Code':'HSN_Code',
                        'Status':'status'
                    }
                    
                    
                    for col in required_columns.keys():
                        if col not in df.columns:
                            return HttpResponse(f"Missing required column: {col}")
                        

                    rows_with_error = []
                    for index, row in df.iterrows():
                        with transaction.atomic():
                            try:
                                
                                color = Color.objects.get(color_name=row['Color'])
                                packaging_m = packaging.objects.get(packing_material=row['Packing'])
                                unit_name = Unit_Name_Create.objects.get(unit_name=row['Unit Name'])
                                fabric_finish = FabricFinishes.objects.get(fabric_finish=row['Fabric Finishes'])
                                fabric_group = Fabric_Group_Model.objects.get(fab_grp_name=row['Fabric Group'])
                                gst_instance = gst.objects.get(gst_percentage=row['GST'])
                                
                                material_name = Item_Creation.objects.filter(item_name=row['Raw Material Name']).first()
                                material_code = Item_Creation.objects.filter(Material_code=row['Material Code']).first()

                                if not material_name and not material_code:
                                    
                                    Item_Creation.objects.create(item_name=row['Raw Material Name'],
                                                                            Material_code= row['Material Code'],
                                                                            Item_Color = color,
                                                                            Item_Packing = packaging_m,
                                                                            unit_name_item = unit_name,
                                                                            Units =  row['Units'],
                                                                            Panha =  row['Panha'],
                                                                            Fabric_nonfabric = row['Fabric or Non Fabric'],
                                                                            Item_Fabric_Finishes = fabric_finish,
                                                                            Fabric_Group = fabric_group,
                                                                            Item_Creation_GST = gst_instance,
                                                                            HSN_Code = row['HSN Code'],
                                                                            status = row['Status']
                                                                            )
                                else:
                                    logger.error(f'Duplicate  material_name{material_name} or material_code {material_code}')
                                    raise ValidationError(f'Duplicate  material_name{material_name} or material_code {material_code}')


                            except ValidationError as ve:
                                logger.error(f'validation error {ve}')
                                rows_with_error.append(row)

                            except Exception as e:
                                rows_with_error.append(row)
                                messages.error(request, f"Error processing row {index + 1}: {str(e)}")
                                logger.error(f"Error processing row {index + 1}: {str(e)}")
                                

                    if rows_with_error:

                        
                        list_of_lists = [df.values.tolist() for df in rows_with_error]


                        wb = Workbook()
                        default_sheet = wb['Sheet']
                        wb.remove(default_sheet) 

                        wb.create_sheet('raw_material_create_with_errors')

                        sheet1 = wb.worksheets[0]
                        headers =  ['Raw Material Name', 'Material Code','Color', 'Packing','Unit Name','Units','Panha', 
                                'Fabric or Non Fabric','Fabric Finishes','Fabric Group','GST','HSN Code','Status']
    
                        sheet1.append(headers)

                        for row in list_of_lists:
                            sheet1.append(row)


                        fileoutput = BytesIO()
                        wb.save(fileoutput)
        
                        
                        response = HttpResponse(fileoutput.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        file_name = 'raw_material_create_with_errors'
                        response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
                        logging.error(f"Item saved successfully with errors no of errors {len(rows_with_error)}")
                        messages.error(request, f"Item saved successfully with errors no of errors {len(rows_with_error)}")
                        return response
                            
                    messages.success(request, "Item saved successfully")
                    return redirect('item-list')
            
                else:
                    messages.error(request, 'Invalid file format. Please upload a valid Excel file.')
                    logger.error('Invalid file format-item excel. Please upload a valid Excel file.')
                    raise ValidationError('Invalid file format-item excel. Please upload a valid Excel file.')
        
        except Exception as e:
            messages.error(request, f'Invalid file format. Please upload a valid Excel file.{e}')
            logger.error(f'Invalid file format-item excel. Please upload a valid Excel file.{e}')
            return redirect('item-list')
    else:
        messages.error(request, "Invalid request method")
        



def finished_goods_model_wise_report(request,ref_id):
    
    if ref_id:
        labour_work_in_instances = labour_work_in_master.objects.filter(
        labour_voucher_number__labour_workout_master_instance__purchase_order_cutting_master__purchase_order_id__product_reference_number__Product_Refrence_ID=ref_id).annotate(
        total_qty = Sum('l_w_in_products__return_pcs'))
        
        data_list = []

        product_instance = get_object_or_404(Product,Product_Refrence_ID=ref_id)  

        for instance in labour_work_in_instances:
            dict_to_append = {
                'voucher_number' : instance.voucher_number,
                'date' : instance.created_date.replace(tzinfo=None).strftime("%d %B %Y"),
                'name': instance.labour_voucher_number.labour_name.name,
                'description' : f'Labour Work In - {instance.labour_voucher_number.labour_name.name}',
                'L_W_I' : instance.total_qty,
                'challan_No' : instance.labour_voucher_number.id,
                'LWI_ID' : instance.id,
                'LWO_Total' : instance.labour_voucher_number.total_process_pcs,
                'Repair_In' : '0',
                'sale' : '0',
                'Repair_Out' : '0',
                'net_closing_stock' :'0'}
            
            data_list.append(dict_to_append)

        initial_sorted_data = sorted(data_list, key = itemgetter('date'), reverse=False)

    return render(request, 'reports/finishedgoodsmodelwisereport.html',{'data_list':initial_sorted_data , 'product_instance':product_instance})



def lwo_and_lwi_report_vendor_wise(request):

    vendor_name = request.POST.get("vendor_name")

    if vendor_name:

        Approval_Pending_Qty = labour_workout_childs.objects.filter(id = OuterRef('id')).annotate(total_pending = Sum(
            'labour_work_in_master__l_w_in_products__pending_for_approval')).values('total_pending')
        
        Approve_Qty = labour_workout_childs.objects.filter(id = OuterRef('id')).annotate(total_approve = Sum(
            'labour_work_in_master__l_w_in_products__approved_qty')).values('total_approve')
        
        Labour_Work_In = labour_workout_childs.objects.filter(id=OuterRef('id')).annotate(total_workin = Sum(
            'labour_work_in_master__total_return_pcs')).values('total_workin')


        queryset = labour_workout_childs.objects.filter(labour_name__name__icontains=vendor_name).annotate(
                    Total_Balance_to_Vendor = Sum('labour_workout_child_items__labour_w_in_pending'),
                    Total_Pending_Qty=Subquery(Approval_Pending_Qty),
                    Total_Approve_Qty=Subquery(Approve_Qty),
                    Total_Labour_Work_In = Subquery(Labour_Work_In))

    else:
        Approval_Pending_Qty = labour_workout_childs.objects.filter(id = OuterRef('id')).annotate(total_pending = Sum('labour_work_in_master__l_w_in_products__pending_for_approval')).values('total_pending')
        Approve_Qty = labour_workout_childs.objects.filter(id = OuterRef('id')).annotate(total_approve = Sum('labour_work_in_master__l_w_in_products__approved_qty')).values('total_approve')
        Labour_Work_In = labour_workout_childs.objects.filter(id=OuterRef('id')).annotate(total_workin = Sum('labour_work_in_master__total_return_pcs')).values('total_workin')


        queryset = labour_workout_childs.objects.all().annotate(
                    Total_Balance_to_Vendor = Sum('labour_workout_child_items__labour_w_in_pending'),
                    Total_Pending_Qty=Subquery(Approval_Pending_Qty),
                    Total_Approve_Qty=Subquery(Approve_Qty),
                    Total_Labour_Work_In = Subquery(Labour_Work_In))
        
        

    return render(request,'reports/lwo_and_lwi_report_vendor_wise.html',{'queryset':queryset,'vendor_name':vendor_name})




def finished_goods_sorting_list(request):
    
    finished_goods_purchase_voucher_instances = product_purchase_voucher_master.objects.all().annotate(
        total_recieved_qty= Sum('product_purchase_voucher_items__quantity_total'), 
        total_qc_qty=Sum('product_purchase_voucher_items__qc_recieved_qty'),
        total_diff_qty = Sum('product_purchase_voucher_items__diffrence_qty'),
        total_boxex_qty = Sum('product_purchase_voucher_items__product_name__Product__Product_QtyPerBox')
        ).select_related('finished_godowns', 'party_name').order_by('created_date')

    finished_goods_transfer_m_instances = Finished_goods_Stock_TransferMaster.objects.filter(transnfer_cancelled=False).annotate(
    total_recieved_qty=Sum('finished_goods_transfer_records__product_quantity_transfer'), 
    total_qc_qty=Sum('finished_goods_transfer_records__qc_recieved_qty'),
    total_diff_qty=Sum('finished_goods_transfer_records__diffrence_qty'),
    total_boxex_qty=Sum('finished_goods_transfer_records__product__Product__Product_QtyPerBox')
    ).select_related('source_warehouse', 'destination_warehouse').order_by('created_date')
   
    

    finished_goods_all_records = list(finished_goods_purchase_voucher_instances) + list(finished_goods_transfer_m_instances)

    sorted_data = sorted(finished_goods_all_records, key=attrgetter('created_date'), reverse = False)

    return render(request,'finished_product/finishedgoodssortinglist.html',{'sorted_data':sorted_data})






def warehouse_navigator(request):
    warehouses = Finished_goods_warehouse.objects.prefetch_related('warehouses__zones__racks').all()
    return render(request,'finished_product/warehouse_navigator.html',{'warehouses':warehouses})













def picklist_product_ajax(request):
    try:
        product_name_typed = request.GET.get('productnamevalue')

        if not product_name_typed:
            return JsonResponse({'error': 'Please enter a search term.'}, status=400)

        logger.info(f"Search initiated by {request.user}: {product_name_typed}")

        products = Product_bin_quantity_through_table.objects.filter(
            Q(product__PProduct_SKU__icontains=product_name_typed) |
            Q(product__PProduct_color__color_name__icontains=product_name_typed) |
            Q(product__Product__Model_Name__icontains=product_name_typed),
            product_quantity__gt=0
        ).order_by('created_date').values(
            'product__PProduct_SKU',
            'product__PProduct_color__color_name',
            'product__Product__Model_Name',
            'product_quantity',
            'product__Product__Product_Refrence_ID',
            'product__PProduct_image',
            'bin__id',
            'bin__bin_name'
        )

        final_data = {}

        # Retrieve session-stored bin data
        temp_bins = request.session.get('temp_bins', {})

        for item in products:
            product_sku = item['product__PProduct_SKU']
            bin_id = item['bin__id']
            bin_name = item['bin__bin_name']
            
            # Fetch reserved quantity
            reserved_qty = Picklist_products_list.objects.filter(
                product__PProduct_SKU=product_sku,
                picklist_master_instance__status="Pending"
            ).aggregate(total_reserved=Sum('product_quantity'))['total_reserved'] or 0

            # Adjust available quantity
            # available_qty = max(item['product_quantity'] - reserved_qty, 0)
            available_qty = max(item['product_quantity'], 0)

            # Apply session-stored bin adjustments
            key = f"{product_sku}_{bin_id}"
            if key in temp_bins:
                available_qty = int(temp_bins[key]['balance_qty'])

            if available_qty > 0:
                bin_data = {bin_id: [bin_name, available_qty]}

                if product_sku in final_data:
                    final_data[product_sku][2] += available_qty
                    final_data[product_sku][3].append(bin_data)
                else:
                    final_data[product_sku] = [
                        item['product__Product__Model_Name'],
                        item['product__PProduct_color__color_name'],
                        available_qty,
                        [bin_data],
                        reserved_qty,
                        item['product__Product__Product_Refrence_ID'],
                        item['product__PProduct_image']
                    ]

        logger.info(f"Final Product Data: {final_data}")
        print('final_data = ', final_data)
        return JsonResponse({'products': final_data}, status=200)

    except Exception as e:
        logger.error(f"Error in picklist_product_ajax: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while processing your request.'}, status=500)




def bin_quantity_ajax(request):
    logger.info('Temp bin quantity function called using sessions')

    try:
        # Fetch and validate request parameters
        sku = request.GET.get('sku')
        binName = request.GET.get('binName')
        binQty = request.GET.get('binQty')
        productQty = request.GET.get('productQty')

        if not sku or not binName:
            return JsonResponse({"status": "error", "message": "Missing required parameters"}, status=400)

        try:
            binQty = int(binQty) if binQty and binQty.isdigit() else 0
            productQty = int(productQty) if productQty and productQty.isdigit() else 0
        except ValueError:
            return JsonResponse({"status": "error", "message": "Invalid quantity values"}, status=400)

        # Initialize session storage if not present
        if 'temp_bins' not in request.session:
            request.session['temp_bins'] = {}

        temp_bins = request.session['temp_bins']

        # Create a unique key for SKU and binName
        key = f"{sku}_{binName}"

        if key in temp_bins:
            previous_qty = temp_bins[key]['product_qty']
            temp_bins[key]['product_qty'] = productQty
            temp_bins[key]['bin_qty'] = binQty

            # Corrected balance calculation
            temp_bins[key]['balance_qty'] = max(0, binQty - productQty)

            logger.info(f"Updated existing record in session: SKU={sku}, Bin={binName}, New Product Qty={productQty}, New Balance Qty={temp_bins[key]['balance_qty']}")
        else:
            temp_bins[key] = {
                "sku": sku,
                "bin_name": binName,
                "bin_qty": binQty,
                "product_qty": productQty,
                "balance_qty": max(0, binQty - productQty),
            }
            logger.info(f"Created new record in session: SKU={sku}, Bin={binName}, Bin Qty={binQty}, Product Qty={productQty}, Balance Qty={temp_bins[key]['balance_qty']}")

        # Save the updated session
        request.session.modified = True

        return JsonResponse({
            "status": "success",
            "bin_qty": temp_bins[key]['bin_qty'],
            "product_qty": temp_bins[key]['product_qty'],
            "balance_qty": temp_bins[key]['balance_qty']
        }, status=200)

    except Exception as e:
        logger.error(f"Error in bin_quantity_ajax: {str(e)}", exc_info=True)
        return JsonResponse({"status": "error", "message": "An error occurred while processing the request."}, status=500)



def create_update_picklist(request, p_id=None):
    
    if 'temp_bins' in request.session:
        del request.session['temp_bins']
        request.session.modified = True
    
    if p_id:
        voucher_instance = get_object_or_404(Picklist_voucher_master, id=p_id)
        master_form = Picklistvouchermasterform(request.POST or None, instance=voucher_instance)
        formset = picklistcreateformsetupdate(request.POST or None, instance=voucher_instance)
        types = None
    else:
        voucher_instance = None
        master_form = Picklistvouchermasterform()
        formset = picklistcreateformset()
        types = ledgerTypes.objects.all()

    if request.method == "POST":
        logger.info(f"Received POST data: {request.POST}")
        
        master_form = Picklistvouchermasterform(request.POST, instance=voucher_instance)
        formset = picklistcreateformsetupdate(request.POST, instance=voucher_instance) if p_id else picklistcreateformset(request.POST)

        if not master_form.is_valid():
            logger.error(f"Master Form Errors: {master_form.errors}")

        for form in formset:
            if not form.is_valid():
                logger.error(f"Form Errors: {form.errors}")

        if master_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    master_form_instance = master_form.save(commit=False)
                    master_form_instance.c_user = request.user
                    master_form_instance.save()

                    
                    # Handle deleted forms (Restoring bin quantity)
                    for form in formset.deleted_forms:
                        if form.instance.pk:
                            sku = form.instance.product
                            binName = form.instance.bin_number
                            productQty = form.instance.product_quantity

                            bin_object, _ = Product_bin_quantity_through_table.objects.get_or_create(product=sku, bin=binName)
                            bin_object.product_quantity += productQty  # Restore the quantity
                            bin_object.save()

                            form.instance.delete()

                    # Handle new or updated forms (Deducting bin quantity)
                    formset.forms = [form for form in formset.forms if form.has_changed()]
                    for form in formset:
                        if not form.cleaned_data.get('DELETE'):
                            form_instance = form.save(commit=False)
                            form_instance.picklist_master_instance = master_form_instance

                            sku = form_instance.product.PProduct_SKU
                            bin_id = form_instance.bin_number
                            qty = form_instance.product_quantity

                            bin_qty_object, _ = Product_bin_quantity_through_table.objects.get_or_create(product=sku, bin=bin_id)
                            if bin_qty_object.product_quantity >= qty:
                                bin_qty_object.product_quantity -= qty
                                bin_qty_object.save()
                                form_instance.save()
                            else:
                                logger.warning(f"Insufficient stock for SKU={sku}, Bin={bin_id}. Required={qty}, Available={bin_qty_object.product_quantity}")
                                return JsonResponse({"status": "error", "message": "Insufficient stock"}, status=400)

                    total_qty = master_form_instance.picklist_products_list.aggregate(total=Sum('product_quantity'))['total'] or 0
                    master_form_instance.total_qty = total_qty
                    master_form_instance.save()

                    return redirect('all-picklists-list')

            except Exception as e:
                logger.error(f"Error while processing picklist: {str(e)}", exc_info=True)
                return JsonResponse({"status": "error", "message": "An error occurred while processing the request."}, status=500)

    return render(request, 'finished_product/createupdatepicklist.html', {'master_form': master_form, 'formset': formset,'types':types})




def delete_form_quantity_revert(request):
    """
    Handles reverting bin quantity when a product is deleted from the picklist.
    Updates session-based bin data to ensure real-time stock updates.
    """
    logger.info('delete_form_quantity_revert function called')

    try:
        sku = request.GET.get('skus')
        bin_id = request.GET.get('binName')
        product_qty = request.GET.get('productQty')

        if not sku or not bin_id:
            return JsonResponse({"error": "Missing required parameters"}, status=400)

        try:
            product_qty = int(product_qty) if product_qty and product_qty.isdigit() else 0
        except ValueError:
            return JsonResponse({"error": "Invalid quantity value"}, status=400)

        # Retrieve or initialize session-stored bin data
        temp_bins = request.session.get('temp_bins', {})

        key = f"{sku}_{bin_id}"

        if key in temp_bins:
            temp_bins[key]['balance_qty'] += product_qty
        else:
            temp_bins[key] = {'balance_qty': product_qty}

        # Save updated bin data back to the session
        request.session['temp_bins'] = temp_bins
        request.session.modified = True

        logger.info(f"Updated session bin data: SKU={sku}, Bin={bin_id}, Bin Qty={temp_bins[key]['balance_qty']}")

        return JsonResponse({"success": "Bin quantity updated successfully"}, status=200)

    except Exception as e:
        logger.error(f"Error in delete_form_quantity_revert: {str(e)}", exc_info=True)
        return JsonResponse({"error": "An error occurred while processing the request."}, status=500)





def deletepicklist(request,pl_id):
    picklist = Picklist_voucher_master.objects.get(pk=pl_id)

    if picklist.status != "Pending":
        messages.error(request, "Only Pending picklists can be deleted.")
        return redirect("all-picklists-list")
    
    picklist_product_instance = Picklist_products_list.objects.filter(picklist_master_instance = pl_id)
    for instance in picklist_product_instance:
        bin_qty_object = Product_bin_quantity_through_table.objects.get(bin=instance.bin_number, product=instance.product)
        bin_qty_object.product_quantity = instance.product_quantity + bin_qty_object.product_quantity
        bin_qty_object.save()
    picklist.delete()
    return redirect('all-picklists-list')




def all_picklists_list(request):
    all_picklists = Picklist_voucher_master.objects.all().annotate(scanned_qty=Sum("picklist_process_in_outward__balance_qty")).order_by('created_date')
    return render(request,'finished_product/allpicklists.html',{'all_picklists':all_picklists})





def picklist_view(request,pl_id):
    picklist_number=get_object_or_404(Picklist_voucher_master,pk=pl_id)
    picklist_data = Picklist_products_list.objects.filter(picklist_master_instance=pl_id)
    return render(request,'finished_product/picklist_view.html',{'picklist_data':picklist_data,'picklist_number':picklist_number})





def download_picklist_pdf(request,pl_id):
 # Get the Picklist_voucher_master instance
    picklist = get_object_or_404(Picklist_voucher_master, id=pl_id)

    # Get all related Picklist_products_list instances
    products_list = picklist.picklist_products_list.all()

    # Prepare context for the PDF, including absolute URLs for images
    context = {
        'picklist': picklist,
        'products_list': []
    }

    # Build the absolute URLs for the images and add them to the context
    for data in products_list:
        product_image_url = request.build_absolute_uri(data.product.PProduct_image.url)
        context['products_list'].append({
            'product': data.product,
            'product_image_url': product_image_url,
            'product_quantity': data.product_quantity,
            'bin_name': data.bin_number.bin_name,
            'rack_name': data.bin_number.rack_finished_name.rack_name,
            'zone_name': data.bin_number.rack_finished_name.zone_finished_name.zone_name
        })

    # Render the HTML template for the PDF
    html = render_to_string('finished_product/picklist_pdf_template.html', context)

    # Create a response object to serve the PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="picklist_{picklist.picklist_no}.pdf"'

    # Generate the PDF from the HTML
    pisa_status = pisa.CreatePDF(html, dest=response)

    # Check if the PDF was successfully created
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response



def download_delivery_challan_pdf(request,dc_id):

    challan = get_object_or_404(DeliveryChallanMaster, id=dc_id)

    products_list = challan.deliverychallanproducts_set.all()

    
    logo_file_path = finders.find('images/NAME.png')
    signature_file_path = finders.find('images/SIGN.png')

    logo_base64 = ''
    signature_base64 = ''

    if logo_file_path:
        with open(logo_file_path, 'rb') as image_file:
            logo_base64 = base64.b64encode(image_file.read()).decode('utf-8')

    if signature_file_path:
        with open(signature_file_path, 'rb') as image_file:
            signature_base64 = base64.b64encode(image_file.read()).decode('utf-8')

    context = {
        'challan': challan,
        'products_list': [],
        'logo_base64': logo_base64,
        'signature_base64': signature_base64,
    }

    for data in products_list:
        context['products_list'].append({
            'product': data.product_name.Product.Model_Name,
            'ref_no':data.product_name.Product.Product_Refrence_ID,
            'product_quantity': data.quantity,
        })

    # Render the HTML template for the PDF
    html = render_to_string('production/delivery_challan_pdf_template.html', context)

    # Create a response object to serve the PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="delivery_challan_no_{challan.delivery_challan_no}.pdf"'

    # Generate the PDF from the HTML
    pisa_status = pisa.CreatePDF(html, dest=response)

    # Check if the PDF was successfully created
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    
    return response




def download_picklist_excel(request,pl_id):
    # Get the Picklist_voucher_master instance
    picklist = get_object_or_404(Picklist_voucher_master, id=pl_id)

    # Get all related Picklist_products_list instances
    products_list = picklist.picklist_products_list.all()

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Picklist Products"

    # Add Picklist_voucher_master data
    ws.append(["Picklist No:", picklist.picklist_no])
    ws.append(["Created by:", picklist.c_user.username])
    ws.append(["Created Date:", picklist.created_date.date()])
    ws.append([])  # Empty line for separation

    # Set headers for products list
    headers = ["Product", "Image", "Zone", "Rack", "Bin", "Quantity"]
    ws.append(headers)

    # Style the headers (bold font)
    for cell in ws[6]:
        cell.font = Font(bold=True)

    # Fill in the product data
    for data in products_list:
        product_name = data.product.Product.Model_Name
        product_image_url = data.product.PProduct_image.url
        zone_name = data.bin_number.rack_finished_name.zone_finished_name.zone_name
        rack_name = data.bin_number.rack_finished_name.rack_name
        bin_name = data.bin_number.bin_name
        product_quantity = data.product_quantity

        # Add data to the row (excluding image, since we can't add it to Excel directly)
        row = [product_name, product_image_url, zone_name, rack_name, bin_name, product_quantity]
        ws.append(row)

    # Create a response object to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="picklist_{picklist.picklist_no}.xlsx"'

    # Save the workbook to the response object
    wb.save(response)

    return response



def outward_picklist_no_ajax(request):

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
    
        picklistNo = request.GET.get('picklistNo')

        try:

            picklist_qty = Picklist_products_list.objects.filter(picklist_master_instance__picklist_no=picklistNo).aggregate(total_qty=Sum('product_quantity'))

            picklist_data = Picklist_products_list.objects.filter(picklist_master_instance__picklist_no=picklistNo).select_related('product', 'picklist_master_instance')
            
            try:
                dict_to_send = {}


                for i in picklist_data:
                    if i.picklist_master_instance.status == 'Pending':
                        sku = i.product.PProduct_SKU
                        id = i.picklist_master_instance.id
                        
                        if picklistNo not in dict_to_send:
                            dict_to_send[picklistNo] = []

                        
                        check_sku = next((j for j in dict_to_send[picklistNo] if j['sku'] == sku), None)

                        if check_sku:
                            check_sku['qty'] += i.product_quantity  # Add quantity if SKU exists
                        else:
                            
                            dict_to_send[picklistNo].append({
                                'sku': sku,
                                'model_name': i.product.Product.Model_Name,
                                'color': i.product.PProduct_color.color_name,
                                'qty': i.product_quantity,
                                'id':i.id
                            })
                    else:
                        message.error(request,"can't use this picklist")
                    

                return JsonResponse({"status": "success", "picklistDict": dict_to_send, "picklistNo":picklistNo, "picklisQty": picklist_qty['total_qty'],'id':id}, status=200)

            except Exception as e:
                print(e)
                return JsonResponse({"status": "error", 'message':'Picklist close'}, status=400)
                
        except Exception as e:
            return JsonResponse({"error": f"An error occurred while processing the request: {str(e)}"}, status=500)



def outward_scan_serial_no_process(request):
    print("in scan def")
    try:
        serialNo = request.GET.get('serialNo')
        picklistNo = request.GET.get('pickNo')
        outwardNo = request.GET.get('outwardNo')

        print('serialNo = ',serialNo)

        if not serialNo:
            return JsonResponse({'error': 'Please enter a search term.'}, status=400)

        
        check_if_exist = outward_products.objects.filter(outward_no__outward_no = outwardNo, unique_serial_no=serialNo).exists()

        if check_if_exist:
            return JsonResponse({'error': 'Serial No already exists'}, status=400)
        

        filtered_product = finishedgoodsbinallocation.objects.get(unique_serial_no = serialNo, outward_done = False)

        list_to_send = []

        product_image_url = request.build_absolute_uri(filtered_product.product.PProduct_image.url)

        list_to_send.append(
        [filtered_product.product.Product.Product_Refrence_ID,
        filtered_product.product.PProduct_SKU,
        filtered_product.product.Product.Model_Name,
        filtered_product.product.PProduct_color.color_name,
        filtered_product.unique_serial_no,
        filtered_product.bin_number.bin_name,
        1,
        product_image_url,
        filtered_product.bin_number.id,
        picklistNo])

        print('list_to_send = ',list_to_send)
        return JsonResponse({'products': list_to_send,'message':f"{serialNo} added successfully"}, status=200)
        
    except Exception as e:
        return JsonResponse({'error': f"An error occurred: {str(e)}"}, status=500)


from decimal import Decimal
def decimal_to_float(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError("Type not serializable")


from urllib.parse import urlencode


def outward_scan_product_create(request,o_id=None):
    
    if 'products_data' in request.session:
        del request.session['products_data']
        request.session.modified = True

    if o_id:
        outward_instance = get_object_or_404(outward_product_master,pk=o_id)
        master_form = Outwardproductmasterform(request.POST or None,instance=outward_instance)
        formset = OutwardProductupdateFormSet(request.POST or None,instance=outward_instance)
        picklist_formset = PicklistProcessInOutwardFormsetupdate(request.POST or None,instance=outward_instance)

        picklist_data_queryset = Picklist_process_in_outward.objects.filter(outward_no=o_id).prefetch_related(
            'picklist__picklist_products_list__product'
        )

        dict_to_send = {}

        for data in picklist_data_queryset:
            picklist_id = data.picklist.picklist_no

            if picklist_id not in dict_to_send:
                dict_to_send[picklist_id] = []

            for product_list in data.picklist.picklist_products_list.all():
                product_sku = product_list.product.PProduct_SKU
                qty = product_list.product_quantity
                model_name = product_list.product.Product.Model_Name
                color = product_list.product.PProduct_color.color_name

                found = False
                for item in dict_to_send[picklist_id]:
                    if item['sku'] == product_sku:
                        item['qty'] += qty
                        found = True
                        break
                

                if not found:
                    dict_to_send[picklist_id].append({
                        'sku': product_sku,
                        'model_name': model_name,
                        'color': color,
                        'qty': qty
                    })

    else:
        outward_instance = None
        master_form = Outwardproductmasterform()
        picklist_formset = PicklistProcessInOutwardFormset()
        formset = OutwardProductcreateFormSet()
        dict_to_send = ''
    

    if request.method == 'POST':
        
        master_form = Outwardproductmasterform(request.POST or None,instance=outward_instance)
        formset = OutwardProductupdateFormSet(request.POST or None,instance=outward_instance)
        picklist_formset = PicklistProcessInOutwardFormsetupdate(request.POST or None, instance=outward_instance)


        if not master_form.is_valid():
            print("Form Errors:", master_form.errors)

        if not formset.is_valid():
            for form in formset:
                if not form.is_valid():
                    print("Form Errors:", form.errors)


        if master_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():

                    master_form_instance = master_form.save(commit=False)
                    master_form_instance.c_user = request.user
                    master_form_instance.save()

                    for form in formset.deleted_forms:
                        if form.instance.pk:
                            form.instance.delete()
                    
                    
                    formset.forms = [form for form in formset if form.has_changed()]

                    for form in formset:
                        if not form.cleaned_data.get('DELETE'):
                            form_instance = form.save(commit=False)

                            form_instance.outward_no = master_form_instance

                            serial_no = form_instance.unique_serial_no
                            
                            try:
                                inward_instance = finishedgoodsbinallocation.objects.get(unique_serial_no=serial_no)
                                
                                # Update outward_done safely
                                inward_instance.outward_done = True
                                inward_instance.save(update_fields=['outward_done'])  # Skips bin validation

                            except finishedgoodsbinallocation.DoesNotExist:
                                print(f"Error: No record found with unique_serial_no = {serial_no}")

                            except Exception as e:
                                print(f"Unexpected error: {e}")

                            form_instance.save()

                    if not picklist_formset.is_valid():
                        for form in picklist_formset:
                            if not form.is_valid():
                                print("Picklist Form Errors:", form.errors)


                    if picklist_formset.is_valid():

                        formset.forms = [form for form in formset if form.has_changed()]

                        for form in picklist_formset:
                            picklist_form_instance = form.save(commit=False)
                            picklist_form_instance.outward_no = master_form_instance
                            picklist_form_instance.save()


                sales_voucher = sales_voucher_master_outward_scan.objects.filter(outward_no=master_form_instance).first()

                if sales_voucher:
                    products = {}

                    post_data = request.POST.dict()

                    total_forms = int(post_data.get('outward_product-TOTAL_FORMS', 0))

                    products[master_form_instance.outward_no] = []

                    for i in range(total_forms):
                        # Get form data
                        product = post_data.get(f'outward_product-{i}-product')
                        quantity = int(post_data.get(f'outward_product-{i}-quantity', 0))
                        delete_flag = post_data.get(f'outward_product-{i}-DELETE', False)

                        if delete_flag:
                            print(f"Product {product} is marked for deletion, skipping.")
                            continue

                        print(f"Processing product: {product}, Quantity: {quantity}")

                        
                        found = False
                        for p in products[master_form_instance.outward_no]:
                            if p['product'] == product:
                                p['quantity'] += quantity
                                found = True
                                break

                        if not found:
                            product_info = PProduct_Creation.objects.get(PProduct_SKU=product)
                            products[master_form_instance.outward_no].append({
                                'product_ref': post_data.get(f'outward_product-{i}-product_RefNo'),
                                'product_name': post_data.get(f'outward_product-{i}-product_name_value'),
                                'product': product,
                                'color': post_data.get(f'product_color_{i}'),
                                'quantity': quantity,
                                'mrp': decimal_to_float(product_info.Product.Product_MRP),
                                'customer_price': decimal_to_float(product_info.Product.Product_SalePrice_CustomerPrice),
                                'gst': product_info.Product.Product_GST.gst_percentage
                            })

                    request.session['products_data'] = products

                    return redirect(f'/salesvouchercreateupdateforwarehouse/{sales_voucher.id}/EDIT/')

                
                else:
                    products = {}

                    post_data = request.POST.dict()
                    
                    total_forms = int(post_data.get('outward_product-TOTAL_FORMS', 0))

                    products[master_form_instance.outward_no] = []

                    for i in range(total_forms):
                        product = post_data.get(f'outward_product-{i}-product')
                        quantity = int(post_data.get(f'outward_product-{i}-quantity', 0))
                        delete_flag = post_data.get(f'outward_product-{i}-DELETE', False)  # Check if the form is marked for deletion

                        if delete_flag:  # If the product is marked for deletion, skip it
                            print(f"Product {product} is marked for deletion, skipping.")
                            continue  # Skip this product and don't count in quantity

                        print(f"Processing product: {product}, Quantity: {quantity}")
                        

                        found = False
                        for p in products[master_form_instance.outward_no]:
                            if p['product'] == product:
                                p['quantity'] += quantity
                                found = True
                                break
                            

                        if not found:
                            product_info = PProduct_Creation.objects.get(PProduct_SKU = product)
                            products[master_form_instance.outward_no].append({
                                'product_ref': post_data.get(f'outward_product-{i}-product_RefNo'),
                                'product_name': post_data.get(f'outward_product-{i}-product_name_value'),
                                'product': product,
                                'color': post_data.get(f'product_color_{i}'),
                                'quantity': quantity,
                                'mrp': decimal_to_float(product_info.Product.Product_MRP),
                                'customer_price': decimal_to_float(product_info.Product.Product_SalePrice_CustomerPrice),
                                'gst': product_info.Product.Product_GST.gst_percentage
                            })
                    
                    request.session['products_data'] = products

                    return redirect('/salesvouchercreateupdateforwarehouse/')
                
            except Exception as e:
                print(e)
    
    return render(request,'finished_product/outward_scan_product_create.html',{'master_form':master_form,'formset': formset, 'picklist_formset':picklist_formset ,'dict_to_send':dict_to_send})






def sales_voucher_create_update_for_warehouse(request, s_id=None, action=None):

    outward_number = None

    product_list = None

    party_name = Ledger.objects.filter(under_group__account_sub_group='Sundry Debtors')

    warehouse_names = Finished_goods_warehouse.objects.all()
    
    salesman_list = Salesman_info.objects.all()

    if s_id and action == 'EDIT':

        print("********* IN EDIT MODE FROM OUTWARD *******")

        page_name = 'Edit Sales Invoice'

        products = request.session.get('products_data', None)

        if products:
            for key, val in products.items():
                outward_number = key
                print('outward_number = ', outward_number)
                product_list = val
                break

            sales_product_info = sales_voucher_outward_scan.objects.filter(sales_voucher_master = s_id).values('product_name__Product__Product_Refrence_ID','product_name__Product__Model_Name','product_name__PProduct_SKU','product_name__PProduct_color__color_name','quantity','trade_disct','spl_disct')

            sales_product_info_list = list(sales_product_info)

            merged_product_list = []

            for product in product_list:
                sku = product['product']
                matched = False
                
                for sales_product in sales_product_info_list:
                    if sku == str(sales_product['product_name__PProduct_SKU']):
                        merged_product_list.append({
                            'product_ref': product['product_ref'],
                            'product_name': product['product_name'],
                            'product': product['product'],
                            'color': product['color'],
                            'quantity': product['quantity'],
                            'mrp': product['mrp'],
                            'customer_price': product['customer_price'],
                            'gst': product['gst'],
                            'trade_disct': sales_product['trade_disct'],
                            'spl_disct': sales_product['spl_disct'],
                        })
                        matched = True
                        break
                
                if not matched:
                    merged_product_list.append({
                        'product_ref': product['product_ref'],
                        'product_name': product['product_name'],
                        'product': product['product'],
                        'color': product['color'],
                        'quantity': product['quantity'],
                        'mrp': product['mrp'],
                        'customer_price': product['customer_price'],
                        'gst': product['gst'],
                        'trade_disct': '',
                        'spl_disct': '',
                    })

        voucher_instance = sales_voucher_master_outward_scan.objects.get(id=s_id)

        master_form = Salesvouchermasteroutwardscanform(request.POST or None, instance=voucher_instance)

        salesvoucherfromscanupdateformset = inlineformset_factory(sales_voucher_master_outward_scan, sales_voucher_outward_scan, form=SalesvoucheroutwardscanForm, extra=len(merged_product_list),  can_delete=True)

        formset = salesvoucherfromscanupdateformset(request.POST or None, initial=merged_product_list)


    elif s_id and action == 'VIEW':

        print("********* ONLY IN VIEW MODE *******")

        voucher_instance = sales_voucher_master_outward_scan.objects.get(id=s_id)

        master_form = Salesvouchermasteroutwardscanform(instance=voucher_instance)

        page_name = 'Sales Invoice'

        merged_product_list = []

        salesvoucherfromscanupdateformset = inlineformset_factory(sales_voucher_master_outward_scan, sales_voucher_outward_scan, form=SalesvoucheroutwardscanForm, extra=0,can_delete=True)

        formset = salesvoucherfromscanupdateformset(instance=voucher_instance)

    else:

        print("********* IN CREATE MODE *******")

        products = request.session.get('products_data', None)

        if products:
            for key, val in products.items():
                outward_number = key
                product_list = val
                break

        master_form = Salesvouchermasteroutwardscanform(initial={'outward_no':outward_number})

        salesvoucherfromscancreateformset = inlineformset_factory(sales_voucher_master_outward_scan, sales_voucher_outward_scan, form=SalesvoucheroutwardscanForm, extra=len(product_list), can_delete=True)

        formset = salesvoucherfromscancreateformset(initial=product_list)

        page_name = 'Create Sales Invoice'

    if request.method == "POST":

        if s_id and action == 'EDIT':

            print("IN EDIT SAVE")

            master_form = Salesvouchermasteroutwardscanform(request.POST or None, instance=voucher_instance)

            old_instances = sales_voucher_outward_scan.objects.filter(sales_voucher_master=voucher_instance)

            for old_instance in old_instances:
            
                warehouse_instance = Product_warehouse_quantity_through_table.objects.get(warehouse=old_instance.sales_voucher_master.selected_warehouse,product=old_instance.product_name)

                warehouse_instance.quantity += old_instance.quantity 
                warehouse_instance.save()

            old_instances.delete()

            formset = salesvoucherfromscanupdateformset(request.POST or None)

            if not master_form.is_valid():
                print("Form Errors:", master_form.errors)

            master_form_instance = master_form.save(commit=False)           
            master_form_instance.save()

            if not formset.is_valid():
                for form in formset:
                    if not form.is_valid():
                        print("Form Errors:", form.errors)

            for form in formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()

            for form in formset:
                if not form.cleaned_data.get('DELETE'):
                    form_instance = form.save(commit=False)
                    voucher_instance = sales_voucher_master_outward_scan.objects.get(id=s_id)
                    form_instance.sales_voucher_master = voucher_instance
                    form_instance.save()

            if 'products_data' in request.session:
                del request.session['products_data']
                request.session.modified = True

            return redirect('sales-voucher-list-warehouse')
        
        else:

            print("IN CREATE SAVE")

            master_form = Salesvouchermasteroutwardscanform(request.POST or None)
            formset = salesvoucherfromscancreateformset(request.POST or None)

            if not master_form.is_valid():
                print("Form Errors:", master_form.errors)

            if not formset.is_valid():
                for form in formset:
                    if not form.is_valid():
                        print("Form Errors:", form.errors)

            if master_form.is_valid() and formset.is_valid():
                try:
                    with transaction.atomic():

                        master_form_instance = master_form.save(commit=False)
                        o_id = outward_product_master.objects.get(outward_no=outward_number)
                        master_form_instance.outward_no = o_id
                        master_form_instance.save()

                        for form in formset.deleted_forms:
                            if form.instance.pk:
                                form.instance.delete()

                        for form in formset:
                            if not form.cleaned_data.get('DELETE'):
                                form_instance = form.save(commit=False)
                                form_instance.sales_voucher_master = master_form_instance
                                form_instance.save()

                        if 'products_data' in request.session:
                            del request.session['products_data']
                            request.session.modified = True

                        return redirect('sales-voucher-list-warehouse')
                    
                except Exception as e:
                    print(e)
    return render(request, 'accounts/salesvouchercreateupdateforwarehouse.html', {
            'master_form': master_form,
            'formset': formset,
            'page_name': page_name,
            'party_name': party_name,
            'warehouse_names': warehouse_names,
            'salesman_list':salesman_list
        })





def outward_scan_product_list(request):

    sale_qty_subqeury = sales_voucher_master_outward_scan.objects.filter(outward_no=OuterRef('pk')).values('outward_no').annotate(total_sale_qty=Sum('sales_voucher_outward_scan__quantity')).values('total_sale_qty')
    
    picklist_qty_subqeury = Picklist_process_in_outward.objects.filter(outward_no=OuterRef('pk')).values('outward_no').annotate(total_qty = Sum('picklist__picklist_products_list__product_quantity')).values('total_qty')

    outward_list = outward_product_master.objects.all().annotate(
        total_qty=Sum('outward_product__quantity'),
        picklist_total_qty=Subquery(picklist_qty_subqeury),
        remaining_qty=F('picklist_total_qty') - F('total_qty'),
        sale_total_qty=Subquery(sale_qty_subqeury)).order_by('created_date')
    

    return render(request,'finished_product/outward_scan_product_list.html',{'outward_list':outward_list})




def salesvoucherlistwarehouse(request):

    if 'products_data' in request.session:
        del request.session['products_data']

    sales_list = sales_voucher_master_outward_scan.objects.all().select_related(
        'outward_no',
        'party_name',
        'selected_warehouse',
        'salesman'
        ).prefetch_related('sales_voucher_outward_scan').values(
            'id',
            'created_date',
            'outward_no__id',
            'outward_no__outward_no',
            'salesman__salesman_name',
            'salesman__id',
            'ledger_type',
            'sale_no',
            'party_name__name',
            'party_name__id',
            'grand_total'
            ).annotate(total_qty = Sum('sales_voucher_outward_scan__quantity'))


    sales_return = sales_return_voucher_master.objects.all().select_related(
        'sales_voucher_master',
        'sales_return_inward_instance',
        'party_name',
        'selected_warehouse',
        'salesman'
        ).prefetch_related('sales_return_voucher').values(
            'id',
            'sales_voucher_master__id',
            'sales_return_inward_instance__id',
            'created_date',
            'salesman__salesman_name',
            'salesman__id',
            'ledger_type',
            'sales_return_inward_instance__sales_return_no',
            'grand_total',
            'party_name__name','party_name__id',
            ).annotate(total_qty = Sum('sales_return_voucher__quantity'))
    
    final_list = sorted(
        chain(sales_list, sales_return), key=lambda x: x['created_date']
    )

    print('final_list = ',final_list)

    return render(request,'accounts/sales_list_warehouse.html',{'final_list':final_list})



def sales_voucher_view_sort_with_salesman(request,id):

    sales_vouchers_queryset = sales_voucher_master_outward_scan.objects.filter(salesman__id = id).select_related(
        'outward_no',
        'party_name',
        'selected_warehouse',
        'salesman'
        ).values(
            'created_date',
            'sale_no',
            'ledger_type',
            'party_name__name',
            'grand_total'
            ).annotate(total_qty = Sum('sales_voucher_outward_scan__quantity'))

    sales_return_vouchers_queryset = sales_return_voucher_master.objects.filter(salesman__id = id).select_related(
        'sales_voucher_master',
        'sales_return_inward_instance',
        'party_name',
        'selected_warehouse',
        'salesman'
        ).values(
            'created_date',
            'sales_return_inward_instance__sales_return_no',
            'ledger_type',
            'party_name__name',
            'grand_total'
            ).annotate(total_qty = Sum('sales_return_voucher__quantity'))
    

    final_list = sorted(chain(sales_vouchers_queryset, sales_return_vouchers_queryset), key=lambda x: x['created_date'])
        
    return render(request,'accounts/sales_voucher_view_sort_with_salesman.html',{'final_list':final_list,})



def sales_voucher_view_sort_with_partyname(request,id):

    sales_vouchers_queryset = sales_voucher_master_outward_scan.objects.filter(party_name__id = id).select_related(
        'outward_no',
        'party_name',
        'selected_warehouse',
        'salesman'
        ).values(
            'created_date',
            'sale_no',
            'ledger_type',
            'party_name__name',
            'grand_total'
            ).annotate(total_qty = Sum('sales_voucher_outward_scan__quantity'))

    sales_return_vouchers_queryset = sales_return_voucher_master.objects.filter(party_name__id = id).select_related(
        'sales_voucher_master',
        'sales_return_inward_instance',
        'party_name',
        'selected_warehouse',
        'salesman'
        ).values(
            'created_date',
            'sales_return_inward_instance__sales_return_no',
            'ledger_type',
            'party_name__name',
            'grand_total'
            ).annotate(total_qty = Sum('sales_return_voucher__quantity'))

    final_list = sorted(chain(sales_vouchers_queryset, sales_return_vouchers_queryset), key=lambda x: x['created_date'])

    return render(request,'accounts/sales_voucher_view_sort_with_partyname.html',{'final_list':final_list,}) 




def ref_no_search_ajax(request):
    try:
        ref_no_typed = request.GET.get('productnamevalue')
        print(ref_no_typed)
        if not ref_no_typed:
            return JsonResponse({'error': 'Please enter a search term.'}, status=400)
        
        logger.info(f"Search initiated by {request.user}: {ref_no_typed}")

        try:
            reference_no = PProduct_Creation.objects.filter(Q(Product__Product_Refrence_ID__icontains = ref_no_typed) | Q(Product__Model_Name__icontains = ref_no_typed)).values('Product__Product_Refrence_ID','Product__id','Product__Model_Name')

            if reference_no.exists():

                reference_no_dict = {product['Product__id']: [product.get('Product__Model_Name', ''),product.get('Product__Product_Refrence_ID', '')] for product in reference_no}

                print('reference_no_dict = ',reference_no_dict)

                logger.info(f"Search results for {ref_no_typed}: {reference_no_dict}")

                return JsonResponse({'reference_no': reference_no_dict}, status=200)
            
            return JsonResponse({'error': 'No items found.'}, status=404)
        
        except Exception as e:
            print(e)

    except Exception as e:
        logger.error(f"Error in ref_no_search_ajax: {str(e)}")
        return JsonResponse({"error": "An error occurred while processing the request."}, status=500)



def party_name_search_ajax(request):
    try:
        party_name_typed = request.GET.get('productnamevalue')

        if not party_name_typed:
            return JsonResponse({'error': 'Please enter a search term.'}, status=400)
        
        logger.info(f"Search initiated by {request.user} : {party_name_typed}")

        try:

            party_names = Ledger.objects.filter(name__icontains = party_name_typed).values('id','name')

            if party_names.exists():

                party_name_dict = {data['id']: [data.get('name', '')] for data in party_names}

                logger.info(f"Search results for {party_name_typed}: {party_name_dict}")

                return JsonResponse({'party_name': party_name_dict}, status=200)
            
            return JsonResponse({'error': 'No items found.'}, status=404)

        except Exception as e:
            print(e)

    except Exception as e:
        logger.error(f"Error in party_name_search_ajax: {str(e)}")
        return JsonResponse({"error": "An error occurred while processing the request."}, status=500)



def otward_data_for_sale_return_ajax(request):
    try:
        sale_no = request.GET.get('saleNo')

        print('sale_no = ', sale_no)

        if sales_return_inward.objects.filter(sales_voucher_master__sale_no=sale_no).exists():
            return JsonResponse({"error": f"Sales return already exists for sale no {sale_no}."},status=400)

        outward_queryset = sales_voucher_master_outward_scan.objects.filter(sale_no = sale_no).values('outward_no__outward_product__product__PProduct_SKU','outward_no__outward_product__product__PProduct_image','outward_no__outward_product__product__Product__Product_Refrence_ID','outward_no__outward_product__product__PProduct_color__color_name','outward_no__outward_product__product__Product__Model_Name','outward_no__outward_product__bin_number__bin_name','outward_no__outward_product__quantity','outward_no__outward_product__unique_serial_no','party_name__name'
        )

        list_to_sent_for_outward_data = [] 

        for data in outward_queryset:
            list_to_sent_for_outward_data.append({
                'PProduct_SKU': data['outward_no__outward_product__product__PProduct_SKU'],
                'PProduct_image': data['outward_no__outward_product__product__PProduct_image'],
                'Product_Refrence_ID': data['outward_no__outward_product__product__Product__Product_Refrence_ID'],
                'PProduct_color': data['outward_no__outward_product__product__PProduct_color__color_name'],
                'Model_Name': data['outward_no__outward_product__product__Product__Model_Name'],
                'Bin_Name': data['outward_no__outward_product__bin_number__bin_name'],
                'Quantity': data['outward_no__outward_product__quantity'],
                'unique_serial_no': data['outward_no__outward_product__unique_serial_no']
            })

        list_to_send_for_sales_voucher_data = []

        try:
            sales_voucher = sales_voucher_master_outward_scan.objects.get(sale_no=sale_no)

            list_to_send_for_sales_voucher_data.append({
                'sale_id':sales_voucher.id,
                'party_name': sales_voucher.party_name.id,
                'party_name_value':sales_voucher.party_name.name,
                'warehouse': sales_voucher.selected_warehouse.id,
                'warehouse_value': sales_voucher.selected_warehouse.warehouse_name_finished,
            })

        except ObjectDoesNotExist:
            print(f"Sales voucher with sale_no {sale_no} not found.")

        return JsonResponse({'list_to_sent': list_to_sent_for_outward_data , 'list_to_send_for_sales_voucher_data':list_to_send_for_sales_voucher_data}, status=200)

    except Exception as e:
        logger.error(f"Error in otward_data_for_sale_return_ajax: {str(e)}")
        return JsonResponse({"error": "An error occurred while processing the request."}, status=500)



def process_serial_no_for_return_sales_ajax(request):
    try:
        serial_no = request.GET.get('serialNo')
        sale_no = request.GET.get('saleNo')

        try:
            outward_no = get_object_or_404(sales_voucher_master_outward_scan,sale_no=sale_no)
            check_if_exist = outward_products.objects.get(outward_no=outward_no.outward_no, unique_serial_no = serial_no)
            
            if check_if_exist:
                
                product_scanned_sku = check_if_exist.product.PProduct_SKU

                product_instance = PProduct_Creation.objects.get(PProduct_SKU=product_scanned_sku)

                model_name = product_instance.Product.Model_Name if product_instance.Product.Model_Name else None
                product_name = product_instance.Product.Product_Name if product_instance.Product.Product_Name else None
                product_sku = product_instance.PProduct_SKU
                product_color = product_instance.PProduct_color.color_name if product_instance.PProduct_color else None
                product_image = product_instance.PProduct_image.url if product_instance.PProduct_image else None

                product_main_cats = product_instance.Product.product_cats.all()

                main_cats_all = [x.SubCategory_id.product_main_category for x in product_main_cats]

                session_bin_data = request.session.get('bin_data', {})

                bins_related_to_product = []


                for bin_obj in finished_product_warehouse_bin.objects.filter(sub_catergory_id__in=main_cats_all):

                    product_count = finishedgoodsbinallocation.objects.filter(bin_number=bin_obj, outward_done=False).count()

                    temp_session_qty = session_bin_data.get(str(bin_obj.id), 0)
                    total_count = product_count + temp_session_qty
                    
                    find_product_count = finishedgoodsbinallocation.objects.filter(bin_number=bin_obj, outward_done=False, product = product_sku).count()

                    bins_related_to_product.append({
                            'bin_id': bin_obj.id,
                            'bin_name': bin_obj.bin_name,
                            'bin_size': bin_obj.product_size_in_bin,
                            'products_in_bin': total_count,
                            'find_product_count': find_product_count
                        })

                    

                    bins_related_to_product = sorted(bins_related_to_product, key=lambda x : x['find_product_count'], reverse=True)

                    print(find_product_count)
                return JsonResponse({
                        'model_name': model_name,
                        'product_name': product_name,
                        'product_sku': product_sku,
                        'bin_to_dict': bins_related_to_product,
                        'product_color': product_color,
                        'product_image': product_image,
                        'message': f'Serial No {serial_no} processed successfully.'},status=200)
            else:
                message.error(f"This Serial no {sale_no} does not exist in sales voucher")
        except ObjectDoesNotExist as e:
            print('This Serial no does not exist in sales voucher')
            logger.error(f"This Serial no {sale_no} does not exist in sales voucher: {str(e)}")
            

    except Exception as e:
        logger.error(f"Error in process_serial_no_for_return_sales_ajax: {str(e)}")
        return JsonResponse({"error": "An error occurred while processing the request."}, status=500)



def return_product_with_bin_ajax(request):
    try:
        product_sku = request.GET.get('productSku')
        bin_no = request.GET.get('binNo')
        serial_no = request.GET.get('serialno')

        bin_obj = get_object_or_404(finished_product_warehouse_bin, pk=bin_no)
        product_instance = PProduct_Creation.objects.get(PProduct_SKU=product_sku)

        session_bin_data = request.session.get('bin_data', {})

        bin_capacity = bin_obj.product_size_in_bin
        product_count = finishedgoodsbinallocation.objects.filter(bin_number=bin_no, outward_done=False).count()

        temp_assigned = session_bin_data.get(str(bin_no), 0)
        total_count = product_count + temp_assigned

        if total_count + 1 > bin_capacity:
            return JsonResponse({'error': f"Cannot add to bin '{bin_obj.bin_name}'. Bin capacity exceeded!"}, status=400)

        session_bin_data[str(bin_no)] = temp_assigned + 1
        request.session['bin_data'] = session_bin_data
        request.session.modified = True

        return JsonResponse({
            'reference_no': product_instance.Product.Product_Refrence_ID or None,
            'model_name': product_instance.Product.Model_Name or None,
            'product_name': product_instance.Product.Product_Name or None,
            'product_sku': product_instance.PProduct_SKU,
            'product_color': product_instance.PProduct_color.color_name if product_instance.PProduct_color else None,
            'product_image': product_instance.PProduct_image.url if product_instance.PProduct_image else None,
            'bin_id': bin_no,
            'bin_name': bin_obj.bin_name,
            'serialno': serial_no,
            'message': f"Product added to bin {bin_obj.bin_name} successfully."
        }, status=200)

    except Exception as e:
        logger.error(f"Error in return_product_with_bin_ajax: {str(e)}")
        return JsonResponse({"error": "An error occurred while processing the request."}, status=500)




def sales_return_inward_to_bin(request, r_id=None):

    for key in ['bin_data', 'product_data']:
        if key in request.session:
            del request.session[key]
            request.session.modified = True

    instance_queryset = sales_return_inward.objects.filter(id=r_id).first()
    master_form = salesreturninwardmasterform(request.POST or None, instance=instance_queryset)

    if instance_queryset:
        formset = sales_return_product_formset_update(request.POST or None, instance=instance_queryset)

        outward_queryset = sales_voucher_master_outward_scan.objects.filter(
            sale_no=instance_queryset.sales_voucher_master.sale_no
        ).values(
            'outward_no__outward_product__product__PProduct_SKU',
            'outward_no__outward_product__product__PProduct_image',
            'outward_no__outward_product__product__Product__Product_Refrence_ID',
            'outward_no__outward_product__product__PProduct_color__color_name',
            'outward_no__outward_product__product__Product__Model_Name',
            'outward_no__outward_product__bin_number__bin_name',
            'outward_no__outward_product__quantity',
            'outward_no__outward_product__unique_serial_no',
        )

        list_to_sent_for_outward_data = [
            {
                'PProduct_SKU': data['outward_no__outward_product__product__PProduct_SKU'],
                'PProduct_image': data['outward_no__outward_product__product__PProduct_image'],
                'Product_Refrence_ID': data['outward_no__outward_product__product__Product__Product_Refrence_ID'],
                'PProduct_color': data['outward_no__outward_product__product__PProduct_color__color_name'],
                'Model_Name': data['outward_no__outward_product__product__Product__Model_Name'],
                'Bin_Name': data['outward_no__outward_product__bin_number__bin_name'],
                'Quantity': data['outward_no__outward_product__quantity'],
                'unique_serial_no': data['outward_no__outward_product__unique_serial_no'],
            }
            for data in outward_queryset
        ]

    else:
        formset = sales_return_product_formset(request.POST or None)
        list_to_sent_for_outward_data = None

    
    if request.method == 'POST':
        master_form = salesreturninwardmasterform(request.POST or None, instance=instance_queryset)
        formset = sales_return_product_formset(request.POST or None, instance=instance_queryset)

        if master_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    master_form_instance = master_form.save(commit=False)
                    master_form_instance.save()

                    sale_id = master_form_instance.sales_voucher_master.id
                    sale_return_id = master_form_instance.id

                    
                    for form in formset.deleted_forms:
                        if form.instance.pk:
                            form.instance.delete()

                    formset.forms = [form for form in formset.forms if form.has_changed()]
                    for form in formset:
                        if not form.cleaned_data.get('DELETE'):
                            form_instance = form.save(commit=False)
                            form_instance.sales_return_inward_instance = master_form_instance
                            form_instance.save()

                            bin_no = form_instance.bin_number
                            serial_no = form_instance.unique_serial_no
                            product = form_instance.product

                            #update bin in inward
                            inward_data, created = finishedgoodsbinallocation.objects.get_or_create(unique_serial_no=serial_no)

                            if not created:
                                finishedgoodsbinallocation.objects.filter(unique_serial_no=serial_no).update(bin_number=bin_no,outward_done=False)

                            #increase bin qantity in Product_bin_quantity_through_table
                            bin_qty_increase, created = Product_bin_quantity_through_table.objects.get_or_create(
                                product=product, bin=bin_no,
                                defaults={'product_quantity': 0}  # Ensures a default value instead of NULL
                            )

                            # If the record already existed, ensure product_quantity is not None
                            if bin_qty_increase.product_quantity is None:
                                bin_qty_increase.product_quantity = 0

                            bin_qty_increase.product_quantity += 1
                            bin_qty_increase.save()


                    sales_return_master = sales_return_voucher_master.objects.filter(sales_voucher_master=sale_id,sales_return_inward_instance=sale_return_id).exists()

                    if sales_return_master:

                        print("URL FOR UPDATE >>>")

                        sales_return_master = sales_return_voucher_master.objects.get(sales_voucher_master=sale_id,sales_return_inward_instance=sale_return_id)
                        
                        sale_return_voucher_id = sales_return_master.id

                        products = {}

                        post_data = request.POST.dict()

                        total_forms = int(post_data.get('sales_return_product_set-TOTAL_FORMS', 0))

                        products[master_form_instance.sales_return_no] = []

                        for i in range(total_forms):

                            product = post_data.get(f'sales_return_product_set-{i}-product')
                            quantity = int(post_data.get(f'sales_return_product_set-{i}-scan_qty', 0))

                            found = False
                            for p in products[master_form_instance.sales_return_no]:
                                if p['product'] == product:
                                    p['quantity'] += quantity
                                    found = True
                                    break

                            if not found:

                                product_info = PProduct_Creation.objects.get(PProduct_SKU = product)
                                
                                products[master_form_instance.sales_return_no].append({
                                'product_ref': post_data.get(f'sales_return_product_set-{i}-product_RefNo'),
                                'product_name': post_data.get(f'sales_return_product_set-{i}-product_name_value'),
                                'product': product,
                                'color': post_data.get(f'product_color_{i}'),
                                'quantity': quantity,
                                'mrp': decimal_to_float(product_info.Product.Product_MRP),
                                'customer_price': decimal_to_float(product_info.Product.Product_SalePrice_CustomerPrice),
                                'gst': product_info.Product.Product_GST.gst_percentage
                                })
                        print('products = ',products)
                        request.session['products_data'] = products
                        
                        return redirect(reverse('sales-return-voucher-create-update', args=[sale_id, sale_return_id,sale_return_voucher_id,'update_srv']))
                    else:
                        print("URL FOR CREATE !!!")
                        return redirect(reverse('sales-return-voucher-create-update', args=[sale_id, sale_return_id]))

            except Exception as e:
                print(f"Error: {e}")

    return render(request, 'accounts/sales_return_inward.html', {
        'master_form': master_form,
        'formset': formset,
        'list_to_sent_for_outward_data': list_to_sent_for_outward_data,
    })




def sales_return_voucher_create_update(request,s_id=None, sr_id=None, sv_id=None, action=None):

    master_form_data = get_object_or_404(sales_return_inward, id=sr_id)

    master_form_data_instance = None

    if sv_id:
        
        master_form_data_instance = get_object_or_404(sales_return_voucher_master, id=sv_id)

        master_form = sales_return_voucher_master_form(instance = master_form_data_instance)

        if action == 'update_srv':

            session_products = request.session.get('products_data', None)

            if session_products:
                for key, val in session_products.items():
                    sale_return_id = key
                    product_list2 = val
                    break

                sales_product_info = sales_return_voucher.objects.filter(sales_return_master = sv_id).values('product_name__Product__Product_Refrence_ID','product_name__Product__Model_Name','product_name__PProduct_SKU','product_name__PProduct_color__color_name','quantity','trade_disct','spl_disct')

                sales_product_info_list = list(sales_product_info)

                merged_product_list = []

                for product in product_list2:
                    sku = product['product']
                    matched = False

                    for sales_product in sales_product_info_list:
                        if sku == str(sales_product['product_name__PProduct_SKU']):
                            merged_product_list.append({
                                'product_ref': product['product_ref'],
                                'product_name': product['product_name'],
                                'product': product['product'],
                                'color': product['color'],
                                'quantity': product['quantity'],
                                'mrp': product['mrp'],
                                'customer_price': product['customer_price'],
                                'gst': product['gst'],
                                'trade_disct': sales_product['trade_disct'],
                                'spl_disct': sales_product['spl_disct'],
                            })
                            matched = True
                            break
                    
                    if not matched:
                        merged_product_list.append({
                            'product_ref': product['product_ref'],
                            'product_name': product['product_name'],
                            'product': product['product'],
                            'color': product['color'],
                            'quantity': product['quantity'],
                            'mrp': product['mrp'],
                            'customer_price': product['customer_price'],
                            'gst': product['gst'],
                            'trade_disct': sales_product['trade_disct'],
                            'spl_disct': sales_product['spl_disct'],
                        })

                print('merged_product_list = ',merged_product_list)

                sales_return_voucher_formset_create = inlineformset_factory(sales_return_voucher_master, sales_return_voucher,form=sales_return_voucher_form, extra=len(merged_product_list), can_delete=True)
            
                formset = sales_return_voucher_formset_create(initial=merged_product_list, prefix="sale_return_forms")

        else:

            sales_return_voucher_formset_view = inlineformset_factory(sales_return_voucher_master, sales_return_voucher,form=sales_return_voucher_form, extra=0, can_delete=True)

            formset = sales_return_voucher_formset_view(instance = master_form_data_instance, prefix="sale_return_forms")
    else:

        products = sales_return_product.objects.filter(sales_return_inward_instance=sr_id)

        product_list = []

        for data in products:
            sku = data.product.PProduct_SKU
            quantity = data.scan_qty

            found = False
            for item in product_list:
                if item['product'] == sku:
                    item['quantity'] += quantity
                    found = True
                    break

            if not found:
                sale_objects = sales_voucher_outward_scan.objects.filter(
                    sales_voucher_master=s_id, product_name=sku
                ).first()

                if sale_objects:
                    product_list.append({
                        'ref_no': data.product.Product.Product_Refrence_ID,
                        'product': sku,
                        'product_name': data.product.Product.Model_Name,
                        'color': data.product.PProduct_color.color_name,
                        'quantity': quantity,
                        'mrp': decimal_to_float(data.product.Product.Product_MRP),
                        'customer_price': decimal_to_float(data.product.Product.Product_SalePrice_CustomerPrice),
                        'gst': data.product.Product.Product_GST.gst_percentage,
                        'trade_disct': sale_objects.trade_disct,
                        'spl_disct': sale_objects.spl_disct,
                        'cash_disct': decimal_to_float(sale_objects.sales_voucher_master.cash_disct),
                    })

        master_form = sales_return_voucher_master_form()
        
        sales_return_voucher_formset_create = inlineformset_factory(sales_return_voucher_master, sales_return_voucher,form=sales_return_voucher_form, extra=len(product_list), can_delete=True)

        formset = sales_return_voucher_formset_create(initial=product_list, prefix="sale_return_forms")

    if request.method == 'POST':

        print(request.POST)

        master_form = sales_return_voucher_master_form(request.POST or None, instance = master_form_data_instance)

        formset = sales_return_voucher_formset_create(request.POST or None,instance = master_form_data_instance, prefix="sale_return_forms")

        if master_form.is_valid() and formset.is_valid():

            try:

                with transaction.atomic():

                    if action == 'update_srv':

                        old_instances = sales_return_voucher.objects.filter(sales_return_master=master_form_data_instance)

                        for old_instance in old_instances:

                            warehouse_instance = Product_warehouse_quantity_through_table.objects.get(warehouse=old_instance.sales_return_master.sales_voucher_master.selected_warehouse,product=old_instance.product_name)

                            warehouse_instance.quantity -= old_instance.quantity 
                            warehouse_instance.save()

                            old_instance.delete()

                    master_form_instance = master_form.save(commit=False)
                    master_form_instance.save()

                    if formset.is_valid():

                        for form in formset.deleted_forms:
                            if form.instance.pk:
                                form.instance.delete()

                        selected_warehouse = master_form_instance.sales_voucher_master.selected_warehouse

                        formset.forms = [form for form in formset.forms if form.has_changed()]

                        for form in formset:
                            instance = form.save(commit=False)
                            instance.sales_return_master = master_form_instance
                            
                            warehouse = instance.sales_return_master.sales_voucher_master.selected_warehouse
                            product = instance.product_name
                            qty = instance.quantity

                            instance.save()

                            warehouse_instance_new = Product_warehouse_quantity_through_table.objects.get(warehouse = warehouse,product = product)

                            warehouse_instance_new.quantity += qty
                            warehouse_instance_new.save()

                        return(redirect('sales-voucher-list-warehouse'))
                    
            except Exception as e:
                print(f"Error saving formset: {e}")

    return render(request, 'accounts/sales_return_create_update.html', {'master_form_data': master_form_data,'formset': formset,'master_form':master_form,})




def sale_return_list(request):

    total_sale_qty_subquery = sales_voucher_outward_scan.objects.filter(sales_voucher_master=OuterRef('sales_voucher_master')).values('sales_voucher_master').annotate(total_sale_qty=Sum('quantity')).values('total_sale_qty')

    grand_total_subquery = sales_return_voucher_master.objects.filter(
    sales_return_inward_instance=OuterRef('pk')).values('grand_total')[:1]
    
    queryset = sales_return_inward.objects.annotate(total_qty=Sum('sales_return_product__scan_qty'),total_sale_qty=Subquery(total_sale_qty_subquery),grand_total=Subquery(grand_total_subquery)).order_by('created_date')

    return render(request,'accounts/sales_return_list.html',{'queryset':queryset})




def delivery_challan_product_ajax(request):
    try:
        product_name_typed = request.GET.get('nameValue')
        godown_name = request.GET.get('selected_godown')

        print(godown_name)

        if not godown_name:
            return JsonResponse({'error': 'Please select godown first'}, status=400)

        logger.info(f"Search initiated by {request.user}: {product_name_typed}")

        products = product_godown_quantity_through_table.objects.filter(
            Q(product_color_name__Product__Model_Name__icontains = product_name_typed) |
            Q(product_color_name__PProduct_color__color_name__icontains= product_name_typed) | 
            Q(product_color_name__Product__Product_Refrence_ID__icontains=product_name_typed)
            ).filter(godown_name = godown_name).values(
                'product_color_name__Product__Model_Name',
                'product_color_name__PProduct_SKU',
                'product_color_name__PProduct_color__color_name',
                'quantity',
                'product_color_name__Product__Product_Refrence_ID',
                'product_color_name__PProduct_image'
                ).exclude(quantity=0)

        print(products)

        product_list = {}

        for product in products:
            sku = product['product_color_name__PProduct_SKU'] 
            
            product_list[sku] = [product['product_color_name__PProduct_color__color_name'],product['product_color_name__Product__Model_Name'],product['product_color_name__Product__Product_Refrence_ID'],product['product_color_name__PProduct_image'],product['quantity']]

        return JsonResponse({'products': product_list}, safe=False)
    

    except Exception as e:
        logger.error(f"Error in delivery_challan_product_ajax: {e}")
        return JsonResponse({'error': 'An error occurred while fetching data.'}, status=500)




def delivery_challan_create_update(request, d_id=None):

    party_names = Ledger.objects.all()
    godowns = Godown_finished_goods.objects.all()
    
    if d_id:
        d_instance = DeliveryChallanMaster.objects.get(id = d_id)
        master_form = DeliveryChallanMasterForm(request.POST or None,instance = d_instance)
        formset = DeliveryChallanProductsUpdateFormset(instance = d_instance)

        products = product_godown_quantity_through_table.objects.all().values('product_color_name__PProduct_SKU','quantity')
    
        product_list = []

        for product in products:
            sku = product['product_color_name__PProduct_SKU']
            product_list.append({sku : product['quantity']})
            

    else:
        d_instance = None
        product_list=None
        master_form = DeliveryChallanMasterForm(request.POST or None, instance = d_instance)
        formset = DeliveryChallanProductsCreateFormset()

    if request.method == 'POST':
        
        print(request.POST)

        master_form = DeliveryChallanMasterForm(request.POST or None, instance = d_instance)

        if d_id:
            formset = DeliveryChallanProductsCreateFormset(request.POST or None, instance = d_instance)
        else:
            formset = DeliveryChallanProductsUpdateFormset(request.POST or None)

        if master_form.is_valid() and formset.is_valid():

            try:

                with transaction.atomic():

                    if not master_form.is_valid():
                        print("Form Errors:", master_form.errors)
                    
                    master_form_instance = master_form.save(commit=False)
                    master_form_instance.save()

                    selected_godown = master_form_instance.selected_godown

                    if not formset.is_valid():
                        for form in formset:
                            if not form.is_valid():
                                print("Form Errors:", form.errors)

                    if formset.is_valid():
                        for form in formset.deleted_forms:
                            if form.instance.pk:

                                is_linked = sales_voucher_finish_Goods.objects.filter(challan=form.instance).exists()

                                if is_linked:
                                    messages.error(request, f"Cannot delete product {form.instance.product_name.Product.Model_Name} as it is linked to a sales voucher.")
                                    continue

                                product_name = form.instance.product_name
                                product_qty = form.instance.quantity

                                del_obj, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name=product_name)

                                del_obj.quantity += product_qty
                                del_obj.save()

                                form.instance.delete()  

                        for form in formset:
                            if not form.cleaned_data.get('DELETE'):
                                if form.has_changed():
                                    instance = form.save(commit=False)
                                    instance.delivery_challan = master_form_instance
                                    instance.balance_qty = form.cleaned_data['quantity']
                                    
                                    old_qty = form.initial.get('quantity', 0)
                                    new_qty = form.cleaned_data['quantity']

                                    if form.has_changed() and 'quantity' in form.changed_data:

                                        if old_qty:

                                            obj, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name=instance.product_name)

                                            obj.quantity += old_qty
                                            obj.quantity -= new_qty  
                                            obj.save()

                                    instance.save()

                    return(redirect('delivery-challan-list'))
                
            except Exception as e:
                print(f"Error saving formset: {e}")

    return render(request,'production/delivery_challan_create_update.html',{'master_form':master_form,'formset':formset,'party_names':party_names,'godowns':godowns,'product_list':product_list})






def delivery_challan_process_for_sale_voucher(request):
    
    try:
        d_challan_no = request.GET.get('challanNo')

        if not d_challan_no:
            return JsonResponse({"error": "challanNo parameter is required"}, status=400)

        d_id = get_object_or_404(DeliveryChallanMaster, delivery_challan_no=d_challan_no)

        total_data = DeliveryChallanProducts.objects.filter(delivery_challan=d_id).aggregate(
            total_qty=Sum('quantity'),
            total_balance=Sum('balance_qty')
        )

        # print('total_data = ',total_data)

        d_challan_data = {
            "delivery_challan_no": d_challan_no,
            "id": d_id.id,
            "godown_id":d_id.selected_godown.id,
            "godown_name":d_id.selected_godown.godown_name_finished,
            "total_qty": total_data['total_qty'] or 0,
            "balance_qty": total_data['total_balance'] or 0
        }

        # print('d_challan_data = ',d_challan_data)

        total_product_data = DeliveryChallanProducts.objects.filter(delivery_challan=d_id)

        d_challan_product_data = {}

        for i in total_product_data:
            challan_no = i.delivery_challan.delivery_challan_no

            if challan_no not in d_challan_product_data:
                d_challan_product_data[challan_no] = []

            d_challan_product_data[challan_no].append({
                'product_name': i.product_name.Product.Model_Name,
                'product_sku': i.product_name.PProduct_SKU,
                'color': i.product_name.PProduct_color.color_name,
                'qty': i.quantity,
                'balance_qty': i.balance_qty
            })

        # print(d_challan_product_data)

        return JsonResponse({"d_challan_data":d_challan_data,"products": d_challan_product_data},status=200)

    except Exception as e:
        print(f"Error in processing delivery challan: {e}")
        return JsonResponse({"error": str(e)}, status=500)





def product_transfer_to_warehouse_ajax(request):
    
    godown_id = request.GET.get('godown_id')
    delivery_challan_id = request.GET.get('deliveryChallan')

    
    if delivery_challan_id:

        print('in delivery_challan and godown_id')

        try:

            filtered_product = list(DeliveryChallanProducts.objects.filter(delivery_challan = delivery_challan_id).annotate(total_qty = Sum('balance_qty')).values('product_name__Product__Product_Name','product_name__PProduct_SKU','product_name__PProduct_color__color_name','quantity','product_name__Product__Model_Name','product_name__Product__Product_Refrence_ID','product_name__Product__Product_UOM','product_name__Product__Product_MRP','product_name__Product__Product_SalePrice_CustomerPrice','product_name__Product__Product_GST__gst_percentage','total_qty','id','balance_qty','delivery_challan__delivery_challan_no'))

            if filtered_product:

                dict_to_send = {}

                for query in filtered_product:
                    ref_no = query.get('product_name__Product__Product_Refrence_ID')
                    p_sku = query.get('product_name__PProduct_SKU')
                    product_name = query.get('product_name__Product__Product_Name')
                    product_model_name = query.get('product_name__Product__Model_Name')
                    color = query.get('product_name__PProduct_color__color_name')
                    uom = query.get('product_name__Product__Product_UOM')
                    qty = query.get('balance_qty')
                    mrp = query.get('product_name__Product__Product_MRP')
                    customer_price = query.get('product_name__Product__Product_SalePrice_CustomerPrice')
                    gst = query.get('product_name__Product__Product_GST__gst_percentage')
                    total_qty = query.get('total_qty')
                    id = query.get('id')
                    d_challan_no = query.get('delivery_challan__delivery_challan_no')

                    dict_to_send[p_sku] = [product_name,color,qty,product_model_name,ref_no,uom,mrp,customer_price,gst,total_qty,id,d_challan_no]

                print('dict_to_send = ',dict_to_send)

                return JsonResponse({'filtered_product':dict_to_send})
        
        except Exception as e:
            return JsonResponse({'error': 'No items found.'}, status=404)
    else:

        print('in godown_id')

        try:
            filtered_product = list(product_godown_quantity_through_table.objects.filter(
            godown_name__id = godown_id).values('product_color_name__Product__Product_Name','product_color_name__PProduct_SKU','product_color_name__PProduct_color__color_name','quantity','product_color_name__Product__Model_Name','product_color_name__Product__Product_Refrence_ID','product_color_name__Product__Product_UOM','product_color_name__Product__Product_MRP','product_color_name__Product__Product_SalePrice_CustomerPrice','product_color_name__Product__Product_GST__gst_percentage'))

            if filtered_product:
                dict_to_send = {}

                for query in filtered_product:
                    ref_no = query.get('product_color_name__Product__Product_Refrence_ID')
                    p_sku = query.get('product_color_name__PProduct_SKU')
                    product_name = query.get('product_color_name__Product__Product_Name')
                    product_model_name = query.get('product_color_name__Product__Model_Name')
                    color = query.get('product_color_name__PProduct_color__color_name')
                    uom = query.get('product_color_name__Product__Product_UOM')
                    qty = query.get('quantity')
                    mrp = query.get('product_color_name__Product__Product_MRP')
                    customer_price = query.get('product_color_name__Product__Product_SalePrice_CustomerPrice')
                    gst = query.get('product_color_name__Product__Product_GST__gst_percentage')
                    
                    dict_to_send[p_sku] = [product_name,color,qty,product_model_name,ref_no,uom,mrp,customer_price,gst]
                
                return JsonResponse({'filtered_product':dict_to_send})
            
            else:
                raise ObjectDoesNotExist('product not found')

        except ObjectDoesNotExist as oe:
            return JsonResponse({'error': 'No items found.'}, status=404)
        
        except Exception as e:
            return JsonResponse({'error': 'No items found.'}, status=404)
    




def salesvouchercreateupdate(request,s_id=None,dc_id=None):
    
    if dc_id == 0:
        dc_id = None
    
    if s_id == 0:
        s_id = None

    salesvouchercreateformset = inlineformset_factory(sales_voucher_master_finish_Goods,sales_voucher_finish_Goods,form = salesvoucherfinishGoodsForm,extra=1, can_delete=True)

    SalesVoucherDeliveryChallanFormset = inlineformset_factory(sales_voucher_master_finish_Goods, SalesVoucherDeliveryChallan,form=SalesVoucherDeliveryChallanForm, extra=1, can_delete=True)

    party_name = Ledger.objects.filter(under_group__account_sub_group = 'Sundry Debtors')

    dict_to_send = None


    if dc_id:

        print("IN DC ")

        challan_product_queryset = DeliveryChallanProducts.objects.filter(delivery_challan = dc_id).values('product_name__Product__Product_Name','product_name__PProduct_SKU','product_name__PProduct_color__color_name','quantity','product_name__Product__Model_Name','product_name__Product__Product_Refrence_ID','product_name__Product__Product_UOM','product_name__Product__Product_MRP','product_name__Product__Product_SalePrice_CustomerPrice','product_name__Product__Product_GST__gst_percentage','id','balance_qty','delivery_challan__delivery_challan_no','id')

        print("challan_product_queryset = ",challan_product_queryset)

        product_initial_data = []

        for product in challan_product_queryset:
            
            product_initial_data.append({
                'challan': product['id'],
                'challanValue': product['delivery_challan__delivery_challan_no'],
                'product_name_value': product['product_name__Product__Model_Name'],
                'color':product['product_name__PProduct_color__color_name'],
                'sku':product['product_name__PProduct_SKU'],
                'quantity': product['balance_qty'],
                'mrp': product['product_name__Product__Product_MRP'],
                'c_price':product['product_name__Product__Product_SalePrice_CustomerPrice'],
                'gst':product['product_name__Product__Product_GST__gst_percentage'],
            })
        
        d_challan_product_data = None
        voucher_instance = None

        master_form = salesvouchermasterfinishGoodsForm()

        salesvouchercreateformset = inlineformset_factory(sales_voucher_master_finish_Goods,sales_voucher_finish_Goods,form = salesvoucherfinishGoodsForm,extra=len(product_initial_data), can_delete=True)

        formset = salesvouchercreateformset(initial = product_initial_data)
        

        total_data = DeliveryChallanProducts.objects.filter(delivery_challan=dc_id).aggregate(total_qty=Sum('quantity'),total_balance=Sum('balance_qty'))

        d_id = get_object_or_404(DeliveryChallanMaster, id=dc_id)

        d_challan_initial_data = []

        d_challan_initial_data.append({
            "delivery_challan_no": d_id.delivery_challan_no,
            "challan_id": d_id.id,
            "total_qty": total_data['total_qty'] or 0,
            "balance_qty": 0
        })

        print(d_challan_initial_data)
        
        SalesVoucherDeliveryChallanFormset = inlineformset_factory(sales_voucher_master_finish_Goods, SalesVoucherDeliveryChallan,form=SalesVoucherDeliveryChallanForm, extra=len(d_challan_initial_data), can_delete=True)

        delivery_challan_formset = SalesVoucherDeliveryChallanFormset(initial = d_challan_initial_data)

        page_name = 'Create Sales Invoice'


    elif s_id:

        print("IN s_id ")

        voucher_instance = sales_voucher_master_finish_Goods.objects.get(id=s_id)

        master_form = salesvouchermasterfinishGoodsForm(request.POST or None,instance=voucher_instance)

        formset = salesvoucherupdateformset(request.POST or None,instance=voucher_instance)

        delivery_challan_formset = SalesVoucherDeliveryChallanFormsetupdate(request.POST or None, instance=voucher_instance)

        page_name = 'Edit Sales Invoice'

        total_product_data = SalesVoucherDeliveryChallan.objects.filter(sales_voucher=s_id)

        d_challan_product_data = {}

        for i in total_product_data:
            challan_no = i.delivery_challan.delivery_challan_no
            if challan_no not in d_challan_product_data:
                d_challan_product_data[challan_no] = []
            
            challan_products = i.delivery_challan.deliverychallanproducts_set.all()
            
            for j in challan_products:
                d_challan_product_data[challan_no].append({
                    'sku': j.product_name.PProduct_SKU,
                    'model_name': j.product_name.Product.Model_Name,
                    'color': j.product_name.PProduct_color.color_name,
                    'qty': j.quantity,
                    'balance_qty': j.balance_qty
                })

    else:
        ("IN ELSE")
        d_challan_product_data = None
        voucher_instance = None
        master_form = salesvouchermasterfinishGoodsForm()
        formset = salesvouchercreateformset()
        delivery_challan_formset = SalesVoucherDeliveryChallanFormset()
        page_name = 'Create Sales Invoice'


    if request.method == "POST":

        print(request.POST)

        if s_id:
            voucher_instance = sales_voucher_master_finish_Goods.objects.get(id=s_id)
            master_form = salesvouchermasterfinishGoodsForm(request.POST, instance=voucher_instance)
            formset = salesvoucherupdateformset(request.POST, instance=voucher_instance)
            delivery_challan_formset = SalesVoucherDeliveryChallanFormsetupdate(request.POST, instance=voucher_instance)
        else:
            voucher_instance = None
            master_form = salesvouchermasterfinishGoodsForm(request.POST)
            formset = salesvouchercreateformset(request.POST)
            delivery_challan_formset = SalesVoucherDeliveryChallanFormset(request.POST)

        formset.forms = [form for form in formset.forms if form.has_changed()]

        if not master_form.is_valid():
            print("Form Errors:", master_form.errors)

        if not formset.is_valid():
            for form in formset:
                if not form.is_valid():
                    print("Form Errors:", form.errors)

        

        if master_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    master_form_instance = master_form.save(commit=False)
                    master_form_instance.save()

                    if delivery_challan_formset.is_valid():
                        for form in delivery_challan_formset:
                            if form.is_valid():
                                if not form.cleaned_data.get('DELETE'):
                                    form_instance = form.save(commit=False)
                                    form_instance.sales_voucher = master_form_instance

                                    challan_id = form.cleaned_data.get('delivery_challan')
                                    balance_qty = form.cleaned_data.get('balance_qty')

                                    sales_voucher_delivery_challan_relation_obj = SalesVoucherDeliveryChallan.objects.filter(delivery_challan = challan_id).update(balance_qty=balance_qty)
                                    

                                    form_instance.save()
                    else:
                        print("Delivery Challan Formset Errors:", delivery_challan_formset.errors)



                    for form in formset.deleted_forms:
                        if form.instance.pk:
                            product = form.instance.product_name
                            product_qty = form.instance.quantity   
                            delivery_challan_master = form.cleaned_data.get('challan')

                            challan_product_entry = DeliveryChallanProducts.objects.get(delivery_challan=delivery_challan_master.delivery_challan,product_name=product)

                            if challan_product_entry:
                                challan_product_entry.balance_qty += product_qty
                                challan_product_entry.save()
                            
                            form.instance.delete()


                    for form in formset:
                        if not form.cleaned_data.get('DELETE'):
                            form_instance = form.save(commit=False)
                            form_instance.sales_voucher_master = master_form_instance
                            


                            if form.has_changed() and 'quantity' in form.changed_data:
                                old_quantity = form.initial.get('quantity')
                                new_quantity = form.cleaned_data.get('quantity')
                                product = form.cleaned_data.get('product_name')
                                delivery_challan_master = form.cleaned_data.get('challan')

                                if old_quantity:
                                    
                                    challan_product_entry = DeliveryChallanProducts.objects.get(delivery_challan=delivery_challan_master.delivery_challan,product_name=product)

                                    if challan_product_entry:
                                        challan_product_entry.balance_qty += old_quantity
                                        challan_product_entry.balance_qty -= new_quantity
                                        challan_product_entry.save()

                            form_instance.save()
                            
                    return redirect('sales-voucher-list')
            except Exception as e:
                print(e) 

    return render(request,'accounts/sales_invoice.html',{'master_form':master_form,'formset':formset,'page_name':page_name,'party_name':party_name,'dict_to_send':dict_to_send,'delivery_challan_formset':delivery_challan_formset,'d_challan_product_data':d_challan_product_data})





def delivery_challan_list(request):

    #Delivery challan total_qty and balance_qty
    delivary_challan_list = DeliveryChallanMaster.objects.annotate(total_qty=Sum('deliverychallanproducts__quantity'),total_balance_qty=Sum('deliverychallanproducts__balance_qty')).order_by('created_date')

    #stock transfer queryset
    warehouse_product_transfer_list = Finished_goods_Stock_TransferMaster.objects.all().annotate(all_qc_qty=Sum('finished_goods_transfer_records__qc_recieved_qty'),total_recieved_qty=Sum('finished_goods_transfer_records__product_quantity_transfer')).order_by('created_date')

    #Labour workin create qty (subquery)
    product_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(total_labour_workin_qty_sum = Coalesce(Sum('return_pcs'), 0)).values('total_labour_workin_qty_sum')

    #Labour workin pending for approval(subquery)
    product_pending_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(total_labour_workin_pen_qty_sum = Coalesce(Sum('pending_for_approval'), 0)).values('total_labour_workin_pen_qty_sum')
    
    #Labour workin total approve qty (subquery)
    product_approve_queryset_subquery = labour_work_in_product_to_item.objects.filter(product_sku = OuterRef('PProduct_SKU')).values('product_sku').annotate(total_labour_workin_aprv_qty_sum = Coalesce(Sum('approved_qty'), 0)).values('total_labour_workin_aprv_qty_sum')

    # delivery challan total qty (subquery)
    delivery_challan_subquery = DeliveryChallanProducts.objects.filter(product_name__PProduct_SKU = OuterRef('PProduct_SKU')).values('product_name__PProduct_SKU').annotate(total_challan_qty_sum = Sum('quantity'),total_challan_bal_qty_sum = Sum('quantity')).values('total_challan_qty_sum')

    # lwo processed pcs total qty (subquery)
    total_lwo_queryset_subquery = product_to_item_labour_child_workout.objects.filter(
    product_sku=Cast(OuterRef('PProduct_SKU'), output_field=CharField())).values('product_sku').annotate(total_processed_qty=Sum('processed_pcs')).values('total_processed_qty')

    #pending for lwi (subquery)
    total_lwi_balance_queryset_subquery = product_to_item_labour_child_workout.objects.filter(
    product_sku=Cast(OuterRef('PProduct_SKU'), output_field=CharField())).values('product_sku').annotate(total_bal_qty=Sum('labour_w_in_pending')).values('total_bal_qty')

    # stock transfer qty (subquery)
    total_stock_trf_queryset_subquery = Finished_goods_transfer_records.objects.filter(product__PProduct_SKU = OuterRef('PProduct_SKU')).values('product__PProduct_SKU').annotate(total_trf_qty = Sum('product_quantity_transfer')).values('total_trf_qty')

    product_queryset = PProduct_Creation.objects.all().annotate(
        total_qty = Sum('godown_colors__quantity'),
        total_labour_workin_qty = Subquery(product_queryset_subquery),
        total_labour_workin_pending_qty = Subquery(product_pending_queryset_subquery),
        total_labour_workin_approve_qty = Subquery(product_approve_queryset_subquery),
        total_challan_qty = Subquery(delivery_challan_subquery),
        total_lwo_qty = Subquery(total_lwo_queryset_subquery),
        total_trf_qty = Subquery(total_stock_trf_queryset_subquery),
        total_bal_for_lwi = Subquery(total_lwi_balance_queryset_subquery)).filter(
            Q(total_qty__gt=0) | 
            Q(total_labour_workin_qty__gt=0) | 
            Q(total_labour_workin_pending_qty__gt=0) | 
            Q(total_labour_workin_approve_qty__gt=0) |  
            Q(total_challan_qty__gt=0)).order_by('Product__Model_Name').select_related('Product','PProduct_color')
    
    
    return render(request,'production/delivery_challan_list.html',{'delivary_challan_list':delivary_challan_list,'product_queryset':product_queryset,'page_name':'Delivery challan','warehouse_product_transfer_list':warehouse_product_transfer_list})



def delete_delivery_challan(request, pk):
    logger.info('delete_delivery_challan function called')

    check_sale_exist = SalesVoucherDeliveryChallan.objects.filter(delivery_challan=pk).exists()
    if check_sale_exist:
        messages.error(request, "Can't delete. Sale has been created for this delivery challan.")
        return redirect('delivery-challan-list')

    try:
        challan = get_object_or_404(DeliveryChallanMaster, pk=pk)
        logger.info(f'delete_delivery_challan object {challan.delivery_challan_no}')

        selected_godown = challan.selected_godown

        related_products = challan.deliverychallanproducts_set.all()
        for product in related_products:
            obj, created = product_godown_quantity_through_table.objects.get_or_create(godown_name = selected_godown,product_color_name=product.product_name)
            obj.quantity += product.quantity
            obj.save()

        
        challan.delete()

        messages.success(request, "Delivery Challan deleted successfully.")
        logger.info(f"delete_delivery_challan object {challan.delivery_challan_no} deleted successfully")

    except Exception as e:
        logger.error(f'Error deleting Delivery Challan {pk}: {e}')
        messages.error(request, f"Error deleting Delivery Challan: {e}")

    return redirect('delivery-challan-list')


def sales_scan_product_dynamic_ajax(request):
    
    try:
        
        serialNo = request.GET.get('serialNo')
        warhouseId = request.GET.get('warhouseId')

        if not serialNo:
            return JsonResponse({'error': 'Please enter a search term.'}, status=400)

        if sales_voucher_finish_Goods.objects.filter(unique_serial_no=serialNo).exists():
            
            return JsonResponse(
                {'error': f'The serial number "{serialNo}" already exists. Please enter a different one.'},
                status=400
            )
        
        filtered_product = list(finishedgoodsbinallocation.objects.filter(unique_serial_no = serialNo).values('product__Product__Model_Name','product__PProduct_color__color_name','product__PProduct_SKU','unique_serial_no','product__Product__Product_MRP','product__Product__Product_SalePrice_CustomerPrice','product__Product__Product_GST__gst_percentage'))


        if filtered_product:
            list_to_send = []

            for query in filtered_product:
                p_sku = query.get('product__PProduct_SKU')
                serial_no = query.get('unique_serial_no')
                product_model_name = query.get('product__Product__Model_Name')
                color = query.get('product__PProduct_color__color_name')
                gst = query.get('product__Product__Product_GST__gst_percentage')
                mrp = query.get('product__Product__Product_MRP')
                customer_price = query.get('product__Product__Product_SalePrice_CustomerPrice')


                list_to_send.append([p_sku,product_model_name,color,serial_no,gst,mrp,customer_price])
            
            return JsonResponse({'products': list_to_send,'message':f"{serialNo} added successfully"}, status=200)
        
        return JsonResponse({'error': 'No items found.'}, status=404)
    
    except Exception as e:
        return JsonResponse({'error': f"An error occurred: {str(e)}"}, status=500)



def salesvoucherlist(request):
    sales_list = sales_voucher_master_finish_Goods.objects.all().order_by('created_date')
    return render(request,'accounts/sales_list.html',{'sales_list':sales_list,'page_name':'Sales list'})





def salesvoucherdelete(request, pk):
    try:
        with transaction.atomic():
            sales_instance = get_object_or_404(sales_voucher_master_finish_Goods, pk=pk)

            transfer_records = sales_voucher_finish_Goods.objects.filter(sales_voucher_master=pk)

            for record in transfer_records:

                challan = record.challan.delivery_challan
                product_name = record.product_name
                product_quantity = record.quantity

                if challan:
                    challan_qty_value = DeliveryChallanProducts.objects.filter(
                        delivery_challan=challan, product_name=product_name
                    ).first()
                    if challan_qty_value:
                        challan_qty_value.balance_qty += product_quantity
                        challan_qty_value.save()

            sales_instance.delete()

            SalesVoucherDeliveryChallan.objects.filter(sales_voucher=pk).delete()

    except Exception as e:
        print(f"Error during deletion: {e}")
        messages.error(request, "Error deleting the sales voucher.")
        

    return redirect('sales-voucher-list')